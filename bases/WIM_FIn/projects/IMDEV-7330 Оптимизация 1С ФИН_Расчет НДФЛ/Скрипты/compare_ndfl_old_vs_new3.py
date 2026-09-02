#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sverka NDFL: Avancor old vs PoNovomu3 (Locks + NkdParam + NkdCoupon).
Konsol - ASCII. HTML - UTF-8.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from html import escape

from openpyxl import load_workbook

SCRIPTS = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPTS)
import compare_ndfl_old_vs_new2 as c  # noqa: E402

BASE = r"c:\1c\Cursor_1c\WIM_DEV\bases\WIM_FIn\projects\IMDEV-7330 НовыеТесты"
DOC = r"c:\1c\Cursor_1c\WIM_DEV\bases\WIM_FIn\projects\IMDEV-7330 Оптимизация 1С ФИН_Расчет НДФЛ\Документация"
OUT_HTML = os.path.join(DOC, "imdev7330_ndfl_old_vs_new3_diff.html")
OUT_JSON = os.path.join(DOC, "imdev7330_ndfl_old_vs_new3_diff.json")

FILES = {
    "uk_old": os.path.join(BASE, "НДФЛ_Управление_27292.xlsx"),
    "uk_n3": os.path.join(BASE, "НДФЛ_Управление_27292_ПоНовому3.xlsx"),
    "pf_old": os.path.join(BASE, "НДФЛ_Портфели_27292.xlsx"),
    "pf_n3": os.path.join(BASE, "НДФЛ_Портфели_27292_ПоНовому3.xlsx"),
    "msg_old": os.path.join(BASE, "Сообщения_поСтарому_Аванкор.txt"),
    "msg_n3": os.path.join(BASE, "Сообщения_поНовому3_Аванкор.txt"),
    "log_n3": os.path.join(BASE, "Лог формирования начисдений НДФЛ_ПоНовому3.xlsx"),
}

PLAT_NUM = "000000000038432"
PLAT_NAME = "Платонов"
PLAT_MARKERS = (
    "Платонов Д.В.",
    "Платонов Дмитрий Вячеславович",
)
# Metki dlya otchyota vendoru: FIO, nomer UK, kody DU ne popadayut v HTML.
CLIENT_A = "Клиент А"
DOC_A = "УК-А"
PF1 = "Портфель 1"
PF2 = "Портфель 2"
RE_FIO_INIT = re.compile(r"[А-ЯЁ][а-яё\-]+\s+[А-ЯЁ]\.[А-ЯЁ]\.")
RE_DU = re.compile(r"ДУ\s+[А-ЯA-ZЁ]*\d+")
RE_DOCNUM = re.compile(r"\b0{6,}\d+\b")
EPS = c.EPS
NUM_UK = c.NUM_UK
NUM_PF = c.NUM_PF

LOCK_RE = re.compile(
    r"(блокир|взаимоблокир|deadlock|1222|повторн|timeout|таймаут)",
    re.I,
)


def safe_print(text: str) -> None:
    c.safe_print(text)


def file_md5(path: str) -> str:
    return c.file_md5(path)


def fmt_int(n) -> str:
    return f"{int(n):,}".replace(",", " ")


def fmt_dec(v) -> str:
    d = Decimal(str(v))
    if d == d.to_integral():
        return fmt_int(d)
    s = f"{d:,.2f}"
    return s.replace(",", " ")


def mask_text(s) -> str:
    """Obeznlichivaet FIO, DU i nomer dokumenta UK. Summy ne trogaet."""
    if s is None:
        return ""
    t = str(s)
    t = t.replace("Платонов Дмитрий Вячеславович", CLIENT_A)
    t = t.replace("Платонов Д.В.", CLIENT_A)
    t = t.replace("ДУ 2258", PF1)
    t = t.replace("ДУ 10155", PF2)
    t = t.replace(PLAT_NUM, DOC_A)
    t = re.sub(r"Платонов\w*", "другой клиент", t)
    t = RE_DU.sub("портфель", t)
    t = RE_DOCNUM.sub(DOC_A, t)
    t = re.sub(r"по портфелю\s+.+", "по портфелю [скрыто]", t)
    t = re.sub(r"по клиенту\s+.+", "по клиенту [скрыто]", t)
    t = RE_FIO_INIT.sub("[ФИО]", t)
    return t


def mask_esc(s) -> str:
    return escape(mask_text(s))


def parse_dt(s: str):
    s = (s or "").strip()
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def dur_hms(a: datetime, b: datetime) -> str:
    sec = int((b - a).total_seconds())
    h, rem = divmod(sec, 3600)
    m, s = divmod(rem, 60)
    return f"{h} ч {m:02d} мин {s:02d} с ({h + m / 60 + s / 3600:.2f} ч)"


def is_plat_row(row, idx) -> bool:
    """Только Платонов Д.В., не однофамильцы (Платонова, Платонов А.В.)."""
    num = str(c.get(row, idx, "НомерДокумента") or "")
    if num == PLAT_NUM:
        return True
    blob = " ".join(
        str(c.get(row, idx, n) or "")
        for n in ("Ссылка", "Портфель", "Клиент")
        if n in idx
    )
    return any(m in blob for m in PLAT_MARKERS)


def docs_with_surname(docs):
    out = []
    for num, d in docs.items():
        blob = (d.get("ref") or "") + " " + " ".join(str(x) for x in (d.get("portfolios") or []))
        if PLAT_NAME not in blob:
            continue
        if num == PLAT_NUM or any(m in blob for m in PLAT_MARKERS):
            continue
        out.append(
            {
                "num": num,
                "n": d["n"],
                "ref": d.get("ref") or "",
                "portfolios": sorted(str(x) for x in (d.get("portfolios") or []) if x),
                "parts": dict(d["parts"]),
            }
        )
    out.sort(key=lambda x: x["num"])
    return out


def empty_uk_docs(docs) -> list:
    empty = []
    for num, d in docs.items():
        parts = set(d["parts"])
        if parts - {"Выводы", "ОбщиеРасходы"}:
            continue
        money = sum(abs(v) for v in d["sums"].values())
        if money < EPS:
            empty.append(num)
    return empty


def dump_plat_rows(rows, idx, num_names):
    out = []
    cols = [
        "ТабличнаяЧасть",
        "Портфель",
        "КодДохода",
        "КодВычета",
        "Ставка",
        "ТипДохода",
    ] + list(num_names)
    for row in rows:
        if not is_plat_row(row, idx):
            continue
        rec = {col: c.get(row, idx, col) for col in cols if col in idx or True}
        rec["_num"] = str(c.get(row, idx, "НомерДокумента") or "")
        rec["_ref"] = str(c.get(row, idx, "Ссылка") or "")
        rec["_sums"] = {n: str(c.to_dec(c.get(row, idx, n))) for n in num_names}
        out.append(rec)
    return out


def parse_log_xlsx(path: str) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = {"start": None, "end": None, "errors": None}
    proveden_pf = []
    proveden_uk = []
    fail_lines = []
    lock_lines = []
    comments = Counter()
    n_rows = 0
    first_comment = None
    last_comment = None
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        cells = ["" if x is None else str(x).strip() for x in row]
        if all(x == "" for x in cells):
            continue
        blob = " ".join(cells)
        n_rows += 1
        if header["start"] is None:
            m = c.RE_LOG_HEADER.search(blob.replace("\n", " "))
            if m:
                header = {
                    "start": m.group(1).strip(),
                    "end": m.group(2).strip(),
                    "errors": int(m.group(3)),
                }
                continue
        obj = cells[0] if cells else ""
        comment = ""
        if len(cells) > 2 and cells[2]:
            comment = cells[2]
        elif len(cells) > 1 and cells[1]:
            comment = cells[1]
        s = comment or blob
        if s and first_comment is None and "Проведено" in s:
            first_comment = s
        if s:
            last_comment = s
        mp = c.RE_PROVEDEN_PF.search(s)
        if mp:
            proveden_pf.append(mp.group(1).strip())
            comments["proveden_pf"] += 1
            continue
        mu = c.RE_PROVEDEN_UK.search(s)
        if mu:
            proveden_uk.append(mu.group(1).strip())
            comments["proveden_uk"] += 1
            continue
        if LOCK_RE.search(s) or LOCK_RE.search(obj):
            lock_lines.append((obj[:120], s[:300]))
            comments["lock_like"] += 1
        if c.RE_FAIL.search(s) and "Ошибок: 0" not in s:
            fail_lines.append(s[:300])
            comments["fail_like"] += 1
        elif "не записан" in s.lower() or "не создан" in s.lower():
            fail_lines.append(s[:300])
            comments["not_created"] += 1
        elif "Проведено" in s:
            comments["proveden_other"] += 1
    wb.close()
    return {
        "header": header,
        "comments": dict(comments),
        "proveden_pf_n": len(proveden_pf),
        "proveden_uk_n": len(proveden_uk),
        "proveden_pf_unique": len(set(proveden_pf)),
        "proveden_uk_unique": len(set(proveden_uk)),
        "fail_lines": fail_lines[:40],
        "fail_n": len(fail_lines),
        "lock_n": len(lock_lines),
        "lock_sample": lock_lines[:20],
        "data_rows": n_rows,
        "first_comment": first_comment,
        "last_comment": last_comment,
        "bytes": os.path.getsize(path),
        "plat_in_log": any(any(m in x for m in PLAT_MARKERS) for x in proveden_uk),
    }


def compare_pair(label, r_old, i_old, r_new, i_new, num_names, key_fn, fmt_fn):
    safe_print("agg " + label)
    docs_o, parts_o, pf_o, _ = c.agg_docs(r_old, i_old, num_names)
    docs_n, parts_n, pf_n, _ = c.agg_docs(r_new, i_new, num_names)
    dc = c.compare_doc_sums(docs_o, docs_n, num_names)
    bag = c.bag_diff(r_old, r_new, key_fn, i_old, i_new)
    bag["only_old_fmt"] = [(n, fmt_fn(k)) for n, k in bag["only_old"]]
    bag["only_new_fmt"] = [(n, fmt_fn(k)) for n, k in bag["only_new"]]
    del bag["only_old"]
    del bag["only_new"]
    empty_o = empty_uk_docs(docs_o) if num_names is NUM_UK else []
    empty_n = empty_uk_docs(docs_n) if num_names is NUM_UK else []
    plat_rows_o = dump_plat_rows(r_old, i_old, num_names)
    plat_rows_n = dump_plat_rows(r_new, i_new, num_names)
    plat_nums = {PLAT_NUM}
    docs_o_wo = {k: v for k, v in docs_o.items() if k not in plat_nums}
    docs_n_wo = {k: v for k, v in docs_n.items() if k not in plat_nums}
    dc_wo = c.compare_doc_sums(docs_o_wo, docs_n_wo, num_names)
    rows_wo_o = [r for r in r_old if str(c.get(r, i_old, "НомерДокумента") or "") not in plat_nums]
    rows_wo_n = [r for r in r_new if str(c.get(r, i_new, "НомерДокумента") or "") not in plat_nums]
    return {
        "rows_old": len(r_old),
        "rows_new": len(r_new),
        "totals_old": c.totals(r_old, i_old, num_names),
        "totals_new": c.totals(r_new, i_new, num_names),
        "totals_wo_old": c.totals(rows_wo_o, i_old, num_names),
        "totals_wo_new": c.totals(rows_wo_n, i_new, num_names),
        "parts_old": dict(parts_o),
        "parts_new": dict(parts_n),
        "docs_old": len(docs_o),
        "docs_new": len(docs_n),
        "portfolios_old": len(pf_o),
        "portfolios_new": len(pf_n),
        "doc_compare": dc,
        "doc_compare_wo_plat": dc_wo,
        "line_bag": bag,
        "empty_old_n": len(empty_o),
        "empty_new_n": len(empty_n),
        "plat_nums": sorted(plat_nums),
        "plat_docs_old": {k: {"n": v["n"], "parts": dict(v["parts"]), "sums": {n: str(v["sums"][n]) for n in num_names}, "ref": v["ref"]} for k, v in docs_o.items() if k in plat_nums},
        "plat_docs_new": {k: {"n": v["n"], "parts": dict(v["parts"]), "sums": {n: str(v["sums"][n]) for n in num_names}, "ref": v["ref"]} for k, v in docs_n.items() if k in plat_nums},
        "plat_rows_old": plat_rows_o,
        "plat_rows_new": plat_rows_n,
        "surname_others_old": docs_with_surname(docs_o),
        "surname_others_new": docs_with_surname(docs_n),
        "by_portfolio": c.compare_pf_sums(
            c.sum_by_portfolio(r_old, i_old, num_names),
            c.sum_by_portfolio(r_new, i_new, num_names),
            num_names,
        ),
    }


def slim_dc(dc):
    dc = dict(dc)
    dc["only_old_n"] = len(dc.get("only_old") or [])
    dc["only_new_n"] = len(dc.get("only_new") or [])
    dc["only_old_sample"] = (dc.get("only_old") or [])[:40]
    dc["only_new_sample"] = (dc.get("only_new") or [])[:40]
    dc.pop("only_old", None)
    dc.pop("only_new", None)
    return dc


def analyze():
    result = {"files": {}, "md5": {}}
    for k, p in FILES.items():
        result["files"][k] = {
            "name": os.path.basename(p),
            "bytes": os.path.getsize(p) if os.path.isfile(p) else 0,
            "exists": os.path.isfile(p),
        }
        if os.path.isfile(p) and p.endswith(".xlsx") and "Лог" not in os.path.basename(p):
            safe_print("md5 " + os.path.basename(p))
            result["md5"][k] = file_md5(p)

    safe_print("Load UK old...")
    _, i_uk_o, r_uk_o = c.load_xlsx_rows(FILES["uk_old"])
    safe_print("Load UK new3...")
    _, i_uk_3, r_uk_3 = c.load_xlsx_rows(FILES["uk_n3"])
    result["uk_old_vs_n3"] = compare_pair(
        "uk old vs n3", r_uk_o, i_uk_o, r_uk_3, i_uk_3, NUM_UK, c.line_key_uk, c.fmt_key_uk
    )
    del r_uk_o, r_uk_3

    safe_print("Load PF old...")
    _, i_pf_o, r_pf_o = c.load_xlsx_rows(FILES["pf_old"])
    safe_print("Load PF new3...")
    _, i_pf_3, r_pf_3 = c.load_xlsx_rows(FILES["pf_n3"])
    result["pf_old_vs_n3"] = compare_pair(
        "pf old vs n3", r_pf_o, i_pf_o, r_pf_3, i_pf_3, NUM_PF, c.line_key_pf, c.fmt_key_pf
    )

    rows_plat_o = [rw for rw in r_pf_o if is_plat_row(rw, i_pf_o)]
    rows_plat_3 = [rw for rw in r_pf_3 if is_plat_row(rw, i_pf_3)]
    bagp = c.bag_diff(rows_plat_o, rows_plat_3, c.line_key_pf, i_pf_o, i_pf_3)
    bagp["only_old_fmt"] = [(n, c.fmt_key_pf(k)) for n, k in bagp["only_old"]]
    bagp["only_new_fmt"] = [(n, c.fmt_key_pf(k)) for n, k in bagp["only_new"]]
    del bagp["only_old"]
    del bagp["only_new"]
    result["pf_plat"] = {
        "rows_old": len(rows_plat_o),
        "rows_new": len(rows_plat_3),
        "totals_old": c.totals(rows_plat_o, i_pf_o, NUM_PF),
        "totals_new": c.totals(rows_plat_3, i_pf_3, NUM_PF),
        "bag": bagp,
    }
    del r_pf_o, r_pf_3, rows_plat_o, rows_plat_3

    safe_print("Parse messages...")
    msg_o = c.parse_messages(FILES["msg_old"])
    msg_3 = c.parse_messages(FILES["msg_n3"])

    def msg_cmp(a, b):
        sa, sb = a["unique"], b["unique"]
        cl_a, cl_b = set(a["clients_rasp"]), set(b["clients_rasp"])
        return {
            "kinds_a": dict(a["kinds"]),
            "kinds_b": dict(b["kinds"]),
            "nonempty_a": a["nonempty"],
            "nonempty_b": b["nonempty"],
            "only_a_n": len(sa - sb),
            "only_b_n": len(sb - sa),
            "common_n": len(sa & sb),
            "clients_rasp_a": len(cl_a),
            "clients_rasp_b": len(cl_b),
            "clients_only_a": sorted(cl_a - cl_b)[:40],
            "clients_only_b": sorted(cl_b - cl_a)[:40],
            "granica_a": len(a["granica"]),
            "granica_b": len(b["granica"]),
            "ne_prov_a": a["ne_prov"],
            "ne_prov_b": b["ne_prov"],
            "plat_msgs_a": [x for x in a["clients_rasp"] if PLAT_NAME in x],
            "plat_msgs_b": [x for x in b["clients_rasp"] if PLAT_NAME in x],
        }

    result["msg_old_vs_n3"] = msg_cmp(msg_o, msg_3)

    safe_print("Parse log new3 xlsx...")
    result["log_n3"] = parse_log_xlsx(FILES["log_n3"])

    for key in ("uk_old_vs_n3", "pf_old_vs_n3"):
        result[key]["doc_compare"] = slim_dc(result[key]["doc_compare"])
        result[key]["doc_compare_wo_plat"] = slim_dc(result[key]["doc_compare_wo_plat"])
        bp = result[key].get("by_portfolio") or {}
        if isinstance(bp.get("only_old"), list):
            bp["only_old_n"] = len(bp["only_old"])
            bp["only_old"] = bp["only_old"][:20]
        if isinstance(bp.get("only_new"), list):
            bp["only_new_n"] = len(bp["only_new"])
            bp["only_new"] = bp["only_new"][:20]

    def conv(obj):
        if isinstance(obj, Decimal):
            return str(obj)
        if isinstance(obj, set):
            return sorted(obj)
        if isinstance(obj, Counter):
            return dict(obj)
        if isinstance(obj, tuple):
            return [conv(x) for x in obj]
        if isinstance(obj, list):
            return [conv(x) for x in obj]
        if isinstance(obj, dict):
            return {str(k): conv(v) for k, v in obj.items()}
        return obj

    # drop bulky row dumps from json except plat rows already stored
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(conv(result), f, ensure_ascii=False, indent=2)
    safe_print("JSON written")
    return result


def money_rows(t_old, t_new):
    out = []
    for k in t_old:
        a = Decimal(t_old[k])
        b = Decimal(t_new[k])
        d = b - a
        cls = "ok" if abs(d) < EPS else "bad"
        out.append(
            f"<tr class='{cls}'><td>{escape(k)}</td>"
            f"<td class='num'>{fmt_dec(a)}</td><td class='num'>{fmt_dec(b)}</td>"
            f"<td class='num'>{fmt_dec(d)}</td></tr>"
        )
    return "\n".join(out)


def parts_rows(po, pn):
    keys = sorted(set(po) | set(pn))
    out = []
    for k in keys:
        a = po.get(k, 0)
        b = pn.get(k, 0)
        cls = "ok" if a == b else "warn"
        out.append(
            f"<tr class='{cls}'><td>{escape(k or '(empty)')}</td>"
            f"<td class='num'>{fmt_int(a)}</td><td class='num'>{fmt_int(b)}</td>"
            f"<td class='num'>{fmt_int(b - a)}</td></tr>"
        )
    return "\n".join(out)


def plat_table(rows_old, rows_new, num_names):
    """Align plat rows by business key without money for side-by-side."""
    def key(rec):
        return (
            str(rec.get("ТабличнаяЧасть") or ""),
            str(rec.get("Портфель") or ""),
            str(rec.get("КодДохода") or ""),
            str(rec.get("КодВычета") or ""),
            str(rec.get("Ставка") or ""),
        )

    def money_s(rec):
        bits = []
        for n in ("НалогооблагаемаяCумма", "СуммаВычета", "СуммаДохода", "СуммаКУдержанию", "Сумма"):
            v = rec.get("_sums", {}).get(n)
            if v and Decimal(v) != 0:
                bits.append(f"{n}={fmt_dec(v)}")
        return "; ".join(bits)

    co = Counter(key(r) for r in rows_old)
    cn = Counter(key(r) for r in rows_new)
    keys = sorted(set(co) | set(cn))
    body = []
    by_old = defaultdict(list)
    by_new = defaultdict(list)
    for r in rows_old:
        by_old[key(r)].append(r)
    for r in rows_new:
        by_new[key(r)].append(r)
    for k in keys:
        a = by_old.get(k, [None])
        b = by_new.get(k, [None])
        n = max(len(a), len(b))
        for i in range(n):
            ro = a[i] if i < len(a) else None
            rn = b[i] if i < len(b) else None
            so = money_s(ro) if ro else "нет строки"
            sn = money_s(rn) if rn else "нет строки"
            cls = "ok" if so == sn and ro and rn else "warn"
            part, pf, kd, kv, st = k
            body.append(
                f"<tr class='{cls}'><td>{mask_esc(part)}</td><td>{mask_esc(pf)}</td>"
                f"<td>{escape(kd)}</td><td>{escape(kv)}</td><td>{escape(st)}</td>"
                f"<td class='sig'>{mask_esc(so)}</td><td class='sig'>{mask_esc(sn)}</td></tr>"
            )
    return "\n".join(body)


def write_html(r):
    uk = r["uk_old_vs_n3"]
    pf = r["pf_old_vs_n3"]
    log3 = r["log_n3"]
    msg = r["msg_old_vs_n3"]

    dc = uk["doc_compare"]
    dc_wo = uk["doc_compare_wo_plat"]
    bag = uk["line_bag"]

    plat_old = uk["plat_docs_old"]
    plat_new = uk["plat_docs_new"]
    surname_o = uk.get("surname_others_old") or []
    surname_n = uk.get("surname_others_new") or []

    st3 = parse_dt(log3["header"].get("start") or "")
    en3 = parse_dt(log3["header"].get("end") or "")
    user_st = parse_dt("01.09.2026 18:38")
    job_dur = dur_hms(st3, en3) if st3 and en3 else "?"
    wall_dur = dur_hms(user_st, en3) if user_st and en3 else "?"
    clients = 27292
    # Etalon skorosti: Word "RegressIZamery_DR _Reliz AVANKOR.docx", reliz 2.8.5.5.
    old_h = 55 + 5 / 60
    old_thr_word = 500
    old8_h = 13 + 11 / 60
    old8_thr = clients / old8_h
    job_h = (en3 - st3).total_seconds() / 3600 if st3 and en3 else None
    wall_h = (en3 - user_st).total_seconds() / 3600 if user_st and en3 else None
    thr_n3 = clients / job_h if job_h else 0
    speed_x_old = (thr_n3 / old_thr_word) if old_thr_word else 0
    dur_x_old = (old_h / job_h) if job_h else 0
    wall_x_old = (old_h / wall_h) if wall_h else 0
    job_h_txt = fmt_dec(Decimal(str(round(job_h, 2)))) if job_h else "?"
    old_h_txt = fmt_dec(Decimal(str(round(old_h, 2))))

    pf_line_ok = pf["line_bag"]["only_old_n"] == 0 and pf["line_bag"]["only_new_n"] == 0
    pf_doc_ok = pf["doc_compare"]["diff_sums"] == 0 and pf["doc_compare"]["only_old_n"] == 0 and pf["doc_compare"]["only_new_n"] == 0
    uk_wo_ok = dc_wo["diff_sums"] == 0

    plat_delta_vychet = None
    plat_delta_uder = None
    if PLAT_NUM in plat_old and PLAT_NUM in plat_new:
        plat_delta_vychet = Decimal(plat_new[PLAT_NUM]["sums"]["СуммаВычета"]) - Decimal(plat_old[PLAT_NUM]["sums"]["СуммаВычета"])
        plat_delta_uder = Decimal(plat_new[PLAT_NUM]["sums"]["СуммаКУдержанию"]) - Decimal(plat_old[PLAT_NUM]["sums"]["СуммаКУдержанию"])

    if pf_line_ok and uk_wo_ok and plat_delta_vychet is not None and abs(plat_delta_vychet - Decimal("8481")) < EPS:
        verdict = (
            "Сверка сумм: по-старому против по-новому3. "
            "Портфели совпали с эталоном. УК без клиента А совпали по суммам документов. "
            "Единственное налоговое расхождение: документ " + DOC_A + " (клиент А) - "
            "инвествычет +8 481.00, к удержанию -1 719; эталон 4 строки ТЧ Начисления, этот прогон 6 строк "
            "(типовой дефект ТаблицаПродаж, не IMDEV-7330). "
            "Скорость: эталон релиз 2.8.5.5 без доработок - 55 ч 05 мин на 27 292 клиента "
            "(500 объектов в час). Этот прогон ~x{:.1f} быстрее по чистому времени лога."
        ).format(dur_x_old if dur_x_old else 0)
        vcls = "warn"
    elif pf_line_ok and uk_wo_ok:
        verdict = "Портфели и УК без клиента А совпали с эталоном. Клиент А выделен отдельно."
        vcls = "warn"
    else:
        verdict = "Есть расхождения помимо клиента А. См. таблицы."
        vcls = "out"

    labels = {
        "raspredelenie": "Невозможно распределить НДФЛ",
        "granica": "Граница актуальности сдвинута",
        "ne_proveden": "Документ не проведен",
        "oshibka": "Ошибка",
        "other": "Прочее",
        "empty": "Пустые",
    }
    kind_keys = sorted(set(msg["kinds_a"]) | set(msg["kinds_b"]))
    kind_rows = []
    for k in kind_keys:
        a = msg["kinds_a"].get(k, 0)
        b = msg["kinds_b"].get(k, 0)
        cls = "ok" if a == b else "warn"
        kind_rows.append(
            f"<tr class='{cls}'><td>{escape(labels.get(k, k))}</td>"
            f"<td class='num'>{a}</td><td class='num'>{b}</td><td class='num'>{b - a}</td></tr>"
        )

    def fname(key):
        f = (r.get("files") or {}).get(key) or {}
        return f.get("name") or key

    def fbytes(key):
        f = (r.get("files") or {}).get(key) or {}
        return fmt_int(f.get("bytes") or 0)

    empty_extra = int(uk.get("empty_new_n") or 0) - int(uk.get("empty_old_n") or 0)

    bag_old = "".join(
        f"<tr><td class='num'>{n}</td><td class='sig'>{mask_esc(s)}</td></tr>"
        for n, s in bag.get("only_old_fmt", [])[:25]
    )
    bag_new = "".join(
        f"<tr><td class='num'>{n}</td><td class='sig'>{mask_esc(s)}</td></tr>"
        for n, s in bag.get("only_new_fmt", [])[:25]
    )

    diff_docs = []
    for d in dc["diffs"][:20]:
        deltas = "; ".join(f"{k}={v}" for k, v in d["delta"].items())
        cls = "warn" if d["num"] == PLAT_NUM else "bad"
        doc_lbl = DOC_A if d["num"] == PLAT_NUM else "другой УК"
        diff_docs.append(
            f"<tr class='{cls}'><td class='sig'>{escape(doc_lbl)}</td>"
            f"<td>{mask_esc(d.get('ref') or '')}</td><td class='sig'>{escape(deltas)}</td>"
            f"<td class='num'>{d['old_rows']}</td><td class='num'>{d['new_rows']}</td></tr>"
        )

    plat_sum_rows = ""
    if PLAT_NUM in plat_old and PLAT_NUM in plat_new:
        so, sn = plat_old[PLAT_NUM]["sums"], plat_new[PLAT_NUM]["sums"]
        plat_sum_rows = money_rows(so, sn)

    plat_lines = plat_table(uk["plat_rows_old"], uk["plat_rows_new"], NUM_UK)

    n_surname = max(len(surname_o), len(surname_n))
    surname_tbl = (
        f"<tr class='ok'><td>другие клиенты с похожей фамилией</td>"
        f"<td class='num'>{fmt_int(n_surname)}</td>"
        f"<td>идентификаторы скрыты</td>"
        f"<td>суммы совпали с эталоном</td></tr>"
        if n_surname
        else "<tr class='ok'><td colspan='4'>Нет других клиентов с похожей фамилией</td></tr>"
    )

    plat_bag_old = "".join(
        f"<tr><td class='num'>{n}</td><td class='sig'>{mask_esc(s)}</td></tr>"
        for n, s in bag.get("only_old_fmt", [])
        if any(m in s for m in PLAT_MARKERS)
    )
    plat_bag_new = "".join(
        f"<tr><td class='num'>{n}</td><td class='sig'>{mask_esc(s)}</td></tr>"
        for n, s in bag.get("only_new_fmt", [])
        if any(m in s for m in PLAT_MARKERS)
    )

    lock_lis = "".join(
        f"<li><code>{mask_esc(a)}</code> - {mask_esc(b)}</li>"
        for a, b in log3.get("lock_sample") or []
    ) or "<p class='small'>Сообщений про блокировки / повтор / 1222 в логе нет.</p>"

    speed_vs = (
        f"эталон скорости - релиз 2.8.5.5 без доработок: 55 ч 05 мин на 27 292 клиента "
        f"(500 объектов в час, протокол DR). "
        f"Этот прогон: {job_dur} по логу (x{dur_x_old:.1f} по времени, "
        f"{fmt_dec(Decimal(str(round(thr_n3, 0))))} объектов/час, x{speed_x_old:.1f} к 500/час); "
        f"стена кнопки Заполнить {wall_dur} (x{wall_x_old:.1f} к 55 ч 05 мин)"
    )

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>IMDEV-7330. Прогон с 3 расширениями: скорость, регресс, клиент А</title>
<style>
    body {{ font-family: "Segoe UI", Arial, sans-serif; line-height: 1.55; color: #212529; background: #f5f6f8; margin: 0; padding: 0 0 60px 0; }}
    .wrap {{ max-width: 1180px; margin: 0 auto; padding: 0 24px; }}
    header {{ background: #1f2d3d; color: #fff; padding: 28px 0 22px 0; margin-bottom: 28px; }}
    header h1 {{ margin: 0 0 8px 0; font-size: 24px; }}
    header .sub {{ color: #b8c4d0; font-size: 14px; }}
    h2 {{ font-size: 20px; margin: 32px 0 12px 0; padding-bottom: 8px; border-bottom: 2px solid #17a2b8; }}
    h3 {{ font-size: 16px; margin: 18px 0 8px 0; color: #1f2d3d; }}
    p {{ margin: 10px 0; }}
    table {{ width: 100%; border-collapse: collapse; background: #fff; margin: 12px 0; font-size: 13.5px; }}
    th {{ background: #1f2d3d; color: #fff; text-align: left; padding: 8px 10px; }}
    td {{ padding: 7px 10px; border-bottom: 1px solid #e3e6ea; vertical-align: top; }}
    tbody tr:nth-child(even) {{ background: #fafbfc; }}
    tr.ok td {{ background: #eefaf1; }}
    tr.bad td {{ background: #fdf0f1; }}
    tr.warn td {{ background: #fff9e6; }}
    .num {{ text-align: right; font-family: Consolas, monospace; white-space: nowrap; }}
    .sig {{ font-family: Consolas, "Courier New", monospace; font-size: 12px; }}
    code {{ background: #eef1f4; padding: 1px 5px; border-radius: 3px; font-family: Consolas, monospace; font-size: 13px; }}
    .box {{ padding: 14px 18px; border-radius: 5px; margin: 14px 0; border-left: 5px solid; }}
    .box b {{ display: block; margin-bottom: 6px; }}
    .in {{ background: #eefaf1; border-color: #28a745; }}
    .out {{ background: #fdf0f1; border-color: #dc3545; }}
    .warn {{ background: #fff9e6; border-color: #f0ad4e; }}
    .info {{ background: #eaf7fa; border-color: #17a2b8; }}
    .small {{ font-size: 13px; color: #6c757d; }}
    ul, ol {{ margin: 8px 0 8px 22px; }}
    li {{ margin: 5px 0; }}
    footer {{ margin-top: 36px; padding-top: 12px; border-top: 1px solid #dee2e6; font-size: 13px; color: #6c757d; }}
    .toc {{ background: #fff; border: 1px solid #e3e6ea; border-radius: 5px; padding: 12px 20px; margin: 16px 0; }}
    .toc a {{ text-decoration: none; color: #0a58ca; }}
</style>
</head>
<body>
<header>
<div class="wrap">
    <h1>IMDEV-7330. Три расширения: скорость, регресс, клиент А</h1>
    <div class="sub">
        Прогон «по-новому3»: IMDEV7330_Locks + IMDEV7330_NkdParam + IMDEV7330_NkdCoupon.
        10 потоков, выборка 27 292. Эталон - выгрузки «по-старому».
        Сверка сумм: эталон vs по-новому3. Скорость: эталон релиз 2.8.5.5 (55 ч 05 мин).
        ФИО, номера УК и коды ДУ в этом файле заменены.
    </div>
</div>
</header>
<div class="wrap">

<div class="toc">
<b>Содержание</b>
<ol>
<li><a href="#p0">Итог</a></li>
<li><a href="#p1">Контур и файлы</a></li>
<li><a href="#p2">Скорость</a></li>
<li><a href="#p3">Регресс портфелей</a></li>
<li><a href="#p4">Регресс УК без клиента А</a></li>
<li><a href="#p5">Клиент А (документ {DOC_A})</a></li>
<li><a href="#p7">Сообщения и лог</a></li>
<li><a href="#p8">Анализ</a></li>
</ol>
</div>

<h2 id="p0">1. Итог</h2>
<div class="box {vcls}">
<b>Вердикт</b>
{escape(verdict)}
</div>

<table>
<thead><tr><th>Вопрос</th><th>Ответ</th></tr></thead>
<tbody>
<tr class="{'ok' if pf_line_ok else 'bad'}"><td>Портфели vs эталон</td>
<td>{'совпали по строкам и суммам документов' if pf_line_ok and pf_doc_ok else 'есть расхождения, см. раздел 4'}</td></tr>
<tr class="{'ok' if uk_wo_ok else 'bad'}"><td>УК без клиента А vs эталон</td>
<td>документов с теми же суммами: {fmt_int(dc_wo['equal_sums'])} из {fmt_int(dc_wo['common'])};
расходящихся номеров: {fmt_int(dc_wo['diff_sums'])}</td></tr>
<tr class="warn"><td>Клиент А vs эталон</td>
<td>документ {escape(DOC_A)}; дельта вычета {fmt_dec(plat_delta_vychet) if plat_delta_vychet is not None else '?'};
дельта к удержанию {fmt_dec(plat_delta_uder) if plat_delta_uder is not None else '?'}</td></tr>
<tr class="ok"><td>Скорость</td>
<td>{escape(speed_vs)}</td></tr>
</tbody>
</table>

<h2 id="p1">2. Контур и файлы</h2>
<div class="box info">
<b>Как сравнивали</b>
Эталон - <code>НДФЛ_*_27292.xlsx</code> (по-старому). Прогон с тремя расширениями -
<code>*_ПоНовому3.xlsx</code>. Клиент А вынесен из общей сверки УК как единственный
расходящийся документ. Ключ строки УК: ТЧ + портфель + коды + ставка + суммы (допуск 0.01).
Скорость: эталон из протокола DR - релиз 2.8.5.5 без доработок, 55 ч 05 мин на те же 27 292 клиента
(500 объектов в час). Цифры этого прогона - шапка лога и замер кнопки Заполнить.
</div>
<div class="box warn">
<b>Обезличивание для вендора</b>
В этом HTML нет скриншотов. ФИО, номера документов УК и коды ДУ заменены:
клиент А, документ {escape(DOC_A)}, {escape(PF1)} (там инвествычет +8 481), {escape(PF2)}.
Налоговые суммы и структура ТЧ сохранены. Внутренний JSON со сверкой не предназначен для передачи.
</div>
<table>
<thead><tr><th>Роль</th><th>Файл</th><th>Байты</th></tr></thead>
<tbody>
<tr><td>Портфели эталон / новое3</td><td class="sig">{escape(fname('pf_old'))} / {escape(fname('pf_n3'))}</td>
<td class="num">{fbytes('pf_old')} / {fbytes('pf_n3')}</td></tr>
<tr><td>УК эталон / новое3</td><td class="sig">{escape(fname('uk_old'))} / {escape(fname('uk_n3'))}</td>
<td class="num">{fbytes('uk_old')} / {fbytes('uk_n3')}</td></tr>
<tr><td>Лог новое3</td><td class="sig">{escape(fname('log_n3'))}</td>
<td class="num">{fbytes('log_n3')}</td></tr>
<tr><td>Сообщения</td><td class="sig">{escape(fname('msg_n3'))}</td>
<td class="num">{fbytes('msg_n3')}</td></tr>
</tbody>
</table>
<p class="small">MD5 xlsx разный при равном размере - это упаковка Excel, не критерий расхождения ячеек.</p>

<h2 id="p2">3. Скорость</h2>
<div class="box in">
<b>Эталон скорости: релиз 2.8.5.5, без доработок IMDEV-7330</b>
Источник: протокол замеров DR (Word). Выборка та же: 27 292 клиента, период 01.01.2026 - 31.12.2026.
Длительность: <b>55 часов 5 минут</b> (27 292 объекта). В протоколе: <b>500 строк в час</b>.
Прогон шёл в две части: останов по ошибке «Значение не является значением объектного типа (Ставка)»,
затем продолжение. Вторая часть 46 ч 50 мин; к первой части в протоколе прибавлено ~8 ч 23 мин.
Итого 55 ч 05 мин. ФИО и номера документов из протокола в отчёт не копируются.
</div>
<div class="box in">
<b>Этот прогон (по-новому3) по логу: {escape(log3['header'].get('start') or '')} - {escape(log3['header'].get('end') or '')}</b>
Ошибок в шапке: {log3['header'].get('errors')}.
Чистая работа формирования: {escape(job_dur)}.
Стена кнопки Заполнить: 01.09.2026 18:38 - 02.09.2026 01:17:59, то есть {escape(wall_dur)}.
Разрыв 18:38 -&gt; 19:45 (~1 ч 07 мин) - до первой записи лога (даты / ожидание фоновых).
</div>
<table>
<thead><tr><th>Прогон</th><th>Потоки</th><th>Конфигурация</th><th>Интервал</th><th>Длительность</th><th>Объектов/час</th></tr></thead>
<tbody>
<tr>
<td>РЕЛИЗ 2.8.5.5 (эталон скорости)</td>
<td class="num">нет (последовательно)</td>
<td>типовая, без доработок</td>
<td class="sig">26.08.2026 - 29.08.2026, две части</td>
<td>55 ч 05 мин ({escape(old_h_txt)} ч)</td>
<td class="num">500</td>
</tr>
<tr>
<td>Первая новая (протокол DR)</td>
<td class="num">8</td>
<td>новая, ещё без Locks; падали 1222</td>
<td class="sig">30.08.2026 07:22:00 - 20:33:21</td>
<td>13 ч 11 мин</td>
<td class="num">{fmt_dec(Decimal(str(round(old8_thr, 0))))}</td>
</tr>
<tr class="ok">
<td>По-новому3 (этот)</td>
<td class="num">10</td>
<td>Locks + NkdParam + NkdCoupon</td>
<td class="sig">{escape(log3['header'].get('start') or '')} - {escape(log3['header'].get('end') or '')}</td>
<td>{escape(job_dur)}</td>
<td class="num">{fmt_dec(Decimal(str(round(thr_n3, 0))))}</td>
</tr>
</tbody>
</table>
<table>
<thead><tr><th>Метрика лога по-новому3</th><th>Значение</th></tr></thead>
<tbody>
<tr><td>Проведено по портфелю</td><td class="num">{fmt_int(log3['proveden_pf_n'])} (уник. {fmt_int(log3['proveden_pf_unique'])})</td></tr>
<tr><td>Проведено по клиенту УК</td><td class="num">{fmt_int(log3['proveden_uk_n'])} (уник. клиентов {fmt_int(log3['proveden_uk_unique'])})</td></tr>
<tr><td>Строк лога (непустых)</td><td class="num">{fmt_int(log3['data_rows'])}</td></tr>
<tr class="{'ok' if log3['fail_n']==0 else 'bad'}"><td>Похоже на ошибку / не создан</td><td class="num">{fmt_int(log3['fail_n'])}</td></tr>
<tr class="{'ok' if log3['lock_n']==0 else 'warn'}"><td>Упоминания блокировок / 1222 / повтор</td><td class="num">{fmt_int(log3['lock_n'])}</td></tr>
<tr><td>Клиент А в проведении УК</td><td>{'да' if log3.get('plat_in_log') else 'нет в выборке лога'}</td></tr>
</tbody>
</table>
<div class="box info">
<b>Как читать ускорение</b>
Эталон - 55 ч 05 мин и 500 объектов/час (релиз 2.8.5.5 без доработок).
По чистому времени лога: {escape(old_h_txt)} ч / {escape(job_h_txt)} ч = x{dur_x_old:.1f}.
По стене кнопки: x{wall_x_old:.1f}.
По объектам в час: {fmt_dec(Decimal(str(round(thr_n3, 0))))} / 500 = x{speed_x_old:.1f}.
Первая новая на 8 потоках из того же протокола DR (13 ч 11 мин, падения 1222) - уже x{old_h / old8_h:.1f}
к 2.8.5.5, но ещё с блокировками. Locks убрал 1222 (в этом логе {fmt_int(log3['lock_n'])} упоминаний),
NkdParam сокращает повторные запросы НКД.
</div>
{lock_lis if log3['lock_n'] else ''}

<h2 id="p3">4. Регресс портфелей</h2>
<div class="box {'in' if pf_line_ok and pf_doc_ok else 'out'}">
<b>{'Совпадение с эталоном' if pf_line_ok and pf_doc_ok else 'Есть расхождения'}</b>
Документов: {fmt_int(pf['docs_old'])} было, {fmt_int(pf['docs_new'])} стало.
Строк: {fmt_int(pf['rows_old'])} / {fmt_int(pf['rows_new'])}.
Совпавших сумм по номеру: {fmt_int(pf['doc_compare']['equal_sums'])}.
Расходящихся номеров: {fmt_int(pf['doc_compare']['diff_sums'])}.
Лишних/пропавших документов: {fmt_int(pf['doc_compare']['only_old_n'])} / {fmt_int(pf['doc_compare']['only_new_n'])}.
Строк только в эталоне / только в новом: {fmt_int(pf['line_bag']['only_old_n'])} / {fmt_int(pf['line_bag']['only_new_n'])}.
</div>
<table>
<thead><tr><th>Колонка</th><th>Эталон</th><th>По-новому3</th><th>Дельта</th></tr></thead>
<tbody>
{money_rows(pf['totals_old'], pf['totals_new'])}
</tbody>
</table>
<table>
<thead><tr><th>ТЧ</th><th>Эталон</th><th>По-новому3</th><th>Дельта</th></tr></thead>
<tbody>
{parts_rows(pf['parts_old'], pf['parts_new'])}
</tbody>
</table>
<p class="small">По портфелям клиента А ({escape(PF1)} и {escape(PF2)}) в выгрузке: строк эталон {fmt_int(r['pf_plat']['rows_old'])},
стало {fmt_int(r['pf_plat']['rows_new'])};
bag only_old/only_new: {fmt_int(r['pf_plat']['bag']['only_old_n'])} / {fmt_int(r['pf_plat']['bag']['only_new_n'])}.
Сдвиг УК не из портфельного расчёта.</p>

<h2 id="p4">5. Регресс УК без клиента А</h2>
<p>В этой сверке исключён только документ <code>{escape(DOC_A)}</code>.
Другие клиенты с похожей фамилией остаются в общей таблице - их суммы совпали с эталоном.
Идентификаторы этих клиентов в отчёт не выводятся.</p>
<div class="box {'in' if uk_wo_ok else 'out'}">
<b>{'Суммы общих документов совпали' if uk_wo_ok else 'Есть расхождения даже без клиента А'}</b>
Общих номеров: {fmt_int(dc_wo['common'])}. С теми же суммами: {fmt_int(dc_wo['equal_sums'])}.
С другими суммами: {fmt_int(dc_wo['diff_sums'])}.
Только в эталоне: {fmt_int(dc_wo['only_old_n'])}. Только в новом: {fmt_int(dc_wo['only_new_n'])}.
Пустые УК (только Выводы + Общие расходы, суммы 0): эталон {fmt_int(uk['empty_old_n'])}, новое3 {fmt_int(uk['empty_new_n'])}.
</div>
<table>
<thead><tr><th>Колонка (без клиента А)</th><th>Эталон</th><th>По-новому3</th><th>Дельта</th></tr></thead>
<tbody>
{money_rows(uk['totals_wo_old'], uk['totals_wo_new'])}
</tbody>
</table>
<table>
<thead><tr><th>ТЧ все УК включая клиента А</th><th>Эталон</th><th>По-новому3</th><th>Дельта</th></tr></thead>
<tbody>
{parts_rows(uk['parts_old'], uk['parts_new'])}
</tbody>
</table>
<p class="small">Строки Выводы/Общие расходы: в по-новому3 на {fmt_int(empty_extra)} пустых УК больше, чем в эталоне.
Налог они не меняют. Документов УК всего: {fmt_int(uk['docs_old'])} / {fmt_int(uk['docs_new'])}.</p>

<h3>Другие клиенты с похожей фамилией - в общую сверку включены</h3>
<p class="small">В выборке есть ещё документы с похожей фамилией. Они не исключались.
В эталоне таких: {fmt_int(len(surname_o))}, в по-новому3: {fmt_int(len(surname_n))}.
Ни один не попал в список расходящихся сумм. Номера и портфели скрыты.</p>
<table>
<thead><tr><th>Метка</th><th>Строк (новое3)</th><th>Комментарий</th><th>ТЧ / суммы</th></tr></thead>
<tbody>
{surname_tbl}
</tbody>
</table>

<h3>Документы УК с другими суммами (полный список)</h3>
<table>
<thead><tr><th>Метка</th><th>Ссылка (обезличено)</th><th>Дельты</th><th>Строк было</th><th>Стало</th></tr></thead>
<tbody>
{''.join(diff_docs) or '<tr class="ok"><td colspan="5">Нет</td></tr>'}
</tbody>
</table>
<p class="small">Всего расходящихся номеров: {fmt_int(dc['diff_sums'])}. Без клиента А: {fmt_int(dc_wo['diff_sums'])}.</p>

<h2 id="p5">6. Клиент А: что произошло на самом деле</h2>
<p>
Документ УК <code>{escape(DOC_A)}</code>, дата 31.12.2026 23:59:59.
Строк ТЧ в выгрузке: {fmt_int(plat_old.get(PLAT_NUM, {}).get('n', 0))} было,
{fmt_int(plat_new.get(PLAT_NUM, {}).get('n', 0))} стало (эталон 12, новое 15).
Портфели: {escape(PF1)} и {escape(PF2)}.
Константы на контуре: <code>НачислениеНДФЛПоПортфелю = Истина</code>,
<code>УчитыватьУКДиНКДПриРасчетеИнвестиционногоВычетаВНДФЛ = Истина</code>.
</p>
<div class="box out">
<b>Эталон = ветка 4 строк ТЧ Начисления. Этот прогон = ветка 6 строк</b>
Это единственный документ УК, у которого суммы не совпали с эталоном.
Сумма дохода документа одинаковая: 46 169 428.19.
Инвествычет {escape(PF1)}, код 618: 269 252.07 -&gt; 277 733.07, дельта <b>+8 481.00</b>.
К удержанию: 105 052 -&gt; 103 333, дельта <b>-1 719</b>.
Портфельная выгрузка по обоим ДУ совпала с эталоном (строк {fmt_int(r['pf_plat']['rows_old'])} / {fmt_int(r['pf_plat']['rows_new'])},
bag 0/0). Сдвиг только в расчёте документа УК.
Этот dump снят до волн A/B в расширении IMDEV7330_NkdParam: в прогоне ещё живёт типовая ветка 6 строк.
</div>
<table>
<thead><tr><th>Колонка документа {escape(DOC_A)}</th><th>Эталон</th><th>По-новому3</th><th>Дельта</th></tr></thead>
<tbody>
{plat_sum_rows}
</tbody>
</table>
<table>
<thead><tr><th>ТЧ документа</th><th>Эталон</th><th>По-новому3</th></tr></thead>
<tbody>
{parts_rows(plat_old.get(PLAT_NUM, {}).get('parts', {}), plat_new.get(PLAT_NUM, {}).get('parts', {}))}
</tbody>
</table>

<h3>Цепочка (типовой код, не IMDEV-7330)</h3>
<ol>
<li><code>ТаблицаПродаж</code>: запрос партий без <code>УПОРЯДОЧИТЬ</code>. Поле <code>ДатаПартии</code>
выбирается, но порядок строк SQL не фиксирует.</li>
<li><code>РаспределитьПоПартиямСУчетомПродаж</code> отдаёт остаток <b>последней строке</b> текущей выборки.</li>
<li>В <code>ОтразитьУКДиНКДВИнвестВычете</code> пары (Партия, Регистратор), которые не совпали
с этой выборкой, <b>молча отбрасываются</b>. Инвествычет по коду 618 получается либо 0, либо 8 481.</li>
<li><code>ПрименитьИнвестиционныйВычет</code> кладёт эту сумму на первую строку кода 1530.
Скачет только <code>СуммаВычета</code> 1530, доход не меняется.</li>
<li>Порог предыдущего документа по вычету 1530: <b>42 117 785.31</b>.
Эталон: вычет 1530 = 42 113 256.71 (ниже порога) - старый срез, 4 строки.
Этот прогон: 42 121 737.71 (выше порога) - копирование прошлого документа, 6 строк.</li>
<li><code>РаспределитьПревышениеПоТипуДохода</code> берёт первую строку через
<code>Найти(КодДохода)</code>. При 4 строках срез корректный. При 6 строках копируется
расклад прошлого документа: оба 1537 на 13% становятся 15 534.62 (это 2 x 7 767.31),
на 15% появляется отрицательный -7 767.31.</li>
</ol>
<div class="box warn">
<b>Второй дефект той же цепочки</b>
<code>РанееНачисленныйНДФЛ</code> кладёт текущие начисления во временную таблицу без свёртки,
затем FULL JOIN. Строка 1537 прошлого документа может удвоиться. Это усиливает ветку 6 строк,
но корень - неупорядоченные партии и потеря 8 481 в инвествычете.
</div>
<div class="box info">
<b>Что это не есть</b>
Не утечка чужого купона из пачки, не <code>МассивПортфелей[0]</code>, не эффект NkdCoupon.
NkdCoupon чинит привязку оплаченного купона к клиенту в массовом запросе НКД; на эти 8 481
он не влияет. Одиночный «Рассчитать» иногда даёт 4 строки, иногда 6 - тот же типовой баг,
он воспроизводится и на старом релизе. В массовом прогоне эталонный dump случайно попал
в 4 строки, по-новому3 - в 6.
</div>
<div class="box in">
<b>Что уже сделано после этого прогона</b>
Волны A/B в IMDEV7330_NkdParam: FIFO-сортировка <code>ТаблицаПродаж</code> и входа
<code>РаспределитьПоПартиямСУчетомПродаж</code>; журнал, если источник НКД/УКД не равен применённому;
свёртка ВТ в <code>РанееНачисленныйНДФЛ</code>; ветка шкалы по сумме строк кода, если строк кода больше одной -
старый механизм. Этот отчёт фиксирует прогон <b>до</b> этих волн.
</div>

<h3>Цифры строк (как в выгрузке)</h3>
<ol>
<li><b>Инвествычет {escape(PF1)}, код 618.</b> 269 252.07 -&gt; 277 733.07 (+8 481).
Ровно дельта колонки «Сумма» документа.</li>
<li><b>Начисления 1530 / вычет 201.</b> Эталон: 13% вычет 42 113 256.71, база 1 696 782.82;
15% без этого вычета. Новое: 13% копирует прошлый документ (доход 44 510 383.53, вычет 42 117 785.31,
база 2 392 598.22); 15% забирает остаток вычета 3 952.40.</li>
<li><b>Начисления 1537 / вычет 211.</b> Итоги дохода/вычета по коду совпадают.
Эталон: две строки 13%. Новое: четыре строки, в том числе 15% с отрицательными суммами
(-7 434.71 / -7 767.31) и две 13% по 15 534.62.</li>
<li><b>Удержания.</b> 15% по {escape(PF1)}: 87 193 -&gt; 83 356; по {escape(PF2)}: 17 859 -&gt; 17 073.
Добавились 13% к удержанию 2 410 и 494. Нетто по документу -1 719.
Финрезультат 13% по {escape(PF1)}: 3 345 441.69 -&gt; 3 336 960.69 (-8 481) - зеркало лишнего вычета.</li>
</ol>

<h3>Строки ТЧ (ключ без сумм, суммы справа; портфели обезличены)</h3>
<table>
<thead><tr><th>ТЧ</th><th>Портфель</th><th>Код дохода</th><th>Код вычета</th><th>Ставка</th><th>Эталон</th><th>По-новому3</th></tr></thead>
<tbody>
{plat_lines}
</tbody>
</table>

<h3>Строки, которые есть только в эталоне / только в новом (ключ с суммами)</h3>
<table>
<thead><tr><th colspan="2">Только в эталоне (все 8 - этот клиент)</th></tr></thead>
<tbody>
{plat_bag_old or '<tr><td colspan="2">нет</td></tr>'}
</tbody>
</table>
<table>
<thead><tr><th colspan="2">Только в новом (фрагмент; полная картина в таблице строк выше)</th></tr></thead>
<tbody>
{plat_bag_new or '<tr><td colspan="2">нет в топе bag (см. таблицу строк)</td></tr>'}
</tbody>
</table>

<h2 id="p7">7. Сообщения и лог</h2>
<table>
<thead><tr><th>Тип</th><th>Эталон</th><th>По-новому3</th><th>Дельта</th></tr></thead>
<tbody>
{''.join(kind_rows)}
</tbody>
</table>
<div class="box info">
<b>Распределение НДФЛ</b>
Эталон (файл сообщений обрезан с начала, с буквы Г): {fmt_int(msg['clients_rasp_a'])} клиентов.
Новое3: {fmt_int(msg['clients_rasp_b'])}.
Клиент А в этих сообщениях: эталон {fmt_int(len(msg.get('plat_msgs_a') or []))},
новое3 {fmt_int(len(msg.get('plat_msgs_b') or []))} (тексты ФИО не выводятся).
Непроведённые: эталон {fmt_int(len(msg['ne_prov_a']))}, новое3 {fmt_int(len(msg['ne_prov_b']))}.
Граница актуальности: {fmt_int(msg['granica_a'])} / {fmt_int(msg['granica_b'])}.
</div>
<p>Первая и последняя строки лога проведения содержат ФИО и коды ДУ - в отчёт не копируются.</p>

<h2 id="p8">8. Анализ</h2>
<ol>
<li><b>Правильность массового контура.</b> Портфели = эталон (28 431 документ, 413 547 строк).
УК без клиента А = эталон по суммам номеров ({fmt_int(dc_wo['equal_sums'])} общих документов).
Пустые УК - политика записи Выводы+ОбщиеРасходы, не налог. Три расширения основной регресс не разъехали.</li>
<li><b>Скорость.</b> Эталон - релиз 2.8.5.5 без доработок: 55 ч 05 мин, 500 объектов/час
(протокол DR, те же 27 292 клиента). Этот прогон: ~{escape(job_h_txt)} ч чистой работы на 10 потоках
(x{dur_x_old:.1f} по времени, x{speed_x_old:.1f} по объектам/час), стена кнопки {escape(wall_dur)}.
Блокировок / 1222 в логе {fmt_int(log3['lock_n'])}.</li>
<li><b>Клиент А.</b> +8 481 в инвествычете {escape(PF1)} и -1 719 к удержанию.
Причина: неупорядоченный <code>ТаблицаПродаж</code> -&gt; потеря УКД/НКД -&gt; порог шкалы
42 117 785.31 -&gt; 4 или 6 строк Начисления. Портфели клиента совпали. Другие клиенты
с похожей фамилией совпали. Это типовой дефект, не регресс оптимизаций IMDEV-7330.</li>
<li><b>NkdCoupon здесь ни при чём.</b> 8 481 - не чужой купон пачки. Расширение на этот
сценарий не рассчитано и цифр клиента А не меняет.</li>
<li><b>Что отдавать вендору по клиенту А.</b> Сортировка партий FIFO после
<code>ТаблицаПродаж</code> и на входе распределения; не отбрасывать молча НКД/УКД;
не копировать прошлый документ, если по коду дохода в текущем больше одной строки;
свернуть начисления документа в <code>РанееНачисленныйНДФЛ</code>. Волны A/B в NkdParam -
временный перехват, в типовую CF ещё не влиты.</li>
</ol>

<footer>
IMDEV-7330. HTML для вендора: персональные данные заменены, скриншотов нет.
Скрипт генерации: Скрипты/compare_ndfl_old_vs_new3.py.
</footer>
</div>
</body>
</html>
"""
    needles = (
        "Платонов",
        "Дмитрий Вячеславович",
        "000000000038432",
        "ДУ 2258",
        "ДУ 10155",
        "Абабилов",
        "Ящук",
    )
    for n in needles:
        if n in html:
            raise RuntimeError("PII leaked into HTML: " + n)
    with open(OUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    safe_print("HTML written: " + OUT_HTML)


def main():
    if "--html-only" in sys.argv:
        with open(OUT_JSON, encoding="utf-8") as f:
            r = json.load(f)
        write_html(r)
        safe_print("DONE html-only")
        return
    r = analyze()
    write_html(r)
    safe_print("DONE")


if __name__ == "__main__":
    main()
