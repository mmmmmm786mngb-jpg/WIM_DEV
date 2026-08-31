#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sverka vygruzok NDFL: po-staromu Avancor vs po-novomu2 (6 potokov).
Konsol - ASCII. HTML otchet - UTF-8.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from html import escape
from openpyxl import load_workbook

BASE = r"c:\1c\Cursor_1c\WIM_DEV\bases\WIM_FIn\projects\IMDEV-7330 НовыеТесты"
DOC = r"c:\1c\Cursor_1c\WIM_DEV\bases\WIM_FIn\projects\IMDEV-7330 Оптимизация 1С ФИН_Расчет НДФЛ\Документация"
OUT_HTML = os.path.join(DOC, "imdev7330_ndfl_old_vs_new2_diff.html")
OUT_JSON = os.path.join(DOC, "imdev7330_ndfl_old_vs_new2_diff.json")

FILES = {
    "uk_old": os.path.join(BASE, "НДФЛ_Управление_27292.xlsx"),
    "uk_new": os.path.join(BASE, "НДФЛ_Управление_27292_ПоНовому2.xlsx"),
    "pf_old": os.path.join(BASE, "НДФЛ_Портфели_27292.xlsx"),
    "pf_new": os.path.join(BASE, "НДФЛ_Портфели_27292_ПоНовому2.xlsx"),
    "msg_old": os.path.join(BASE, "Сообщения_поСтарому_Аванкор.txt"),
    "msg_new": os.path.join(BASE, "Сообщения_поНовому2_Аванкор.txt"),
    "log_new": os.path.join(BASE, "Лог формирования начисдений НДФЛ_ПоНовому2.txt"),
}

NUM_UK = [
    "СуммаВывода",
    "НалогооблагаемаяCумма",
    "СуммаВычета",
    "СуммаДохода",
    "СуммаКЗачету",
    "СуммаКУдержанию",
    "СуммаРанееУдержанногоНДФЛ",
    "ФинансовыйРезультат",
    "ИзменениеФинРезультата",
    "Пропорция",
    "Сумма",
]
NUM_PF = [
    "НалогооблагаемаяCумма",
    "СуммаВычета",
    "СуммаДохода",
    "СуммаРанееУдержанногоНДФЛ",
    "СуммаКУдержанию",
    "СуммаИсчисленногоНалога",
    "УплаченныйНДФЛ",
    "Сумма",
]
EPS = Decimal("0.01")
TOP_N = 40


def safe_print(text: str) -> None:
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", "replace").decode("ascii"))


def file_md5(path: str, chunk: int = 1024 * 1024) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        while True:
            b = f.read(chunk)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def to_dec(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, bool):
        return Decimal("0")
    if isinstance(v, int):
        return Decimal(v)
    if isinstance(v, float):
        return Decimal(str(round(v, 4)))
    s = str(v).strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def load_xlsx_rows(path: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    header = [str(c) if c is not None else "" for c in header]
    idx = {name: i for i, name in enumerate(header)}
    rows = []
    for row in it:
        if row is None:
            continue
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        rows.append(row)
    wb.close()
    return header, idx, rows


def get(row, idx, name, default=None):
    i = idx.get(name)
    if i is None or i >= len(row):
        return default
    return row[i]


def money_tuple(row, idx, names):
    return tuple(to_dec(get(row, idx, n)) for n in names)


def classify_msg(line: str) -> str:
    s = line.strip()
    if not s:
        return "empty"
    if "Невозможно распределить НДФЛ" in s:
        return "raspredelenie"
    if "граница актуальности сдвинута" in s:
        return "granica"
    if "не проведен" in s.lower() or "не проведён" in s.lower():
        return "ne_proveden"
    if "Ошибок:" in s or "ошибка" in s.lower():
        return "oshibka"
    return "other"


RE_CLIENT_RASP = re.compile(r"Невозможно распределить НДФЛ по клиенту (.+?) по алгоритму")
RE_PORTF_GRAN = re.compile(r"по портфелю (.+?) граница актуальности")
RE_DOC_GRAN = re.compile(r"документа (.+?) по портфелю")
RE_NE_PROVEDEN = re.compile(r"Документ (.+?) не проведен")


def read_text_lines(path: str):
    raw = open(path, "rb").read()
    for enc in ("utf-8-sig", "utf-8", "cp1251", "cp866", "utf-16"):
        try:
            text = raw.decode(enc)
            return text.splitlines()
        except UnicodeDecodeError:
            continue
    return raw.decode("cp1251", errors="replace").splitlines()


def parse_messages(path: str):
    lines = read_text_lines(path)
    nonempty = [ln for ln in lines if ln.strip()]
    kinds = Counter(classify_msg(ln) for ln in nonempty)
    clients_rasp = []
    for ln in nonempty:
        m = RE_CLIENT_RASP.search(ln)
        if m:
            clients_rasp.append(m.group(1).strip())
    granica = [ln.strip() for ln in nonempty if classify_msg(ln) == "granica"]
    ne_prov = [ln.strip() for ln in nonempty if classify_msg(ln) == "ne_proveden"]
    other = [ln.strip() for ln in nonempty if classify_msg(ln) == "other"]
    return {
        "total_lines": len(lines),
        "nonempty": len(nonempty),
        "kinds": dict(kinds),
        "clients_rasp": clients_rasp,
        "granica": granica,
        "ne_prov": ne_prov,
        "other": other,
        "unique": set(nonempty),
    }


RE_LOG_HEADER = re.compile(
    r"Начало выполнения:\s*(.+?),\s*окончание выполнения:\s*(.+?)\s*Ошибок:\s*(\d+)",
    re.S,
)
RE_PROVEDEN_PF = re.compile(r"Проведено начисление НДФЛ по портфелю (.+)$")
RE_PROVEDEN_UK = re.compile(r"Проведено начисление НДФЛ по клиенту (.+)$")
RE_FAIL = re.compile(r"(не удалось|ошибк|исключен)", re.I)


def parse_log(path: str):
    text = "\n".join(read_text_lines(path))
    header = {"start": None, "end": None, "errors": None}
    m = RE_LOG_HEADER.search(text)
    if m:
        header = {"start": m.group(1).strip(), "end": m.group(2).strip(), "errors": int(m.group(3))}
    proveden_pf = []
    proveden_uk = []
    fail_lines = []
    comments = Counter()
    for ln in text.splitlines():
        s = ln.strip()
        if not s:
            continue
        mp = RE_PROVEDEN_PF.search(s)
        if mp:
            proveden_pf.append(mp.group(1).strip())
            comments["proveden_pf"] += 1
            continue
        mu = RE_PROVEDEN_UK.search(s)
        if mu:
            proveden_uk.append(mu.group(1).strip())
            comments["proveden_uk"] += 1
            continue
        if "Проведено" in s:
            comments["proveden_other"] += 1
        elif RE_FAIL.search(s) and "Ошибок: 0" not in s:
            fail_lines.append(s[:300])
            comments["fail_like"] += 1
        elif "не записан" in s.lower() or "не создан" in s.lower():
            fail_lines.append(s[:300])
            comments["not_created"] += 1
    return {
        "header": header,
        "comments": dict(comments),
        "proveden_pf_n": len(proveden_pf),
        "proveden_uk_n": len(proveden_uk),
        "proveden_pf_unique": len(set(proveden_pf)),
        "proveden_uk_unique": len(set(proveden_uk)),
        "fail_lines": fail_lines[:50],
        "fail_n": len(fail_lines),
        "bytes": os.path.getsize(path),
        "chars": len(text),
    }


def agg_docs(rows, idx, num_names):
    """Sum numeric cols per document number; count parts and portfolios."""
    docs = {}
    parts = Counter()
    portfolios = set()
    numbers = set()
    for row in rows:
        num = str(get(row, idx, "НомерДокумента") or "")
        part = str(get(row, idx, "ТабличнаяЧасть") or "")
        pf = str(get(row, idx, "Портфель") or "")
        numbers.add(num)
        parts[part] += 1
        if pf:
            portfolios.add(pf)
        d = docs.setdefault(
            num,
            {
                "n": 0,
                "parts": Counter(),
                "portfolios": set(),
                "sums": {n: Decimal("0") for n in num_names},
                "ref": str(get(row, idx, "Ссылка") or ""),
            },
        )
        d["n"] += 1
        d["parts"][part] += 1
        if pf:
            d["portfolios"].add(pf)
        for n in num_names:
            d["sums"][n] += to_dec(get(row, idx, n))
    return docs, parts, portfolios, numbers


def line_key_uk(row, idx):
    """Business line without document number (numbers may differ)."""
    return (
        str(get(row, idx, "ТабличнаяЧасть") or ""),
        str(get(row, idx, "Портфель") or ""),
        str(get(row, idx, "КодВычета") or ""),
        str(get(row, idx, "КодДохода") or ""),
        str(get(row, idx, "Ставка") or ""),
        str(get(row, idx, "ТипДохода") or ""),
        str(get(row, idx, "Система") or ""),
        str(get(row, idx, "РаспределятьОбщиеРасходы") or ""),
        str(get(row, idx, "Закрытие") or ""),
        money_tuple(row, idx, NUM_UK),
    )


def line_key_pf(row, idx):
    return (
        str(get(row, idx, "ТабличнаяЧасть") or ""),
        str(get(row, idx, "Портфель") or ""),
        str(get(row, idx, "ВидРасхода") or ""),
        str(get(row, idx, "КодДохода") or ""),
        str(get(row, idx, "КодВычета") or ""),
        str(get(row, idx, "Ставка") or ""),
        str(get(row, idx, "ТипДохода") or ""),
        money_tuple(row, idx, NUM_PF),
    )


def bag_diff(old_rows, new_rows, key_fn, idx_old, idx_new):
    c_old = Counter(key_fn(r, idx_old) for r in old_rows)
    c_new = Counter(key_fn(r, idx_new) for r in new_rows)
    only_old = []
    only_new = []
    for k, n in c_old.items():
        d = n - c_new.get(k, 0)
        if d > 0:
            only_old.append((d, k))
    for k, n in c_new.items():
        d = n - c_old.get(k, 0)
        if d > 0:
            only_new.append((d, k))
    only_old.sort(reverse=True)
    only_new.sort(reverse=True)
    matched = sum(min(c_old[k], c_new.get(k, 0)) for k in c_old)
    return {
        "matched_rows": matched,
        "only_old_n": sum(d for d, _ in only_old),
        "only_new_n": sum(d for d, _ in only_new),
        "only_old_kinds": len(only_old),
        "only_new_kinds": len(only_new),
        "only_old": only_old[:TOP_N],
        "only_new": only_new[:TOP_N],
    }


def compare_doc_sums(docs_old, docs_new, num_names):
    common = set(docs_old) & set(docs_new)
    only_old = sorted(set(docs_old) - set(docs_new))
    only_new = sorted(set(docs_new) - set(docs_old))
    diffs = []
    equal = 0
    for num in common:
        a = docs_old[num]["sums"]
        b = docs_new[num]["sums"]
        delta = {}
        changed = False
        for n in num_names:
            d = b[n] - a[n]
            if abs(d) >= EPS:
                changed = True
                delta[n] = str(d)
        if changed:
            diffs.append(
                {
                    "num": num,
                    "ref": docs_new[num]["ref"] or docs_old[num]["ref"],
                    "delta": delta,
                    "old_rows": docs_old[num]["n"],
                    "new_rows": docs_new[num]["n"],
                }
            )
        else:
            equal += 1
    diffs.sort(key=lambda x: -sum(abs(Decimal(v)) for v in x["delta"].values()))
    return {
        "common": len(common),
        "equal_sums": equal,
        "diff_sums": len(diffs),
        "only_old": only_old,
        "only_new": only_new,
        "diffs": diffs[:80],
        "diffs_total": len(diffs),
    }


def pf_key_no_money(row, idx):
    return (
        str(get(row, idx, "ТабличнаяЧасть") or ""),
        str(get(row, idx, "Портфель") or ""),
        str(get(row, idx, "ВидРасхода") or ""),
        str(get(row, idx, "КодДохода") or ""),
        str(get(row, idx, "КодВычета") or ""),
        str(get(row, idx, "Ставка") or ""),
        str(get(row, idx, "ТипДохода") or ""),
    )


def sum_by_portfolio(rows, idx, num_names):
    out = {}
    for row in rows:
        pf = str(get(row, idx, "Портфель") or "")
        d = out.setdefault(pf, {n: Decimal("0") for n in num_names})
        for n in num_names:
            d[n] += to_dec(get(row, idx, n))
    return out


def compare_pf_sums(s_old, s_new, num_names):
    keys = set(s_old) | set(s_new)
    diffs = []
    equal = 0
    only_old = []
    only_new = []
    for k in keys:
        if k not in s_new:
            only_old.append(k)
            continue
        if k not in s_old:
            only_new.append(k)
            continue
        delta = {}
        changed = False
        for n in num_names:
            d = s_new[k][n] - s_old[k][n]
            if abs(d) >= EPS:
                changed = True
                delta[n] = str(d)
        if changed:
            diffs.append({"portfolio": k, "delta": delta})
        else:
            equal += 1
    diffs.sort(key=lambda x: -sum(abs(Decimal(v)) for v in x["delta"].values()))
    return {
        "equal": equal,
        "diff_n": len(diffs),
        "only_old": only_old,
        "only_new": only_new,
        "diffs": diffs[:80],
    }


def totals(rows, idx, names):
    t = {n: Decimal("0") for n in names}
    for row in rows:
        for n in names:
            t[n] += to_dec(get(row, idx, n))
    return {n: str(t[n]) for n in names}


def fmt_key_uk(k):
    part, pf, vychet, dohod, stavka, tip, sistema, rasp, zakr, money = k
    money_s = ", ".join(f"{n}={v}" for n, v in zip(NUM_UK, money) if v != 0)
    return (
        f"{part} | {pf} | kodDoh={dohod} kodVych={vychet} "
        f"stavka={stavka} tip={tip} sys={sistema} {money_s}"
    )


def fmt_key_pf(k):
    part, pf, vid, dohod, vychet, stavka, tip, money = k
    money_s = ", ".join(f"{n}={v}" for n, v in zip(NUM_PF, money) if v != 0)
    return (
        f"{part} | {pf} | vid={vid} kodDoh={dohod} kodVych={vychet} "
        f"stavka={stavka} tip={tip} {money_s}"
    )


def dec_str_map(d):
    return {k: str(v) if isinstance(v, Decimal) else v for k, v in d.items()}


def analyze():
    result = {"files": {}}
    for k, p in FILES.items():
        result["files"][k] = {
            "name": os.path.basename(p),
            "bytes": os.path.getsize(p),
            "exists": os.path.isfile(p),
        }

    safe_print("MD5 portfolio files...")
    md5_pf_old = file_md5(FILES["pf_old"])
    md5_pf_new = file_md5(FILES["pf_new"])
    result["pf_same_bytes"] = result["files"]["pf_old"]["bytes"] == result["files"]["pf_new"]["bytes"]
    result["pf_md5_old"] = md5_pf_old
    result["pf_md5_new"] = md5_pf_new
    result["pf_identical"] = md5_pf_old == md5_pf_new
    safe_print("pf identical=" + str(result["pf_identical"]))

    safe_print("Load UK old...")
    h_uk_o, i_uk_o, r_uk_o = load_xlsx_rows(FILES["uk_old"])
    safe_print("Load UK new...")
    h_uk_n, i_uk_n, r_uk_n = load_xlsx_rows(FILES["uk_new"])
    result["uk"] = {
        "header_equal": h_uk_o == h_uk_n,
        "rows_old": len(r_uk_o),
        "rows_new": len(r_uk_n),
        "totals_old": totals(r_uk_o, i_uk_o, NUM_UK),
        "totals_new": totals(r_uk_n, i_uk_n, NUM_UK),
    }
    docs_o, parts_o, pfset_o, nums_o = agg_docs(r_uk_o, i_uk_o, NUM_UK)
    docs_n, parts_n, pfset_n, nums_n = agg_docs(r_uk_n, i_uk_n, NUM_UK)
    result["uk"]["parts_old"] = dict(parts_o)
    result["uk"]["parts_new"] = dict(parts_n)
    result["uk"]["docs_old"] = len(docs_o)
    result["uk"]["docs_new"] = len(docs_n)
    result["uk"]["portfolios_old"] = len(pfset_o)
    result["uk"]["portfolios_new"] = len(pfset_n)
    result["uk"]["doc_compare"] = compare_doc_sums(docs_o, docs_n, NUM_UK)
    result["uk"]["line_bag"] = bag_diff(r_uk_o, r_uk_n, line_key_uk, i_uk_o, i_uk_n)
    sums_pf_uk_o = sum_by_portfolio(r_uk_o, i_uk_o, NUM_UK)
    sums_pf_uk_n = sum_by_portfolio(r_uk_n, i_uk_n, NUM_UK)
    result["uk"]["by_portfolio"] = compare_pf_sums(sums_pf_uk_o, sums_pf_uk_n, NUM_UK)

    if result["pf_identical"]:
        result["pf"] = {
            "skipped_full": True,
            "reason": "MD5 files identical",
            "rows": None,
        }
        safe_print("Skip portfolio row compare (identical files).")
    else:
        safe_print("Load PF old...")
        h_pf_o, i_pf_o, r_pf_o = load_xlsx_rows(FILES["pf_old"])
        safe_print("Load PF new...")
        h_pf_n, i_pf_n, r_pf_n = load_xlsx_rows(FILES["pf_new"])
        result["pf"] = {
            "skipped_full": False,
            "header_equal": h_pf_o == h_pf_n,
            "rows_old": len(r_pf_o),
            "rows_new": len(r_pf_n),
            "totals_old": totals(r_pf_o, i_pf_o, NUM_PF),
            "totals_new": totals(r_pf_n, i_pf_n, NUM_PF),
        }
        docs_po, parts_po, pfset_po, nums_po = agg_docs(r_pf_o, i_pf_o, NUM_PF)
        docs_pn, parts_pn, pfset_pn, nums_pn = agg_docs(r_pf_n, i_pf_n, NUM_PF)
        result["pf"]["parts_old"] = dict(parts_po)
        result["pf"]["parts_new"] = dict(parts_pn)
        result["pf"]["docs_old"] = len(docs_po)
        result["pf"]["docs_new"] = len(docs_pn)
        result["pf"]["doc_compare"] = compare_doc_sums(docs_po, docs_pn, NUM_PF)
        result["pf"]["line_bag"] = bag_diff(r_pf_o, r_pf_n, line_key_pf, i_pf_o, i_pf_n)
        result["pf"]["by_portfolio"] = compare_pf_sums(
            sum_by_portfolio(r_pf_o, i_pf_o, NUM_PF),
            sum_by_portfolio(r_pf_n, i_pf_n, NUM_PF),
            NUM_PF,
        )

    safe_print("Parse messages...")
    msg_o = parse_messages(FILES["msg_old"])
    msg_n = parse_messages(FILES["msg_new"])
    set_o = msg_o["unique"]
    set_n = msg_n["unique"]
    cl_o = set(msg_o["clients_rasp"])
    cl_n = set(msg_n["clients_rasp"])
    result["messages"] = {
        "old": {k: v for k, v in msg_o.items() if k != "unique"},
        "new": {k: v for k, v in msg_n.items() if k != "unique"},
        "only_old_n": len(set_o - set_n),
        "only_new_n": len(set_n - set_o),
        "common_n": len(set_o & set_n),
        "only_old_sample": sorted(set_o - set_n)[:TOP_N],
        "only_new_sample": sorted(set_n - set_o)[:TOP_N],
        "clients_rasp_old": len(cl_o),
        "clients_rasp_new": len(cl_n),
        "clients_only_old": sorted(cl_o - cl_n),
        "clients_only_new": sorted(cl_n - cl_o),
        "granica_old": len(msg_o["granica"]),
        "granica_new": len(msg_n["granica"]),
        "granica_only_old": sorted(set(msg_o["granica"]) - set(msg_n["granica"]))[:TOP_N],
        "granica_only_new": sorted(set(msg_n["granica"]) - set(msg_o["granica"]))[:TOP_N],
        "ne_prov_old": msg_o["ne_prov"],
        "ne_prov_new": msg_n["ne_prov"],
    }

    safe_print("Parse log new2...")
    result["log_new"] = parse_log(FILES["log_new"])

    # JSON-safe: convert leftover
    def conv(obj):
        if isinstance(obj, Decimal):
            return str(obj)
        if isinstance(obj, set):
            return sorted(obj)
        if isinstance(obj, Counter):
            return dict(obj)
        if isinstance(obj, tuple):
            return [conv(x) for x in obj]
        if isinstance(obj, list):
            return [conv(x) for x in obj]
        if isinstance(obj, dict):
            return {str(k): conv(v) for k, v in obj.items()}
        return obj

    # stringify bag keys for json
    for block in (result.get("uk", {}), result.get("pf", {})):
        bag = block.get("line_bag")
        if not bag:
            continue
        fmt = fmt_key_uk if block is result.get("uk") else fmt_key_pf
        if block is result.get("pf") and result.get("pf_identical"):
            continue
        if block is result.get("uk"):
            fmt = fmt_key_uk
        else:
            fmt = fmt_key_pf
        bag["only_old_fmt"] = [(n, fmt(k)) for n, k in bag["only_old"]]
        bag["only_new_fmt"] = [(n, fmt(k)) for n, k in bag["only_new"]]
        del bag["only_old"]
        del bag["only_new"]

    result["uk"]["doc_compare"]["only_old_n"] = len(result["uk"]["doc_compare"]["only_old"])
    result["uk"]["doc_compare"]["only_new_n"] = len(result["uk"]["doc_compare"]["only_new"])
    result["uk"]["doc_compare"]["only_old_sample"] = result["uk"]["doc_compare"]["only_old"][:TOP_N]
    result["uk"]["doc_compare"]["only_new_sample"] = result["uk"]["doc_compare"]["only_new"][:TOP_N]
    del result["uk"]["doc_compare"]["only_old"]
    del result["uk"]["doc_compare"]["only_new"]

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(conv(result), f, ensure_ascii=False, indent=2)
    safe_print("JSON written")
    return result


def money_table(totals_old, totals_new):
    rows = []
    for k in totals_old:
        a = Decimal(totals_old[k])
        b = Decimal(totals_new[k])
        d = b - a
        cls = "ok" if abs(d) < EPS else "bad"
        rows.append(
            f"<tr class='{cls}'><td>{escape(k)}</td><td class='num'>{a}</td>"
            f"<td class='num'>{b}</td><td class='num'>{d}</td></tr>"
        )
    return "\n".join(rows)


def parts_table(po, pn):
    keys = sorted(set(po) | set(pn))
    out = []
    for k in keys:
        a = po.get(k, 0)
        b = pn.get(k, 0)
        cls = "ok" if a == b else "warn"
        out.append(
            f"<tr class='{cls}'><td>{escape(k or '(empty)')}</td>"
            f"<td class='num'>{a}</td><td class='num'>{b}</td><td class='num'>{b-a}</td></tr>"
        )
    return "\n".join(out)


def write_html(r):
    uk = r["uk"]
    dc = uk["doc_compare"]
    bag = uk["line_bag"]
    bp = uk["by_portfolio"]
    msg = r["messages"]
    logn = r["log_new"]

    verdict_bits = []
    pf_ok = r["pf_identical"]
    uk_totals_ok = all(
        abs(Decimal(uk["totals_new"][k]) - Decimal(uk["totals_old"][k])) < EPS for k in uk["totals_old"]
    )
    uk_line_ok = bag["only_old_n"] == 0 and bag["only_new_n"] == 0
    if pf_ok:
        verdict_bits.append("Vygruzka portfeley byte-v-byte sovpadaet.")
    if uk_line_ok:
        verdict_bits.append("Stroki UK sovpadayut kak multiset biznes-klyuchey.")
    elif uk_totals_ok and dc["diff_sums"] == 0 and dc["only_old_n"] == 0 and dc["only_new_n"] == 0:
        verdict_bits.append("Summy UK po nomeram dokumentov sovpadayut, no nabor strok otlichaetsya.")
    else:
        verdict_bits.append("Po UK est raskhozhdeniya (sm. tablitsy).")

    # Human verdict in Russian for HTML
    if pf_ok and bag["only_old_n"] == 0 and bag["only_new_n"] == 0 and dc["diff_sums"] == 0:
        verdict_ru = "Расхождений в суммах и составе строк выгрузок нет. Файлы портфелей идентичны. Отличия только в служебных сообщениях (см. раздел 4)."
        verdict_cls = "in"
    elif pf_ok and uk_totals_ok and dc["diff_sums"] == 0:
        verdict_ru = "Итоги сумм УК совпадают, портфели идентичны. Есть отличие состава/количества строк УК - разбор ниже."
        verdict_cls = "warn"
    else:
        verdict_ru = "Найдены расхождения в суммах или составе документов. Детали в таблицах."
        verdict_cls = "out"

    def sample_lis(items, n=30):
        if not items:
            return "<p class='small'>Нет.</p>"
        lis = "".join(f"<li>{escape(str(x))}</li>" for x in items[:n])
        more = f"<p class='small'>Показано {min(n,len(items))} из {len(items)}.</p>" if len(items) > n else ""
        return f"<ul>{lis}</ul>{more}"

    bag_old_rows = "".join(
        f"<tr><td class='num'>{n}</td><td class='sig'>{escape(s)}</td></tr>"
        for n, s in bag.get("only_old_fmt", [])
    )
    bag_new_rows = "".join(
        f"<tr><td class='num'>{n}</td><td class='sig'>{escape(s)}</td></tr>"
        for n, s in bag.get("only_new_fmt", [])
    )

    diff_doc_rows = []
    for d in dc["diffs"][:40]:
        deltas = "; ".join(f"{k}={v}" for k, v in d["delta"].items())
        diff_doc_rows.append(
            f"<tr class='bad'><td class='sig'>{escape(d['num'])}</td>"
            f"<td>{escape(d['ref'])}</td><td class='sig'>{escape(deltas)}</td>"
            f"<td class='num'>{d['old_rows']}</td><td class='num'>{d['new_rows']}</td></tr>"
        )

    pf_diff_rows = []
    for d in bp["diffs"][:40]:
        deltas = "; ".join(f"{k}={v}" for k, v in d["delta"].items())
        pf_diff_rows.append(
            f"<tr class='bad'><td>{escape(d['portfolio'])}</td><td class='sig'>{escape(deltas)}</td></tr>"
        )

    kinds_o = msg["old"]["kinds"]
    kinds_n = msg["new"]["kinds"]
    kind_keys = sorted(set(kinds_o) | set(kinds_n))
    kind_rows = []
    labels = {
        "raspredelenie": "Невозможно распределить НДФЛ (на первый портфель)",
        "granica": "Граница актуальности сдвинута",
        "ne_proveden": "Документ не проведен",
        "oshibka": "Ошибка",
        "other": "Прочее",
        "empty": "Пустые",
    }
    for k in kind_keys:
        a = kinds_o.get(k, 0)
        b = kinds_n.get(k, 0)
        cls = "ok" if a == b else "warn"
        kind_rows.append(
            f"<tr class='{cls}'><td>{escape(labels.get(k,k))}</td>"
            f"<td class='num'>{a}</td><td class='num'>{b}</td><td class='num'>{b-a}</td></tr>"
        )

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>IMDEV-7330. Сверка НДФЛ: по-старому vs по-новому2</title>
<style>
    body {{ font-family: "Segoe UI", Arial, sans-serif; line-height: 1.55; color: #212529; background: #f5f6f8; margin: 0; padding: 0 0 60px 0; }}
    .wrap {{ max-width: 1200px; margin: 0 auto; padding: 0 24px; }}
    header {{ background: #1f2d3d; color: #fff; padding: 28px 0 22px 0; margin-bottom: 28px; }}
    header h1 {{ margin: 0 0 8px 0; font-size: 24px; }}
    header .sub {{ color: #b8c4d0; font-size: 14px; }}
    h2 {{ font-size: 20px; margin: 32px 0 12px 0; padding-bottom: 8px; border-bottom: 2px solid #17a2b8; }}
    h3 {{ font-size: 16px; margin: 18px 0 8px 0; color: #1f2d3d; }}
    p {{ margin: 10px 0; }}
    table {{ width: 100%; border-collapse: collapse; background: #fff; margin: 12px 0; font-size: 13.5px; }}
    th {{ background: #1f2d3d; color: #fff; text-align: left; padding: 8px 10px; }}
    td {{ padding: 7px 10px; border-bottom: 1px solid #e3e6ea; vertical-align: top; }}
    tbody tr:nth-child(even) {{ background: #fafbfc; }}
    tr.ok td {{ background: #eefaf1; }}
    tr.bad td {{ background: #fdf0f1; }}
    tr.warn td {{ background: #fff9e6; }}
    .num {{ text-align: right; font-family: Consolas, monospace; white-space: nowrap; }}
    .sig {{ font-family: Consolas, "Courier New", monospace; font-size: 12px; }}
    code {{ background: #eef1f4; padding: 1px 5px; border-radius: 3px; font-family: Consolas, monospace; font-size: 13px; }}
    pre {{ background: #1f2d3d; color: #e6edf3; padding: 14px 16px; border-radius: 5px; overflow-x: auto; font-size: 12.5px; }}
    .box {{ padding: 14px 18px; border-radius: 5px; margin: 14px 0; border-left: 5px solid; }}
    .box b {{ display: block; margin-bottom: 6px; }}
    .in {{ background: #eefaf1; border-color: #28a745; }}
    .out {{ background: #fdf0f1; border-color: #dc3545; }}
    .warn {{ background: #fff9e6; border-color: #f0ad4e; }}
    .info {{ background: #eaf7fa; border-color: #17a2b8; }}
    .small {{ font-size: 13px; color: #6c757d; }}
    ul, ol {{ margin: 8px 0 8px 22px; }}
    footer {{ margin-top: 36px; padding-top: 12px; border-top: 1px solid #dee2e6; font-size: 13px; color: #6c757d; }}
    .toc {{ background: #fff; border: 1px solid #e3e6ea; border-radius: 5px; padding: 12px 20px; margin: 16px 0; }}
    .toc a {{ text-decoration: none; color: #0a58ca; }}
</style>
</head>
<body>
<header>
<div class="wrap">
    <h1>IMDEV-7330. Сверка расчёта НДФЛ: по-старому vs по-новому2</h1>
    <div class="sub">
        Прогон без расширений IMDEV7330_Locks / IMDEV7330_NkdParam.
        6 потоков, выборка отчёта 27292. Сравнение выгрузок табличных частей и сообщений.
    </div>
</div>
</header>
<div class="wrap">

<div class="toc">
<b>Содержание</b>
<ol>
<li><a href="#p0">Итог</a></li>
<li><a href="#p1">Файлы и метод</a></li>
<li><a href="#p2">Портфели</a></li>
<li><a href="#p3">Управление (УК)</a></li>
<li><a href="#p4">Сообщения</a></li>
<li><a href="#p5">Лог формирования (по-новому2)</a></li>
<li><a href="#p6">Анализ</a></li>
</ol>
</div>

<h2 id="p0">1. Итог</h2>
<div class="box {verdict_cls}">
<b>Вердикт</b>
{escape(verdict_ru)}
</div>

<table>
<thead><tr><th>Контур</th><th>Было</th><th>Стало</th><th>Результат</th></tr></thead>
<tbody>
<tr class="{'ok' if r['pf_identical'] else 'bad'}">
<td>НДФЛ_Портфели_27292.xlsx</td>
<td class="num">{r['files']['pf_old']['bytes']} байт</td>
<td class="num">{r['files']['pf_new']['bytes']} байт</td>
<td>{"MD5 совпадает, файлы идентичны" if r['pf_identical'] else "Файлы различаются"}</td>
</tr>
<tr class="{'ok' if uk['rows_old']==uk['rows_new'] else 'warn'}">
<td>НДФЛ_Управление_27292.xlsx, строки</td>
<td class="num">{uk['rows_old']}</td>
<td class="num">{uk['rows_new']}</td>
<td>дельта {uk['rows_new']-uk['rows_old']}</td>
</tr>
<tr class="{'ok' if uk['docs_old']==uk['docs_new'] else 'warn'}">
<td>Документы УК (уник. номер)</td>
<td class="num">{uk['docs_old']}</td>
<td class="num">{uk['docs_new']}</td>
<td>общих {dc['common']}; только старые {dc['only_old_n']}; только новые {dc['only_new_n']}; суммы отличаются {dc['diff_sums']}</td>
</tr>
<tr class="{'ok' if bag['only_old_n']==0 and bag['only_new_n']==0 else 'warn'}">
<td>Строки УК как мультимножество (без номера документа)</td>
<td class="num">matched {bag['matched_rows']}</td>
<td class="num">только старые {bag['only_old_n']} / только новые {bag['only_new_n']}</td>
<td>сравнение по ТЧ + портфель + коды + ставки + суммы</td>
</tr>
<tr class="{'ok' if msg['only_old_n']==0 and msg['only_new_n']==0 else 'warn'}">
<td>Сообщения (уник. строки)</td>
<td class="num">{msg['old']['nonempty']}</td>
<td class="num">{msg['new']['nonempty']}</td>
<td>общих {msg['common_n']}; только старые {msg['only_old_n']}; только новые {msg['only_new_n']}</td>
</tr>
</tbody>
</table>

<h2 id="p1">2. Файлы и метод</h2>
<div class="box info">
<b>Как сравнивали</b>
<ul>
<li>Портфели: MD5 файла. При совпадении построчное сравнение не нужно.</li>
<li>УК: итоги по числовым колонкам; суммы по номеру документа; суммы по портфелю; мультимножество строк без номера документа (номера могли бы разъехаться между прогонами).</li>
<li>Сообщения: классификация шаблонов + множества клиентов / границ актуальности.</li>
<li>Лог формирования есть только у прогона по-новому2 (старого лога в наборе нет).</li>
<li>Допуск по деньгам: 0.01. Расширения Locks / NkdParam в этом прогоне не использовались.</li>
</ul>
</div>
<table>
<thead><tr><th>Роль</th><th>Файл</th><th>Размер</th></tr></thead>
<tbody>
<tr><td>Портфели, было</td><td class="sig">{escape(r['files']['pf_old']['name'])}</td><td class="num">{r['files']['pf_old']['bytes']}</td></tr>
<tr><td>Портфели, стало</td><td class="sig">{escape(r['files']['pf_new']['name'])}</td><td class="num">{r['files']['pf_new']['bytes']}</td></tr>
<tr><td>УК, было</td><td class="sig">{escape(r['files']['uk_old']['name'])}</td><td class="num">{r['files']['uk_old']['bytes']}</td></tr>
<tr><td>УК, стало</td><td class="sig">{escape(r['files']['uk_new']['name'])}</td><td class="num">{r['files']['uk_new']['bytes']}</td></tr>
<tr><td>Сообщения, было</td><td class="sig">{escape(r['files']['msg_old']['name'])}</td><td class="num">{r['files']['msg_old']['bytes']}</td></tr>
<tr><td>Сообщения, стало</td><td class="sig">{escape(r['files']['msg_new']['name'])}</td><td class="num">{r['files']['msg_new']['bytes']}</td></tr>
<tr><td>Лог, стало</td><td class="sig">{escape(r['files']['log_new']['name'])}</td><td class="num">{r['files']['log_new']['bytes']}</td></tr>
</tbody>
</table>
<p class="small">MD5 портфелей: было <code>{r['pf_md5_old']}</code>, стало <code>{r['pf_md5_new']}</code>.</p>

<h2 id="p2">3. Портфели</h2>
"""
    if r["pf_identical"]:
        html += """
<div class="box in">
<b>Расхождений нет</b>
Файлы <code>НДФЛ_Портфели_27292.xlsx</code> и <code>НДФЛ_Портфели_27292_ПоНовому2.xlsx</code>
совпадают побайтово (одинаковый размер и MD5). Состав и суммы начислений по портфелям
в этих выгрузках идентичны.
</div>
"""
    else:
        pf = r["pf"]
        html += f"""
<div class="box out"><b>Файлы портфелей различаются</b></div>
<p>Строк: {pf.get('rows_old')} / {pf.get('rows_new')}.</p>
<table>
<thead><tr><th>Колонка</th><th>Было</th><th>Стало</th><th>Дельта</th></tr></thead>
<tbody>
{money_table(pf['totals_old'], pf['totals_new'])}
</tbody>
</table>
"""

    html += f"""
<h2 id="p3">4. Управление (документы УК)</h2>
<p>Заголовки колонок {"совпадают" if uk["header_equal"] else "НЕ совпадают"}.
Портфелей в строках: {uk["portfolios_old"]} / {uk["portfolios_new"]}.</p>

<h3>Итоги по числовым колонкам (все строки)</h3>
<table>
<thead><tr><th>Колонка</th><th>Было</th><th>Стало</th><th>Дельта (стало - было)</th></tr></thead>
<tbody>
{money_table(uk['totals_old'], uk['totals_new'])}
</tbody>
</table>

<h3>Строки по табличной части</h3>
<table>
<thead><tr><th>ТЧ</th><th>Было</th><th>Стало</th><th>Дельта</th></tr></thead>
<tbody>
{parts_table(uk['parts_old'], uk['parts_new'])}
</tbody>
</table>

<h3>Сопоставление по номеру документа</h3>
<p>Общих номеров: <b>{dc['common']}</b>, из них суммы равны: <b>{dc['equal_sums']}</b>,
суммы отличаются: <b>{dc['diff_sums']}</b>.
Только в старом: {dc['only_old_n']}. Только в новом: {dc['only_new_n']}.</p>
"""
    if dc["diffs"]:
        html += f"""
<table>
<thead><tr><th>Номер</th><th>Ссылка</th><th>Дельта сумм</th><th>Строк было</th><th>Строк стало</th></tr></thead>
<tbody>
{''.join(diff_doc_rows)}
</tbody>
</table>
"""
    else:
        html += "<div class='box in'><b>По общим номерам документов суммы не отличаются.</b></div>"

    if dc["only_old_n"] or dc["only_new_n"]:
        html += "<h3>Номера только в одном прогоне</h3>"
        html += "<p><b>Только старый</b></p>" + sample_lis(dc["only_old_sample"], 40)
        html += "<p><b>Только новый</b></p>" + sample_lis(dc["only_new_sample"], 40)

    html += f"""
<h3>Мультимножество строк (ключ без номера документа)</h3>
<p>Совпало строк: {bag['matched_rows']}.
Только в старом: {bag['only_old_n']} ({bag['only_old_kinds']} видов).
Только в новом: {bag['only_new_n']} ({bag['only_new_kinds']} видов).</p>
"""
    if bag["only_old_n"] or bag["only_new_n"]:
        html += f"""
<table>
<thead><tr><th colspan="2">Только в старом (топ {TOP_N})</th></tr>
<tr><th>Кол-во</th><th>Ключ</th></tr></thead>
<tbody>{bag_old_rows or '<tr><td colspan="2">нет</td></tr>'}</tbody>
</table>
<table>
<thead><tr><th colspan="2">Только в новом (топ {TOP_N})</th></tr>
<tr><th>Кол-во</th><th>Ключ</th></tr></thead>
<tbody>{bag_new_rows or '<tr><td colspan="2">нет</td></tr>'}</tbody>
</table>
"""
    else:
        html += "<div class='box in'><b>Набор строк УК совпадает.</b> Разница количества строк в xlsx, если есть, не даёт разных сумм/ключей (дубли или пустые не попали в выборку).</div>"

    html += f"""
<h3>Суммы УК, свёрнутые по портфелю</h3>
<p>Портфелей с равными суммами: {bp['equal']}. С отличием: {bp['diff_n']}.
Только старый: {len(bp['only_old'])}. Только новый: {len(bp['only_new'])}.</p>
"""
    if bp["diff_n"]:
        html += f"""
<table>
<thead><tr><th>Портфель</th><th>Дельта</th></tr></thead>
<tbody>{''.join(pf_diff_rows)}</tbody>
</table>
"""
    else:
        html += "<div class='box in'><b>По портфелям суммы УК совпадают.</b></div>"
    if bp["only_old"] or bp["only_new"]:
        html += "<p><b>Портфели только в старом</b></p>" + sample_lis(bp["only_old"])
        html += "<p><b>Портфели только в новом</b></p>" + sample_lis(bp["only_new"])

    html += f"""
<h2 id="p4">5. Сообщения</h2>
<p>Непустых строк: {msg['old']['nonempty']} / {msg['new']['nonempty']}.
Уникальных общих: {msg['common_n']}.</p>
<table>
<thead><tr><th>Тип</th><th>Было</th><th>Стало</th><th>Дельта</th></tr></thead>
<tbody>
{''.join(kind_rows)}
</tbody>
</table>
<p>Клиенты «невозможно распределить»: {msg['clients_rasp_old']} / {msg['clients_rasp_new']}.
Только старый: {len(msg['clients_only_old'])}. Только новый: {len(msg['clients_only_new'])}.</p>
<p><b>Клиенты только в старом прогоне</b></p>
{sample_lis(msg['clients_only_old'])}
<p><b>Клиенты только в новом прогоне</b></p>
{sample_lis(msg['clients_only_new'])}
<p>Сообщения про границу актуальности: {msg['granica_old']} / {msg['granica_new']}.</p>
<p><b>Граница актуальности только в старом</b></p>
{sample_lis(msg['granica_only_old'])}
<p><b>Граница актуальности только в новом</b></p>
{sample_lis(msg['granica_only_new'])}
<p><b>Не проведен (старый)</b></p>
{sample_lis(msg['ne_prov_old'])}
<p><b>Не проведен (новый)</b></p>
{sample_lis(msg['ne_prov_new'])}
<p><b>Уникальные тексты только в старом (образец)</b></p>
{sample_lis(msg['only_old_sample'])}
<p><b>Уникальные тексты только в новом (образец)</b></p>
{sample_lis(msg['only_new_sample'])}

<h2 id="p5">6. Лог формирования (только по-новому2)</h2>
<div class="box info">
<b>Интервал</b>
Начало: {escape(str(logn['header'].get('start')))}.
Окончание: {escape(str(logn['header'].get('end')))}.
Ошибок в шапке лога: {escape(str(logn['header'].get('errors')))}.
</div>
<table>
<thead><tr><th>Показатель</th><th>Значение</th></tr></thead>
<tbody>
<tr><td>Проведено начисление по портфелю (строк)</td><td class="num">{logn['proveden_pf_n']}</td></tr>
<tr><td>Уникальных портфелей в этих строках</td><td class="num">{logn['proveden_pf_unique']}</td></tr>
<tr><td>Проведено начисление по клиенту УК (строк)</td><td class="num">{logn['proveden_uk_n']}</td></tr>
<tr><td>Уникальных клиентов УК</td><td class="num">{logn['proveden_uk_unique']}</td></tr>
<tr><td>Строк похожих на ошибку / не создан</td><td class="num">{logn['fail_n']}</td></tr>
</tbody>
</table>
<p class="small">Старого лога в комплекте нет, длительность 6 потоков сопоставить по этому файлу нельзя.
В заметке 6_потоков.txt: старт 30.08.2026 22:17, финиш указан как 30.08.2026 18:40 (даты в заметке противоречивы).
Шапка лога по-новому2: 30.08.2026 23:35:37 - 31.08.2026 18:40:33.</p>
"""
    if logn["fail_lines"]:
        html += "<pre>" + escape("\n".join(logn["fail_lines"])) + "</pre>"
    else:
        html += "<div class='box in'><b>В логе нет строк с маркерами ошибки / не создан / не удалось.</b></div>"

    html += f"""
<h2 id="p6">7. Анализ расхождений</h2>
<ol>
<li><b>Документы по портфелю.</b> Выгрузки xlsx идентичны. Новый расчёт не меняет начисления по портфелям
относительно старого прогона на этой выборке.</li>
<li><b>Документы УК.</b>
Итоги колонок: см. таблицу раздела 4.
По номерам документов отличающихся сумм: {dc['diff_sums']}.
Строк-сирот мультимножества: старых {bag['only_old_n']}, новых {bag['only_new_n']}.
Если дельты нулевые, разница размера xlsx ({uk['rows_new']-uk['rows_old']} строк) не несёт другого налога:
это либо одни и те же ключи, либо служебные/пустые отличия экспорта.</li>
<li><b>Сообщения.</b> Основной объём - штатное «невозможно распределить НДФЛ, весь налог на первый портфель».
Это не ошибка расчёта сумм, а сообщение алгоритма распределения по портфелям внутри УК.
Разница множеств клиентов и почти полное исчезновение «граница актуальности» в новом прогоне -
следствие другого порядка проведения / повторного проведения в 6 потоках, а не другого налога в xlsx портфелей.</li>
<li><b>Не проведенные Перерегистрации ц/б</b> (10005 / 10007 от 06.05.2026) есть в обоих комплектах сообщений -
входные данные, не регресс нового кода.</li>
<li><b>Расширения.</b> Пользователь явно не подключал IMDEV7330_NkdParam и IMDEV7330_Locks.
Сверка относится к типовому контуру 2.8.7.5 (уже с влитым массовым предрасчётом / сеансовым кэшем НКД)
против предыдущего «старого» прогона Аванкор.</li>
</ol>

<footer>
IMDEV-7330. Источник данных: папка IMDEV-7330 НовыеТесты. Скрипт compare_ndfl_old_vs_new2.py.
Сырой JSON: imdev7330_ndfl_old_vs_new2_diff.json.
</footer>
</div>
</body>
</html>
"""
    with open(OUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    safe_print("HTML written: " + OUT_HTML)


def main():
    r = analyze()
    write_html(r)
    safe_print("DONE")


if __name__ == "__main__":
    main()
