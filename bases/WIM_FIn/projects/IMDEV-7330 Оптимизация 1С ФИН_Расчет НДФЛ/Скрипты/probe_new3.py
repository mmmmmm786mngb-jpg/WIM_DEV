#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Probe New3 files: log xlsx, word excerpt, xlsx headers."""

from pathlib import Path
from zipfile import ZipFile
import re
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\WIM_FIn\projects\IMDEV-7330 НовыеТесты")
OUT = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\WIM_FIn\projects\IMDEV-7330 Оптимизация 1С ФИН_Расчет НДФЛ\Скрипты\probe_new3_out.txt")


def safe(s):
    return str(s).encode("ascii", "replace").decode("ascii")


lines = []


def log(s):
    lines.append(s)
    print(s.encode("ascii", "replace").decode("ascii"))


# --- log xlsx ---
logp = BASE / "Лог формирования начисдений НДФЛ_ПоНовому3.xlsx"
wb = load_workbook(logp, read_only=True, data_only=True)
log(f"LOG sheets: {wb.sheetnames}")
for sn in wb.sheetnames:
    ws = wb[sn]
    log(f"--- sheet {sn} dims={ws.dimensions} ---")
    n = 0
    for row in ws.iter_rows(values_only=True):
        log(repr(row)[:500])
        n += 1
        if n >= 8:
            break
wb.close()

# --- UK / PF headers ---
for name in [
    "НДФЛ_Управление_27292.xlsx",
    "НДФЛ_Управление_27292_ПоНовому3.xlsx",
    "НДФЛ_Портфели_27292.xlsx",
    "НДФЛ_Портфели_27292_ПоНовому3.xlsx",
]:
    p = BASE / name
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = next(ws.iter_rows(values_only=True))
    log(f"XLSX {name} sheet={wb.sheetnames[0]} header_len={len(hdr)}")
    log(" | ".join(str(c) for c in hdr))
    wb.close()

# --- word ---
docx = BASE / "Тестирование ИТ расширений.docx"
ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
with ZipFile(docx) as z:
    xml = z.read("word/document.xml")
root = ET.fromstring(xml)
paras = []
for p in root.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p"):
    texts = [t.text or "" for t in p.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t")]
    s = "".join(texts).strip()
    if s:
        paras.append(s)
log(f"WORD paras={len(paras)}")
marker = None
for i, s in enumerate(paras):
    if "3 расширения" in s.lower() or "три расширения" in s.lower() or "Запустим" in s:
        marker = i
        log(f"MARKER i={i}: {s[:200]}")
if marker is None:
    for i, s in enumerate(paras):
        if "IMDEV7330_NkdCoupon" in s or "NkdCoupon" in s:
            marker = i
            log(f"COUPON i={i}: {s[:200]}")
start = 0 if marker is None else max(0, marker - 3)
chunk = paras[start:]
log(f"WORD from {start}, {len(chunk)} paras")
for s in chunk:
    log(s)

OUT.write_text("\n".join(lines), encoding="utf-8")
log(f"wrote {OUT}")
