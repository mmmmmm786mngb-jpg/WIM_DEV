#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Сборка KudaVstavitiZamer_DU15.xlsx из HTML 04_kuda_vstaviti_zamer.html по формату КО.xlsx."""

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_DIR = Path(__file__).resolve().parents[1]
HTML_PATH = PROJECT_DIR / "Документация" / "04_kuda_vstaviti_zamer.html"
KO_PATH = PROJECT_DIR / "КО.xlsx"
OUT_PATH = PROJECT_DIR / "Документация" / "KudaVstavitiZamer_DU15.xlsx"


def strip_tags(text: str) -> str:
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(
        r'<div[^>]*class="target-block"[^>]*>',
        "",
        text,
        flags=re.I,
    )
    text = re.sub(
        r'<span[^>]*class="target-label"[^>]*>(.*?)</span>',
        r"\1: ",
        text,
        flags=re.I | re.DOTALL,
    )
    text = re.sub(r"</div>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = text.replace("&mdash;", "-")
    text = text.replace("&gt;", ">")
    text = text.replace("&lt;", "<")
    text = text.replace("&nbsp;", " ")
    text = re.sub(r"\s+\n", "\n", text)
    text = re.sub(r"\n\s+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def cell_period(td_html: str) -> str:
    for pattern, default in (
        (r'tag-daily[^"]*"[^>]*>([^<]+)', "Ежедн."),
        (r'tag-quarterly[^"]*"[^>]*>([^<]+)', None),
        (r'tag-manual[^"]*"[^>]*>([^<]+)', None),
        (r'tag-cfg[^"]*"[^>]*>([^<]+)', None),
        (r'tag-wait[^"]*"[^>]*>([^<]+)', None),
    ):
        match = re.search(pattern, td_html)
        if match:
            return match.group(1).strip() if default is None else default
    return strip_tags(td_html)


def parse_html(html: str) -> list:
    rows_raw = re.findall(r"<tr[^>]*>.*?</tr>", html, re.DOTALL | re.IGNORECASE)
    data_rows = []

    for tr_full in rows_raw:
        if "group-row" in tr_full:
            inner = re.sub(r"^<tr[^>]*>", "", tr_full, flags=re.I)
            inner = re.sub(r"</tr>\s*$", "", inner, flags=re.I)
            group = strip_tags(re.sub(r'<td[^>]*colspan[^>]*>', "", inner))
            data_rows.append(("group", group))
            continue

        if 'class="num"' not in tr_full:
            continue

        inner = re.sub(r"^<tr[^>]*>", "", tr_full, flags=re.I)
        inner = re.sub(r"</tr>\s*$", "", inner, flags=re.I)
        tds = re.findall(r"<td[^>]*>(.*?)</td>", inner, re.DOTALL | re.IGNORECASE)
        if len(tds) < 7:
            continue

        num = strip_tags(tds[0])
        op = strip_tags(tds[1])
        period = cell_period(tds[2])
        obj = strip_tags(tds[3])
        cmd = strip_tags(tds[4])
        rz = strip_tags(tds[5])
        desc = strip_tags(tds[6])
        desc = re.sub(r"https://jira/browse/([A-Z]+-\d+)", r"\1", desc)

        launch_parts = []
        if period:
            launch_parts.append(f"Периодичность: {period}")
        if cmd:
            launch_parts.append(f"Команда: {cmd}")
        if rz:
            launch_parts.append(f"РЗ: {rz}")
        launch = "\n".join(launch_parts)

        data_rows.append(("row", int(num), op, obj, launch, desc))

    return data_rows


def build_workbook(data_rows: list) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Куда вставить замер"

    ko_wb = openpyxl.load_workbook(KO_PATH)
    ko_ws = ko_wb[ko_wb.sheetnames[0]]
    for col in range(1, 6):
        letter = get_column_letter(col)
        if ko_ws.column_dimensions[letter].width:
            ws.column_dimensions[letter].width = ko_ws.column_dimensions[letter].width

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 55

    header_fill = PatternFill("solid", fgColor="17A2B8")
    header_font = Font(bold=True, color="FFFFFF")
    group_fill = PatternFill("solid", fgColor="2C3E50")
    group_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="DEE2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["B1"] = "Куда вставить замер производительности - 20 ключевых операций ДУ 15"
    ws["B1"].font = Font(bold=True, size=14, color="17A2B8")
    ws.merge_cells("B1:E1")
    ws["B2"] = "IMDEV-8927 (анализ), IMDEV-8471 (замеры). Источник: 04_kuda_vstaviti_zamer.html"
    ws.merge_cells("B2:E2")

    headers = ["№", "Ключевая операция", "Объект", "Описание запуска", "Процедуры замеров"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    row_idx = 4
    for item in data_rows:
        if item[0] == "group":
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
            cell = ws.cell(row=row_idx, column=1, value=item[1])
            cell.fill = group_fill
            cell.font = group_font
            cell.alignment = Alignment(vertical="center")
            row_idx += 1
            continue

        _, num, op, obj, launch, desc = item
        values = [num, op, obj, launch, desc]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = wrap
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top")
        row_idx += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:E{row_idx - 1}"
    return wb


def main() -> None:
    html = HTML_PATH.read_text(encoding="utf-8")
    data_rows = parse_html(html)
    wb = build_workbook(data_rows)
    wb.save(OUT_PATH)
    op_count = sum(1 for item in data_rows if item[0] == "row")
    print(f"OK: {op_count} operations, file: {OUT_PATH}")


if __name__ == "__main__":
    main()
