#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""IMDEV-9393: расчёт текущей и прогнозной длительности ключевых операций ДУ 1.5.

Читает выгрузки регистра ЗамерыВремени (ПРОД и разработческая база),
считает по каждой ключевой операции длительность на полном объёме,
строит две оценки роста на 250 000 договоров и определяет предел
по числу договоров. Результат складывает в JSON для сборки отчётов.

Запуск из корня репозитория:
    python "bases\\Wim_Du\\projects\\IMDEV-9393 ...\\Скрипты\\calc_durations.py"
"""

import collections
import datetime
import json
import os
import re
import statistics
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from operations_registry import (  # noqa: E402
    DOGOVOROV_PLAN, DOGOVOROV_SEYCHAS, KRATNOST, OPERATIONS,
)

PROJECT = os.path.join(
    'bases', 'Wim_Du', 'projects',
    'IMDEV-9393 Создать общую таблицу текущей длительности ключевых операций')
PROD_XLSX = os.path.join(
    'bases', 'Wim_Du', 'projects',
    'IMDEV-9391 ПРоверить ключевые операции замеры', 'ЗамерыВремени.xlsx')
DEV_XLSX = os.path.join(PROJECT, 'ЗамерыВечернихРегламентовРазработчика.xlsx')
OUT_JSON = os.path.join(PROJECT, 'Тестирование', 'reports', 'durations_9393.json')

EPOCH = datetime.datetime(1, 1, 1)
# Замеры одного прогона многоэтапной операции разнесены по времени не более
# чем на этот интервал - по нему этапы собираются в один цикл.
CYCLE_GAP = datetime.timedelta(minutes=45)


# --------------------------------------------------------------------------- #
#  Чтение выгрузок
# --------------------------------------------------------------------------- #

def to_dt(value):
    """Дата 1С в миллисекундах от 0001-01-01 -> datetime."""
    try:
        return EPOCH + datetime.timedelta(seconds=float(value) / 1000.0)
    except (TypeError, ValueError, OverflowError):
        return None


def parse_local(value):
    if isinstance(value, datetime.datetime):
        return value
    text = str(value).strip()
    for fmt in ('%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M'):
        try:
            return datetime.datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def dop_inf(comment):
    if not comment:
        return ''
    try:
        return str(json.loads(str(comment)).get('ДопИнф', ''))
    except (ValueError, AttributeError):
        found = re.search(r'"ДопИнф"\s*:\s*"(.*?)"', str(comment))
        return found.group(1) if found else str(comment)[:200]


def load(path):
    """Читает выгрузку регистра в список словарей."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    next(rows)
    records = collections.defaultdict(list)
    for row in rows:
        if not row[0]:
            continue
        try:
            seconds = float(row[4])
        except (TypeError, ValueError):
            continue
        try:
            weight = float(row[5]) if row[5] is not None else 0.0
        except (TypeError, ValueError):
            weight = 0.0
        records[str(row[0]).strip()].append({
            'sec': seconds,
            'weight': weight,
            'start': to_dt(row[1]),
            'finish': to_dt(row[8]),
            'local': parse_local(row[10]),
            'user': str(row[9] or ''),
            'error': str(row[11] or '').strip() == 'Да',
            'inf': dop_inf(row[6]),
        })
    workbook.close()
    return records


# --------------------------------------------------------------------------- #
#  Отбор замеров и агрегация длительности
# --------------------------------------------------------------------------- #

def is_scheduled(record):
    """Регламентный контур - запуск под служебной учётной записью."""
    return record['user'].lower().startswith('svc_')


def select(records, spec):
    """Отбирает замеры операции по контуру, объёму и времени суток."""
    picked = []
    for key in spec['keys']:
        for record in records.get(key, []):
            # Аварийные замеры отражают не длительность работы, а таймаут отказа.
            if record['error']:
                continue
            if spec.get('contour') == 'rz' and not is_scheduled(record):
                continue
            if record['weight'] < spec.get('min_weight', 0):
                continue
            hours = spec.get('hours')
            if hours and record['local'] and not hours[0] <= record['local'].hour <= hours[1]:
                continue
            picked.append(dict(record, key=key))
    return picked


def build_cycles(picked):
    """Собирает этапы одного прогона в цикл по близости во времени."""
    ordered = sorted((r for r in picked if r['local']), key=lambda r: r['local'])
    cycles, current = [], []
    for record in ordered:
        if current and record['local'] - current[-1]['local'] > CYCLE_GAP:
            cycles.append(current)
            current = []
        current.append(record)
    if current:
        cycles.append(current)
    return cycles


def duration_stats(picked, method):
    """Возвращает длительность в секундах и пояснение, как она получена."""
    if not picked:
        return None, None, 0, ''

    if method == 'single':
        values = sorted(r['sec'] for r in picked)
        return (statistics.median(values), values, len(values),
                'медиана по %d замерам' % len(values))

    if method == 'unit':
        # Массового прогона нет: считаем удельную стоимость единицы объёма
        # и разворачиваем её на текущую базу с учётом числа потоков.
        return None, None, len(picked), 'пересчёт от удельной стоимости договора'

    if method == 'cycle':
        cycles = build_cycles(picked)
        per_day = collections.defaultdict(float)
        for cycle in cycles:
            total = sum(r['sec'] for r in cycle)
            day = cycle[0]['local'].date()
            per_day[day] = max(per_day[day], total)
        values = sorted(per_day.values())
        return (statistics.median(values), values, len(values),
                'сумма этапов цикла, максимальный цикл за день, медиана по %d дням' % len(values))

    if method == 'span':
        per_day = collections.defaultdict(list)
        for record in picked:
            if record['local']:
                per_day[record['local'].date()].append(record)
        values = []
        for day_records in per_day.values():
            starts = [r['start'] for r in day_records if r['start']]
            finishes = [r['finish'] for r in day_records if r['finish']]
            if starts and finishes:
                span = (max(finishes) - min(starts)).total_seconds()
                if span > 0:
                    values.append(span)
        values.sort()
        if not values:
            return None, None, 0, ''
        return (statistics.median(values), values, len(values),
                'календарное окно массового прогона за день, медиана по %d дням' % len(values))

    return None, None, 0, ''


def typical_weight(picked):
    weights = [r['weight'] for r in picked if r['weight'] > 0]
    return statistics.median(weights) if weights else None


def start_time(picked, method):
    """Типичное время начала операции: медиана момента старта по замерам."""
    moments = [minute for _day, minute in schedule_moments(picked, method)]
    if not moments:
        return None
    middle = int(statistics.median(sorted(moments)))
    return '%d:%02d' % (middle // 60, middle % 60)


def schedule_moments(picked, method):
    """Минуты от полуночи для старта каждого прогона."""
    moments = []
    source = build_cycles(picked) if method in ('cycle', 'span') else [[r] for r in picked]
    for cycle in source:
        first = min((r for r in cycle if r['local']), key=lambda r: r['local'], default=None)
        if first is None:
            continue
        began = first['local'] - datetime.timedelta(seconds=first['sec'])
        moments.append((began.date(), began.hour * 60 + began.minute))
    return moments


def fmt_hhmm(minutes):
    minutes = int(round(minutes)) % (24 * 60)
    return '%d:%02d' % (minutes // 60, minutes % 60)


def schedule_from_picked(picked, method, override=None):
    """Расписание запусков: первое время, частота, последнее время.

    Параметры:
        picked   - отобранные замеры
        method   - метод агрегации операции
        override - готовая строка из реестра, если задана явно

    Возвращаемое значение:
        Строка вида "22:00, 1 раз, 22:00" или "15:00, 10 мин, 23:50".
    """
    if override:
        return override
    moments = schedule_moments(picked, method)
    if not moments:
        return None

    by_day = collections.defaultdict(list)
    for day, minute in moments:
        by_day[day].append(minute)

    firsts, lasts, gaps = [], [], []
    for day_minutes in by_day.values():
        ordered = sorted(day_minutes)
        firsts.append(ordered[0])
        lasts.append(ordered[-1])
        for left, right in zip(ordered, ordered[1:]):
            delta = right - left
            if delta > 0:
                gaps.append(delta)

    first = fmt_hhmm(statistics.median(firsts))
    last = fmt_hhmm(statistics.median(lasts))
    runs_per_day = statistics.median([len(v) for v in by_day.values()])
    # Один типичный прогон в день - частота "1 раз", даже если в выборке
    # есть отдельные повторные запуски в другие часы.
    if runs_per_day <= 1.5 or not gaps:
        freq = '1 раз'
        last = first
    else:
        gap = int(round(statistics.median(gaps)))
        for step in (5, 10, 15, 30, 60, 120):
            if abs(gap - step) <= step * 0.35:
                gap = step
                break
        freq = '%d мин' % gap
    return '%s, %s, %s' % (first, freq, last)


# --------------------------------------------------------------------------- #
#  Модель роста
# --------------------------------------------------------------------------- #

def cluster_by_volume(points, tolerance=0.25):
    """Группирует замеры по близким значениям объёма.

    Замеры одной и той же операции идут пачками на характерных объёмах
    (единичный договор, тестовая пачка, полная база). Кластеризация позволяет
    отделить эти уровни друг от друга и взять по каждому устойчивую медиану.

    Параметры:
        points    - список пар (объём, длительность в секундах)
        tolerance - относительный разрыв, с которого начинается новый кластер

    Возвращаемое значение:
        Список словарей {weight, sec, count}, упорядоченный по возрастанию объёма.
    """
    clusters = []
    current = []
    for weight, seconds in sorted(points):
        if current and weight > current[-1][0] * (1.0 + tolerance):
            clusters.append(current)
            current = []
        current.append((weight, seconds))
    if current:
        clusters.append(current)
    return [{
        'weight': statistics.median([p[0] for p in group]),
        'sec': statistics.median([p[1] for p in group]),
        'count': len(group),
    } for group in clusters]


def fit_model(records, spec):
    """Модель роста "постоянная часть + предельная стоимость единицы объёма".

    Наклон берётся по двум САМЫМ КРУПНЫМ уровням объёма, а не по всему
    диапазону: на малых объёмах преобладают накладные расходы, и наклон
    по ним завышает прогноз в разы. Нас интересует предельная стоимость
    единицы объёма именно на большом объёме - её и экстраполируем.

    Единица объёма берётся из веса замера и не всегда равна договору
    (бывают сделки, платежи, строки отчёта). Пересчёт такого объёма
    на целевую клиентскую базу делается отдельно - см. target_volume.

    Модель строится только для операций, у которых один замер соответствует
    одному прогону (метод 'single'). У многоэтапных операций длительность - это
    сумма этапов за день, а точки для подгонки - отдельные этапы: величины
    разного масштаба, сопоставлять их нельзя.

    Параметры:
        records - все замеры по ключевым операциям
        spec    - описание операции из реестра

    Возвращаемое значение:
        Словарь с параметрами модели либо Неопределено, если данных не хватает.
    """
    if spec.get('method') != 'single':
        return None
    points = []
    for key in spec['keys']:
        for record in records.get(key, []):
            if record['weight'] > 0 and record['sec'] > 0 and not record['error']:
                points.append((record['weight'], record['sec']))
    if len(points) < 6:
        return None

    clusters = cluster_by_volume(points)
    if len(clusters) < 2:
        return None
    high, low = clusters[-1], clusters[-2]

    # Уровни объёма должны различаться заметно, иначе наклон - это шум.
    if high['weight'] < low['weight'] * 1.5:
        return None
    slope = (high['sec'] - low['sec']) / (high['weight'] - low['weight'])
    if slope <= 0:
        return None

    # Переменная часть должна объяснять существенную долю длительности на
    # верхнем наблюдаемом объёме. Если почти всё время уходит на постоянную
    # работу, наблюдаемый диапазон ничего не говорит о поведении при росте,
    # и экстраполировать такой наклон нельзя.
    if slope * high['weight'] / high['sec'] < 0.2:
        return None

    const = max(low['sec'] - slope * low['weight'], 0.0)
    unit = spec.get('volume_unit', 'ед.')
    return {
        'const_sec': const,
        'per_unit_sec': slope,
        'unit': unit,
        'points': len(points),
        'low': [low['weight'], low['sec'], low['count']],
        'high': [high['weight'], high['sec'], high['count']],
        'note': 'предельная стоимость единицы объёма по двум верхним уровням: '
                '%.0f %s -> %.0f с (n=%d) и %.0f %s -> %.0f с (n=%d), '
                'то есть %.4f с на единицу плюс постоянные %.0f с'
                % (low['weight'], unit, low['sec'], low['count'],
                   high['weight'], unit, high['sec'], high['count'], slope, const),
    }


def unit_cost(picked, unit_is_run=False):
    """Удельная стоимость одного договора.

    Параметры:
        picked      - отобранные замеры операции
        unit_is_run - Истина, если один замер соответствует одному договору
                      (например, построение одного отчёта); иначе объём берётся
                      из веса замера

    Возвращаемое значение:
        Число - секунд на один договор, либо Неопределено при отсутствии данных.
    """
    if unit_is_run:
        values = [r['sec'] for r in picked if r['sec'] > 0]
    else:
        values = [r['sec'] / r['weight'] for r in picked if r['weight'] > 0 and r['sec'] > 0]
    return statistics.median(values) if values else None


def target_volume(row):
    """Объём работы операции при целевой клиентской базе.

    Если вес замера - это договоры, целевой объём известен точно. Для прочих
    единиц (сделки, платежи, строки отчёта) объём считается пропорциональным
    числу договоров и растёт с той же кратностью.

    Параметры:
        row - рассчитанная строка таблицы

    Возвращаемое значение:
        Число единиц объёма либо Неопределено, если типовой объём неизвестен.
    """
    if row.get('volume_unit') == 'договоров':
        return DOGOVOROV_PLAN
    weight = row.get('weight_typical')
    return weight * KRATNOST if weight else None


def contracts_limit(row):
    """При каком числе договоров операция упирается в своё окно выполнения.

    Параметры:
        row - рассчитанная строка таблицы (длительность, модель, окно)

    Возвращаемое значение:
        Число договоров либо Неопределено для операций, не зависящих от базы.
    """
    if not row.get('depends') or not row.get('sec_now'):
        return None
    window_sec = float(row['window_min']) * 60.0
    model = row.get('model')
    if model and model['per_unit_sec'] > 0:
        # Сколько единиц объёма влезает в окно.
        volume = max((window_sec - model['const_sec']) / model['per_unit_sec'], 0.0)
        if row.get('volume_unit') == 'договоров':
            return volume
        weight = row.get('weight_typical')
        if weight:
            return DOGOVOROV_SEYCHAS * volume / weight
    # Без модели - простая пропорция от текущего объёма договоров.
    return max(DOGOVOROV_SEYCHAS * window_sec / row['sec_now'], 0.0)


# --------------------------------------------------------------------------- #
#  Основной расчёт
# --------------------------------------------------------------------------- #

def main():
    prod = load(PROD_XLSX)
    dev = load(DEV_XLSX)
    print('ПРОД: %d ключевых операций, разработческая база: %d' % (len(prod), len(dev)))

    result, by_id = [], {}
    for spec in OPERATIONS:
        if 'group' in spec:
            result.append({'kind': 'group', 'name': spec['group']})
            continue

        records = dev if spec.get('source') == 'dev' else prod
        picked = select(records, spec)
        seconds, values, count, how = duration_stats(picked, spec.get('method'))
        weight = typical_weight(picked)
        model = fit_model(records, spec) if seconds else None
        # Контроль состоятельности: модель, посчитанная на текущем объёме, должна
        # воспроизводить измеренную сейчас длительность. Если расходится в разы,
        # значит отобранные для длительности замеры и точки подгонки описывают
        # разные режимы работы - такой модели доверять нельзя.
        if model and weight:
            predicted = model['const_sec'] + model['per_unit_sec'] * weight
            if not seconds / 2.5 <= predicted <= seconds * 2.5:
                model = None

        # Время запуска можно брать из другого контура (например длительность с
        # разработческой базы, а расписание - с ПРОД, где часы старта достоверны).
        schedule_records = prod if spec.get('schedule_source') == 'prod' else records
        schedule_picked = (
            select(schedule_records, spec)
            if spec.get('schedule_source') == 'prod' else picked)

        row = {
            'kind': 'op',
            'id': spec['id'],
            'name': spec['name'],
            'periodicity': spec['periodicity'],
            'object': spec['object'],
            'source': 'Разработческая база' if spec.get('source') == 'dev' else 'ПРОД',
            'keys': spec['keys'],
            'threads_now': spec['threads_now'],
            'threads_max': spec['threads_max'],
            'threads_note': spec['threads_note'],
            'threads_reserve': spec.get('threads_reserve', ''),
            'depends': spec.get('depends', False),
            'depends_note': spec.get('depends_note', ''),
            'volume_unit': spec.get('volume_unit', ''),
            'window_min': spec['window_min'],
            'window_note': spec.get('window_note', ''),
            'comment': spec.get('comment', ''),
            'estimate': spec.get('estimate', False),
            'no_measurement': spec.get('no_measurement', False),
            'analog_note': spec.get('analog_note', ''),
            'measurements': count,
            'how': how,
            'weight_typical': weight,
            'sec_now': seconds,
            'min_now': seconds / 60.0 if seconds else None,
            'min_max': None,
            'start_time': (start_time(schedule_picked, spec.get('method'))
                           if schedule_picked else None),
            'schedule': schedule_from_picked(
                schedule_picked, spec.get('method'), spec.get('schedule'))
                if schedule_picked or spec.get('schedule') else None,
            'model': model,
            'unit_sec': (unit_cost(picked, spec.get('unit_is_run', False))
                         if spec.get('method') == 'unit' else None),
            'unit_donor': spec.get('unit_donor'),
            'analog_id': spec.get('analog_id'),
        }
        if values:
            row['sec_min'] = min(values)
            row['sec_max'] = max(values)
            row['min_max'] = max(values) / 60.0
        # Замер в код не вставлен: ничего не подставляем, в отчёте будет прочерк.
        if row['no_measurement']:
            row['sec_now'] = None
            row['min_now'] = None
            row['min_max'] = None
            row['how'] = 'замер в код не вставлен'
            row['unit_sec'] = None
            row['unit_donor'] = None
            row['analog_id'] = None
            row['model'] = None
            row['schedule'] = None
        result.append(row)
        by_id[spec['id']] = row

    # Операции без массового прогона: разворачиваем удельную стоимость договора
    # на всю базу с учётом числа потоков. Только если замер в коде есть
    # (свои ключи), но массового прогона в выгрузке нет.
    for row in result:
        if row.get('kind') != 'op' or row.get('sec_now') is not None:
            continue
        if row.get('no_measurement'):
            continue
        unit = row.get('unit_sec')
        if unit is None and row.get('unit_donor'):
            donor = by_id.get(row['unit_donor'])
            unit = donor.get('unit_sec') if donor else None
            if unit is not None:
                row['unit_sec'] = unit
                row['how'] = ('пересчёт от удельной стоимости договора операции "%s"'
                              % donor['name'])
        if unit is None:
            continue
        threads = max(row.get('threads_now') or 1, 1)
        per_contract = unit / threads
        row['sec_now'] = per_contract * DOGOVOROV_SEYCHAS
        row['min_now'] = row['sec_now'] / 60.0
        row['estimate'] = True
        row['model'] = {
            'const_sec': 0.0,
            'per_unit_sec': per_contract,
            'unit': 'договоров',
            'points': row.get('measurements', 0),
            'note': 'удельная стоимость одного договора %.2f с, при %d потоках это '
                    '%.3f с на договор; развёрнуто на всю базу'
                    % (unit, threads, per_contract),
        }
        row['volume_unit'] = 'договоров'

    # Аналогия по классу операций - только если замер в коде есть, но данных нет.
    # Если замера в коде нет (no_measurement) - не подставляем чужие цифры.
    for row in result:
        if row.get('kind') != 'op' or row.get('sec_now') is not None:
            continue
        if row.get('no_measurement'):
            continue
        donor = by_id.get(row.get('analog_id'))
        if not donor or not donor.get('sec_now'):
            continue
        row['sec_now'] = donor['sec_now']
        row['min_now'] = donor['min_now']
        row['estimate'] = True
        row['how'] = 'оценка по аналогии с операцией "%s"' % donor['name']
        row['weight_typical'] = donor.get('weight_typical')
        row['model'] = donor.get('model')
        row['start_time'] = row.get('start_time') or donor.get('start_time')

    # Прогноз на 250 000 договоров: две оценки рядом - линейная x10 по просьбе
    # руководителя и расчётная по предельной стоимости договора.
    # Операции, не зависящие от числа договоров (в т.ч. сделки), не масштабируем.
    for row in result:
        if row.get('kind') != 'op' or not row.get('sec_now'):
            continue
        base = row['sec_now']
        if not row['depends']:
            row['sec_linear'] = base
            row['sec_model'] = base
        else:
            row['sec_linear'] = base * KRATNOST
            model = row.get('model')
            volume = target_volume(row)
            row['sec_model'] = (
                model['const_sec'] + model['per_unit_sec'] * volume
                if model and volume else None)
            row['target_volume'] = volume
        row['min_linear'] = row['sec_linear'] / 60.0 if row['sec_linear'] else None
        row['min_model'] = row['sec_model'] / 60.0 if row['sec_model'] else None
        worst = max(v for v in (row['sec_linear'], row['sec_model']) if v)
        row['hours_worst'] = worst / 3600.0
        row['limit_contracts'] = contracts_limit(row)

    payload = {
        'meta': {
            'built': datetime.datetime.now().strftime('%d.%m.%Y %H:%M'),
            'dogovorov_seychas': DOGOVOROV_SEYCHAS,
            'dogovorov_plan': DOGOVOROV_PLAN,
            'kratnost': KRATNOST,
            'prod_source': PROD_XLSX,
            'dev_source': DEV_XLSX,
            'prod_ops': len(prod),
        },
        'rows': result,
    }
    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    with open(OUT_JSON, 'w', encoding='utf-8') as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=1)

    ops = [r for r in result if r.get('kind') == 'op']
    print('Операций в таблице: %d, с замерами: %d'
          % (len(ops), sum(1 for r in ops if r.get('sec_now'))))
    for row in ops:
        print('  %-58s now=%-9s x10=%-9s model=%-9s limit=%s'
              % (row['name'][:58],
                 _m(row.get('min_now')), _m(row.get('min_linear')),
                 _m(row.get('min_model')),
                 ('%.0f тыс' % (row['limit_contracts'] / 1000)
                  if row.get('limit_contracts') else '-')))
    print('->', OUT_JSON)


def _m(value):
    return '-' if value is None else '%.1f' % value


if __name__ == '__main__':
    main()
