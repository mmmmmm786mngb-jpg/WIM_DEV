#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""IMDEV-9393: интерактивный What-If Excel (вариант от худшего замера).

Файл KeyOperationsDuration_WhatIf_Max.xlsx:
  - колонка "Прогноз расчётный" отсутствует;
  - все масштабирования от "Длительность макс, мин";
  - "Длительность сейчас" серая, только справочно;
  - "Окно, мин" и "БУДЕТ ПОТОКОВ" - жёлтые, редактируемые;
  - "Предел при текущих потоках" - голубая формула:
        Предел_тыс = (база_сейчас_тыс) * Окно / Длительность_макс
    (линейная пропорция; пересчитывается при смене окна).

Запуск из корня репозитория:
    python "bases\\Wim_Du\\projects\\IMDEV-9393 ...\\Скрипты\\build_whatif_xlsx.py"
"""

import json
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from build_report import (  # noqa: E402
    cell_values, verdict,
)
from operations_registry import (  # noqa: E402
    DOGOVOROV_PLAN, DOGOVOROV_SEYCHAS,
)

PROJECT = os.path.join(
    'bases', 'Wim_Du', 'projects',
    'IMDEV-9393 Создать общую таблицу текущей длительности ключевых операций')
IN_JSON = os.path.join(PROJECT, 'Тестирование', 'reports', 'durations_9393.json')
OUT_XLSX = os.path.join(
    PROJECT, 'Документация', 'KeyOperationsDuration_WhatIf_Max.xlsx')

TARGET_K = DOGOVOROV_PLAN / 1000.0  # 250
BASE_K = DOGOVOROV_SEYCHAS / 1000.0  # 25

# Базовые колонки: без прогноза расчётного; предел и x10 - формулами.
BASE_COLUMNS = [
    ('no', '#', 4),
    ('name', 'Ключевая операция', 40),
    ('periodicity', 'Периодичность', 11),
    ('schedule', 'Время запуска', 18),
    ('object', 'Объект', 28),
    ('source', 'Источник', 12),
    ('threads_now', 'Потоков сейчас', 9),
    ('threads_will', 'БУДЕТ ПОТОКОВ', 10),
    ('min_now', 'Длительность сейчас, мин', 11),
    ('min_max', 'Длительность макс, мин', 10),
    ('min_linear', 'Прогноз x10 линейный, мин', 11),
    ('depends', 'Зависит от клиентов', 10),
    ('limit', 'Предел при текущих потоках, тыс.', 12),
    ('window', 'Окно, мин', 8),
]

# Расчётные колонки: формулы Excel, голубая подсветка.
CALC_COLUMNS = [
    ('quota', 'РАСЧЕТНОЕ КВО, тыс. договоров', 12),
    ('dur_will', 'Длительность при БУДЕТ ПОТОКОВ, мин', 12),
    ('speedup', 'Ускорение, раз', 9),
    ('slack', 'Запас до окна, мин', 10),
    ('ok250', 'Выдержит 250 тыс.?', 11),
    ('need250', 'Нужно потоков для 250 тыс.', 11),
]

COMMENT_COL = ('comment', 'Комментарий', 45)

# Предел и x10 тоже голубые (формулы от макс / окна).
FORMULA_BASE_KEYS = {'min_linear', 'limit'}

FILL = {
    'crit': PatternFill('solid', fgColor='FDECEA'),
    'warn': PatternFill('solid', fgColor='FFF8E1'),
    'ok': PatternFill('solid', fgColor='EAF7EE'),
    'none': PatternFill('solid', fgColor='E9ECEF'),
    'pending': PatternFill('solid', fgColor='FFE0B2'),
    'grp': PatternFill('solid', fgColor='D6E2F0'),
    'head': PatternFill('solid', fgColor='17365D'),
    'head_will': PatternFill('solid', fgColor='EF6C00'),
    'head_calc': PatternFill('solid', fgColor='0277BD'),
    'will': PatternFill('solid', fgColor='FFEB3B'),
    'will_lock': PatternFill('solid', fgColor='FFF9C4'),
    'calc': PatternFill('solid', fgColor='B3E5FC'),
    'ref': PatternFill('solid', fgColor='BDBDBD'),
}
THIN = Side(style='thin', color='C8D0DA')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def scales_with_threads(row):
    """Можно ли менять потоки и ждать пропорционального эффекта.

    Параметры:
        row - строка операции из JSON

    Возвращаемое значение:
        Булево - Истина, если есть пул потоков или запас в коде.
    """
    if row.get('no_measurement') or row.get('no_forecast'):
        return False
    if not row.get('depends'):
        return False
    now = int(row.get('threads_now') or 1)
    maximum = int(row.get('threads_max') or now)
    reserve = bool(row.get('threads_reserve'))
    return now > 1 or maximum > now or reserve


def has_limit_inputs(row, min_max_value, window_value):
    """Есть ли данные для линейного предела от макс-длительности."""
    if row.get('no_measurement') or row.get('no_forecast'):
        return False
    if not row.get('depends'):
        return False
    if not isinstance(min_max_value, (int, float)) or min_max_value <= 0:
        return False
    if not isinstance(window_value, (int, float)) or window_value <= 0:
        return False
    return True


def to_number(value, numeric_keys, key):
    """Приводит значение ячейки к числу, где это уместно."""
    if key not in numeric_keys:
        return value
    if value in ('-', 'нет замеров', 'не зависит', 'нужен замер', '> 1 000', 'от сделок'):
        return None if key == 'limit' else value
    if isinstance(value, str) and 'дог' in value:
        try:
            return float(value.split()[0].replace(',', '.'))
        except ValueError:
            return value
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def build():
    with open(IN_JSON, encoding='utf-8') as handle:
        payload = json.load(handle)

    rows = payload['rows']
    meta = payload['meta']
    all_cols = BASE_COLUMNS + CALC_COLUMNS + [COMMENT_COL]
    col_index = {key: index for index, (key, _t, _w) in enumerate(all_cols, start=1)}

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = 'Сценарий по потокам (макс)'

    total_cols = len(all_cols)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    sheet['A1'] = (
        'IMDEV-9393. What-If от ХУДШЕГО замера (Длительность макс). '
        'Жёлтые: "БУДЕТ ПОТОКОВ", "Окно, мин" и "Длительность макс". Голубые - формулы. '
        'Серая "Длительность сейчас" - справочно. Цель: %.0f тыс. договоров. '
        'Сформировано %s.'
        % (TARGET_K, meta['built']))
    sheet['A1'].font = Font(bold=True, size=12, color='17365D')

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    sheet['A2'] = (
        'Формулы: '
        'Предел_тыс = база_тыс × Окно / Длительность_макс '
        '(база обычно %.0f; для загрузки/исполнения сделок - сделки на худшем '
        'замере/1000, допущение база x10 -> сделки x10, 1 сделка = 1 договор); '
        'Прогноз x10 = Длительность_макс × 10; '
        'КВО = Предел × (БУДЕТ ПОТОКОВ / Потоков сейчас); '
        'Длительность при БУДЕТ = Длительность_макс × (Потоков сейчас / БУДЕТ); '
        'Ускорение = БУДЕТ / Сейчас; '
        'Запас = Окно − Длительность при БУДЕТ; '
        'Выдержит 250 тыс. = КВО >= 250; '
        'Нужно потоков = CEILING(Потоков сейчас × 250 / Предел). '
        'Смена окна или длительности макс сразу пересчитывает предел и связанные колонки.'
        % BASE_K)
    sheet['A2'].font = Font(size=9, color='455A64')
    sheet['A2'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 56

    header = 4
    for index, (key, title, width) in enumerate(all_cols, start=1):
        cell = sheet.cell(row=header, column=index, value=title)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(wrap_text=True, vertical='bottom')
        cell.border = BORDER
        if key in ('threads_will', 'window', 'min_max'):
            cell.fill = FILL['head_will']
        elif key in FORMULA_BASE_KEYS or key in {k for k, _t, _w in CALC_COLUMNS}:
            cell.fill = FILL['head_calc']
        elif key == 'min_now':
            cell.fill = PatternFill('solid', fgColor='616161')
        else:
            cell.fill = FILL['head']
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = sheet.cell(row=header + 1, column=3)
    sheet.row_dimensions[header].height = 36

    numeric_base = {
        'threads_now', 'min_now', 'min_max', 'window',
    }

    line = header
    number = 0
    data_rows = []

    for row in rows:
        line += 1
        if row.get('kind') == 'group':
            cell = sheet.cell(row=line, column=1, value=row['name'])
            cell.font = Font(bold=True, color='17365D', size=11)
            sheet.merge_cells(start_row=line, start_column=1,
                              end_row=line, end_column=total_cols)
            for index in range(1, total_cols + 1):
                sheet.cell(row=line, column=index).fill = FILL['grp']
                sheet.cell(row=line, column=index).border = BORDER
            continue

        number += 1
        mark = verdict(row)
        scales = scales_with_threads(row)
        values = cell_values(row, number)
        # Original cell_values: no,name,periodicity,schedule,object,source,
        # threads_now,threads_max,min_now,min_max,min_linear,min_model,depends,
        # limit,window,comment
        mapping = {
            'no': values[0],
            'name': values[1],
            'periodicity': values[2],
            'schedule': values[3],
            'object': values[4],
            'source': values[5],
            'threads_now': values[6],
            'threads_will': int(row.get('threads_now') or 1),
            'min_now': values[8],
            'min_max': values[9],
            'depends': values[12],
            'window': values[14],
            'comment': values[15],
        }
        if row.get('real_name'):
            mapping['name'] = '%s\nреальное название: %s' % (row['name'], row['real_name'])

        min_max_num = to_number(mapping['min_max'], {'min_max'}, 'min_max')
        window_num = to_number(mapping['window'], {'window'}, 'window')
        can_limit = has_limit_inputs(row, min_max_num, window_num)

        # База для линейного предела, тыс.:
        # - при volume_as_contracts: объём на худшем замере (1 ед. = 1 договор);
        # - иначе текущая клиентская база 25 тыс.
        if (row.get('volume_as_contracts') and row.get('weight_at_max')
                and float(row['weight_at_max']) > 0):
            base_k_row = float(row['weight_at_max']) / 1000.0
        else:
            base_k_row = BASE_K

        r = line
        c_now = get_column_letter(col_index['threads_now'])
        c_will = get_column_letter(col_index['threads_will'])
        c_max = get_column_letter(col_index['min_max'])
        c_lim = get_column_letter(col_index['limit'])
        c_win = get_column_letter(col_index['window'])
        c_dur_will = get_column_letter(col_index['dur_will'])
        c_quota = get_column_letter(col_index['quota'])

        for key, _title, _w in BASE_COLUMNS:
            if key == 'min_linear':
                if isinstance(min_max_num, (int, float)):
                    value = '=IF(ISNUMBER({m}{r}),{m}{r}*10,"-")'.format(m=c_max, r=r)
                else:
                    value = '-'
            elif key == 'limit':
                if can_limit:
                    # Линейный предел: база_тыс * окно / длительность_макс
                    value = (
                        '=IF(AND(ISNUMBER({m}{r}),ISNUMBER({w}{r}),{m}{r}>0),'
                        '{base}*{w}{r}/{m}{r},"-")'
                        .format(m=c_max, w=c_win, r=r, base=base_k_row)
                    )
                else:
                    value = '-'
            else:
                value = mapping[key]
                if key in numeric_base:
                    value = to_number(value, numeric_base, key)

            cell = sheet.cell(row=line, column=col_index[key], value=value)
            cell.border = BORDER
            cell.alignment = Alignment(
                wrap_text=key in ('name', 'object', 'schedule'),
                vertical='top',
                horizontal='right' if key in (
                    'no', 'threads_now', 'threads_will', 'min_now', 'min_max',
                    'min_linear', 'limit', 'window') else 'left')

            if key == 'threads_will':
                cell.fill = FILL['will'] if scales else FILL['will_lock']
                cell.font = Font(bold=True, size=11, color='E65100')
                cell.number_format = '0'
            elif key in ('window', 'min_max'):
                cell.fill = FILL['will']
                cell.font = Font(bold=True, size=11, color='E65100')
                if key == 'window' and isinstance(value, (int, float)):
                    cell.number_format = '0'
                if key == 'min_max' and isinstance(value, float):
                    cell.number_format = '0.0'
            elif key == 'min_now':
                cell.fill = FILL['ref']
                cell.font = Font(size=10, color='424242', italic=True)
                if isinstance(value, float):
                    cell.number_format = '0.0'
            elif key in FORMULA_BASE_KEYS:
                cell.fill = FILL['calc']
                cell.font = Font(size=10, color='01579B')
                cell.number_format = '0.0'
            elif key == 'name':
                cell.fill = FILL[mark]
                cell.font = Font(bold=True, size=10)
            else:
                cell.fill = FILL[mark]
                cell.font = Font(size=10)

        ok_scale = (
            'AND(ISNUMBER({lim}{r}),ISNUMBER({now}{r}),ISNUMBER({will}{r}),'
            '{now}{r}>0,{will}{r}>0)'
            .format(lim=c_lim, now=c_now, will=c_will, r=r))

        # Длительность и масштаб - от Длительность макс.
        formulas = {
            'quota': (
                '=IF({ok},{lim}{r}*{will}{r}/{now}{r},IF(ISNUMBER({lim}{r}),{lim}{r},"-"))'
                if scales else
                '=IF(ISNUMBER({lim}{r}),{lim}{r},"-")'
            ).format(ok=ok_scale, lim=c_lim, will=c_will, now=c_now, r=r),
            'dur_will': (
                '=IF({ok},{mx}{r}*{now}{r}/{will}{r},IF(ISNUMBER({mx}{r}),{mx}{r},"-"))'
                if scales else
                '=IF(ISNUMBER({mx}{r}),{mx}{r},"-")'
            ).format(ok=ok_scale, mx=c_max, now=c_now, will=c_will, r=r),
            'speedup': (
                '=IF({ok},{will}{r}/{now}{r},1)'
                if scales else
                '=1'
            ).format(ok=ok_scale, will=c_will, now=c_now, r=r),
            'slack': (
                '=IF(AND(ISNUMBER({win}{r}),ISNUMBER({dw}{r})),{win}{r}-{dw}{r},"-")'
                .format(win=c_win, dw=c_dur_will, r=r)
            ),
            'ok250': (
                '=IF(ISNUMBER({q}{r}),IF({q}{r}>={tgt},"Да","Нет"),"-")'
                .format(q=c_quota, r=r, tgt=TARGET_K)
            ),
            'need250': (
                '=IF({ok},ROUNDUP({now}{r}*{tgt}/{lim}{r},0),"-")'
                if scales else
                '="-"'
            ).format(ok=ok_scale, now=c_now, lim=c_lim, r=r, tgt=TARGET_K),
        }

        for key in ['quota', 'dur_will', 'speedup', 'slack', 'ok250', 'need250']:
            cell = sheet.cell(row=line, column=col_index[key], value=formulas[key])
            cell.fill = FILL['calc']
            cell.border = BORDER
            cell.font = Font(size=10, color='01579B')
            cell.alignment = Alignment(horizontal='right', vertical='top')
            if key in ('quota', 'dur_will', 'speedup', 'slack'):
                cell.number_format = '0.0'

        comment_cell = sheet.cell(
            row=line, column=col_index['comment'], value=mapping['comment'])
        comment_cell.fill = FILL[mark]
        comment_cell.border = BORDER
        comment_cell.alignment = Alignment(wrap_text=True, vertical='top')
        comment_cell.font = Font(size=9)

        data_rows.append((line, scales, can_limit))

    last_data = line
    sheet.auto_filter.ref = 'A%d:%s%d' % (
        header, get_column_letter(total_cols), last_data)

    ok250_letter = get_column_letter(col_index['ok250'])
    ok_range = '%s%d:%s%d' % (ok250_letter, header + 1, ok250_letter, last_data)
    sheet.conditional_formatting.add(
        ok_range,
        CellIsRule(operator='equal', formula=['"Нет"'],
                   fill=PatternFill('solid', fgColor='FFCDD2'),
                   font=Font(color='B71C1C', bold=True)))
    sheet.conditional_formatting.add(
        ok_range,
        CellIsRule(operator='equal', formula=['"Да"'],
                   fill=PatternFill('solid', fgColor='C8E6C9'),
                   font=Font(color='1B5E20', bold=True)))

    # ------------------------------------------------------------------ Легенда
    legend = book.create_sheet('Как пользоваться')
    legend.column_dimensions['A'].width = 28
    legend.column_dimensions['B'].width = 110
    blocks = [
        ('Вариант Max',
         'Этот файл считает сценарий от ХУДШЕГО замера (колонка "Длительность макс"). '
         'Исходный KeyOperationsDuration_WhatIf.xlsx не меняется.'),
        ('Жёлтые колонки',
         'Редактируйте "БУДЕТ ПОТОКОВ", "Окно, мин" и "Длительность макс". '
         'Смена окна или макс-длительности сразу пересчитывает предел и все голубые колонки.'),
        ('Серая колонка',
         '"Длительность сейчас" - справочно (медиана). В расчёты What-If не входит.'),
        ('Предел при текущих потоках, тыс.',
         'Формула: база_тыс × Окно / Длительность_макс. '
         'Обычные операции: база = %.0f тыс. договоров. '
         'Загрузка и исполнение сделок: база = сделки на худшем замере / 1000 '
         '(допущение: база x10 -> сделки x10, 1 сделка = 1 договор).'
         % BASE_K),
        ('Прогноз x10 линейный',
         'Длительность_макс × 10.'),
        ('РАСЧЕТНОЕ КВО',
         'Предел × (БУДЕТ ПОТОКОВ / Потоков сейчас).'),
        ('Длительность при БУДЕТ ПОТОКОВ',
         'Длительность_макс × (Потоков сейчас / БУДЕТ ПОТОКОВ).'),
        ('Выдержит 250 тыс.?',
         'Да, если РАСЧЕТНОЕ КВО >= 250.'),
        ('Важно',
         'Оценочная модель для обсуждения. Реальное ускорение от потоков нужно '
         'подтверждать замером.'),
    ]
    legend['A1'] = 'Как пользоваться (вариант от макс-длительности)'
    legend['A1'].font = Font(bold=True, size=14, color='0277BD')
    line = 2
    for title, text in blocks:
        line += 1
        a = legend.cell(row=line, column=1, value=title)
        a.font = Font(bold=True, size=10, color='01579B')
        a.fill = FILL['calc']
        a.alignment = Alignment(vertical='top', wrap_text=True)
        b = legend.cell(row=line, column=2, value=text)
        b.font = Font(size=10)
        b.alignment = Alignment(wrap_text=True, vertical='top')
        legend.row_dimensions[line].height = 48

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    book.save(OUT_XLSX)
    print('What-if Max XLSX ->', OUT_XLSX)
    print('Data rows:', len(data_rows),
          'scalable:', sum(1 for _r, s, _h in data_rows if s),
          'with limit formula:', sum(1 for _r, _s, h in data_rows if h))


if __name__ == '__main__':
    build()
