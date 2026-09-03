#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""IMDEV-9393: сборка итоговой таблицы длительности ключевых операций ДУ 1.5.

Читает результат расчёта (durations_9393.json) и формирует два документа
в папке Документация проекта:
    KeyOperationsDuration.html - отчёт с методикой, подсветкой и пояснениями;
    KeyOperationsDuration.xlsx - та же таблица для вставки в переписку.

Запуск из корня репозитория:
    python "bases\\Wim_Du\\projects\\IMDEV-9393 ...\\Скрипты\\build_report.py"
"""

import json
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT = os.path.join(
    'bases', 'Wim_Du', 'projects',
    'IMDEV-9393 Создать общую таблицу текущей длительности ключевых операций')
IN_JSON = os.path.join(PROJECT, 'Тестирование', 'reports', 'durations_9393.json')
DOCS = os.path.join(PROJECT, 'Документация')
OUT_HTML = os.path.join(DOCS, 'KeyOperationsDuration.html')
OUT_XLSX = os.path.join(DOCS, 'KeyOperationsDuration.xlsx')

# Порог, ниже которого предел по договорам считается критичным: система
# должна выдержать целевые 250 000 договоров.
LIMIT_CRIT = 250000
LIMIT_WARN = 500000
# Предел выше этого значения показываем как "не ограничивает" - точность
# расчёта на таких величинах не имеет практического смысла.
LIMIT_HUGE = 1000000

COLUMNS = [
    ('no', '#', 4),
    ('name', 'Ключевая операция', 42),
    ('periodicity', 'Периодичность', 12),
    ('schedule', 'Время запуска (первое, частота, последнее)', 22),
    ('object', 'Объект (обработка / регламент)', 32),
    ('source', 'Источник замера', 14),
    ('threads_now', 'Потоков сейчас', 8),
    ('threads_max', 'Потоков макс', 8),
    ('min_now', 'Длительность сейчас, мин', 11),
    ('min_max', 'Длительность макс, мин', 11),
    ('min_linear', 'Прогноз x10 линейный, мин', 11),
    ('min_model', 'Прогноз расчётный, мин', 11),
    ('depends', 'Зависит от числа клиентов', 11),
    ('limit', 'Предел, тыс. договоров', 11),
    ('window', 'Окно, мин', 8),
    ('comment', 'Комментарий', 55),
]


# --------------------------------------------------------------------------- #
#  Подготовка значений
# --------------------------------------------------------------------------- #

def verdict(row):
    """Категория критичности операции.

    Параметры:
        row - строка таблицы из результата расчёта

    Возвращаемое значение:
        Строка: 'crit', 'warn', 'ok', 'none' (нет замера в коде)
        либо 'pending' (нет квартальной выгрузки - прогноз не публикуем).
    """
    if row.get('no_measurement') and row.get('sec_now') is None:
        return 'none'
    if row.get('no_forecast'):
        return 'pending'
    if not row.get('depends'):
        return 'ok'
    limit = row.get('limit_contracts')
    if limit is None:
        return 'warn'
    if limit < LIMIT_CRIT:
        return 'crit'
    if limit < LIMIT_WARN:
        return 'warn'
    return 'ok'


def fmt_min(value):
    if value is None:
        return '-'
    if value < 1:
        return '%.2f' % value
    if value < 100:
        return '%.1f' % value
    return '%.0f' % value


def fmt_limit(row):
    """Предел по числу договоров в тысячах, с отсечкой заведомо больших значений."""
    if row.get('no_measurement') or row.get('no_forecast'):
        return '-'
    if not row.get('depends'):
        return 'не зависит'
    limit = row.get('limit_contracts')
    if limit is None:
        return 'нужен замер'
    if limit >= LIMIT_HUGE:
        return '> 1 000'
    return '%.0f' % (limit / 1000.0)


def fmt_depends(row):
    if row.get('no_measurement') and row.get('sec_now') is None:
        return 'Да' if row.get('depends') else 'Нет'
    if not row.get('depends'):
        return 'Нет'
    if row.get('volume_unit') == 'сделок':
        return 'Да (через сделки)'
    return 'Да'


def cell_values(row, number):
    """Готовит список значений строки в порядке COLUMNS."""
    name = row['name']
    if row.get('no_measurement') and row.get('sec_now') is None:
        dash = 'нет замеров'
        return [
            number,
            name,
            row['periodicity'],
            '-',
            row['object'],
            row['source'],
            row['threads_now'],
            row['threads_max'],
            dash,
            dash,
            dash,
            dash,
            fmt_depends(row),
            '-',
            '%.0f' % row['window_min'] if row.get('window_min') is not None else '-',
            row.get('comment', ''),
        ]
    if row.get('no_forecast'):
        contracts = row.get('weight_typical')
        now_text = fmt_min(row.get('min_now'))
        if row.get('min_now') is not None and contracts:
            now_text = '%s (ср. %.0f дог./день)' % (now_text, contracts)
        max_text = fmt_min(row.get('min_max'))
        return [
            number,
            name,
            row['periodicity'],
            row.get('schedule') or '-',
            row['object'],
            row['source'],
            row['threads_now'],
            row['threads_max'],
            now_text,
            max_text,
            '-',
            '-',
            fmt_depends(row),
            '-',
            '%.0f' % row['window_min'] if row.get('window_min') is not None else '-',
            row.get('comment', ''),
        ]
    return [
        number,
        name,
        row['periodicity'],
        row.get('schedule') or '-',
        row['object'],
        row['source'],
        row['threads_now'],
        row['threads_max'],
        fmt_min(row.get('min_now')),
        fmt_min(row.get('min_max')),
        fmt_min(row.get('min_linear')),
        fmt_min(row.get('min_model')),
        fmt_depends(row),
        fmt_limit(row),
        '%.0f' % row['window_min'],
        row.get('comment', ''),
    ]


def methodics(row):
    """Строка пояснения: как получена цифра и на чём основана модель."""
    parts = []
    if row.get('how'):
        parts.append('Длительность: %s.' % row['how'])
    if row.get('measurements'):
        parts.append('Замеров в выгрузке: %d.' % row['measurements'])
    if row.get('weight_typical'):
        parts.append('Типовой объём замера: %.0f %s.'
                     % (row['weight_typical'], row.get('volume_unit', 'ед.')))
    model = row.get('model')
    if model and model.get('note'):
        parts.append('Модель роста: %s.' % model['note'])
    if row.get('depends_note'):
        parts.append('Зависимость от базы: %s' % row['depends_note'])
    if row.get('threads_note'):
        parts.append('Потоки: %s' % row['threads_note'])
    if row.get('threads_reserve'):
        parts.append('Резерв по потокам: %s' % row['threads_reserve'])
    if row.get('window_note'):
        parts.append('Окно: %s' % row['window_note'])
    if row.get('analog_note'):
        parts.append(row['analog_note'])
    if row.get('keys'):
        parts.append('Ключевые операции регистра: %s.' % '; '.join(row['keys']))
    return ' '.join(parts)


# --------------------------------------------------------------------------- #
#  HTML
# --------------------------------------------------------------------------- #

HTML_HEAD = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>IMDEV-9393. Длительность ключевых операций Розничного ДУ</title>
<style>
 body {font-family: "Segoe UI", Tahoma, Arial, sans-serif; margin: 0; padding: 0 0 60px;
       background: #f4f6f8; color: #222; font-size: 14px;}
 .wrap {max-width: 1900px; margin: 0 auto; padding: 0 24px;}
 header {background: linear-gradient(135deg, #17365d 0%%, #2e6da4 100%%); color: #fff;
         padding: 32px 24px; margin-bottom: 24px;}
 header h1 {margin: 0 0 8px; font-size: 26px;}
 header .sub {opacity: .85; font-size: 15px;}
 h2 {color: #17365d; border-bottom: 2px solid #2e6da4; padding-bottom: 6px; margin-top: 34px;}
 h3 {color: #2e6da4; margin-bottom: 6px;}
 .cards {display: flex; flex-wrap: wrap; gap: 14px; margin: 18px 0;}
 .card {background: #fff; border-radius: 6px; padding: 16px 20px; min-width: 170px;
        box-shadow: 0 1px 3px rgba(0,0,0,.12); border-left: 4px solid #17a2b8;}
 .card.crit {border-left-color: #dc3545;}
 .card.warn {border-left-color: #ffc107;}
 .card.ok {border-left-color: #28a745;}
 .card .val {font-size: 28px; font-weight: 600; color: #17365d;}
 .card .cap {font-size: 12px; color: #666; text-transform: uppercase; letter-spacing: .4px;}
 .box {background: #fff; border-radius: 6px; padding: 18px 22px; margin: 16px 0;
       box-shadow: 0 1px 3px rgba(0,0,0,.10);}
 .box.info {background: #e7f6f8; border-left: 4px solid #17a2b8;}
 .box.warn {background: #fff8e1; border-left: 4px solid #ffc107;}
 .box.blue {background: #e8f0fb; border-left: 4px solid #2e6da4;}
 .box.viol {background: #f3e9fa; border-left: 4px solid #7b4fa8;}
 table {border-collapse: collapse; width: 100%%; background: #fff; font-size: 12.5px;
        box-shadow: 0 1px 3px rgba(0,0,0,.10);}
 th {background: #17365d; color: #fff; padding: 9px 7px; text-align: left;
     font-weight: 600; font-size: 11.5px; vertical-align: bottom; position: sticky; top: 0;}
 td {padding: 7px; border-bottom: 1px solid #e3e7ec; vertical-align: top;}
 tr.grp td {background: #d6e2f0; font-weight: 700; color: #17365d; font-size: 13px;
            text-transform: uppercase; letter-spacing: .3px;}
 tr.crit td.mark {background: #fdecea; border-left: 3px solid #dc3545;}
 tr.warn td.mark {background: #fff8e1; border-left: 3px solid #ffc107;}
 tr.ok   td.mark {background: #eaf7ee; border-left: 3px solid #28a745;}
 tr.pending td {background: #ffe0b2;}
 tr.pending td.mark {background: #ffcc80; border-left: 3px solid #ef6c00;}
 td.num {text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums;}
 td.name {font-weight: 600;}
 .real-name {display: block; font-weight: 400; font-size: 11px; color: #5c6b7a;
             margin-top: 2px;}
 .tag {display: inline-block; padding: 1px 7px; border-radius: 9px; font-size: 10.5px;
       font-weight: 600; white-space: nowrap;}
 .tag.est {background: #fff3cd; color: #856404; border: 1px solid #ffe08a;}
 .tag.dev {background: #f3e9fa; color: #5c3480; border: 1px solid #d9c2ec;}
 .tag.nom {background: #e9ecef; color: #495057; border: 1px solid #ced4da;}
 .tag.pending {background: #ef6c00; color: #fff;}
 .tag.crit {background: #dc3545; color: #fff;}
 .tag.warn {background: #ffc107; color: #4d3b00;}
 .tag.ok {background: #28a745; color: #fff;}
 tr.none td.mark {background: #e9ecef; border-left: 3px solid #6c757d;}
 tr.none td {color: #6c757d;}
 pre {background: #f7f9fb; border: 1px solid #dde3ea; border-radius: 4px;
      padding: 10px 12px; overflow-x: auto; font-size: 12px; margin: 8px 0;
      white-space: pre-wrap; font-family: Consolas, "Courier New", monospace;}
 code {background: #eef2f6; padding: 1px 5px; border-radius: 3px;
       font-family: Consolas, "Courier New", monospace; font-size: 12.5px;}
 ul {margin: 8px 0 8px 18px; padding: 0;}
 li {margin: 4px 0;}
 .det {font-size: 12px; color: #555; line-height: 1.55;}
 .det b {color: #17365d;}
 footer {text-align: center; color: #888; font-size: 12px; margin-top: 36px;}
</style>
</head>
<body>
<header>
 <div class="wrap">
  <h1>Текущая и прогнозная длительность ключевых операций Розничного ДУ (ДУ 1.5)</h1>
  <div class="sub">IMDEV-9393 &middot; сформировано %(built)s &middot;
   сценарий роста: %(now_k)s тыс. &rarr; %(plan_k)s тыс. договоров (x%(kratnost)s)</div>
 </div>
</header>
<div class="wrap">
"""


def html_escape(text):
    return (str(text).replace('&', '&amp;').replace('<', '&lt;')
            .replace('>', '&gt;').replace('"', '&quot;'))


def build_html(payload):
    meta = payload['meta']
    rows = payload['rows']
    ops = [r for r in rows if r.get('kind') == 'op']

    crit = [r for r in ops if verdict(r) == 'crit']
    warn = [r for r in ops if verdict(r) == 'warn']
    independent = [r for r in ops if not r.get('depends') and not r.get('no_measurement')]
    estimates = [r for r in ops if r.get('estimate') and not r.get('no_measurement')]
    no_meas = [r for r in ops if r.get('no_measurement')]

    out = [HTML_HEAD % {
        'built': html_escape(meta['built']),
        'now_k': '%.0f' % (meta['dogovorov_seychas'] / 1000.0),
        'plan_k': '%.0f' % (meta['dogovorov_plan'] / 1000.0),
        'kratnost': '%.0f' % meta['kratnost'],
    }]

    # ---------------------------------------------------------------- Введение
    out.append("""
<h2>1. Задача и что сделано</h2>
<div class="box">
 <p>Технический архитектор запросил единую таблицу по образцу таблицы МО Розница: по каждой
 ключевой операции Розничного ДУ показать текущую длительность, число потоков сейчас
 и максимально возможное, прогноз длительности при росте клиентской базы в 10 раз
 и предельное число договоров, которое операция выдерживает в своём окне.</p>
 <p>В таблице %(ops)d операций: 17 из списка технического архитектора и %(extra)d операций,
 выявленных дополнительно при разборе выгрузки замеров и кода.</p>
</div>
<div class="box blue">
 <b>Источники данных</b>
 <ul>
  <li><b>ПРОД, регистр ЗамерыВремени</b> &mdash; выгрузка по задаче IMDEV-9391.
      В выгрузке %(prod_ops)d различных ключевых операций.</li>
  <li><b>Разработческая база</b> &mdash; замеры вечерних регламентов после оптимизации.
      Используется только для вечерних регламентов: там уже стоит оптимизация,
      которая скоро уйдёт на ПРОД.</li>
  <li><b>IMDEV-9153 (DR)</b> &mdash; замер полного объёма начисления вознаграждения
      (параллельная версия).</li>
  <li><b>Исходный код конфигурации Wim_Du</b> &mdash; из него взято фактическое
      и предельное число потоков по каждому объекту.</li>
 </ul>
</div>
""" % {
        'ops': len(ops),
        'extra': len(ops) - 17,
        'prod_ops': meta['prod_ops'],
    })

    # --------------------------------------------------------------- Методика
    out.append("""
<h2>2. Методика расчёта</h2>

<h3>2.1. Как определялась текущая длительность</h3>
<div class="box">
 <p>Регистр ЗамерыВремени содержит замеры и с полного объёма базы, и с единичных
 договоров (отладка, ручные перезапуски). Прямая медиана по всем замерам даёт
 заниженную цифру, поэтому по каждой операции применялись отборы:</p>
 <ul>
  <li><b>Контур запуска.</b> Для регламентных операций брались только запуски под
      служебной учётной записью (<code>svc_*</code>) &mdash; это и есть работа по расписанию.</li>
  <li><b>Минимальный объём.</b> Для регламентов отбирались замеры с весом от 15 000
      договоров, то есть прогоны по всей базе, а не по тестовой пачке.</li>
  <li><b>Время суток.</b> Для двух сеансов загрузки котировок (утренний и финальный)
      замеры разделялись по часу запуска.</li>
  <li><b>Аварийные замеры исключены</b> &mdash; они показывают не длительность работы,
      а время до отказа (например, таймаут веб-сервиса).</li>
 </ul>
 <p>Дальше длительность считалась одним из четырёх способов:</p>
 <ul>
  <li><b>Медиана</b> &mdash; для операций, пишущих один замер за прогон.</li>
  <li><b>Сумма этапов цикла</b> &mdash; для многоэтапных операций (получение из Tibco,
      разбор, создание документов). Этапы одного прогона собираются по близости во
      времени (разрыв не более 45 минут), внутри дня берётся самый тяжёлый цикл,
      затем медиана по дням.</li>
  <li><b>Календарное окно</b> &mdash; от первого старта до последнего завершения за день.</li>
  <li><b>Пересчёт от удельной стоимости договора</b> &mdash; для операций, у которых
      массового прогона в выгрузке нет вообще. Берётся стоимость обработки одного
      договора и разворачивается на всю базу с учётом числа потоков.</li>
 </ul>
</div>

<h3>2.2. Две оценки прогноза рядом</h3>
<div class="box info">
 <p>В таблице сознательно две колонки прогноза, потому что они отвечают на разные вопросы.</p>
 <ul>
  <li><b>Прогноз x10 линейный</b> &mdash; текущая длительность, умноженная на 10, как
      просил технический архитектор. Простая и проверяемая цифра.</li>
  <li><b>Прогноз расчётный</b> &mdash; линейная модель по фактическим замерам:
      <code>Длительность = ПостояннаяЧасть + СтоимостьДоговора * ЧислоДоговоров</code>.
      Наклон берётся по <b>двум самым крупным уровням объёма</b>, а не по всему диапазону:
      на малых объёмах преобладают накладные расходы, и наклон по ним завышает прогноз
      в разы. Нас интересует стоимость договора именно на большом объёме.</li>
 </ul>
 <p>Расхождение двух оценок &mdash; это не ошибка, а полезный сигнал. Если расчётный
 прогноз ниже линейного, операция масштабируется лучше линейной (работают пачки и
 многопоточность). Если выше &mdash; хуже линейной, и это самый опасный случай.
 В колонке "Прогноз x10" и "Прогноз расчётный" две оценки стоят рядом для сравнения.</p>
</div>

<h3>2.3. Потоки: сейчас и максимум</h3>
<div class="box warn">
 <p><b>Важная оговорка по колонке "Потоков макс".</b> В таблице МО Розница максимум
 потоков &mdash; это проверенное значение (5 &rarr; 20). По ДУ такого подтверждения нет,
 поэтому в колонку поставлен <b>только предел, подтверждённый кодом</b>: жёстко зашитое
 значение, ограничение интерфейса или значение константы. Там, где предел упирается
 не в код, а в ресурсы кластера, максимум приравнен к текущему значению, а резерв
 описан в пояснении к операции.</p>
 <p>Сделано так намеренно: если подставить в максимум допустимые типом константы
 999 потоков, прогноз получится оптимистичнее текущего замера, что бессмысленно.
 Поэтому прогноз считается <b>без предположения об увеличении потоков</b> &mdash; это
 консервативная и честная оценка. Ускорение за счёт потоков надо подтверждать
 нагрузочным замером, и это отдельная задача.</p>
</div>

<h3>2.4. Предел по числу договоров</h3>
<div class="box">
 <p>Предел &mdash; это число договоров, при котором операция упирается в своё окно
 выполнения. Считается из модели: <code>(Окно - ПостояннаяЧасть) / СтоимостьДоговора</code>,
 а при отсутствии модели &mdash; пропорцией от текущего объёма. Окно задавалось по
 расписанию регламентного задания или по целевым значениям: для всех сверок
 окно не менее 60 минут; для прочих операций - по расписанию и целям IMDEV-8927.</p>
 <p>Подсветка строк: <span class="tag crit">красный</span> предел ниже целевых
 250 тыс. договоров, <span class="tag warn">жёлтый</span> предел от 250 до 500 тыс.
 (запас менее двукратного), <span class="tag ok">зелёный</span> запас достаточный
 либо операция не зависит от числа клиентов,
 <span class="tag pending">оранжевый</span> нет квартальной выгрузки
 (показано среднее за день, прогноз полного прогона не публикуется),
 <span class="tag nom">серый</span>
 замер в код не вставлен (прочерк, без расчёта).</p>
 <p><b>Загрузка и исполнение сделок</b> зависят от числа сделок. Допущение для прогноза:
 средняя интенсивность сделок на договор постоянна, поэтому при росте базы x10 объём
 сделок тоже x10 (прогноз x10 применяется). Для исполнения сделок окно: не более
 2 часов на 200 тыс. сделок.</p>
</div>
""")

    # -------------------------------------------------------------- Статистика
    out.append("""
<h2>3. Сводка</h2>
<div class="cards">
 <div class="card"><div class="val">%d</div><div class="cap">операций в таблице</div></div>
 <div class="card crit"><div class="val">%d</div><div class="cap">не выдержат 250 тыс. договоров</div></div>
 <div class="card warn"><div class="val">%d</div><div class="cap">запас менее двукратного</div></div>
 <div class="card ok"><div class="val">%d</div><div class="cap">не зависят от числа клиентов</div></div>
 <div class="card warn"><div class="val">%d</div><div class="cap">оценочные цифры (нужен замер)</div></div>
 <div class="card"><div class="val">%d</div><div class="cap">замер в код не вставлен</div></div>
</div>
""" % (len(ops), len(crit), len(warn), len(independent), len(estimates), len(no_meas)))

    # ----------------------------------------------------------- Главная таблица
    out.append('<h2>4. Итоговая таблица</h2>\n<table>\n<tr>')
    for _, title, _w in COLUMNS:
        out.append('<th>%s</th>' % html_escape(title))
    out.append('</tr>\n')

    number = 0
    for row in rows:
        if row.get('kind') == 'group':
            out.append('<tr class="grp"><td colspan="%d">%s</td></tr>\n'
                       % (len(COLUMNS), html_escape(row['name'])))
            continue
        number += 1
        mark = verdict(row)
        values = cell_values(row, number)
        out.append('<tr class="%s">' % mark)
        for (key, _t, _w), value in zip(COLUMNS, values):
            css = ['mark'] if key == 'no' else []
            if key in ('min_now', 'min_max', 'min_linear', 'min_model',
                       'threads_now', 'threads_max', 'limit', 'window', 'no'):
                css.append('num')
            if key == 'name':
                css.append('name')
            text = html_escape(value)
            if key == 'name':
                if row.get('real_name'):
                    text += ('<span class="real-name">реальное название: %s</span>'
                             % html_escape(row['real_name']))
                if row.get('no_measurement'):
                    text += ' <span class="tag nom">нет замеров</span>'
                elif row.get('no_forecast'):
                    text += ' <span class="tag pending">нет квартальной выгрузки</span>'
                elif row.get('estimate'):
                    text += ' <span class="tag est">оценка</span>'
                if row.get('source') == 'Разработческая база':
                    text += ' <span class="tag dev">dev</span>'
            out.append('<td class="%s">%s</td>' % (' '.join(css), text))
        out.append('</tr>\n')
    out.append('</table>\n')

    # ------------------------------------------------------------ Узкие места
    out.append('<h2>5. Узкие места</h2>\n')
    if crit:
        out.append('<div class="box warn"><b>Не выдерживают целевые 250 тыс. договоров'
                   '</b><ul>')
        for row in sorted(crit, key=lambda r: r.get('limit_contracts') or 0):
            out.append('<li><b>%s</b> &mdash; предел около %s тыс. договоров при окне %s мин. '
                       'Сейчас %s мин (макс %s). Потоков: %s из %s. %s</li>'
                       % (html_escape(row['name']), fmt_limit(row),
                          '%.0f' % row['window_min'], fmt_min(row.get('min_now')),
                          fmt_min(row.get('min_max')),
                          row['threads_now'], row['threads_max'],
                          html_escape(row.get('depends_note', ''))))
        out.append('</ul></div>\n')

    single = [r for r in ops if r['threads_now'] == 1 and r.get('depends')
              and not r.get('no_measurement')]
    out.append('<div class="box viol"><b>Однопоточные операции, зависящие от числа '
               'клиентов (%d)</b><p>Это главный источник риска: рост базы такие операции '
               'принимают на себя целиком, без распараллеливания.</p><ul>' % len(single))
    for row in sorted(single, key=lambda r: -(r.get('sec_now') or 0)):
        out.append('<li><b>%s</b> &mdash; сейчас %s мин, прогноз до %s мин. %s</li>'
                   % (html_escape(row['name']), fmt_min(row.get('min_now')),
                      fmt_min(max(v for v in (row.get('min_linear'), row.get('min_model'))
                                  if v)),
                      html_escape(row.get('threads_note', ''))))
    out.append('</ul></div>\n')

    deals = [r for r in ops if r.get('volume_unit') == 'сделок' and r.get('depends')
             and not r.get('no_measurement')]
    other_indep = independent
    out.append('<div class="box ok" style="background:#eaf7ee;border-left:4px solid #28a745">'
               '<b>Не зависят от числа клиентов (%d)</b>'
               '<p>Прогноз x10 по договорам к ним не применяется.</p><ul>' % len(independent))
    for row in other_indep:
        out.append('<li><b>%s</b> &mdash; %s мин. %s</li>'
                   % (html_escape(row['name']), fmt_min(row.get('min_now')),
                      html_escape(row.get('depends_note', ''))))
    out.append('</ul></div>\n')
    if deals:
        out.append('<div class="box blue"><b>Зависят от числа сделок '
                   '(прогноз через интенсивность на договор)</b>'
                   '<p>Допущение: сделок на договор не меняется, база x10 -> сделки x10. '
                   'Для исполнения окно: не более 2 часов на 200 тыс. сделок.</p><ul>')
        for row in deals:
            out.append('<li><b>%s</b> &mdash; сейчас %s мин, x10 %s мин, предел %s тыс. %s</li>'
                       % (html_escape(row['name']), fmt_min(row.get('min_now')),
                          fmt_min(row.get('min_linear')), fmt_limit(row),
                          html_escape(row.get('depends_note', ''))))
        out.append('</ul></div>\n')

    out.append('<h2>6. Что требует замера</h2>\n')
    if no_meas:
        out.append('<div class="box"><b>Замер в код не вставлен (%d)</b>'
                   '<p>Длительность и прогноз не рассчитываются: в таблице прочерк '
                   '&laquo;нет замеров&raquo;.</p><ul>' % len(no_meas))
        for row in no_meas:
            out.append('<li><b>%s</b> &mdash; %s</li>'
                       % (html_escape(row['name']), html_escape(row.get('comment', ''))))
        out.append('</ul></div>\n')
    out.append('<div class="box warn">'
               '<p>По этим операциям цифра в таблице оценочная (замер в коде есть, '
               'но массового прогона в выгрузке нет). Пометка нужна, чтобы '
               'оценки не ушли в работу как факт.</p><ul>')
    for row in estimates:
        out.append('<li><b>%s</b> &mdash; %s</li>'
                   % (html_escape(row['name']), html_escape(row.get('comment', ''))))
    out.append('</ul></div>\n')

    # ------------------------------------------------------- Детали по операциям
    out.append('<h2>7. Детали расчёта по каждой операции</h2>\n')
    number = 0
    for row in rows:
        if row.get('kind') != 'op':
            continue
        number += 1
        mark = verdict(row)
        tag = {'crit': 'crit', 'warn': 'warn', 'ok': 'ok', 'none': 'nom',
               'pending': 'pending'}[mark]
        label = {'crit': 'не выдержит 250 тыс.',
                 'warn': 'запас менее двукратного',
                 'ok': 'запас достаточный',
                 'none': 'нет замеров',
                 'pending': 'нет квартальной выгрузки'}[mark]
        extra = ''
        if row.get('no_measurement'):
            extra = ' <span class="tag nom">нет замеров</span>'
        elif row.get('no_forecast'):
            extra = ' <span class="tag pending">прогноз не публикуется</span>'
        elif row.get('estimate'):
            extra = ' <span class="tag est">оценка</span>'
        out.append('<div class="box"><h3>%d. %s <span class="tag %s">%s</span>%s</h3>'
                   % (number, html_escape(row['name']), tag, label, extra))
        if row.get('real_name'):
            out.append('<p class="det"><b>Реальное название:</b> %s</p>'
                       % html_escape(row['real_name']))
        out.append('<p class="det"><b>Объект:</b> %s &nbsp;|&nbsp; '
                   '<b>Периодичность:</b> %s &nbsp;|&nbsp; <b>Запуск:</b> %s &nbsp;|&nbsp; '
                   '<b>Источник:</b> %s &nbsp;|&nbsp; '
                   '<b>Потоков:</b> %s из %s</p>'
                   % (html_escape(row['object']), html_escape(row['periodicity']),
                      html_escape(row.get('schedule') or '-'),
                      html_escape(row['source']), row['threads_now'], row['threads_max']))
        if row.get('no_forecast'):
            contracts = row.get('weight_typical')
            out.append('<p class="det"><b>Сейчас (ср. за день):</b> %s мин '
                       '&nbsp;|&nbsp; <b>Ср. договоров/день:</b> %s '
                       '&nbsp;|&nbsp; <b>Макс за день:</b> %s мин '
                       '&nbsp;|&nbsp; <b>Прогноз / предел:</b> не публикуются '
                       '(нет квартальной выгрузки) при окне %s мин</p>'
                       % (fmt_min(row.get('min_now')),
                          ('%.0f' % contracts if contracts else '-'),
                          fmt_min(row.get('min_max')),
                          '%.0f' % row['window_min']))
        else:
            out.append('<p class="det"><b>Сейчас:</b> %s мин &nbsp;|&nbsp; '
                       '<b>Макс из замеров:</b> %s мин &nbsp;|&nbsp; '
                       '<b>x10 линейно:</b> %s мин &nbsp;|&nbsp; <b>расчётно:</b> %s мин '
                       '&nbsp;|&nbsp; <b>предел:</b> %s тыс. договоров при окне %s мин</p>'
                       % (fmt_min(row.get('min_now')), fmt_min(row.get('min_max')),
                          fmt_min(row.get('min_linear')), fmt_min(row.get('min_model')),
                          fmt_limit(row), '%.0f' % row['window_min']))
        out.append('<p class="det">%s</p>' % html_escape(methodics(row)))
        if row.get('comment'):
            out.append('<p class="det"><b>Комментарий:</b> %s</p>'
                       % html_escape(row['comment']))
        out.append('</div>\n')

    out.append('<footer>IMDEV-9393 &middot; расчёт выполнен по выгрузкам регистра '
               'ЗамерыВремени и анализу исходного кода конфигурации Wim_Du</footer>\n'
               '</div>\n</body>\n</html>')

    with open(OUT_HTML, 'w', encoding='utf-8') as handle:
        handle.write(''.join(out))
    return len(ops), len(crit), len(warn), len(estimates)


# --------------------------------------------------------------------------- #
#  XLSX
# --------------------------------------------------------------------------- #

FILL = {
    'crit': PatternFill('solid', fgColor='FDECEA'),
    'warn': PatternFill('solid', fgColor='FFF8E1'),
    'ok': PatternFill('solid', fgColor='EAF7EE'),
    'none': PatternFill('solid', fgColor='E9ECEF'),
    'pending': PatternFill('solid', fgColor='FFE0B2'),
    'grp': PatternFill('solid', fgColor='D6E2F0'),
    'head': PatternFill('solid', fgColor='17365D'),
}
THIN = Side(style='thin', color='C8D0DA')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_xlsx(payload):
    rows = payload['rows']
    meta = payload['meta']
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = 'Ключевые операции'

    sheet['A1'] = ('IMDEV-9393. Длительность ключевых операций Розничного ДУ (ДУ 1.5). '
                   'Рост %d тыс. -> %d тыс. договоров. Сформировано %s'
                   % (meta['dogovorov_seychas'] / 1000, meta['dogovorov_plan'] / 1000,
                      meta['built']))
    sheet['A1'].font = Font(bold=True, size=12, color='17365D')
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    header = 3
    for index, (_key, title, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=header, column=index, value=title)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = FILL['head']
        cell.alignment = Alignment(wrap_text=True, vertical='bottom')
        cell.border = BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = sheet.cell(row=header + 1, column=3)

    line = header
    number = 0
    for row in rows:
        line += 1
        if row.get('kind') == 'group':
            cell = sheet.cell(row=line, column=1, value=row['name'])
            cell.font = Font(bold=True, color='17365D', size=11)
            sheet.merge_cells(start_row=line, start_column=1,
                              end_row=line, end_column=len(COLUMNS))
            for index in range(1, len(COLUMNS) + 1):
                sheet.cell(row=line, column=index).fill = FILL['grp']
            continue

        number += 1
        mark = verdict(row)
        values = cell_values(row, number)
        for index, ((key, _t, _w), value) in enumerate(zip(COLUMNS, values), start=1):
            # Числовые колонки пишем числами, чтобы работали сортировка и фильтры.
            # Для no_forecast в "сейчас" текст вида "4.7 (ср. 20 дог./день)" - оставляем строкой.
            if (key in ('min_now', 'min_max', 'min_linear', 'min_model')
                    and value not in ('-', 'нет замеров')
                    and not isinstance(value, str)):
                value = float(value)
            elif (key in ('min_now', 'min_max', 'min_linear', 'min_model')
                  and isinstance(value, str)
                  and value not in ('-', 'нет замеров')
                  and 'дог' not in value):
                try:
                    value = float(value)
                except ValueError:
                    pass
            elif key == 'limit' and value not in ('не зависит', 'нужен замер', '> 1 000', '-', 'от сделок', 'нет замеров'):
                value = float(value)
            if key == 'name' and row.get('real_name'):
                value = '%s\nреальное название: %s' % (row['name'], row['real_name'])
            cell = sheet.cell(row=line, column=index, value=value)
            cell.fill = FILL[mark]
            cell.border = BORDER
            cell.alignment = Alignment(
                wrap_text=key in ('name', 'object', 'comment', 'schedule'),
                vertical='top',
                horizontal='right' if key in (
                    'no', 'threads_now', 'threads_max', 'min_now', 'min_max', 'min_linear',
                    'min_model', 'limit', 'window') else 'left')
            if key == 'name':
                cell.font = Font(bold=True, size=10)
            else:
                cell.font = Font(size=10)
            if key in ('min_now', 'min_max', 'min_linear', 'min_model'):
                cell.number_format = '0.0'
    sheet.auto_filter.ref = 'A%d:%s%d' % (header, get_column_letter(len(COLUMNS)), line)

    # ---------------------------------------------------- Лист с методикой
    detail = book.create_sheet('Методика и детали')
    detail.column_dimensions['A'].width = 46
    detail.column_dimensions['B'].width = 150
    detail['A1'] = 'Как получена каждая цифра'
    detail['A1'].font = Font(bold=True, size=12, color='17365D')
    line = 2
    for row in rows:
        if row.get('kind') != 'op':
            continue
        line += 1
        name = detail.cell(row=line, column=1, value=row['name'])
        name.font = Font(bold=True, size=10)
        name.alignment = Alignment(wrap_text=True, vertical='top')
        name.fill = FILL[verdict(row)]
        text = detail.cell(row=line, column=2, value=methodics(row))
        text.alignment = Alignment(wrap_text=True, vertical='top')
        text.font = Font(size=10)

    legend = book.create_sheet('Легенда')
    legend.column_dimensions['A'].width = 26
    legend.column_dimensions['B'].width = 120
    rows_legend = [
        ('Подсветка строки', ''),
        ('Красный', 'Предел ниже целевых 250 тыс. договоров - операция не выдержит рост базы.'),
        ('Жёлтый', 'Предел от 250 до 500 тыс. договоров - запас менее двукратного.'),
        ('Зелёный', 'Запас достаточный либо операция не зависит от числа клиентов.'),
        ('Оранжевый', 'Нет квартальной выгрузки массового прогона: показано среднее за день '
         'и среднее число договоров/день; прогноз и предел не публикуются.'),
        ('Серый', 'Замер в код не вставлен: длительность и прогноз не рассчитываются (прочерк).'),
        ('', ''),
        ('Колонки', ''),
        ('Длительность сейчас',
         'Медиана прогонов на полном объёме базы: только регламентный контур, '
         'вес замера от 15 тыс. договоров, аварийные замеры исключены.'),
        ('Длительность макс',
         'Максимум длительности среди тех же отобранных замеров (полный объём, без аварийных).'),
        ('Время запуска',
         'По фактическим замерам: первое типичное время старта за день, частота запусков, '
         'последнее типичное время старта. Формат: ЧЧ:ММ, N мин (или 1 раз), ЧЧ:ММ.'),
        ('Прогноз x10 линейный',
         'Текущая длительность, умноженная на 10. Для загрузки/исполнения сделок '
         'применяется по допущению: интенсивность сделок на договор постоянна. '
         'Не применяется к операциям, не зависящим от числа договоров (котировки, счета ЕРС).'),
        ('Прогноз расчётный',
         'Линейная модель ПостояннаяЧасть + СтоимостьЕдиницы * Объём. '
         'Наклон по двум самым крупным уровням объёма из фактических замеров.'),
        ('Потоков макс',
         'Только предел, подтверждённый кодом: жёсткое значение, ограничение интерфейса '
         'или значение константы. Где предел упирается в ресурсы кластера, а не в код, '
         'максимум приравнен к текущему значению, а резерв описан на листе "Методика и детали".'),
        ('Предел, тыс. договоров',
         'Число договоров, при котором операция упирается в своё окно выполнения. '
         'Для сделок - через пересчёт объёма сделок при постоянной интенсивности на договор '
         '(окно исполнения: 2 ч на 200 тыс. сделок).'),
        ('Пометка "оценка"',
         'Массового прогона в выгрузке нет, цифра получена пересчётом от удельной '
         'стоимости договора. Требуется замер массового прогона.'),
        ('Пометка "нет замеров"',
         'Замер в код не вставлен. Расчёт по аналогии не делается - в ячейках прочерк.'),
    ]
    for index, (left, right) in enumerate(rows_legend, start=1):
        cell = legend.cell(row=index, column=1, value=left)
        cell.font = Font(bold=True, size=10, color='17365D')
        cell.alignment = Alignment(vertical='top')
        text = legend.cell(row=index, column=2, value=right)
        text.alignment = Alignment(wrap_text=True, vertical='top')
        text.font = Font(size=10)
    for name, color in (('Красный', 'FDECEA'), ('Жёлтый', 'FFF8E1'),
                        ('Зелёный', 'EAF7EE'), ('Оранжевый', 'FFE0B2'),
                        ('Серый', 'E9ECEF')):
        for index in range(1, len(rows_legend) + 1):
            if legend.cell(row=index, column=1).value == name:
                legend.cell(row=index, column=1).fill = PatternFill('solid', fgColor=color)

    book.save(OUT_XLSX)


def main():
    with open(IN_JSON, encoding='utf-8') as handle:
        payload = json.load(handle)
    os.makedirs(DOCS, exist_ok=True)
    total, crit, warn, estimates = build_html(payload)
    build_xlsx(payload)
    print('Operations: %d, critical: %d, warning: %d, estimates: %d'
          % (total, crit, warn, estimates))
    print('HTML ->', OUT_HTML)
    print('XLSX ->', OUT_XLSX)


if __name__ == '__main__':
    main()
