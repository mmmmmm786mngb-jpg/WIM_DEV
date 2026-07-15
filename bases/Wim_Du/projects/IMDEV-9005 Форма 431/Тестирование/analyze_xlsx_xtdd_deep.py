#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Deep compare XLSX sheet columns vs XTDD section columns for 0420431."""

from pathlib import Path
import zipfile
import re
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent / "Обработки"
XLSX = next(BASE.glob("*.xlsx"))
XTDD = next(BASE.glob("*.xtdd"))

NS = {"p": "http://www.it.ru/Schemas/Avior/ПУРЦБ"}


def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", "replace").decode("ascii"))


def local(tag):
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def xlsx_headers_and_rows():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            result[name] = {"header": [], "nrows": 0, "sample": []}
            continue
        header = [str(c) if c is not None else "" for c in header]
        nrows = 0
        sample = []
        for r in rows:
            if all(c is None or str(c).strip() == "" for c in r):
                continue
            nrows += 1
            if len(sample) < 2:
                sample.append([c for c in r[:8]])
        result[name] = {"header": header, "nrows": nrows, "sample": sample}
    wb.close()
    return result


def xtdd_sections():
    text = XTDD.read_text(encoding="utf-8-sig")
    # iterative parse for large file - ET is ok at 25MB
    root = ET.fromstring(text)
    reporting = None
    for c in root:
        if local(c.tag) == "Отчетность":
            reporting = c
            break
    if reporting is None:
        raise RuntimeError("no Otchetnost")

    result = {}
    for sec in reporting:
        sname = local(sec.tag)
        if sname == "ДатаПредставления":
            result["_meta_ДатаПредставления"] = {
                "header": ["ДатаПредставления"],
                "nrows": 1,
                "sample": [[(sec.text or "").strip()]],
                "row_tag": "",
            }
            continue
        if sname == "СопроводительноеПисьмо":
            # flat fields
            leaves = []
            for el in sec:
                if list(el) == []:
                    leaves.append(local(el.tag))
            result[sname] = {
                "header": leaves,
                "nrows": 1 if leaves else 0,
                "sample": [],
                "row_tag": "(flat)",
            }
            continue

        # find repeating row children
        child_tags = [local(ch.tag) for ch in sec]
        if not child_tags:
            result[sname] = {"header": [], "nrows": 0, "sample": [], "row_tag": ""}
            continue

        # majority tag is row
        from collections import Counter
        ctr = Counter(child_tags)
        row_tag, _ = ctr.most_common(1)[0]
        headers = []
        nrows = 0
        samples = []
        for ch in sec:
            if local(ch.tag) != row_tag:
                continue
            nrows += 1
            fields = []
            vals = []
            for el in ch:
                if list(el) == []:
                    fields.append(local(el.tag))
                    vals.append((el.text or "").strip())
                else:
                    # nested - flatten one level with prefix?
                    fields.append(local(el.tag) + "(complex)")
                    vals.append("<complex>")
            if not headers:
                headers = fields
            if len(samples) < 2:
                samples.append(vals[:8])
        result[sname] = {
            "header": headers,
            "nrows": nrows,
            "sample": samples,
            "row_tag": row_tag,
        }
    return result


def map_sheet_to_xtdd(sheet_name, xtdd_keys):
    # Excel sheet names match section names mostly
    mapping = {
        "Шапка": "_meta / doc header (not in XTDD sections as TC)",
        "РеестрЦенныхБумаг": "ОтчетРеестрЦенныхБумаг",
    }
    if sheet_name in mapping:
        return mapping[sheet_name]
    if sheet_name in xtdd_keys:
        return sheet_name
    return None


def main():
    safe_print("Loading XLSX (openpyxl)...")
    try:
        xlsx = xlsx_headers_and_rows()
    except Exception as e:
        safe_print(f"openpyxl fail: {e}")
        safe_print("fallback: zip xml parse for headers only")
        xlsx = {}

    safe_print("Loading XTDD...")
    xtdd = xtdd_sections()

    safe_print("")
    safe_print("=== XLSX summary ===")
    for name, info in xlsx.items():
        safe_print(f"[{name}] rows={info['nrows']} cols={len(info['header'])}")
        safe_print(f"  header: {info['header'][:25]}")

    safe_print("")
    safe_print("=== XTDD summary ===")
    for name, info in xtdd.items():
        safe_print(
            f"[{name}] rows={info['nrows']} cols={len(info['header'])} row_tag={info.get('row_tag')}"
        )
        safe_print(f"  header: {info['header'][:25]}")

    safe_print("")
    safe_print("=== COMPARE (same section names) ===")
    xtdd_keys = set(xtdd.keys())
    for sheet, info in xlsx.items():
        target = sheet
        if sheet == "РеестрЦенныхБумаг":
            target = "ОтчетРеестрЦенныхБумаг"
        if sheet == "Шапка":
            safe_print(f"[{sheet}] XLSX-only document header; XTDD has ДатаПредставления + letter")
            continue
        if target not in xtdd:
            safe_print(f"[{sheet}] NO matching XTDD section")
            continue
        xh = info["header"]
        th = xtdd[target]["header"]
        # strip complex markers for compare
        th_clean = [h.replace("(complex)", "") for h in th]
        set_x = set([h for h in xh if h])
        set_t = set([h for h in th_clean if h])
        only_x = sorted(set_x - set_t)
        only_t = sorted(set_t - set_x)
        both = sorted(set_x & set_t)
        safe_print(
            f"[{sheet} <-> {target}] xlsx_rows={info['nrows']} xtdd_rows={xtdd[target]['nrows']} "
            f"common_cols={len(both)} only_xlsx={len(only_x)} only_xtdd={len(only_t)}"
        )
        if only_x[:15]:
            safe_print(f"  only XLSX: {only_x[:15]}")
        if only_t[:15]:
            safe_print(f"  only XTDD: {only_t[:15]}")
        # order similarity: common prefix of headers
        same_order = 0
        for a, b in zip(xh, th_clean):
            if a == b:
                same_order += 1
            else:
                break
        safe_print(f"  header prefix match length: {same_order}/{max(len(xh), len(th_clean))}")


if __name__ == "__main__":
    main()
