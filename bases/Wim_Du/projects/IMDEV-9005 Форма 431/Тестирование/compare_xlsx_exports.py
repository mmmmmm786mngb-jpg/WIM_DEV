#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare document-export XLSX vs XML-export XLSX for 0420431."""

from pathlib import Path
from openpyxl import load_workbook
from collections import Counter

BASE = Path(__file__).resolve().parent / "Test" / "Test"
DOC = BASE / "0420431_000000003_20260101_20260131.xlsx"
XML = BASE / "0420431_XML.xlsx"

# numeric tolerance for float compare
EPS = 1e-6


def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", "replace").decode("ascii"))


def norm(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        # try number with comma
        t = s.replace("\xa0", "").replace(" ", "").replace(",", ".")
        try:
            return float(t)
        except ValueError:
            return s
    # datetime etc
    return v


def values_equal(a, b):
    a, b = norm(a), norm(b)
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if isinstance(a, float) and isinstance(b, float):
        if abs(a - b) <= EPS:
            return True
        # also accept if equal after rounding to 6 decimals
        return round(a, 6) == round(b, 6)
    return a == b


def load_sheet_data(wb, name):
    if name not in wb.sheetnames:
        return None
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"header": [], "data": []}
    header = [str(c) if c is not None else "" for c in rows[0]]
    # Шапка has different structure - special case later
    data = []
    for r in rows[1:]:
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in r):
            continue
        data.append(list(r))
    return {"header": header, "data": data}


def sheet_to_dicts(sheet):
    header = sheet["header"]
    # trim trailing empty header cells
    while header and header[-1] == "":
        header = header[:-1]
    result = []
    for row in sheet["data"]:
        d = {}
        for i, h in enumerate(header):
            if not h:
                continue
            d[h] = row[i] if i < len(row) else None
        result.append(d)
    return header, result


def compare_data_sheets(name, doc_sheet, xml_sheet):
    issues = []
    h1, rows1 = sheet_to_dicts(doc_sheet)
    h2, rows2 = sheet_to_dicts(xml_sheet)

    if h1 != h2:
        only1 = [c for c in h1 if c not in h2]
        only2 = [c for c in h2 if c not in h1]
        order_diff = h1 != h2 and set(h1) == set(h2)
        if only1 or only2:
            issues.append(f"columns differ only_doc={only1[:10]} only_xml={only2[:10]}")
        elif order_diff:
            issues.append("column ORDER differs (set same)")
            # align by name for value compare
        else:
            issues.append(f"headers differ: doc={h1[:8]} xml={h2[:8]}")

    if len(rows1) != len(rows2):
        issues.append(f"row count: doc={len(rows1)} xml={len(rows2)}")

    common_cols = [c for c in h1 if c in set(h2)]
    if not common_cols and h1 and h2:
        return issues, {"rows": (len(rows1), len(rows2))}

    # compare by index (same export order expected)
    n = min(len(rows1), len(rows2))
    mismatches = 0
    samples = []
    for i in range(n):
        for col in common_cols:
            v1 = rows1[i].get(col)
            v2 = rows2[i].get(col)
            if not values_equal(v1, v2):
                mismatches += 1
                if len(samples) < 8:
                    samples.append(
                        f"  row {i+1} col {col}: doc={v1!r} xml={v2!r}"
                    )
    stats = {
        "rows": (len(rows1), len(rows2)),
        "cols_common": len(common_cols),
        "mismatches": mismatches,
        "samples": samples,
        "header_same_order": h1 == h2,
        "header_same_set": set(h1) == set(h2),
    }
    if mismatches:
        issues.append(f"cell mismatches: {mismatches} (on {n} rows x {len(common_cols)} cols)")
    return issues, stats


def compare_shapka(doc_wb, xml_wb):
    issues = []
    sd = load_sheet_data(doc_wb, "Шапка")
    sx = load_sheet_data(xml_wb, "Шапка")
    if sd is None or sx is None:
        return ["Шапка missing"], {}

    # build key->value for rows that look like Реквизит/Значение
    def as_map(sheet):
        m = {}
        for row in sheet["data"]:
            if not row:
                continue
            k = row[0] if len(row) > 0 else None
            v = row[1] if len(row) > 1 else None
            if k is None:
                continue
            m[str(k)] = v
        return m

    md, mx = as_map(sd), as_map(sx)
    # expected: different keys (document vs XML source) - report overlap
    common = set(md) & set(mx)
    only_doc = sorted(set(md) - set(mx))
    only_xml = sorted(set(mx) - set(md))
    mism = []
    for k in sorted(common):
        if not values_equal(md[k], mx[k]):
            mism.append((k, md[k], mx[k]))
    return issues, {
        "doc_keys": len(md),
        "xml_keys": len(mx),
        "common_keys": len(common),
        "only_doc": only_doc[:20],
        "only_xml": only_xml[:20],
        "value_mismatches": mism[:10],
        "doc_rows": len(sd["data"]),
        "xml_rows": len(sx["data"]),
    }


def main():
    safe_print(f"DOC: {DOC.name} exists={DOC.exists()} size={DOC.stat().st_size if DOC.exists() else 0}")
    safe_print(f"XML: {XML.name} exists={XML.exists()} size={XML.stat().st_size if XML.exists() else 0}")
    if not DOC.exists() or not XML.exists():
        return

    wb_doc = load_workbook(DOC, read_only=True, data_only=True)
    wb_xml = load_workbook(XML, read_only=True, data_only=True)

    safe_print("")
    safe_print(f"DOC sheets ({len(wb_doc.sheetnames)}): {wb_doc.sheetnames}")
    safe_print(f"XML sheets ({len(wb_xml.sheetnames)}): {wb_xml.sheetnames}")

    if wb_doc.sheetnames != wb_xml.sheetnames:
        safe_print("WARN: sheet list/order differs")
        only_d = [s for s in wb_doc.sheetnames if s not in wb_xml.sheetnames]
        only_x = [s for s in wb_xml.sheetnames if s not in wb_doc.sheetnames]
        if only_d:
            safe_print(f"  only DOC: {only_d}")
        if only_x:
            safe_print(f"  only XML: {only_x}")
    else:
        safe_print("OK: sheet names and order match")

    safe_print("")
    safe_print("=== Шапка ===")
    _, sh = compare_shapka(wb_doc, wb_xml)
    safe_print(
        f"keys doc/xml/common: {sh.get('doc_keys')}/{sh.get('xml_keys')}/{sh.get('common_keys')}"
    )
    if sh.get("only_doc"):
        safe_print(f"only DOC keys: {sh['only_doc']}")
    if sh.get("only_xml"):
        safe_print(f"only XML keys: {sh['only_xml']}")
    if sh.get("value_mismatches"):
        safe_print("common key value diffs:")
        for k, a, b in sh["value_mismatches"]:
            safe_print(f"  {k}: doc={a!r} xml={b!r}")
    else:
        safe_print("OK: no conflicting shared keys (or no overlap)")

    all_ok = True
    safe_print("")
    safe_print("=== Data sheets ===")
    for name in wb_doc.sheetnames:
        if name == "Шапка":
            continue
        if name not in wb_xml.sheetnames:
            safe_print(f"[{name}] MISSING in XML export")
            all_ok = False
            continue
        d = load_sheet_data(wb_doc, name)
        x = load_sheet_data(wb_xml, name)
        issues, stats = compare_data_sheets(name, d, x)
        status = "OK" if not issues else "DIFF"
        if issues:
            all_ok = False
        safe_print(
            f"[{name}] {status} rows={stats.get('rows')} "
            f"header_order={stats.get('header_same_order')} "
            f"header_set={stats.get('header_same_set')} "
            f"mismatches={stats.get('mismatches')}"
        )
        for iss in issues:
            safe_print(f"  ! {iss}")
        for s in stats.get("samples") or []:
            safe_print(s)

    # sheets only in xml
    for name in wb_xml.sheetnames:
        if name not in wb_doc.sheetnames:
            safe_print(f"[{name}] EXTRA in XML export")
            all_ok = False

    safe_print("")
    if all_ok:
        safe_print("VERDICT: data sheets MATCH (same rows/cols/values within tolerance)")
    else:
        safe_print("VERDICT: differences found - see above")

    wb_doc.close()
    wb_xml.close()


if __name__ == "__main__":
    main()
