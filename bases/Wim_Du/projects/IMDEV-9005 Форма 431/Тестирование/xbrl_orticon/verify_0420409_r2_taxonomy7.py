#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Verify 0420409 R2: taxonomy7 vs etalon."""

import re
import zipfile
from collections import Counter
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431\ОРТИКОН")
E_PATH = BASE / "0420431_409_январь_2026_конвертер.xlsx"
O_PATH = BASE / "XBRL_Orticon_taxonomy7.xlsx"


def N(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ").replace("\n", " ")).strip()


def Nl(v):
    return N(v).lower()


def norm_amt(v):
    s = N(v).replace(" ", "").replace(",", ".")
    if not s:
        return ""
    try:
        f = float(s)
        if abs(f - int(f)) < 1e-9:
            return str(int(f))
        return ("%.2f" % f).rstrip("0").rstrip(".")
    except Exception:
        return s.lower()


def sheet_dim(path, name):
    with zipfile.ZipFile(path) as z:
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid = {r.get("Id"): r.get("Target") for r in rels}
        for sh in wb.findall("m:sheets/m:sheet", NS):
            if sh.get("name") != name:
                continue
            ridv = sh.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            )
            sp = "xl/" + rid[ridv].lstrip("/")
            root = ET.fromstring(z.read(sp))
            dim = root.find("m:dimension", NS)
            ref = dim.get("ref", "A1:Z200") if dim is not None else "A1:Z200"
            m = re.match(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?", ref.upper())

            def col(letters):
                n = 0
                for c in letters:
                    n = n * 26 + ord(c) - 64
                return n

            return int(m.group(4) or m.group(2)), col(m.group(3) or m.group(1))
    return 200, 30


def find_sheet(wb, code, sec):
    needle = "раздел %d" % sec
    for s in wb.sheetnames:
        if code in s and needle in s.lower():
            return s
    return None


def parse_r2(path, sheet_name, is_ours):
    mr, mc = sheet_dim(path, sheet_name)
    if (not mr or mr < 2) and not is_ours:
        mr, mc = 100, 30
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(
        ws.iter_rows(
            min_row=1, max_row=max(mr, 50), max_col=min(mc or 30, 30), values_only=True
        )
    )
    wb.close()

    hdr_i = None
    headers = []
    for i, row in enumerate(rows[:25]):
        vals = [N(v) for v in row]
        j = " ".join(vals).lower()
        n = sum(1 for v in vals if v)
        if n < 4:
            continue
        if "сокращенн" in j or "брокер" in j or ("инн" in j and "валют" in j):
            # skip TOC-like
            if vals[0] in ("TOC",) or vals[0].startswith("0420409"):
                continue
            hdr_i = i
            headers = vals
            break

    col = {}
    for i, h in enumerate(headers):
        hl = h.lower()
        if not hl:
            continue
        if "период" in hl:
            col.setdefault("period", i)
        elif "сокращенн" in hl or ("наименование" in hl and ("брокер" in hl or "кредитн" in hl)):
            col.setdefault("name", i)
        elif "инн" in hl or "tin" in hl:
            col.setdefault("inn", i)
        elif "вид организации" in hl:
            col.setdefault("org", i)
        elif "код страны" in hl:
            col.setdefault("country", i)
        elif "код валюты" in hl or (hl.startswith("код валют")):
            col.setdefault("ccy", i)
        elif "драгоценн" in hl:
            col.setdefault("metal", i)
        elif "идентификатор брокера" in hl:
            col.setdefault("broker_id", i)
        elif "идентификатор строки" in hl:
            col.setdefault("row_id", i)
        elif "возможн" in hl:
            col.setdefault("use", i)
        elif "ограничен" in hl or "распоряжен" in hl:
            col.setdefault("restrict", i)
        elif "остатк" in hl or "сумма" in hl:
            col.setdefault("amt", i)

    print("  header row", hdr_i, "cols:", sorted(col.items()))
    print("  headers:", [h[:50] for h in headers if h][:12])

    data = []
    for i in range((hdr_i or 0) + 1, len(rows)):
        vals = [N(v) for v in rows[i]]
        if not any(vals):
            continue
        filled = [v for v in vals if v]
        if filled and all(re.fullmatch(r"\d+", v) for v in filled) and len(filled) >= 3:
            continue
        if len(filled) == 1 and "идентификатор" in filled[0].lower():
            continue
        # skip axis title rows
        if filled and filled[0].lower().startswith("идентификатор") and len(filled) <= 2:
            continue

        def g(key):
            if key not in col or col[key] >= len(vals):
                return ""
            return vals[col[key]]

        name = g("name")
        inn = g("inn")
        # skip rows without meaningful entity
        if not name and not inn:
            # maybe etalon has name later - try any name-like
            continue

        rec = {
            "name": name,
            "inn": inn,
            "country": g("country"),
            "ccy": g("ccy"),
            "org": g("org"),
            "metal": g("metal"),
            "broker_id": g("broker_id"),
            "row_id": g("row_id"),
            "use": g("use"),
            "restrict": g("restrict"),
            "amt": norm_amt(g("amt")),
            "period": g("period"),
            "raw_amts": [],
            "raw_amt_texts": [],
            "all_vals": [v for v in vals if v],
        }
        for v in vals:
            if not re.search(r"\d", v):
                continue
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", v):
                continue
            if re.fullmatch(r"\d{10}", v) or re.fullmatch(r"\d{1,2}", v):
                continue
            if re.fullmatch(r"\d{3}-[A-Za-z]{3}", v):
                continue
            if re.fullmatch(r"\d{3}", v):
                continue
            if ".." in v:
                continue
            a = norm_amt(v)
            if a and re.match(r"^-?\d", a) and len(a) >= 1:
                # skip tiny row ids if they look like ints 1-20 only without decimal and already as row_id
                if re.fullmatch(r"\d{1,2}", a):
                    continue
                rec["raw_amts"].append(a)
                rec["raw_amt_texts"].append(v)
        data.append(rec)

    return {"headers": headers, "col": col, "rows": data, "hdr_i": hdr_i}


def row_key(r):
    # stable key: inn + ccy + row_id/broker + name fragment
    parts = [
        Nl(r["inn"]),
        Nl(r["ccy"]),
        Nl(r["row_id"]) or Nl(r["broker_id"]),
        Nl(r["name"])[:40],
        Nl(r["org"])[:30],
    ]
    return "|".join(parts)


def fingerprint(r):
    """Bag of identifying + amount values for multiset match."""
    vals = []
    for k in ("inn", "name", "country", "ccy", "org", "use", "row_id", "broker_id"):
        v = Nl(r.get(k, ""))
        if v:
            vals.append(v)
    vals.extend(r["raw_amts"])
    return tuple(sorted(vals))


def main():
    wb_e = load_workbook(E_PATH, read_only=True)
    wb_o = load_workbook(O_PATH, read_only=True)
    e_name = find_sheet(wb_e, "0420409", 2)
    o_name = find_sheet(wb_o, "0420409", 2)
    print("E sheet:", e_name)
    print("O sheet:", o_name)
    for row in wb_o[o_name].iter_rows(min_row=1, max_row=5, max_col=1, values_only=True):
        if row[0] and "Generator" in str(row[0]):
            print("O version:", row[0])
    wb_e.close()
    wb_o.close()

    print("\nParsing etalon R2...")
    E = parse_r2(E_PATH, e_name, False)
    print("Parsing ours R2...")
    O = parse_r2(O_PATH, o_name, True)
    print("E data rows:", len(E["rows"]))
    print("O data rows:", len(O["rows"]))

    print("\n=== Sample E rows ===")
    for r in E["rows"][:8]:
        print(" ", r["name"][:40], r["inn"], r["country"], r["ccy"], r["org"][:40] if r["org"] else "", r["raw_amts"][:3])
    print("=== Sample O rows ===")
    for r in O["rows"][:8]:
        print(" ", r["name"][:40], r["inn"], r["country"], r["ccy"], r["org"][:40] if r["org"] else "", "row_id="+r["row_id"], r["raw_amts"][:3])

    # Header check: broker wording
    print("\n========== CHECKS ==========")
    e_hdr = " | ".join(h for h in E["headers"] if h).lower()
    o_hdr = " | ".join(h for h in O["headers"] if h).lower()
    print("1. Headers (SokrNaim broker wording):")
    print("   E has broker:", "брокер" in e_hdr and "сокращенн" in e_hdr)
    print("   O has broker:", "брокер" in o_hdr and "сокращенн" in o_hdr)
    print("   O has kreditn in sokr:", "кредитн" in o_hdr and "сокращенн" in o_hdr)
    e_sokr = [h for h in E["headers"] if "сокращенн" in h.lower()]
    o_sokr = [h for h in O["headers"] if "сокращенн" in h.lower()]
    print("   E sokr:", e_sokr[0][:90] if e_sokr else None)
    print("   O sokr:", o_sokr[0][:90] if o_sokr else None)

    print("\n2. Row counts: E=%d O=%d" % (len(E["rows"]), len(O["rows"])))

    # INNs
    e_inn = Counter(r["inn"] for r in E["rows"] if r["inn"])
    o_inn = Counter(r["inn"] for r in O["rows"] if r["inn"])
    print("\n3. INNs:")
    print("   E:", dict(e_inn))
    print("   O:", dict(o_inn))
    print("   common INNs:", set(e_inn) & set(o_inn))

    # names
    e_names = Counter(Nl(r["name"]) for r in E["rows"] if r["name"])
    o_names = Counter(Nl(r["name"]) for r in O["rows"] if r["name"])
    print("\n4. Names:")
    print("   E:", dict(e_names))
    print("   O:", dict(o_names))

    # country
    e_c = Counter(r["country"] for r in E["rows"] if r["country"])
    o_c = Counter(r["country"] for r in O["rows"] if r["country"])
    print("\n5. Country:")
    print("   E:", dict(e_c))
    print("   O:", dict(o_c))
    russia = sum(1 for r in O["rows"] if "росси" in Nl(r["country"]))
    print("   O Rossiya words:", russia)

    # currency
    e_ccy = Counter(r["ccy"] for r in E["rows"] if r["ccy"])
    o_ccy = Counter(r["ccy"] for r in O["rows"] if r["ccy"])
    print("\n6. Currency:")
    print("   E:", dict(e_ccy))
    print("   O:", dict(o_ccy))

    # org type
    e_org = Counter(Nl(r["org"])[:60] for r in E["rows"] if r["org"])
    o_org = Counter(Nl(r["org"])[:60] for r in O["rows"] if r["org"])
    print("\n7. Org type:")
    print("   E:", e_org.most_common(5))
    print("   O:", o_org.most_common(5))

    # value multiset
    e_vals = Counter()
    o_vals = Counter()
    for r in E["rows"]:
        for k in ("inn", "name", "country", "ccy", "org", "use", "restrict", "metal"):
            v = Nl(r.get(k, ""))
            if v:
                e_vals[v] += 1
        e_vals.update(r["raw_amts"])
    for r in O["rows"]:
        for k in ("inn", "name", "country", "ccy", "org", "use", "restrict", "metal"):
            v = Nl(r.get(k, ""))
            if v:
                o_vals[v] += 1
        o_vals.update(r["raw_amts"])
    shared = sum(min(e_vals[k], o_vals.get(k, 0)) for k in e_vals)
    e_tot = sum(e_vals.values())
    print("\n8. Value multiset E->O: %d / %d (%.1f%%)" % (
        shared, e_tot, 100.0 * shared / e_tot if e_tot else 0))
    only_e = sorted(((e_vals[k] - o_vals.get(k, 0), k) for k in e_vals if e_vals[k] > o_vals.get(k, 0)), reverse=True)
    only_o = sorted(((o_vals[k] - e_vals.get(k, 0), k) for k in o_vals if o_vals[k] > e_vals.get(k, 0)), reverse=True)
    print("   only/more in E (top):", only_e[:8])
    print("   only/more in O (top):", only_o[:8])

    # triad
    triad = 0
    for r in O["rows"]:
        for t in r["raw_amt_texts"]:
            if re.search(r"\d[\u00a0 ]\d{3}", t):
                triad += 1
    print("\n9. Triad separators in O amounts:", triad)

    # fingerprints
    e_fp = Counter(fingerprint(r) for r in E["rows"])
    o_fp = Counter(fingerprint(r) for r in O["rows"])
    shared_fp = sum(min(e_fp[k], o_fp.get(k, 0)) for k in e_fp)
    print("\n10. Row fingerprints match: %d / %d" % (shared_fp, sum(e_fp.values())))

    # pair by inn+ccy+name
    print("\n11. Pairwise by INN+CCY:")
    e_groups = {}
    for r in E["rows"]:
        k = (r["inn"], Nl(r["ccy"]))
        e_groups.setdefault(k, []).append(r)
    o_groups = {}
    for r in O["rows"]:
        k = (r["inn"], Nl(r["ccy"]))
        o_groups.setdefault(k, []).append(r)
    for k in sorted(set(e_groups) | set(o_groups)):
        er, or_ = e_groups.get(k, []), o_groups.get(k, [])
        print("  %s / %s : E=%d O=%d" % (k[0], k[1], len(er), len(or_)))
        if er and or_:
            # compare names
            if Nl(er[0]["name"]) == Nl(or_[0]["name"]):
                print("    name OK:", er[0]["name"][:40])
            else:
                print("    name DIFF E=%s O=%s" % (er[0]["name"][:40], or_[0]["name"][:40]))
            if er[0]["country"] and or_[0]["country"]:
                print("    country E=%s O=%s %s" % (
                    er[0]["country"], or_[0]["country"],
                    "OK" if er[0]["country"] == or_[0]["country"] else "DIFF"))

    print("\n=== VERDICT ===")
    print("Rows E/O:", len(E["rows"]), len(O["rows"]))
    print("Broker header OK:", "брокер" in o_hdr and "сокращенн" in o_hdr)
    print("Value recall:", "%.1f%%" % (100.0 * shared / e_tot if e_tot else 0))


if __name__ == "__main__":
    main()
