#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare etalon Orticon/CBR Excel vs 1C processing output (fixed dims)."""

from pathlib import Path
from collections import Counter
import re
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
ORT = BASE.parents[1] / "ОРТИКОН"
ETALON = ORT / "0420431_409_январь_2026_конвертер.xlsx"
OURS = ORT / "XBRL_Orticon.xlsx"
OUT_MD = BASE / "compare_1c_vs_etalon_analysis.md"


def norm_name(n):
    return re.sub(r"\s+", " ", (n or "").strip())


def count_nonempty_rows(ws, start_row=1):
    """Count rows with at least one nonempty cell from start_row."""
    count = 0
    max_r = ws.max_row or 0
    for r in range(start_row, max_r + 1):
        row_has = False
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                row_has = True
                break
        if row_has:
            count += 1
    return count


def find_data_start_etalon(ws):
    """
    Etalon layout: meta rows, then often a header block, then data.
    Heuristic: first row where col A looks like an identifier (len>=6 digit/alnum)
    and neighboring cells filled, after row 8.
    """
    max_r = min(ws.max_row or 0, 80)
    for r in range(1, max_r + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        s = str(a).strip()
        # skip TOC/title/schema/period rows
        if s.startswith("http") or s.startswith("T=") or s.startswith("TOC"):
            continue
        if "Раздел" in s and len(s) > 20 and r < 10:
            continue
        # numeric-like id / account / ISIN-ish
        if len(s) >= 6 and any(ch.isdigit() for ch in s):
            filled = sum(
                1
                for c in range(1, min((ws.max_column or 1), 12) + 1)
                if ws.cell(r, c).value not in (None, "")
            )
            if filled >= 2:
                return r
    # fallback: first row with >=3 filled after row 9
    for r in range(10, max_r + 1):
        filled = sum(
            1
            for c in range(1, min((ws.max_column or 1), 20) + 1)
            if ws.cell(r, c).value not in (None, "")
        )
        if filled >= 3:
            return r
    return 11


def etalon_stats(ws):
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    data_start = find_data_start_etalon(ws) if max_row > 1 else 1
    data_rows = max(max_row - data_start + 1, 0) if max_row else 0
    # header candidates: rows just above data_start
    headers = []
    for r in range(max(1, data_start - 4), data_start):
        vals = [ws.cell(r, c).value for c in range(1, min(max_col, 25) + 1)]
        filled = [v for v in vals if v not in (None, "")]
        if len(filled) >= 3:
            headers = filled
    return {
        "max_row": max_row,
        "max_col": max_col,
        "data_start": data_start,
        "data_rows": data_rows,
        "headers": headers,
        "n_headers": len(headers),
    }


def ours_stats(ws):
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    headers = []
    if max_row:
        headers = [
            ws.cell(1, c).value
            for c in range(1, max_col + 1)
            if ws.cell(1, c).value not in (None, "")
        ]
    return {
        "max_row": max_row,
        "max_col": max_col,
        "data_start": 2 if max_row > 1 else 1,
        "data_rows": max(max_row - 1, 0),
        "headers": headers,
        "n_headers": len(headers),
    }


def classify(got, et):
    if et == 0 and got == 0:
        return "EMPTY"
    if et == 0:
        return "OURS_ONLY_DATA"
    ratio = abs(got - et) / max(et, 1)
    if ratio <= 0.05:
        return "CLOSE"
    if ratio <= 0.25:
        return "NEAR"
    if got > et * 1.5:
        return "OURS_MUCH_MORE"
    if got < et * 0.5:
        return "OURS_MUCH_LESS"
    return "DIFF"


def main():
    print("Loading etalon (full)...")
    wb_e = load_workbook(ETALON, read_only=False, data_only=True)
    print("Loading ours (readonly)...")
    wb_o = load_workbook(OURS, read_only=True, data_only=True)

    name_map_e = {norm_name(n): n for n in wb_e.sheetnames}
    name_map_o = {norm_name(n): n for n in wb_o.sheetnames}
    # normalize trailing spaces differences
    # also map ours short names without trailing space
    for k, v in list(name_map_o.items()):
        name_map_o[k.rstrip()] = v
    for k, v in list(name_map_e.items()):
        name_map_e[k.rstrip()] = v

    enames = sorted(set(norm_name(n).rstrip() for n in wb_e.sheetnames))
    onames = sorted(set(norm_name(n).rstrip() for n in wb_o.sheetnames))
    set_e, set_o = set(enames), set(onames)
    common = sorted(set_e & set_o)
    only_e = sorted(set_e - set_o)
    only_o = sorted(set_o - set_e)

    lines = []
    lines.append("# Comparison: etalon vs XBRL_Orticon (1C)")
    lines.append("")
    lines.append(f"- Etalon: `{ETALON.name}` ({ETALON.stat().st_size} bytes)")
    lines.append(f"- Ours: `{OURS.name}` ({OURS.stat().st_size} bytes)")
    lines.append(f"- Sheets etalon: **{len(wb_e.sheetnames)}**, ours: **{len(wb_o.sheetnames)}**")
    lines.append(f"- Exact name matches: **{len(common)}**")
    lines.append(f"- Only etalon: **{len(only_e)}**, only ours: **{len(only_o)}**")
    lines.append("")

    # TOC
    gen = None
    if "TOC" in name_map_o:
        ws = wb_o[name_map_o["TOC"]]
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if row and row[0] == "Generator" and len(row) > 1:
                gen = row[1]
    lines.append("## TOC / generator")
    lines.append(f"- Etalon TOC rows: {wb_e[name_map_e['TOC']].max_row}")
    lines.append(f"- Ours Generator: {gen}")
    lines.append("")

    lines.append("## Sheets only in etalon (no exact name in ours)")
    for n in only_e:
        if n == "_dropDownSheet":
            lines.append(f"- `{n}` — service sheet of converter UI, ignore")
            continue
        st = etalon_stats(wb_e[name_map_e[n]])
        lines.append(
            f"- `{n}`: rows={st['max_row']}, cols={st['max_col']}, data~{st['data_rows']} (from r{st['data_start']})"
        )
    lines.append("")
    lines.append("## Sheets only in ours")
    for n in only_o:
        st = ours_stats(wb_o[name_map_o[n]])
        lines.append(
            f"- `{n}`: rows={st['max_row']}, cols={st['max_col']}, data={st['data_rows']}"
        )
    lines.append("")

    lines.append("## Volume compare on exact name matches")
    lines.append(
        "| Status | Sheet | Etalon data~ | Ours data | Delta | E cols | O cols |"
    )
    lines.append("|---|---|---:|---:|---:|---:|---:|")
    counts = Counter()
    details = []
    for n in common:
        if n == "TOC":
            se = {
                "data_rows": max((wb_e[name_map_e[n]].max_row or 0) - 5, 0),
                "max_col": wb_e[name_map_e[n]].max_column or 0,
                "headers": [],
                "data_start": 1,
            }
            so = ours_stats(wb_o[name_map_o[n]])
        else:
            se = etalon_stats(wb_e[name_map_e[n]])
            so = ours_stats(wb_o[name_map_o[n]])
        status = classify(so["data_rows"], se["data_rows"])
        counts[status] += 1
        delta = so["data_rows"] - se["data_rows"]
        details.append((status, n, se["data_rows"], so["data_rows"], delta, se["max_col"], so["max_col"], se, so))
        lines.append(
            f"| {status} | `{n}` | {se['data_rows']} | {so['data_rows']} | {delta:+d} | {se['max_col']} | {so['max_col']} |"
        )
    lines.append("")
    lines.append(
        "### Status counts: " + ", ".join(f"{k}={v}" for k, v in counts.most_common())
    )
    lines.append("")

    # Group only-etalon by role
    broker_like = [n for n in only_e if n.endswith("пор") or "маржин" in n.lower() or n.endswith("_1") or n.endswith("_2") or n.endswith("_3")]
    lines.append("## Interpretation of etalon-only sheets")
    lines.append(
        "Etalon often splits broker/DU (and margin forms) into separate sheets; "
        "ours currently keeps mainly DU / primary bucket sheets."
    )
    lines.append("")
    for n in only_e:
        if n == "_dropDownSheet":
            continue
        st = etalon_stats(wb_e[name_map_e[n]])
        lines.append(f"- `{n}` data~{st['data_rows']}")
    lines.append("")

    key = [
        "0420409 Раздел 1 Сведения о бан",
        "0420431 Раздел 1 Сведения об ос",
        "0420431 Раздел 4 Сведения о п_1",
        "0420431 Раздел 2 Сведения о п_1",
        "0420431 Раздел 7 Сведения о п_4",
        "0420431 Раздел 7 Сведения о п_5",
        "0420459 Раздел 1 Сведения о цен",
        "0420431 Раздел 1 Сведения об _3",
        "0420431 Раздел 1 Сведения об _2",
    ]
    lines.append("## Key sheets detail")
    for n in key:
        if n not in set_e or n not in set_o:
            lines.append(f"### `{n}` — missing on one side (e={n in set_e}, o={n in set_o})")
            continue
        se = etalon_stats(wb_e[name_map_e[n]])
        so = ours_stats(wb_o[name_map_o[n]])
        lines.append(f"### `{n}`")
        lines.append(
            f"- Etalon: max_row={se['max_row']}, data_start={se['data_start']}, data~{se['data_rows']}, cols={se['max_col']}"
        )
        lines.append(
            f"- Ours: max_row={so['max_row']}, data={so['data_rows']}, cols={so['max_col']}"
        )
        lines.append(f"- Etalon header labels (sample): {se['headers'][:8]}")
        lines.append(f"- Ours headers (concept names): {so['headers'][:10]}")
        # row ratio
        ratio = so["data_rows"] / se["data_rows"] if se["data_rows"] else None
        lines.append(f"- Row ratio ours/etalon: {ratio:.3f}" if ratio else "- Row ratio: n/a")
        lines.append("")

    # ours unique contract sheet
    if "0420431 Раздел 1 Договоры ДУ" in set_o:
        so = ours_stats(wb_o[name_map_o["0420431 Раздел 1 Договоры ДУ"]])
        lines.append("## Extra ours sheet")
        lines.append(
            f"- `0420431 Раздел 1 Договоры ДУ`: data={so['data_rows']} — "
            "likely etalon folds this into another section1 sheet or pivot."
        )
        lines.append("")

    lines.append("## Summary verdict")
    close = counts.get("CLOSE", 0) + counts.get("NEAR", 0)
    lines.append(
        f"- Of {len(common)} exact-name sheets: CLOSE/NEAR={close}, "
        f"MUCH_MORE={counts.get('OURS_MUCH_MORE', 0)}, "
        f"MUCH_LESS={counts.get('OURS_MUCH_LESS', 0)}, DIFF={counts.get('DIFF', 0)}"
    )
    lines.append(
        "- Structural difference: etalon = multi-row RU headers + taxonomy table layout; "
        "ours = flat one-header row with XBRL concept/axis codes."
    )
    lines.append(
        f"- Coverage: ours {len(wb_o.sheetnames)} sheets vs etalon {len(wb_e.sheetnames)} "
        "(etalon has broker/margin/extra splits)."
    )

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")

    print("Exact", len(common), "OnlyE", len(only_e), "OnlyO", len(only_o))
    print("Statuses", dict(counts))
    print("Key volumes:")
    for n in key:
        if n in set_e and n in set_o:
            se = etalon_stats(wb_e[name_map_e[n]])
            so = ours_stats(wb_o[name_map_o[n]])
            print(
                " ",
                n[:40],
                "e",
                se["data_rows"],
                "o",
                so["data_rows"],
                "ratio",
                round(so["data_rows"] / se["data_rows"], 3) if se["data_rows"] else None,
            )
    print("Wrote", OUT_MD)
    wb_e.close()
    wb_o.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
