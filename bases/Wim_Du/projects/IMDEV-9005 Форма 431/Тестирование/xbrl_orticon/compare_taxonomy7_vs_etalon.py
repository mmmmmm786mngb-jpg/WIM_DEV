#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sheet-by-sheet compare: XBRL_Orticon_taxonomy7.xlsx vs etalon converter.
Reports headers and data match per sheet + final summary.
"""

from __future__ import annotations

import re
import sys
import zipfile
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", "replace").decode("ascii"))


def norm_text(v):
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ")
    return re.sub(r"\s+", " ", s).strip()


def norm_cmp(v):
    s = norm_text(v)
    if not s:
        return ""
    if re.fullmatch(r"-?\d+,\d+", s):
        s = s.replace(",", ".")
    if re.fullmatch(r"-?\d+\.0+", s):
        s = s.split(".")[0]
    return s.lower()


def norm_header(v):
    return norm_cmp(v)


def form_code(name):
    m = re.search(r"0\d{6}", name or "")
    return m.group(0) if m else ""


def sheet_family(name):
    code = form_code(name)
    n = (name or "").lower()
    sec = "x"
    for i in range(1, 12):
        if ("раздел %d" % i) in n:
            sec = "r%d" % i
            break
    return "%s|%s" % (code or "other", sec)


def workbook_sheet_paths(z):
    wb = ET.fromstring(z.read("xl/workbook.xml"))
    rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    rid_to = {rel.get("Id"): rel.get("Target") for rel in rels}
    out = {}
    for sh in wb.findall("m:sheets/m:sheet", NS):
        name = sh.get("name")
        rid = sh.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = rid_to.get(rid, "")
        target = "xl/" + target.lstrip("/")
        if target.startswith("xl/xl/"):
            target = target[3:]
        out[name] = target
    return out


def sheet_dimensions_from_zip(z, sheet_path):
    try:
        root = ET.fromstring(z.read(sheet_path))
    except KeyError:
        return None, None
    dim = root.find("m:dimension", NS)
    if dim is None:
        return None, None
    ref = dim.get("ref") or ""
    m2 = re.match(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?", ref.upper())
    if not m2:
        return None, None

    def col_to_idx(letters):
        n = 0
        for ch in letters:
            n = n * 26 + (ord(ch) - 64)
        return n

    r2 = int(m2.group(4) or m2.group(2))
    c2 = col_to_idx(m2.group(3) or m2.group(1))
    return r2, c2


def read_sheet_rows(ws, max_row=None, max_col=None):
    kwargs = {"values_only": True}
    if max_row and max_row > 0:
        kwargs["max_row"] = max_row
    if max_col and max_col > 0:
        kwargs["max_col"] = max_col
    return list(ws.iter_rows(**kwargs))


def parse_sheet(raw_rows):
    if not raw_rows:
        return 0, [], []

    scan = min(15, len(raw_rows))
    best_r, best_n = 0, -1
    for r in range(scan):
        vals = raw_rows[r]
        n = sum(1 for v in vals if norm_text(v))
        a = norm_text(vals[0]) if vals else ""
        if a in ("TOC", "Generator") or a.startswith("http") or a.startswith("Generator:"):
            continue
        if a.startswith("T=") or a.startswith("Z="):
            continue
        if a == "Наименование показателя":
            continue
        if n == 1 and "идентификатор" in a.lower():
            continue
        if n > best_n and n >= 3:
            best_r, best_n = r, n

    headers = [norm_text(v) for v in raw_rows[best_r]]
    while headers and not headers[-1]:
        headers.pop()
    max_c = max(len(headers), 1)
    for r in range(best_r, min(best_r + 5, len(raw_rows))):
        max_c = max(max_c, len(raw_rows[r]))

    data = []
    for r in range(best_r + 1, len(raw_rows)):
        src = raw_rows[r]
        vals = []
        nonempty = 0
        for c in range(max_c):
            v = src[c] if c < len(src) else None
            nv = norm_cmp(v)
            vals.append(nv)
            if nv:
                nonempty += 1
        if nonempty == 0:
            continue
        a0 = vals[0] if vals else ""
        if nonempty == 1 and "идентификатор" in a0:
            continue
        filled = [v for v in vals if v]
        if filled and all(re.fullmatch(r"\d+", v) for v in filled) and len(filled) >= 3:
            continue
        data.append(vals)
    return best_r + 1, headers, data


def value_multiset(rows):
    c = Counter()
    for row in rows:
        for v in row:
            if v:
                c[v] += 1
    return c


def load_all(path):
    safe_print("  load %s ..." % path.name)
    with zipfile.ZipFile(path) as z:
        paths = workbook_sheet_paths(z)
        dims = {name: sheet_dimensions_from_zip(z, sp) for name, sp in paths.items()}

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        max_r, max_c = dims.get(name, (None, None))
        if not max_r or max_r < 2:
            max_r = max(int(wb[name].max_row or 0), 50000)
        if not max_c or max_c < 2:
            max_c = 60
        raw = read_sheet_rows(wb[name], max_row=max_r, max_col=max_c)
        hdr_r, headers, rows = parse_sheet(raw)
        sheets[name] = {
            "hdr_r": hdr_r,
            "headers": headers,
            "headers_norm": [norm_header(h) for h in headers if norm_header(h)],
            "n_cols": max(len(headers), max_c if rows else len(headers)),
            "n_rows": len(rows),
            "rows": rows,
            "values": value_multiset(rows),
        }
        safe_print("    %s: rows=%d cols=%d" % (name[:45], len(rows), sheets[name]["n_cols"]))
    wb.close()
    return sheets


def match_sheets(etalon_names, ours_names, etalon_info, ours_info):
    used_o = set()
    pairs = []
    only_e = []

    def score(a, b):
        fa, fb = sheet_family(a), sheet_family(b)
        s = 0
        if fa == fb and not fa.startswith("other"):
            s += 50
        elif form_code(a) and form_code(a) == form_code(b):
            s += 20
        ta = set(re.findall(r"[a-zа-я0-9]+", a.lower()))
        tb = set(re.findall(r"[a-zа-я0-9]+", b.lower()))
        if ta and tb:
            s += int(30 * len(ta & tb) / max(len(ta | tb), 1))
        er = etalon_info.get(a, {}).get("n_rows", -1)
        or_ = ours_info.get(b, {}).get("n_rows", -1)
        if er >= 0 and or_ >= 0:
            if er == or_:
                s += 40
            elif er > 0 and or_ > 0:
                ratio = min(er, or_) / max(er, or_)
                if ratio >= 0.9:
                    s += 25
                elif ratio >= 0.5:
                    s += 10
            elif er == 0 and or_ == 0:
                s += 15
        return s

    e_order = sorted(
        [n for n in etalon_names if n != "_dropDownSheet"],
        key=lambda n: -etalon_info.get(n, {}).get("n_rows", 0),
    )
    for e in e_order:
        best, best_s = None, -1
        for o in ours_names:
            if o in used_o:
                continue
            sc = score(e, o)
            if sc > best_s:
                best_s, best = sc, o
        if best is not None and best_s >= 25:
            pairs.append((e, best, best_s))
            used_o.add(best)
        else:
            only_e.append(e)
    only_o = [o for o in ours_names if o not in used_o]
    return pairs, only_e, only_o


def compare_headers(e_headers, o_headers):
    e = [h for h in e_headers if h]
    o = [h for h in o_headers if h]
    e_set = set(e)
    o_set = set(o)
    common = e_set & o_set
    only_e = e_set - o_set
    only_o = o_set - e_set

    # fuzzy pairs for wording diffs
    fuzzy = []
    for eh in only_e:
        et = set(re.findall(r"[а-яa-z0-9]{4,}", eh.lower()))
        best, bov = None, 0
        for oh in only_o:
            ot = set(re.findall(r"[а-яa-z0-9]{4,}", oh.lower()))
            ov = len(et & ot)
            if ov > bov:
                bov, best = ov, oh
        if best and bov >= 3:
            fuzzy.append((eh, best))

    exact_match = len(only_e) == 0 and len(only_o) == 0
    near_match = exact_match or (
        len(common) >= max(len(e_set), len(o_set)) * 0.8 and len(fuzzy) <= 2
    )
    return {
        "e_count": len(e),
        "o_count": len(o),
        "common": len(common),
        "only_e": sorted(only_e),
        "only_o": sorted(only_o),
        "fuzzy": fuzzy,
        "exact_match": exact_match,
        "near_match": near_match,
    }


def data_recall(es, os_):
    vc = {"a_total": sum(es["values"].values()), "b_total": sum(os_["values"].values())}
    shared = 0
    for k in set(es["values"]) & set(os_["values"]):
        shared += min(es["values"][k], os_["values"][k])
    recall = 100.0 * shared / vc["a_total"] if vc["a_total"] else (100.0 if vc["b_total"] == 0 else 0.0)
    precision = 100.0 * shared / vc["b_total"] if vc["b_total"] else 0.0
    return recall, precision, shared, vc


@dataclass
class SheetResult:
    idx: int
    e_name: str
    o_name: str
    e_rows: int
    o_rows: int
    recall: float
    precision: float
    headers_ok: bool
    headers_exact: bool
    data_ok: bool
    notes: list = field(default_factory=list)
    header_issues: list = field(default_factory=list)


def classify_sheet(res: SheetResult):
    if res.e_rows == 0 and res.o_rows == 0:
        return "OK_EMPTY"
    if res.headers_exact and res.data_ok:
        return "OK"
    if res.headers_ok and res.data_ok:
        return "OK_MINOR_HDR"
    if res.data_ok and not res.headers_ok:
        return "HDR_DIFF"
    if res.headers_ok and not res.data_ok:
        return "DATA_DIFF"
    if res.recall >= 80 and res.headers_ok:
        return "DATA_NEAR"
    return "FAIL"


def main():
    base = Path(__file__).resolve().parents[2] / "ОРТИКОН"
    etalon_p = base / "0420431_409_январь_2026_конвертер.xlsx"
    ours_p = base / "XBRL_Orticon_taxonomy7.xlsx"
    out_md = Path(__file__).resolve().parent / "compare_taxonomy7_vs_etalon_report.md"

    if not etalon_p.exists() or not ours_p.exists():
        safe_print("Missing files")
        return 1

    safe_print("=== LOAD ===")
    etalon = load_all(etalon_p)
    ours = load_all(ours_p)

    pairs, only_e, only_o = match_sheets(
        list(etalon.keys()), list(ours.keys()), etalon, ours)

    safe_print("")
    safe_print("=== SHEET BY SHEET ===")
    results = []
    for idx, (e_name, o_name, sc) in enumerate(sorted(pairs, key=lambda x: x[0]), 1):
        es, os_ = etalon[e_name], ours[o_name]
        hc = compare_headers(es["headers_norm"], os_["headers_norm"])
        recall, precision, shared, vc = data_recall(es, os_)

        row_diff = abs(es["n_rows"] - os_["n_rows"])
        row_tol = max(5, int(0.05 * max(es["n_rows"], 1)))
        data_ok = recall >= 95.0 or (
            es["n_rows"] == 0 and os_["n_rows"] == 0
        ) or (
            recall >= 85.0 and row_diff <= row_tol
        )

        header_issues = []
        for eh, oh in hc["fuzzy"]:
            el, ol = eh.lower(), oh.lower()
            if ("брокер" in el) != ("брокер" in ol) or ("кредитн" in el) != ("кредитн" in ol):
                header_issues.append("wording: E=%s | O=%s" % (eh[:80], oh[:80]))
        for h in hc["only_e"][:5]:
            if len(h) > 15:
                header_issues.append("only etalon: %s" % h[:90])
        for h in hc["only_o"][:5]:
            if len(h) > 15:
                header_issues.append("only ours: %s" % h[:90])

        notes = []
        if row_diff > row_tol:
            notes.append("rows %d vs %d" % (es["n_rows"], os_["n_rows"]))
        if recall < 95 and es["n_rows"] > 0:
            notes.append("recall %.1f%%" % recall)

        res = SheetResult(
            idx=idx,
            e_name=e_name,
            o_name=o_name,
            e_rows=es["n_rows"],
            o_rows=os_["n_rows"],
            recall=recall,
            precision=precision,
            headers_ok=hc["near_match"],
            headers_exact=hc["exact_match"],
            data_ok=data_ok,
            notes=notes,
            header_issues=header_issues,
        )
        status = classify_sheet(res)
        safe_print("")
        safe_print("[%02d] %s" % (idx, status))
        safe_print("  E: %s" % e_name[:70])
        safe_print("  O: %s" % o_name[:70])
        safe_print("  rows: %d / %d | recall: %.1f%% | precision: %.1f%%" % (
            es["n_rows"], os_["n_rows"], recall, precision))
        safe_print("  headers: exact=%s near=%s (e=%d o=%d common=%d)" % (
            hc["exact_match"], hc["near_match"], hc["e_count"], hc["o_count"], hc["common"]))
        if header_issues:
            for hi in header_issues[:4]:
                safe_print("  HDR: %s" % hi[:120])
        if notes:
            safe_print("  NOTE: %s" % "; ".join(notes))
        res.status = status
        results.append(res)

    # summary
    counts = Counter(r.status for r in results)
    safe_print("")
    safe_print("=== SUMMARY ===")
    safe_print("Matched sheets: %d" % len(results))
    safe_print("Only etalon: %d" % len(only_e))
    safe_print("Only ours: %d" % len(only_o))
    for st in ("OK", "OK_EMPTY", "OK_MINOR_HDR", "HDR_DIFF", "DATA_DIFF", "DATA_NEAR", "FAIL"):
        if counts[st]:
            safe_print("  %s: %d" % (st, counts[st]))

    hdr_ok = sum(1 for r in results if r.headers_ok)
    data_ok_n = sum(1 for r in results if r.data_ok)
    both_ok = sum(1 for r in results if r.headers_ok and r.data_ok)

    safe_print("")
    safe_print("Headers OK (near/exact): %d / %d" % (hdr_ok, len(results)))
    safe_print("Data OK (recall>=95%% or empty): %d / %d" % (data_ok_n, len(results)))
    safe_print("Both headers+data OK: %d / %d" % (both_ok, len(results)))

    problem_hdr = [r for r in results if not r.headers_ok]
    problem_data = [r for r in results if not r.data_ok]

    if problem_hdr:
        safe_print("")
        safe_print("Header problems:")
        for r in problem_hdr:
            safe_print("  - %s" % r.e_name[:50])
    if problem_data:
        safe_print("")
        safe_print("Data problems:")
        for r in problem_data:
            safe_print("  - %s recall=%.1f%% rows=%d/%d" % (
                r.e_name[:50], r.recall, r.e_rows, r.o_rows))

    # write md report
    lines = ["# taxonomy7 vs etalon", ""]
    lines.append("## Summary")
    lines.append("- matched: %d" % len(results))
    lines.append("- headers OK: %d / %d" % (hdr_ok, len(results)))
    lines.append("- data OK: %d / %d" % (data_ok_n, len(results)))
    lines.append("- both OK: %d / %d" % (both_ok, len(results)))
    lines.append("")
    lines.append("| # | status | etalon | ours | e_rows | o_rows | recall% | headers | data | notes |")
    lines.append("|---:|---|---|---|---:|---:|---:|---|---|---|")
    for r in results:
        lines.append(
            "| %d | %s | %s | %s | %d | %d | %.1f | %s | %s | %s |"
            % (
                r.idx,
                r.status,
                r.e_name.replace("|", "/")[:40],
                r.o_name.replace("|", "/")[:40],
                r.e_rows,
                r.o_rows,
                r.recall,
                "OK" if r.headers_ok else "DIFF",
                "OK" if r.data_ok else "DIFF",
                "; ".join(r.notes + r.header_issues[:2])[:120].replace("|", "/"),
            )
        )
    out_md.write_text("\n".join(lines) + "\n", encoding="utf-8")
    safe_print("")
    safe_print("Report: %s" % out_md)
    return 0


if __name__ == "__main__":
    sys.exit(main())
