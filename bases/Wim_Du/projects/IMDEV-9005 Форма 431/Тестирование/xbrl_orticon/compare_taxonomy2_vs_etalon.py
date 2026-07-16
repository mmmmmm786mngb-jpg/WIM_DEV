#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare XBRL_Orticon_taxonomy2.xlsx vs etalon converter Excel."""

from openpyxl import load_workbook
from pathlib import Path
import re
import sys


def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", "replace").decode("ascii"))


def clean_h(v):
    if v is None:
        return None
    return re.sub(r"\s+", " ", str(v)).strip().lower()


def row_vals(ws, r, maxc=25):
    return [ws.cell(r, c).value for c in range(1, maxc + 1)]


def count_nonempty_rows(ws, max_rows=50000):
    filled = 0
    for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, max_rows), max_col=3, values_only=True):
        if any(x is not None and str(x).strip() != "" for x in r):
            filled += 1
    return filled


def header_cyr_stats(wb):
    total = 0
    cyr = 0
    for s in wb.sheetnames:
        if s in ("TOC", "_dropDownSheet"):
            continue
        ws = wb[s]
        best = None
        bestn = -1
        for r in range(1, 9):
            vals = [ws.cell(r, c).value for c in range(1, 30)]
            n = sum(1 for v in vals if v is not None and str(v).strip())
            if n > bestn:
                bestn = n
                best = vals
        if not best:
            continue
        for v in best:
            if v is None:
                continue
            total += 1
            if re.search(r"[А-Яа-яЁё]", str(v)):
                cyr += 1
    return cyr, total


def main():
    base = Path(__file__).resolve().parents[2] / "ОРТИКОН"
    etalon_p = base / "0420431_409_январь_2026_конвертер.xlsx"
    ours_p = base / "XBRL_Orticon_taxonomy2.xlsx"

    safe_print("etalon=%s exists=%s" % (etalon_p.name, etalon_p.exists()))
    safe_print("ours=%s exists=%s" % (ours_p.name, ours_p.exists()))
    if not etalon_p.exists() or not ours_p.exists():
        return 1

    etalon = load_workbook(etalon_p, read_only=True, data_only=True)
    ours = load_workbook(ours_p, read_only=True, data_only=True)

    safe_print("ETALON sheets: %d" % len(etalon.sheetnames))
    safe_print("OURS sheets: %d" % len(ours.sheetnames))
    safe_print("OURS: %s" % ", ".join(ours.sheetnames))

    etalon_set = set(etalon.sheetnames)
    ours_set = set(ours.sheetnames)
    safe_print("exact common sheets: %d" % len(etalon_set & ours_set))
    only_e = sorted(etalon_set - ours_set)
    only_o = sorted(ours_set - etalon_set)
    safe_print("only etalon (%d): %s" % (len(only_e), "; ".join(only_e[:25])))
    safe_print("only ours (%d): %s" % (len(only_o), "; ".join(only_o[:25])))

    name = "0420409 Раздел 1 Сведения о бан"
    ews = etalon[name]
    ows = ours[name]

    safe_print("")
    safe_print("=== 0420409 meta (etalon) ===")
    for r in range(1, 9):
        vals = [v for v in row_vals(ews, r) if v is not None]
        if vals:
            shown = [str(v).replace("\n", " ")[:90] for v in vals[:8]]
            safe_print("R%d: %s" % (r, shown))

    safe_print("")
    safe_print("=== 0420409 meta (ours) ===")
    for r in range(1, 9):
        vals = [v for v in row_vals(ows, r) if v is not None]
        if vals:
            shown = [str(v).replace("\n", " ")[:90] for v in vals[:8]]
            safe_print("R%d: %s" % (r, shown))

    # detect ours header row
    ours_header_row = 6
    for r in range(1, 10):
        v = ows.cell(r, 1).value
        if v and ("SokrNaim" in str(v) or "Сокращенное" in str(v) or "Rek_kred" in str(v)):
            # if row looks like headers
            filled = sum(1 for c in range(1, 15) if ows.cell(r, c).value)
            if filled >= 5:
                ours_header_row = r
                break

    etalon_hdrs = []
    for c in range(1, 25):
        v = ews.cell(7, c).value
        if v:
            etalon_hdrs.append((c, clean_h(v), str(v).replace("\n", " ")[:90]))

    ours_hdrs = []
    for c in range(1, 25):
        v = ows.cell(ours_header_row, c).value
        if v:
            ours_hdrs.append((c, clean_h(v), str(v).replace("\n", " ")[:90]))

    safe_print("")
    safe_print("=== Headers etalon R7 ===")
    for c, _, t in etalon_hdrs:
        safe_print("  C%d: %s" % (c, t))

    safe_print("")
    safe_print("=== Headers ours R%d ===" % ours_header_row)
    for c, _, t in ours_hdrs:
        safe_print("  C%d: %s" % (c, t))

    # Compare: etalon concept headers usually start at C2; ours may start with Rek at C1
    e_concepts = [(c, h, t) for c, h, t in etalon_hdrs if c >= 2]
    o_concepts = list(ours_hdrs)
    start_o = 0
    if o_concepts and (
        o_concepts[0][1].startswith("rek_")
        or "rek_kred" in o_concepts[0][1]
        or "идентификатор банковского" in (o_concepts[0][1] or "")
    ):
        start_o = 1

    safe_print("")
    safe_print("=== Header match (etalon C2+ vs ours after Rek) ===")
    matched = 0
    soft = 0
    for i, (ec, eh, et) in enumerate(e_concepts):
        oi = start_o + i
        if oi >= len(o_concepts):
            safe_print("MISSING ours for etalon C%d: %s" % (ec, et))
            continue
        oc, oh, ot = o_concepts[oi]
        exact = eh == oh
        soft_ok = (eh in oh) or (oh in eh) if eh and oh else False
        if exact:
            matched += 1
            status = "OK"
        elif soft_ok:
            soft += 1
            status = "SOFT"
        else:
            status = "DIFF"
        safe_print("%s etalonC%d vs oursC%d" % (status, ec, oc))
        safe_print("   E: %s" % et)
        safe_print("   O: %s" % ot)
    safe_print(
        "Matched exact=%d soft=%d of %d"
        % (matched, soft, len(e_concepts))
    )

    # types
    data_row_o = ours_header_row + 1
    safe_print("")
    safe_print("=== Types ours R%d ===" % data_row_o)
    for c in range(1, 14):
        val = ows.cell(data_row_o, c).value
        if val is not None:
            safe_print(
                "  C%d: %s %r"
                % (c, type(val).__name__, str(val)[:60])
            )

    safe_print("")
    safe_print("=== Types etalon R11 ===")
    for c in range(1, 14):
        val = ews.cell(11, c).value
        if val is not None:
            safe_print(
                "  C%d: %s %r"
                % (c, type(val).__name__, str(val)[:60])
            )

    # account/INN must be str
    acc_ok = isinstance(ows.cell(data_row_o, 1).value, str)
    # find Nom_schet / INN columns by header
    inn_ok = None
    nom_ok = None
    for c, h, t in ours_hdrs:
        val = ows.cell(data_row_o, c).value
        if h and ("инн" in h or h == "inn_tin"):
            inn_ok = isinstance(val, str)
            safe_print("INN col C%d type=%s val=%r" % (c, type(val).__name__, str(val)[:40]))
        if h and (h.startswith("nom_schet") or h.startswith("номер счета")):
            nom_ok = isinstance(val, str)
            safe_print("Nom col C%d type=%s val=%r" % (c, type(val).__name__, str(val)[:40]))
    safe_print("Account C1 is str: %s" % acc_ok)
    safe_print("INN is str: %s" % inn_ok)
    safe_print("Nom is str: %s" % nom_ok)

    ec, et = header_cyr_stats(etalon)
    oc, ot = header_cyr_stats(ours)
    safe_print("")
    safe_print(
        "Header cyrillic: etalon %d/%d=%.1f%% | ours %d/%d=%.1f%%"
        % (ec, et, 100.0 * ec / et if et else 0, oc, ot, 100.0 * oc / ot if ot else 0)
    )

    safe_print("")
    safe_print("=== Sheet volumes (nonempty rows in first 3 cols) ===")
    for s in ours.sheetnames:
        if s == "TOC":
            continue
        or_ = count_nonempty_rows(ours[s])
        er = count_nonempty_rows(etalon[s]) if s in etalon.sheetnames else None
        if er is None:
            mark = "NEW"
        elif abs(er - or_) <= max(3, int(er * 0.05)):
            mark = "OK"
        else:
            mark = "DIFF"
        safe_print(
            "%-4s %-42s ours=%-6s etalon=%s"
            % (mark, s[:42], or_, er)
        )

    # generator version
    for r in range(1, 8):
        for c in range(1, 4):
            v = ows.cell(r, c).value
            if v and "Generator" in str(v) or (v and "v1." in str(v)):
                safe_print("Generator cell R%dC%d: %s" % (r, c, v))

    # TOC generator
    if "TOC" in ours.sheetnames:
        tw = ours["TOC"]
        for r in range(1, 8):
            vals = [tw.cell(r, c).value for c in range(1, 4)]
            if any(vals):
                safe_print("TOC R%d: %s" % (r, vals))

    return 0


if __name__ == "__main__":
    sys.exit(main())
