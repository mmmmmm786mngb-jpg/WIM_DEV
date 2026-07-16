#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Compare headers E vs O for all multi-section forms.
Also map EP/table keys vs lab alias collisions in taxonomy.
"""

import os
import re
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431\ОРТИКОН")
ETALON = BASE / "0420431_409_январь_2026_конвертер.xlsx"
OURS = BASE / "XBRL_Orticon_taxonomy6.xlsx"
TPL = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431"
    r"\Обработки\внВыгрузкаXBRLОртиконВXLSX\Templates\Таксономия_20251230\Ext\Template.bin"
)


def N(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ").replace("\n", " ")).strip()


def sheet_sec(name):
    m = re.search(r"раздел\s*(\d+)", (name or "").lower())
    return int(m.group(1)) if m else None


def form_code(name):
    m = re.search(r"0\d{6}", name or "")
    return m.group(0) if m else ""


def headers_row(ws):
    rows = list(
        ws.iter_rows(
            min_row=1,
            max_row=30,
            max_col=min(ws.max_column or 60, 60),
            values_only=True,
        )
    )
    best = []
    best_score = -1
    for row in rows:
        vals = [N(v) for v in row]
        nonempty = [v for v in vals if v]
        if len(nonempty) < 3:
            continue
        if all(re.fullmatch(r"\d+", v) for v in nonempty):
            continue
        j = " ".join(nonempty).lower()
        if nonempty[0] in ("TOC",) or nonempty[0].startswith("http") or nonempty[0].startswith("Generator"):
            continue
        if nonempty[0].startswith("T="):
            continue
        score = len(nonempty)
        for w in (
            "наименование",
            "идентификатор",
            "сумма",
            "код ",
            "инн",
            "дата",
            "вид ",
            "количество",
            "стоимость",
        ):
            if w in j:
                score += 3
        if "сокращенн" in j:
            score += 10
        if score > best_score:
            best_score = score
            best = nonempty
    return best


def wording_markers(text):
    t = text.lower()
    marks = []
    for m in (
        "кредитн",
        "брокер",
        "клиринг",
        "депозитар",
        "управляющ",
        "клиент",
        "эмитент",
        "пайщик",
        "акционер",
        "заемщик",
        "контрагент",
        "страхов",
        "банк",
        "нфо",
        "организац",
    ):
        if m in t:
            marks.append(m)
    return tuple(marks)


def compare_sheets():
    wb_e = load_workbook(ETALON, read_only=True, data_only=True)
    wb_o = load_workbook(OURS, read_only=True, data_only=True)

    e_map = defaultdict(dict)  # code -> sec -> sheet
    o_map = defaultdict(dict)
    for s in wb_e.sheetnames:
        if s == "_dropDownSheet":
            continue
        c, sec = form_code(s), sheet_sec(s)
        if c and sec:
            e_map[c][sec] = s
    for s in wb_o.sheetnames:
        c, sec = form_code(s), sheet_sec(s)
        if c and sec:
            o_map[c][sec] = s

    print("=== Multi-section forms present ===")
    for c in sorted(set(e_map) | set(o_map)):
        print(
            c,
            "E secs",
            sorted(e_map.get(c, {})),
            "O secs",
            sorted(o_map.get(c, {})),
        )

    print("\n=== Header wording mismatches (marker swap / substantial) ===")
    issues = []
    for c in sorted(set(e_map) & set(o_map)):
        for sec in sorted(set(e_map[c]) & set(o_map[c])):
            se, so = e_map[c][sec], o_map[c][sec]
            eh, oh = headers_row(wb_e[se]), headers_row(wb_o[so])
            if not eh or not oh:
                continue
            for h in eh:
                if len(h) < 12:
                    continue
                et = set(re.findall(r"[а-яa-z0-9]{4,}", h.lower()))
                best, bov = None, 0
                for ohv in oh:
                    ot = set(re.findall(r"[а-яa-z0-9]{4,}", ohv.lower()))
                    ov = len(et & ot)
                    if ov > bov:
                        bov, best = ov, ohv
                if not best or bov < 3:
                    continue
                if N(h).lower() == N(best).lower():
                    continue
                em, om = wording_markers(h), wording_markers(best)
                # ignore pure length/punctuation diffs without marker change
                marker_diff = set(em) != set(om)
                # also flag if first 40 chars differ a lot but shared topic
                if not marker_diff and h[:50].lower() == best[:50].lower():
                    continue
                if marker_diff or (bov >= 4 and abs(len(h) - len(best)) > 25):
                    issues.append((c, sec, h, best, em, om, marker_diff))

    if not issues:
        print("(none found on taxonomy6 vs etalon with current heuristic)")
    for it in issues:
        c, sec, h, best, em, om, md = it
        print("%s R%d marker_diff=%s" % (c, sec, md))
        print("  E marks", em)
        print("  O marks", om)
        print("  E:", h[:120])
        print("  O:", best[:120])
        print()

    # specifically list section pairs where E has marker X and O has different for same Sok/INN-like
    print("=== Per-section credit/broker/depo marker counts ===")
    for c in sorted(set(e_map) & set(o_map)):
        for sec in sorted(set(e_map[c]) & set(o_map[c])):
            eh = headers_row(wb_e[e_map[c][sec]])
            oh = headers_row(wb_o[o_map[c][sec]])
            def cnt(hs, m):
                return sum(1 for h in hs if m in h.lower())
            line = "%s R%d | E broker=%d credit=%d depo=%d | O broker=%d credit=%d depo=%d" % (
                c,
                sec,
                cnt(eh, "брокер"),
                cnt(eh, "кредитн"),
                cnt(eh, "депозитар"),
                cnt(oh, "брокер"),
                cnt(oh, "кредитн"),
                cnt(oh, "депозитар"),
            )
            # highlight discrepancies
            if (
                (cnt(eh, "брокер") and not cnt(oh, "брокер") and cnt(oh, "кредитн"))
                or (cnt(eh, "кредитн") and not cnt(oh, "кредитн") and cnt(oh, "брокер"))
                or (cnt(eh, "депозитар") != cnt(oh, "депозитар") and abs(cnt(eh, "депозитар") - cnt(oh, "депозитар")) >= 1 and (cnt(eh, "депозитар") or cnt(oh, "депозитар")))
            ):
                line += "  *** SUSPECT ***"
            print(line)

    wb_e.close()
    wb_o.close()


def full_key(name):
    name = os.path.basename(name)
    if name.lower().endswith(".xml"):
        name = name[:-4]
    low = name.lower()
    if low.endswith("-rend"):
        name = name[:-5]
    elif low.endswith("-lab"):
        name = name[:-4]
    return name


def short_key_old(name):
    """Pre-fix short key logic."""
    m = re.search(r"SR_(\d{7})", name, re.I)
    if not m:
        return ""
    code = m.group(1)
    if name.upper().count("SR_") >= 2:
        return code + "_2"
    return code


def short_key_new(name):
    m = re.search(r"SR_(\d{7})", name, re.I)
    if not m:
        return ""
    code = m.group(1)
    up = name.upper()
    if up.count("SR_") >= 2:
        p1 = up.find("SR_")
        rest = up[p1 + 3 :]
        p2 = rest.find("SR_")
        if p2 >= 0:
            abs_p = p1 + 3 + p2
            tail = name[abs_p + 3 :]
            mm = re.match(r"[0-9_]+", tail)
            if mm and len(mm.group(0).rstrip("_")) >= 7:
                return mm.group(0).rstrip("_")
        return code + "_2"
    if "_r2" in name.lower():
        return code + "_r2"
    return code


def taxonomy_collision_report():
    print("\n=== Taxonomy lab collisions OLD short-key vs NEW ===")
    with zipfile.ZipFile(TPL, "r") as z:
        labs = [
            n
            for n in z.namelist()
            if n.endswith("-lab.xml")
            and "/tab/" in n.replace("\\", "/")
            and "purcb" in n
        ]
        by_code = defaultdict(list)
        for n in labs:
            bn = os.path.basename(n)
            m = re.search(r"SR_(\d{7})", bn, re.I)
            if m:
                by_code[m.group(1)].append(bn)

        for code in ("0420409", "0420414", "0420431", "0420459"):
            print("\n--", code, "labs", len(by_code.get(code, [])), "--")
            old_map = defaultdict(list)
            new_map = defaultdict(list)
            full_map = {}
            for bn in sorted(by_code.get(code, [])):
                old_map[short_key_old(bn)].append(bn)
                new_map[short_key_new(bn)].append(bn)
                full_map[bn] = full_key(bn)

            print("OLD short-key collisions (same key, multiple labs):")
            for k, files in sorted(old_map.items()):
                if len(files) > 1:
                    print("  ", k, "x", len(files))
                    for f in files[:8]:
                        print("     ", f)
                    if len(files) > 8:
                        print("      ...", len(files) - 8, "more")

            print("NEW short-key collisions:")
            for k, files in sorted(new_map.items()):
                if len(files) > 1:
                    print("  ", k, "x", len(files))
                    for f in files[:8]:
                        print("     ", f)

            print("Full EP keys unique?", len(full_map) == len(set(full_map.values())))
            # for 0420431 show sample short->full for different sections
            if code == "0420431":
                print("Sample R-section filenames -> new short / full:")
                for bn in sorted(by_code[code]):
                    if re.search(r"_r_?\d", bn.lower()) or "R1" in bn or "R2" in bn:
                        print(
                            " ",
                            bn,
                            "=> short",
                            short_key_new(bn),
                            "| full",
                            full_key(bn),
                        )


if __name__ == "__main__":
    compare_sheets()
    taxonomy_collision_report()
