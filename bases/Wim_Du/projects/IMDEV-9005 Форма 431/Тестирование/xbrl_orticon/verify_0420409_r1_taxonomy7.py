#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Verify 0420409 R1: taxonomy7 vs etalon (same checks as taxonomy6)."""

import re
import zipfile
from collections import Counter
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431\ОРТИКОН")
E_PATH = BASE / "0420431_409_январь_2026_конвертер.xlsx"
O_PATH = BASE / "XBRL_Orticon_taxonomy7.xlsx"


def N(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ").replace("\n", " ")).strip()


def Nl(v):
    return N(v).lower()


def norm_amt(v):
    s = N(v).replace(" ", "").replace(",", ".")
    if not s:
        return ""
    try:
        f = float(s)
        if abs(f - int(f)) < 1e-9:
            return str(int(f))
        return ("%.2f" % f).rstrip("0").rstrip(".")
    except Exception:
        return s.lower()


def sheet_dim(path, name):
    with zipfile.ZipFile(path) as z:
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid = {r.get("Id"): r.get("Target") for r in rels}
        for sh in wb.findall("m:sheets/m:sheet", NS):
            if sh.get("name") != name:
                continue
            ridv = sh.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            )
            sp = "xl/" + rid[ridv].lstrip("/")
            root = ET.fromstring(z.read(sp))
            dim = root.find("m:dimension", NS)
            ref = dim.get("ref", "A1:Z5000") if dim is not None else "A1:Z5000"
            m = re.match(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?", ref.upper())

            def col(letters):
                n = 0
                for c in letters:
                    n = n * 26 + ord(c) - 64
                return n

            return int(m.group(4) or m.group(2)), col(m.group(3) or m.group(1))
    return 5000, 40


def find_sheet(wb, code, sec):
    needle = "раздел %d" % sec
    for s in wb.sheetnames:
        if code in s and needle in s.lower():
            return s
    return None


def parse_bank_sheet(path, sheet_name, is_ours):
    mr, mc = sheet_dim(path, sheet_name)
    if (not mr or mr < 2) and not is_ours:
        mr, mc = 5000, 40
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(
        ws.iter_rows(
            min_row=1, max_row=mr, max_col=min(mc or 40, 40), values_only=True
        )
    )
    wb.close()

    hdr_i = None
    headers = []
    for i, row in enumerate(rows[:25]):
        vals = [N(v) for v in row]
        j = " ".join(vals).lower()
        n = sum(1 for v in vals if v)
        if n < 5:
            continue
        if (
            "сокращенн" in j
            or ("номер счета" in j and "инн" in j)
            or "идентификатор банков" in j
        ):
            hdr_i = i
            headers = vals
            break
    if hdr_i is None:
        return {"error": "no header", "rows": [], "headers": [], "col": {}}

    col = {}
    for i, h in enumerate(headers):
        hl = h.lower()
        if not hl:
            continue
        if "идентификатор банков" in hl:
            col.setdefault("id", i)
        elif "сокращенн" in hl:
            col.setdefault("name", i)
        elif "инн" in hl or ", tin" in hl:
            col.setdefault("inn", i)
        elif "код страны" in hl:
            col.setdefault("country", i)
        elif "номер счета" in hl:
            col.setdefault("acc", i)
        elif "вид счета" in hl:
            col.setdefault("kind", i)
        elif "код валюты" in hl:
            col.setdefault("ccy", i)
        elif "дата открытия" in hl:
            col.setdefault("open", i)
        elif "дата закрытия" in hl:
            col.setdefault("close", i)
        elif "остатка" in hl and "начало" in hl:
            col.setdefault("bal_beg", i)
        elif "остатка" in hl and "конец" in hl:
            col.setdefault("bal_end", i)
        elif "зачислен" in hl:
            col.setdefault("dt", i)
        elif "списан" in hl:
            col.setdefault("kt", i)

    data = []
    for i in range(hdr_i + 1, len(rows)):
        vals = [N(v) for v in rows[i]]
        if not any(vals):
            continue
        filled = [v for v in vals if v]
        if filled and all(re.fullmatch(r"\d+", v) for v in filled) and len(filled) >= 3:
            continue
        if len(filled) == 1 and "идентификатор" in filled[0].lower():
            continue

        acc = ""
        if "acc" in col and col["acc"] < len(vals):
            acc = vals[col["acc"]]
        if not re.fullmatch(r"\d{20}", acc):
            for v in vals:
                if re.fullmatch(r"\d{20}", v):
                    acc = v
                    break
        if not acc and "id" in col and col["id"] < len(vals):
            if re.fullmatch(r"\d{20}", vals[col["id"]]):
                acc = vals[col["id"]]
        if not acc:
            continue

        def g(key):
            if key not in col or col[key] >= len(vals):
                return ""
            return vals[col[key]]

        rec = {
            "acc": acc,
            "name": g("name"),
            "inn": g("inn"),
            "country": g("country"),
            "kind": g("kind"),
            "ccy": g("ccy"),
            "open": g("open")[:10],
            "close": g("close")[:10],
            "bal_beg": norm_amt(g("bal_beg")),
            "bal_end": norm_amt(g("bal_end")),
            "dt": norm_amt(g("dt")),
            "kt": norm_amt(g("kt")),
            "raw_amts": [],
            "raw_amt_texts": [],
        }
        for v in vals:
            if not re.search(r"\d", v):
                continue
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", v):
                continue
            if re.fullmatch(r"\d{10}", v) or re.fullmatch(r"\d{20}", v):
                continue
            if re.fullmatch(r"\d{3}", v):  # country code alone
                continue
            if re.fullmatch(r"\d{3}-[A-Z]{3}", v, re.I):
                continue
            a = norm_amt(v)
            if a and re.match(r"^-?\d", a):
                rec["raw_amts"].append(a)
                rec["raw_amt_texts"].append(v)
        data.append(rec)

    return {"headers": headers, "col": col, "rows": data, "hdr_i": hdr_i}


def main():
    print("Loading sheets...")
    wb_e = load_workbook(E_PATH, read_only=True)
    wb_o = load_workbook(O_PATH, read_only=True)
    e_name = find_sheet(wb_e, "0420409", 1)
    o_name = find_sheet(wb_o, "0420409", 1)
    print("E sheet:", e_name)
    print("O sheet:", o_name)
    # generator version
    for row in wb_o[o_name].iter_rows(min_row=1, max_row=5, max_col=1, values_only=True):
        if row[0] and "Generator" in str(row[0]):
            print("O version:", row[0])
    wb_e.close()
    wb_o.close()

    print("Parsing etalon...")
    E = parse_bank_sheet(E_PATH, e_name, False)
    print("Parsing ours...")
    O = parse_bank_sheet(O_PATH, o_name, True)
    print("E accounts:", len(E["rows"]), "map", sorted(E["col"]))
    print("O accounts:", len(O["rows"]), "map", sorted(O["col"]))

    e_by = {r["acc"]: r for r in E["rows"]}
    o_by = {r["acc"]: r for r in O["rows"]}
    e_acc, o_acc = set(e_by), set(o_by)

    print()
    print("========== CHECKS ==========")
    print(
        "1. Accounts: E=%d O=%d common=%d only_E=%d only_O=%d"
        % (
            len(e_acc),
            len(o_acc),
            len(e_acc & o_acc),
            len(e_acc - o_acc),
            len(o_acc - e_acc),
        )
    )
    if e_acc - o_acc:
        print("   only E:", list(e_acc - o_acc)[:5])
    if o_acc - e_acc:
        print("   only O:", list(o_acc - e_acc)[:5])

    e_c = Counter(r["country"] for r in E["rows"])
    o_c = Counter(r["country"] for r in O["rows"])
    russia_words = sum(1 for r in O["rows"] if "росси" in Nl(r["country"]))
    print()
    print("2. Country codes:")
    print("   E:", dict(e_c.most_common(8)))
    print("   O:", dict(o_c.most_common(8)))
    print("   O with word Rossiya:", russia_words)

    e_ccy = Counter(r["ccy"] for r in E["rows"] if r["ccy"])
    o_ccy = Counter(r["ccy"] for r in O["rows"] if r["ccy"])
    print()
    print("3. Currency:")
    print("   E top:", e_ccy.most_common(12))
    print("   O top:", o_ccy.most_common(12))
    ccy_diff = 0
    for a in sorted(e_acc & o_acc):
        if Nl(e_by[a]["ccy"]) != Nl(o_by[a]["ccy"]):
            ccy_diff += 1
            if ccy_diff <= 5:
                print("   DIFF", a, "E=", e_by[a]["ccy"], "O=", o_by[a]["ccy"])
    print("   ccy mismatches on common:", ccy_diff)

    print()
    print("4. Field diffs on common accounts:")
    for f in ("name", "inn", "country", "kind", "open", "close"):
        n = 0
        samples = []
        for a in e_acc & o_acc:
            ev, ov = Nl(e_by[a][f]), Nl(o_by[a][f])
            if ev != ov:
                n += 1
                if len(samples) < 3:
                    samples.append((a, e_by[a][f][:50], o_by[a][f][:50]))
        print("   %s: diffs=%d" % (f, n))
        for s in samples:
            print("     ", s)

    print()
    print("5. Amounts:")
    e_amts = Counter()
    o_amts = Counter()
    for a in e_acc & o_acc:
        e_amts.update(e_by[a]["raw_amts"])
        o_amts.update(o_by[a]["raw_amts"])
    shared = sum(min(e_amts[k], o_amts.get(k, 0)) for k in e_amts)
    e_tot = sum(e_amts.values())
    print(
        "   amount multiset E->O: %d / %d (%.1f%%)"
        % (shared, e_tot, 100.0 * shared / e_tot if e_tot else 0)
    )
    field_ok = field_tot = 0
    for a in e_acc & o_acc:
        for f in ("bal_beg", "bal_end", "dt", "kt"):
            ev, ov = e_by[a][f], o_by[a][f]
            if not ev and not ov:
                continue
            field_tot += 1
            if ev == ov:
                field_ok += 1
    print("   typed bal/dt/kt matches: %d / %d" % (field_ok, field_tot))

    print()
    print("6. Triad separators in O amount texts:")
    triad = 0
    samples = []
    for r in O["rows"]:
        for t in r["raw_amt_texts"]:
            if re.search(r"\d[\u00a0 ]\d{3}", t):
                triad += 1
                if len(samples) < 5:
                    samples.append(t)
    print("   triad-spaced amount texts:", triad)
    for s in samples:
        print("    ", s)

    print()
    print("7. Example account 40701156403800000001:")
    ex = "40701156403800000001"
    for label, store in (("E", e_by), ("O", o_by)):
        r = store.get(ex)
        if not r:
            print("  %s: NOT FOUND" % label)
        else:
            print(
                "  %s: country=%s ccy=%s name=%s inn=%s kind=%s open=%s"
                % (
                    label,
                    r["country"],
                    r["ccy"],
                    r["name"][:40],
                    r["inn"],
                    r["kind"][:50],
                    r["open"],
                )
            )

    print()
    print("=== VERDICT ===")
    acc_ok = len(e_acc) == len(o_acc) == len(e_acc & o_acc) and len(e_acc) > 0
    print("Accounts equal:", acc_ok, "(%d)" % len(e_acc))
    print(
        "Country code 643 count O:",
        o_c.get("643", 0),
        "of",
        len(O["rows"]),
        "; russia words:",
        russia_words,
    )
    print("Currency mismatches:", ccy_diff)


if __name__ == "__main__":
    main()
