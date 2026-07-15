#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Logical compare converted XLSX vs etalon Orticon converter."""

from pathlib import Path
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
ORT = BASE.parents[1] / "ОРТИКОН"
CONV = BASE / "converted_from_xbrl.xlsx"
ETALON = next(ORT.glob("*конвертер.xlsx"))

# expected close matches (name_substr or exact -> approx rows in etalon data)
CHECKS = [
    ("0420409 Раздел 1 Сведения о бан", 4210, 0.15),
    ("0420431 Раздел 1 Сведения об ос", 53, 0.2),
    ("0420431 Раздел 1 Сведения об _2", 300, 0.5),  # flat vs pivot
    ("0420431 Раздел 1 Сведения об _3", 300, 0.7),
    ("0420431 Раздел 2 Сведения о п_1", 1379, 0.15),
    ("0420431 Раздел 4 Сведения о п_1", 17036, 0.05),
    ("0420431 Раздел 7 Сведения о п_4", 1590, 0.1),
    ("0420431 Раздел 7 Сведения о п_5", 595, 0.1),
    ("0420459 Раздел 1 Сведения о цен", 542, 0.1),
]


def sheet_rows(wb, name):
    if name not in wb.sheetnames:
        return None
    ws = wb[name]
    # count non-empty after header area: for our flat = max_row-1; for etalon skip first 10 meta rows roughly
    return ws.max_row


def main():
    print("CONV", CONV, CONV.exists(), CONV.stat().st_size if CONV.exists() else 0)
    print("ETALON", ETALON.name, ETALON.stat().st_size)
    wb_c = load_workbook(CONV, read_only=True, data_only=True)
    wb_e = load_workbook(ETALON, read_only=False, data_only=True)

    print("\nConverted sheets:", len(wb_c.sheetnames))
    for n in wb_c.sheetnames:
        print(" ", n, "max_row=", wb_c[n].max_row)

    ok = 0
    warn = 0
    print("\n=== CHECKS ===")
    for name, expect_data_rows, tol in CHECKS:
        if name not in wb_c.sheetnames:
            print("MISS", name)
            warn += 1
            continue
        # our converter: row1 header, rest data
        got = max((wb_c[name].max_row or 1) - 1, 0)
        # etalon data rows rough: max_row - ~10 header
        et_name = name if name in wb_e.sheetnames else None
        et_rows = None
        if et_name:
            et_rows = max((wb_e[et_name].max_row or 1) - 10, 0)
        ratio = abs(got - expect_data_rows) / max(expect_data_rows, 1)
        status = "OK" if ratio <= tol else "DIFF"
        if status == "OK":
            ok += 1
        else:
            warn += 1
        print(
            "%s | %s | got=%d expect~%d etalon_data~%s tol=%.0f%%"
            % (status, name, got, expect_data_rows, et_rows, tol * 100)
        )

    # spot check strategy value
    print("\n=== SPOT STRATEGY ===")
    ws = wb_c["0420431 Раздел 1 Сведения об ос"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print("headers", headers[:10])
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
        print(i, row[:8])

    print("\nRESULT: ok=%d warn=%d" % (ok, warn))
    wb_c.close()
    wb_e.close()
    return 0 if warn <= 2 else 1


if __name__ == "__main__":
    raise SystemExit(main())
