#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Inspect etalon vs our enum/member cell values and taxonomy labels."""

from openpyxl import load_workbook
from pathlib import Path
from zipfile import ZipFile
import os
import re
import sys


def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", "replace").decode("ascii"))


def main():
    base = Path(__file__).resolve().parents[2] / "ОРТИКОН"
    etalon = load_workbook(
        base / "0420431_409_январь_2026_конвертер.xlsx",
        read_only=True,
        data_only=True,
    )
    ours = load_workbook(
        base / "XBRL_Orticon_taxonomy2.xlsx",
        read_only=True,
        data_only=True,
    )
    name = "0420409 Раздел 1 Сведения о бан"
    ews = etalon[name]
    ows = ours[name]

    safe_print("=== ETALON headers R7 C2-C9 ===")
    for c in range(2, 10):
        v = ews.cell(7, c).value
        if v:
            safe_print("C%d: %s" % (c, str(v).replace("\n", " ")[:70]))

    safe_print("=== ETALON data samples ===")
    for r in range(11, 20):
        vals = [ews.cell(r, c).value for c in range(1, 10)]
        if vals[0]:
            safe_print("R%d: %s" % (r, vals))

    safe_print("=== OURS headers R6 ===")
    for c in range(1, 10):
        v = ows.cell(6, c).value
        if v:
            safe_print("C%d: %s" % (c, str(v)[:70]))

    safe_print("=== OURS data R7-R9 ===")
    for r in range(7, 10):
        safe_print("R%d: %s" % (r, [ows.cell(r, c).value for c in range(1, 10)]))

    zip_path = Path(os.environ["LOCALAPPDATA"]) / "XBRLConverter" / "Taxonomies" / "20251230.zip"
    with ZipFile(zip_path) as z:
        text = z.read("final_7_1/www.cbr.ru/xbrl/udr/dom/mem-int-label.xml").decode("utf-8")
        for needle in [
            "Valyuta_643",
            "Strana_643",
            "Schet_doveritelnogo_upravleniya",
        ]:
            i = text.find(needle)
            safe_print("--- mem-int-label %s at %s ---" % (needle, i))
            if i < 0:
                continue
            chunk = text[max(0, i - 250) : i + 700]
            labs = re.findall(r'xml:lang="ru"[^>]*>([^<]+)', chunk)
            roles = re.findall(r'xlink:role="([^"]+)"', chunk)
            safe_print("labs: %s" % labs[:6])
            safe_print("roles: %s" % [r.split("/")[-1] for r in roles[:6]])

            # find short label specifically
            m = re.search(
                r'xlink:href="[^"]*%s[^"]*"[\s\S]{0,1200}?</link:labelArc>' % re.escape(needle),
                text,
            )
            if m:
                block = m.group(0)
                pairs = re.findall(
                    r'xlink:role="([^"]+)"[^>]*xml:lang="ru"[^>]*>([^<]+)',
                    block,
                )
                for role, lab in pairs[:8]:
                    safe_print("  %s => %s" % (role.split("/")[-1], lab))

    with ZipFile(zip_path) as z:
        text = z.read("final_7_1/www.cbr.ru/xbrl/udr/dom/mem-int-label.xml").decode("utf-8")
        for needle in [
            "mem-int_Valyuta_643RubRossijskijRublMember",
            "mem-int_Valyuta_156YuanKitajMember",
            "mem-int_Strana_643RusRossiyaMember",
        ]:
            i = text.find(needle)
            safe_print("==== %s at %s" % (needle, i))
            if i < 0:
                # try without mem-int_
                alt = needle.replace("mem-int_", "")
                i = text.find(alt)
                safe_print(" alt %s at %s" % (alt, i))
            if i < 0:
                continue
            chunk = text[i : i + 2500]
            pairs = re.findall(
                r'xlink:role="([^"]+)"[^>]*xml:lang="ru"[^>]*>([^<]+)',
                chunk,
            )
            for role, lab in pairs[:12]:
                safe_print("  %s => %s" % (role.split("/")[-1], lab))

    return 0


if __name__ == "__main__":
    sys.exit(main())
