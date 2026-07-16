#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare section-specific headers (credit vs broker) and taxonomy lab keys."""

import os
import re
import zipfile
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431\ОРТИКОН")
ETALON = BASE / "0420431_409_январь_2026_конвертер.xlsx"
OURS = BASE / "XBRL_Orticon_taxonomy6.xlsx"
TPL = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431"
    r"\Обработки\внВыгрузкаXBRLОртиконВXLSX\Templates\Таксономия_20251230\Ext\Template.bin"
)


def N(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ").replace("\n", " ")).strip()


def sheet_by_form_sec(wb, code, sec):
    needle = "раздел %d" % sec
    for s in wb.sheetnames:
        if code not in s:
            continue
        if needle in s.lower():
            return s
    return None


def get_header_labels(ws):
    rows = list(
        ws.iter_rows(
            min_row=1,
            max_row=25,
            max_col=min(ws.max_column or 50, 50),
            values_only=True,
        )
    )
    for row in rows:
        vals = [N(v) for v in row]
        nonempty = [v for v in vals if v]
        j = " ".join(nonempty)
        if "Сокращенное" in j or "наименование" in j.lower() and len(nonempty) >= 4:
            # skip TOC-like single-cell rows
            if len(nonempty) >= 3:
                return nonempty
    return []


def current_key_from_filename(name):
    """Mirror КлючТаблицыТаксономииИзИмениФайла (simplified)."""
    base = os.path.basename(name)
    m = re.search(r"SR_(\d{7})", base, re.I)
    if not m:
        return ""
    code = m.group(1)
    if base.upper().count("SR_") >= 2:
        return code + "_2"
    return code


def main():
    wb_e = load_workbook(ETALON, read_only=True, data_only=True)
    wb_o = load_workbook(OURS, read_only=True, data_only=True)

    pairs = [
        ("0420409", 1),
        ("0420409", 2),
        ("0420414", 1),
        ("0420414", 2),
        ("0420459", 1),
        ("0420459", 2),
    ]
    for i in range(1, 10):
        pairs.append(("0420431", i))

    print("=== Header credit/broker compare ===")
    for code, sec in pairs:
        se = sheet_by_form_sec(wb_e, code, sec)
        so = sheet_by_form_sec(wb_o, code, sec)
        if not se or not so:
            print("SKIP %s R%d E=%s O=%s" % (code, sec, bool(se), bool(so)))
            continue
        eh = get_header_labels(wb_e[se])
        oh = get_header_labels(wb_o[so])
        e_credit = sum(1 for h in eh if "кредитн" in h.lower())
        o_credit = sum(1 for h in oh if "кредитн" in h.lower())
        e_broker = sum(1 for h in eh if "брокер" in h.lower())
        o_broker = sum(1 for h in oh if "брокер" in h.lower())

        mismatches = []
        for h in eh:
            hl = h.lower()
            et = set(re.findall(r"[а-яa-z0-9]{4,}", hl))
            best = None
            bov = 0
            for ohv in oh:
                ot = set(re.findall(r"[а-яa-z0-9]{4,}", ohv.lower()))
                ov = len(et & ot)
                if ov > bov:
                    bov, best = ov, ohv
            if best is None or bov < 3:
                continue
            if N(h).lower() == N(best).lower():
                continue
            if (("брокер" in hl) != ("брокер" in best.lower())) or (
                ("кредитн" in hl) != ("кредитн" in best.lower())
            ):
                mismatches.append((h, best))

        if mismatches or (e_broker and o_credit and e_broker != o_broker):
            print("%s R%d" % (code, sec))
            print(
                "  counts E credit=%d broker=%d | O credit=%d broker=%d"
                % (e_credit, e_broker, o_credit, o_broker)
            )
            for a, b in mismatches[:8]:
                print("  E:", a[:110])
                print("  O:", b[:110])

    wb_e.close()
    wb_o.close()

    print("\n=== Taxonomy lab key mapping ===")
    with zipfile.ZipFile(TPL, "r") as z:
        names = [
            n
            for n in z.namelist()
            if n.endswith("-lab.xml")
            and "/tab/" in n.replace("\\", "/")
            and "purcb" in n
        ]
        for code in ("0420409", "0420414", "0420431", "0420459"):
            rel = [n for n in names if code in os.path.basename(n)]
            print(code, "labs", len(rel))
            keys = {}
            for n in sorted(rel):
                key = current_key_from_filename(n)
                keys.setdefault(key, []).append(os.path.basename(n))
            for k, files in sorted(keys.items()):
                collision = " COLLISION" if len(files) > 1 else ""
                print("  key=%s%s" % (k, collision))
                for f in files:
                    print("    ", f)

        # sokr texts from 0420409 labs
        print("\n=== Sokr texts in 0420409 lab files ===")
        for n in sorted(names):
            bn = os.path.basename(n)
            if "0420409" not in bn or "_q" in bn:
                continue
            data = z.read(n).decode("utf-8", errors="replace")
            hits = []
            for m in re.finditer(
                r"<link:label[^>]*>([^<]*[Сс]окращенн[^<]*)</link:label>", data
            ):
                txt = re.sub(r"\s+", " ", m.group(1)).strip()
                if txt and txt not in hits:
                    hits.append(txt)
            print(bn, "-> key", current_key_from_filename(n))
            for t in hits[:3]:
                print("  ", t[:120])


if __name__ == "__main__":
    main()
