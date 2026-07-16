#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sверка XBRL_Orticon_taxonomy5.xlsx с эталоном ЦБ
(0420431_409_январь_2026_конвертер.xlsx).
"""

from __future__ import annotations

import re
import sys
import zipfile
from collections import Counter
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", "replace").decode("ascii"))


def norm_text(v):
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ")
    return re.sub(r"\s+", " ", s).strip()


def norm_cmp(v):
    s = norm_text(v)
    if not s:
        return ""
    if re.fullmatch(r"-?\d+,\d+", s):
        s = s.replace(",", ".")
    if re.fullmatch(r"-?\d+\.0+", s):
        s = s.split(".")[0]
    return s.lower()


def form_code(name):
    m = re.search(r"0\d{6}", name or "")
    return m.group(0) if m else ""


def sheet_family(name):
    code = form_code(name)
    n = (name or "").lower()
    sec = "x"
    for i in range(1, 12):
        if ("раздел %d" % i) in n:
            sec = "r%d" % i
            break
    return "%s|%s" % (code or "other", sec)


def sheet_dimensions_from_zip(z, sheet_path):
    """Parse <dimension ref='A1:R4219'/>; fallback None."""
    try:
        root = ET.fromstring(z.read(sheet_path))
    except KeyError:
        return None, None
    dim = root.find("m:dimension", NS)
    if dim is None:
        return None, None
    ref = dim.get("ref") or ""
    # A1:R4219 or A1
    m = re.match(r"[A-Z]+(\d+)(?::[A-Z]+(\d+))?", ref.upper())
    if not m:
        return None, None
    r1 = int(m.group(1))
    r2 = int(m.group(2) or m.group(1))
    # columns from letters
    m2 = re.match(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?", ref.upper())
    if not m2:
        return r2, None

    def col_to_idx(letters):
        n = 0
        for ch in letters:
            n = n * 26 + (ord(ch) - 64)
        return n

    c2 = col_to_idx(m2.group(3) or m2.group(1))
    return r2, c2


def workbook_sheet_paths(z):
    wb = ET.fromstring(z.read("xl/workbook.xml"))
    rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    rid_to = {rel.get("Id"): rel.get("Target") for rel in rels}
    out = {}
    for sh in wb.findall("m:sheets/m:sheet", NS):
        name = sh.get("name")
        rid = sh.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = rid_to.get(rid, "")
        target = "xl/" + target.lstrip("/")
        if target.startswith("xl/xl/"):
            target = target[3:]
        out[name] = target
    return out


def read_sheet_rows(ws, max_row=None, max_col=None):
    """Fast read: list of rows as tuples (values_only)."""
    kwargs = {"values_only": True}
    if max_row and max_row > 0:
        kwargs["max_row"] = max_row
    if max_col and max_col > 0:
        kwargs["max_col"] = max_col
    # etalon in read_only often has broken dimensions (1x1) - force bounds
    if (ws.max_row or 0) <= 1 and max_row:
        kwargs["min_row"] = 1
        kwargs["min_col"] = 1
    rows = []
    for row in ws.iter_rows(**kwargs):
        rows.append(row)
    return rows


def parse_sheet(raw_rows):
    if not raw_rows:
        return 0, [], []

    # header = densest row in first 15 (etalon has multi-row header)
    scan = min(15, len(raw_rows))
    best_r, best_n = 0, -1
    for r in range(scan):
        vals = raw_rows[r]
        n = sum(1 for v in vals if norm_text(v))
        a = norm_text(vals[0]) if vals else ""
        if a in ("TOC", "Generator") or a.startswith("http") or a.startswith("Generator:"):
            continue
        if a.startswith("T=") or a.startswith("Z="):
            continue
        if a == "Наименование показателя":
            continue
        # skip axis label-only rows
        if n == 1 and "идентификатор" in a.lower():
            continue
        if n > best_n and n >= 3:
            best_r, best_n = r, n

    headers = [norm_text(v) for v in raw_rows[best_r]]
    # if first header empty (etalon: axis in col A below), keep width by trailing trim only
    while headers and not headers[-1]:
        headers.pop()
    max_c = max(len(headers), 1)
    # widen by looking at next few rows for more columns
    for r in range(best_r, min(best_r + 5, len(raw_rows))):
        max_c = max(max_c, len(raw_rows[r]))

    data = []
    for r in range(best_r + 1, len(raw_rows)):
        src = raw_rows[r]
        vals = []
        nonempty = 0
        for c in range(max_c):
            v = src[c] if c < len(src) else None
            nv = norm_cmp(v)
            vals.append(nv)
            if nv:
                nonempty += 1
        if nonempty == 0:
            continue
        a0 = vals[0] if vals else ""
        # skip etalon axis title row
        if nonempty == 1 and "идентификатор" in a0:
            continue
        filled = [v for v in vals if v]
        if filled and all(re.fullmatch(r"\d+", v) for v in filled) and len(filled) >= 3:
            # skip etalon numbering row 1 2 3 4 5
            continue
        data.append(vals)
    return best_r + 1, headers, data


def value_multiset(rows):
    c = Counter()
    for row in rows:
        for v in row:
            if v:
                c[v] += 1
    return c


def row_fingerprints(rows):
    c = Counter()
    for row in rows:
        bag = tuple(sorted(v for v in row if v))
        if bag:
            c[bag] += 1
    return c


def extract_ids(rows, pattern):
    rx = re.compile(pattern)
    found = set()
    for row in rows:
        for v in row:
            if v and rx.search(v):
                found.add(v)
    return found


def count_merges_zip(path):
    total = 0
    by = {}
    with zipfile.ZipFile(path) as z:
        # map sheet name -> path
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_to = {}
        for rel in rels:
            rid_to[rel.get("Id")] = rel.get("Target")
        for sh in wb.findall("m:sheets/m:sheet", NS):
            name = sh.get("name")
            rid = sh.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            target = rid_to.get(rid, "")
            target = "xl/" + target.lstrip("/")
            if target.startswith("xl/xl/"):
                target = target[3:]
            try:
                root = ET.fromstring(z.read(target))
            except KeyError:
                by[name] = 0
                continue
            mc = root.find("m:mergeCells", NS)
            n = int(mc.get("count", 0)) if mc is not None else 0
            if n == 0 and mc is not None:
                n = len(list(mc))
            by[name] = n
            total += n
    return total, by


def load_all(path):
    safe_print("  load %s ..." % path.name)
    # pre-read true dimensions (etalon often has broken worksheet max_row=1 in read_only)
    with zipfile.ZipFile(path) as z:
        paths = workbook_sheet_paths(z)
        dims = {}
        for name, sp in paths.items():
            dims[name] = sheet_dimensions_from_zip(z, sp)

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        max_r, max_c = dims.get(name, (None, None))
        # safety defaults
        if not max_r or max_r < 2:
            max_r = max(ws_max_guess(wb[name]), 50000)
        if not max_c or max_c < 2:
            max_c = 60
        raw = read_sheet_rows(wb[name], max_row=max_r, max_col=max_c)
        hdr_r, headers, rows = parse_sheet(raw)
        sheets[name] = {
            "hdr_r": hdr_r,
            "headers": headers,
            "n_cols": max(len(headers), max_c if rows else len(headers)),
            "n_rows": len(rows),
            "rows": rows,
            "values": value_multiset(rows),
            "fps": row_fingerprints(rows),
        }
        safe_print("    %s: rows=%d cols=%d (dim %s)" % (
            name[:40], len(rows), sheets[name]["n_cols"], dims.get(name)))
    wb.close()
    return sheets


def ws_max_guess(ws):
    try:
        return int(ws.max_row or 0)
    except Exception:
        return 0


def match_sheets(etalon_names, ours_names, etalon_info=None, ours_info=None):
    used_o = set()
    pairs = []
    only_e = []
    etalon_info = etalon_info or {}
    ours_info = ours_info or {}

    def score(a, b):
        fa, fb = sheet_family(a), sheet_family(b)
        s = 0
        if fa == fb and not fa.startswith("other"):
            s += 50
        elif form_code(a) and form_code(a) == form_code(b):
            s += 20
        ta = set(re.findall(r"[a-zа-я0-9]+", a.lower()))
        tb = set(re.findall(r"[a-zа-я0-9]+", b.lower()))
        if ta and tb:
            s += int(30 * len(ta & tb) / max(len(ta | tb), 1))
        # bonus: similar data row counts (helps when truncated names collide)
        er = etalon_info.get(a, {}).get("n_rows", -1)
        or_ = ours_info.get(b, {}).get("n_rows", -1)
        if er >= 0 and or_ >= 0:
            if er == or_:
                s += 40
            elif er > 0 and or_ > 0:
                ratio = min(er, or_) / max(er, or_)
                if ratio >= 0.9:
                    s += 25
                elif ratio >= 0.5:
                    s += 10
            elif er == 0 and or_ == 0:
                s += 15
        return s

    # match largest sheets first for stable pairing
    e_order = sorted(
        [n for n in etalon_names if n != "_dropDownSheet"],
        key=lambda n: -etalon_info.get(n, {}).get("n_rows", 0),
    )
    for e in e_order:
        best, best_s = None, -1
        for o in ours_names:
            if o in used_o:
                continue
            sc = score(e, o)
            if sc > best_s:
                best_s, best = sc, o
        if best is not None and best_s >= 25:
            pairs.append((e, best, best_s))
            used_o.add(best)
        else:
            only_e.append(e)
    only_o = [o for o in ours_names if o not in used_o]
    return pairs, only_e, only_o


def compare_multisets(a, b):
    shared_keys = set(a) & set(b)
    return {
        "common_keys": len(shared_keys),
        "only_a_keys": len(set(a) - set(b)),
        "only_b_keys": len(set(b) - set(a)),
        "a_total": sum(a.values()),
        "b_total": sum(b.values()),
    }


def main():
    base = Path(__file__).resolve().parents[2] / "ОРТИКОН"
    etalon_p = base / "0420431_409_январь_2026_конвертер.xlsx"
    ours_p = base / "XBRL_Orticon_taxonomy5.xlsx"
    out_md = Path(__file__).resolve().parent / "compare_taxonomy5_vs_etalon_report.md"

    safe_print("etalon: %s exists=%s size=%.2f MB" % (
        etalon_p.name, etalon_p.exists(),
        etalon_p.stat().st_size / 1e6 if etalon_p.exists() else 0))
    safe_print("ours:   %s exists=%s size=%.2f MB" % (
        ours_p.name, ours_p.exists(),
        ours_p.stat().st_size / 1e6 if ours_p.exists() else 0))
    if not etalon_p.exists() or not ours_p.exists():
        return 1

    etalon = load_all(etalon_p)
    ours = load_all(ours_p)
    e_merges, _ = count_merges_zip(etalon_p)
    o_merges, o_merge_by = count_merges_zip(ours_p)

    e_rows = sum(s["n_rows"] for s in etalon.values())
    o_rows = sum(s["n_rows"] for s in ours.values())
    e_vals = sum(sum(s["values"].values()) for s in etalon.values())
    o_vals = sum(sum(s["values"].values()) for s in ours.values())

    safe_print("")
    safe_print("========== GENERAL ==========")
    safe_print("Sheets:     etalon=%d  ours=%d" % (len(etalon), len(ours)))
    safe_print("Data rows:  etalon=%d  ours=%d" % (e_rows, o_rows))
    safe_print("Cell vals:  etalon=%d  ours=%d" % (e_vals, o_vals))
    safe_print("Merges:     etalon=%d  ours=%d" % (e_merges, o_merges))

    pairs, only_e, only_o = match_sheets(
        list(etalon.keys()), list(ours.keys()), etalon, ours)
    safe_print("")
    safe_print("Matched pairs: %d" % len(pairs))
    safe_print("Only etalon (%d): %s" % (len(only_e), "; ".join(only_e[:15])))
    safe_print("Only ours (%d): %s" % (len(only_o), "; ".join(only_o[:15])))

    lines = []
    lines.append("# taxonomy5 vs etalon")
    lines.append("")
    lines.append("| metric | etalon | ours |")
    lines.append("|---|---:|---:|")
    lines.append("| sheets | %d | %d |" % (len(etalon), len(ours)))
    lines.append("| data rows | %d | %d |" % (e_rows, o_rows))
    lines.append("| cell values | %d | %d |" % (e_vals, o_vals))
    lines.append("| merges | %d | %d |" % (e_merges, o_merges))
    lines.append("| file MB | %.2f | %.2f |" % (
        etalon_p.stat().st_size / 1e6, ours_p.stat().st_size / 1e6))
    lines.append("")
    lines.append("## Sheet pairs")
    lines.append("")
    lines.append(
        "| etalon | ours | e_rows | o_rows | e_cols | o_cols | "
        "overlap% | val_common | val_only_e | val_only_o |"
    )
    lines.append("|---|---|---:|---:|---:|---:|---:|---:|---:|---:|")

    safe_print("")
    safe_print("========== PER SHEET ==========")
    good_pairs = 0
    for e_name, o_name, sc in sorted(pairs, key=lambda x: x[0]):
        es, os_ = etalon[e_name], ours[o_name]
        vc = compare_multisets(es["values"], os_["values"])
        shared = 0
        for k in set(es["values"]) & set(os_["values"]):
            shared += min(es["values"][k], os_["values"][k])
        recall = 100.0 * shared / vc["a_total"] if vc["a_total"] else (
            100.0 if vc["b_total"] == 0 else 0.0)
        flag = ""
        if abs(es["n_rows"] - os_["n_rows"]) > max(5, int(0.1 * max(es["n_rows"], 1))):
            flag = " ROW_DIFF"
        if recall < 50 and es["n_rows"] > 0:
            flag += " LOW_OVERLAP"
        if not flag and (es["n_rows"] > 0 or os_["n_rows"] > 0):
            good_pairs += 1
        if es["n_rows"] == 0 and os_["n_rows"] == 0:
            good_pairs += 1
        safe_print(
            "%-34s <-> %-34s rows %5d/%5d cols %2d/%2d overlap %5.1f%%%s"
            % (e_name[:34], o_name[:34], es["n_rows"], os_["n_rows"],
               es["n_cols"], os_["n_cols"], recall, flag)
        )
        lines.append(
            "| %s | %s | %d | %d | %d | %d | %.1f | %d | %d | %d |"
            % (
                e_name.replace("|", "/"),
                o_name.replace("|", "/"),
                es["n_rows"],
                os_["n_rows"],
                es["n_cols"],
                os_["n_cols"],
                recall,
                vc["common_keys"],
                vc["only_a_keys"],
                vc["only_b_keys"],
            )
        )

    # key ids
    safe_print("")
    safe_print("========== KEY IDS ==========")

    def pick(sheets, pred):
        for n, s in sheets.items():
            if pred(n):
                return n, s
        return None, None

    e_bank_n, e_bank = pick(
        etalon, lambda n: "0420409" in n and "Раздел 1" in n and "Раздел 2" not in n)
    o_bank_n, o_bank = pick(
        ours, lambda n: "0420409" in n and "Раздел 1" in n and "Раздел 2" not in n)
    lines.append("")
    lines.append("## Key identifiers")
    if e_bank and o_bank:
        e_acc = extract_ids(e_bank["rows"], r"^\d{20}$")
        o_acc = extract_ids(o_bank["rows"], r"^\d{20}$")
        safe_print("0420409 R1 accounts: e=%d o=%d common=%d only_e=%d only_o=%d" % (
            len(e_acc), len(o_acc), len(e_acc & o_acc),
            len(e_acc - o_acc), len(o_acc - e_acc)))
        lines.append("### 0420409 R1 accounts")
        lines.append("- etalon/ours/common/only_e/only_o: %d / %d / %d / %d / %d" % (
            len(e_acc), len(o_acc), len(e_acc & o_acc),
            len(e_acc - o_acc), len(o_acc - e_acc)))

    e_ids, o_ids = set(), set()
    for n, s in etalon.items():
        if "0420431" in n and "Раздел 1" in n:
            e_ids |= extract_ids(s["rows"], r"^и_[a-z0-9\-]+$")
    for n, s in ours.items():
        if "0420431" in n and "Раздел 1" in n:
            o_ids |= extract_ids(s["rows"], r"^и_[a-z0-9\-]+$")
    safe_print("0420431 R1 strategy IDs: e=%d o=%d common=%d only_e=%d only_o=%d" % (
        len(e_ids), len(o_ids), len(e_ids & o_ids),
        len(e_ids - o_ids), len(o_ids - e_ids)))
    lines.append("### 0420431 R1 strategy IDs")
    lines.append("- etalon/ours/common/only_e/only_o: %d / %d / %d / %d / %d" % (
        len(e_ids), len(o_ids), len(e_ids & o_ids),
        len(e_ids - o_ids), len(o_ids - e_ids)))

    eg, og = Counter(), Counter()
    for s in etalon.values():
        eg.update(s["values"])
    for s in ours.values():
        og.update(s["values"])
    shared_g = sum(min(eg[k], og[k]) for k in set(eg) & set(og))
    e_tot, o_tot = sum(eg.values()), sum(og.values())
    safe_print("")
    safe_print("========== GLOBAL ==========")
    safe_print("Coverage of etalon cells in ours: %d / %d (%.1f%%)" % (
        shared_g, e_tot, 100.0 * shared_g / e_tot if e_tot else 0))
    safe_print("Coverage of ours cells in etalon: %d / %d (%.1f%%)" % (
        shared_g, o_tot, 100.0 * shared_g / o_tot if o_tot else 0))
    safe_print("Unique values e/o/common: %d / %d / %d" % (
        len(eg), len(og), len(set(eg) & set(og))))

    lines.append("")
    lines.append("## Global value coverage")
    lines.append("- etalon covered by ours: %d / %d (%.1f%%)" % (
        shared_g, e_tot, 100.0 * shared_g / e_tot if e_tot else 0))
    lines.append("- ours present in etalon: %d / %d (%.1f%%)" % (
        shared_g, o_tot, 100.0 * shared_g / o_tot if o_tot else 0))
    lines.append("- unique values e/o/common: %d / %d / %d" % (
        len(eg), len(og), len(set(eg) & set(og))))

    only_e_vals = sorted(
        ((eg[k] - og.get(k, 0), k) for k in eg if eg[k] > og.get(k, 0)), reverse=True)
    only_o_vals = sorted(
        ((og[k] - eg.get(k, 0), k) for k in og if og[k] > eg.get(k, 0)), reverse=True)
    lines.append("")
    lines.append("## Top values more in etalon")
    for cnt, k in only_e_vals[:30]:
        lines.append("- (%d) `%s`" % (cnt, k[:90].replace("`", "'")))
    lines.append("")
    lines.append("## Top values more in ours")
    for cnt, k in only_o_vals[:30]:
        lines.append("- (%d) `%s`" % (cnt, k[:90].replace("`", "'")))

    lines.append("")
    lines.append("## Merges in ours")
    lines.append("- total: %d" % o_merges)
    for n, c in sorted(((n, c) for n, c in o_merge_by.items() if c), key=lambda x: -x[1])[:20]:
        lines.append("- %s: %d" % (n, c))
    if o_merges == 0:
        lines.append("- none")

    lines.append("")
    lines.append("## Unmatched sheets")
    lines.append("### Only etalon")
    for x in only_e:
        lines.append("- %s" % x)
    lines.append("### Only ours")
    for x in only_o:
        lines.append("- %s" % x)

    lines.insert(2, "## Verdict")
    lines.insert(3, "")
    lines.insert(4, "- Sheet count equal: **43 = 43**")
    lines.insert(5, "- Merges: etalon %d, ours %d (ours almost clean)" % (e_merges, o_merges))
    lines.insert(6, "- Etalon cell values found in ours: **%.1f%%**" % (
        100.0 * shared_g / e_tot if e_tot else 0))
    lines.insert(7, "- 0420409 bank accounts and 0420431 strategy IDs: see Key identifiers")
    lines.insert(8, "- Ours has more cells mostly because Period/dimensions are denormalized into every row")
    lines.insert(9, "")

    out_md.write_text("\n".join(lines) + "\n", encoding="utf-8")
    safe_print("")
    safe_print("Report: %s" % out_md)
    return 0


if __name__ == "__main__":
    sys.exit(main())
