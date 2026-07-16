#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""April etalon vs ours - etalon needs non-read_only load."""

from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook
import re

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ")
ETALON = BASE / "409-414-431 апрель 26.xlsx"
OURS = BASE / "XBRL_Orticon_taxonomy.xlsx_АПРЕЛЬ.xlsx"


def looks_num_str(s: str) -> bool:
    t = str(s).strip().replace(" ", "").replace("\u00a0", "")
    if not t or "_" in t or ":" in t:
        return False
    if len(t) >= 8 and t[4:5] == "-" and t[7:8] == "-":
        return False
    if "-" in t and any(ch.isalpha() for ch in t):
        return False
    u = t.replace(",", ".")
    if u.count(".") > 1:
        return False
    body = u.replace(".", "").replace("-", "")
    return body.isdigit() and len(body) > 0


def find_header_row(ws, max_scan=20):
    """CBR often has multi-row headers; find row with 'Идентификатор' or similar."""
    for r in range(1, max_scan + 1):
        vals = []
        for c in range(1, min(20, (ws.max_column or 1) + 1)):
            v = ws.cell(r, c).value
            if v:
                vals.append(str(v))
        joined = " ".join(vals).lower()
        if "идентификатор" in joined or "наименование показателя" in joined:
            # prefer row that looks like column titles for data
            if "идентификатор" in joined:
                return r
    return 6


def find_data_start(ws, header_row):
    # data usually after header; for CBR multi-header may be header_row+1..+5
    for r in range(header_row + 1, header_row + 8):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        s = str(v)
        # skip numbering rows 1 2 3
        if s.isdigit() and len(s) <= 3:
            continue
        if s.lower() in ("наименование показателя",):
            continue
        return r
    return header_row + 1


def analyze_cbr(path):
    wb = load_workbook(path, data_only=True)
    sheets = {}
    print("ETALON sheet dims:")
    for sn in wb.sheetnames:
        ws = wb[sn]
        toc = str(ws.cell(2, 1).value or "")[:120]
        uri = str(ws.cell(3, 1).value or "")[:80]
        print(f"  {sn}: max_row={ws.max_row} max_col={ws.max_column} toc={toc[:60]}")
        if ws.max_row <= 1:
            sheets[sn] = {"toc": toc, "rows": 0, "cols": 0, "raw_okato": 0, "num": 0, "text_num": 0, "hdr_row": 0}
            continue
        hdr_row = find_header_row(ws)
        # for CBR pivot tables, last header row with identifiers is often near bottom of header block
        # scan for last row before data containing Идентификатор
        last_id_row = hdr_row
        for r in range(1, min(25, ws.max_row) + 1):
            v1 = str(ws.cell(r, 1).value or "")
            if "Идентификатор" in v1 or "идентификатор" in v1.lower():
                last_id_row = r
        data_start = last_id_row + 1
        # if next rows are empty of real ids, still count nonempty
        rows = 0
        num = 0
        text_num = 0
        raw_okato = 0
        cols = ws.max_column or 0
        for r in range(data_start, ws.max_row + 1):
            nonempty = False
            for c in range(1, cols + 1):
                v = ws.cell(r, c).value
                if v is None or str(v).strip() == "":
                    continue
                nonempty = True
                if isinstance(v, str) and v.startswith("OKATO") and any(ch.isalpha() for ch in v[5:]):
                    raw_okato += 1
                if isinstance(v, (int, float)):
                    num += 1
                elif isinstance(v, str) and looks_num_str(v):
                    text_num += 1
            if nonempty:
                rows += 1
        sheets[sn] = {
            "toc": toc,
            "uri": uri,
            "rows": rows,
            "cols": cols,
            "raw_okato": raw_okato,
            "num": num,
            "text_num": text_num,
            "hdr_row": last_id_row,
            "data_start": data_start,
        }
    wb.close()
    return sheets


def analyze_ours(path):
    wb = load_workbook(path, read_only=True, data_only=False)
    gen = None
    sheets = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        toc = ""
        rows = 0
        cols = 0
        raw_okato = 0
        num = 0
        text_num = 0
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=False), 1):
            vals = [c.value for c in row]
            if i == 2 and vals and vals[0]:
                toc = str(vals[0])[:120]
            if i == 4 and vals and vals[0] and "Generator" in str(vals[0]):
                gen = str(vals[0])[:100]
            if i == 6:
                cols = sum(1 for c in row if c.value is not None)
        for row in ws.iter_rows(min_row=7, values_only=False):
            if not any(c.value not in (None, "") for c in row):
                continue
            rows += 1
            for c in row:
                v = c.value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith("OKATO") and any(ch.isalpha() for ch in v[5:]):
                    raw_okato += 1
                if isinstance(v, (int, float)) or c.data_type == "n":
                    num += 1
                elif isinstance(v, str) and looks_num_str(v):
                    text_num += 1
        sheets[sn] = {
            "toc": toc,
            "rows": rows,
            "cols": cols,
            "raw_okato": raw_okato,
            "num": num,
            "text_num": text_num,
        }
    wb.close()
    return gen, sheets


def norm_toc(t):
    return " ".join(str(t).lower().split())


def main():
    print("=== Load etalon (full) ===")
    et = analyze_cbr(ETALON)
    print("\n=== Load ours ===")
    gen, ours = analyze_ours(OURS)
    print("OURS gen:", gen)
    print(
        "ET nonempty",
        sum(1 for s in et.values() if s["rows"] > 0),
        "rows",
        sum(s["rows"] for s in et.values()),
    )
    print(
        "OUR nonempty",
        sum(1 for s in ours.values() if s["rows"] > 0),
        "rows",
        sum(s["rows"] for s in ours.values()),
    )

    print("\n=== ETALON nonempty sheets ===")
    for sn, s in et.items():
        if s["rows"] > 0:
            print(
                f"  {s['rows']:6d} cols={s['cols']:3d} hdr={s.get('hdr_row')} "
                f"num={s['num']} text={s['text_num']} okato={s['raw_okato']} | {sn}"
            )
            print(f"         toc={s['toc'][:80]}")

    print("\n=== OURS nonempty sheets ===")
    for sn, s in ours.items():
        if s["rows"] > 0:
            print(
                f"  {s['rows']:6d} cols={s['cols']:3d} num={s['num']} text={s['text_num']} "
                f"okato={s['raw_okato']} | {sn}"
            )

    # match by TOC
    print("\n=== Match by TOC ===")
    et_toc = defaultdict(list)
    our_toc = defaultdict(list)
    for sn, s in et.items():
        if s["rows"] > 0:
            et_toc[norm_toc(s["toc"])].append((sn, s))
    for sn, s in ours.items():
        if s["rows"] > 0 and sn != "TOC":
            our_toc[norm_toc(s["toc"])].append((sn, s))

    for t in sorted(set(et_toc) | set(our_toc)):
        er = sum(x[1]["rows"] for x in et_toc.get(t, []))
        or_ = sum(x[1]["rows"] for x in our_toc.get(t, []))
        mark = "OK" if er == or_ else "DIFF"
        print(f"[{mark}] {er:6d} -> {or_:6d} (d={or_-er:+d}) | {t[:85]}")

    print("\n=== Only etalon TOC ===")
    for t in sorted(set(et_toc) - set(our_toc)):
        print(f"  {sum(x[1]['rows'] for x in et_toc[t]):6d} | {t[:90]}")
    print("=== Only ours TOC ===")
    for t in sorted(set(our_toc) - set(et_toc)):
        print(f"  {sum(x[1]['rows'] for x in our_toc[t]):6d} | {t[:90]}")

    # form codes
    print("\n=== Form totals ===")
    for code in ("0420409", "0420414", "0420431", "0420459"):
        er = sum(s["rows"] for sn, s in et.items() if code in sn or code in s["toc"])
        or_ = sum(s["rows"] for sn, s in ours.items() if code in sn or code in s["toc"])
        eok = sum(s["raw_okato"] for sn, s in et.items() if code in sn or code in s["toc"])
        ook = sum(s["raw_okato"] for sn, s in ours.items() if code in sn or code in s["toc"])
        print(f"{code}: rows ET={er} OUR={or_} delta={or_-er} okato ET={eok} OUR={ook}")

    print("DONE")


if __name__ == "__main__":
    main()
