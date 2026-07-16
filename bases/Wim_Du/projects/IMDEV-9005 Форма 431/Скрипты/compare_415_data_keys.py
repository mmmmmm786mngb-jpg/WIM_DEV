#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from pathlib import Path

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ")
cbr = load_workbook(BASE / "401-414-415-0420437_июнь_2026_кроме 431.xlsx", data_only=True)
ours = load_workbook(BASE / "XBRL_Orticon_taxonomy_июнь.xlsx", read_only=True, data_only=True)

ws = cbr["0420415 Раздел 1 Операции с ц_3"]
print("=== CBR ц_3 headers R6-R12 ===")
for r in range(6, 13):
    vals = []
    for c in range(1, 33):
        v = ws.cell(r, c).value
        if v:
            vals.append(f"{c}:{str(v).replace(chr(10), ' ')[:45]}")
    print(f"R{r}:", vals)

sec, org, stroka = set(), set(), set()
for r in range(13, (ws.max_row or 0) + 1):
    a, b, c = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
    if a:
        sec.add(str(a).strip())
    if b:
        org.add(str(b).strip())
    if c:
        stroka.add(str(c).strip())
print(
    "CBR unique securities",
    len(sec),
    "orgs",
    len(org),
    "stroka",
    len(stroka),
    "rows",
    (ws.max_row or 0) - 12,
)

print("CBR row13 values:")
for i in range(1, 33):
    v = ws.cell(13, i).value
    if v is not None:
        print(f"  c{i} {type(v).__name__}={repr(v)[:60]}")

ws2 = ours["0420415 Раздел 1. Операции с _2"]
print("\n=== OURS _2 headers ===")
h = list(next(ws2.iter_rows(min_row=6, max_row=6, values_only=True)))
for i, x in enumerate(h):
    if x:
        print(f"  {i}: {str(x).replace(chr(10), ' ')[:75]}")

sec2, org2, stroka2 = set(), set(), set()
n = 0
qty_vals = []
for row in ws2.iter_rows(min_row=7, values_only=True):
    if not row or row[0] is None:
        continue
    n += 1
    if len(row) > 3 and row[3]:
        sec2.add(str(row[3]).strip())
    if len(row) > 1 and row[1]:
        org2.add(str(row[1]).strip())
    if len(row) > 2 and row[2]:
        stroka2.add(str(row[2]).strip())
    # col 7 often quantity
    if len(row) > 7 and row[7] is not None:
        qty_vals.append(row[7])

print("OURS rows", n, "unique sec", len(sec2), "orgs", len(org2), "stroka", len(stroka2))
print("OURS qty col7 sample types", type(qty_vals[0]).__name__ if qty_vals else None, qty_vals[:3])

both = sec & sec2
print("securities in both", len(both))
print("only CBR", len(sec - sec2), list(sec - sec2)[:5])
print("only OURS", len(sec2 - sec), list(sec2 - sec)[:5])

# key = security|org
cbr_keys = set()
for r in range(13, (ws.max_row or 0) + 1):
    a, b = ws.cell(r, 1).value, ws.cell(r, 2).value
    if a and b:
        cbr_keys.add(f"{a}|{b}")
ours_keys = set()
for row in ours["0420415 Раздел 1. Операции с _2"].iter_rows(min_row=7, values_only=True):
    if row and row[3] and row[1]:
        ours_keys.add(f"{row[3]}|{row[1]}")
print("keys sec|org both", len(cbr_keys & ours_keys), "cbr", len(cbr_keys), "ours", len(ours_keys))

cbr.close()
ours.close()
