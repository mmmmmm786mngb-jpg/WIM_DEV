#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare April CBR etalon vs our XBRL_Orticon_taxonomy.xlsx_АПРЕЛЬ.xlsx."""

from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ")
ETALON = BASE / "409-414-431 апрель 26.xlsx"
OURS = BASE / "XBRL_Orticon_taxonomy.xlsx_АПРЕЛЬ.xlsx"


def looks_num_str(s: str) -> bool:
    t = str(s).strip().replace(" ", "").replace("\u00a0", "")
    if not t or "_" in t or ":" in t:
        return False
    if len(t) >= 8 and t[4:5] == "-" and t[7:8] == "-":
        return False
    if "-" in t and any(ch.isalpha() for ch in t):
        return False
    u = t.replace(",", ".")
    if u.count(".") > 1:
        return False
    body = u.replace(".", "").replace("-", "")
    return body.isdigit() and len(body) > 0


def sheet_stats(path, read_gen=True):
    wb = load_workbook(path, read_only=True, data_only=False)
    gen = None
    sheets = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        toc = ""
        rows = 0
        cols = 0
        raw_okato = 0
        num_cells = 0
        text_num = 0
        headers = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=False), 1):
            vals = [c.value for c in row]
            if i == 2 and vals and vals[0]:
                toc = str(vals[0])[:120]
            if i == 4 and vals and vals[0] and "Generator" in str(vals[0]):
                gen = str(vals[0])[:100]
            if i == 6:
                headers = [
                    str(c.value).replace("\n", " ")[:50]
                    for c in row
                    if c.value is not None
                ]
                cols = len(headers)
        for row in ws.iter_rows(min_row=7, values_only=False):
            if not any(c.value not in (None, "") for c in row):
                continue
            rows += 1
            for c in row:
                v = c.value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith("OKATO") and any(
                    ch.isalpha() for ch in v[5:]
                ):
                    raw_okato += 1
                if isinstance(v, (int, float)) or c.data_type == "n":
                    num_cells += 1
                elif isinstance(v, str) and looks_num_str(v):
                    text_num += 1
        sheets[sn] = {
            "toc": toc,
            "rows": rows,
            "cols": cols,
            "raw_okato": raw_okato,
            "num": num_cells,
            "text_num": text_num,
            "headers": headers,
        }
    wb.close()
    return gen, sheets


def norm_toc(t: str) -> str:
    return " ".join(str(t).lower().split())


def form_code(text: str) -> str:
    import re

    m = re.search(r"0420\d{3}", str(text))
    return m.group(0) if m else ""


def main():
    print("ETALON", ETALON.exists(), ETALON.name, "MB", round(ETALON.stat().st_size / 1e6, 2))
    print("OURS  ", OURS.exists(), OURS.name, "MB", round(OURS.stat().st_size / 1e6, 2))
    gen_e, et = sheet_stats(ETALON)
    gen_o, ours = sheet_stats(OURS)
    print("ETALON gen:", gen_e)
    print("OURS gen:", gen_o)
    print(
        "ETALON sheets",
        len(et),
        "nonempty",
        sum(1 for s in et.values() if s["rows"] > 0),
        "rows",
        sum(s["rows"] for s in et.values()),
    )
    print(
        "OURS sheets",
        len(ours),
        "nonempty",
        sum(1 for s in ours.values() if s["rows"] > 0),
        "rows",
        sum(s["rows"] for s in ours.values()),
    )

    # by TOC
    print("\n=== Nonempty TOC compare ===")
    et_toc = defaultdict(list)
    our_toc = defaultdict(list)
    for sn, s in et.items():
        if sn == "TOC" or s["rows"] == 0:
            continue
        et_toc[norm_toc(s["toc"] or sn)].append((sn, s))
    for sn, s in ours.items():
        if sn == "TOC" or s["rows"] == 0:
            continue
        our_toc[norm_toc(s["toc"] or sn)].append((sn, s))

    all_toc = sorted(set(et_toc) | set(our_toc))
    for t in all_toc:
        er = sum(x[1]["rows"] for x in et_toc.get(t, []))
        or_ = sum(x[1]["rows"] for x in our_toc.get(t, []))
        eok = sum(x[1]["raw_okato"] for x in et_toc.get(t, []))
        ook = sum(x[1]["raw_okato"] for x in our_toc.get(t, []))
        enum = sum(x[1]["num"] for x in et_toc.get(t, []))
        onum = sum(x[1]["num"] for x in our_toc.get(t, []))
        etx = sum(x[1]["text_num"] for x in et_toc.get(t, []))
        otx = sum(x[1]["text_num"] for x in our_toc.get(t, []))
        mark = "OK" if er == or_ else "DIFF"
        short = (t[:75] if t else "(no toc)")
        print(
            f"[{mark}] rows {er}->{or_} (d={or_-er}) okato {eok}->{ook} "
            f"num {enum}->{onum} textnum {etx}->{otx}"
        )
        print(f"      {short}")
        if er != or_ or eok or ook:
            print("      ET:", [(x[0], x[1]["rows"]) for x in et_toc.get(t, [])])
            print("      OUR:", [(x[0], x[1]["rows"]) for x in our_toc.get(t, [])])

    # only in one side
    print("\n=== TOC only in etalon ===")
    for t in sorted(set(et_toc) - set(our_toc)):
        print(" ", sum(x[1]["rows"] for x in et_toc[t]), t[:90])
    print("=== TOC only in ours ===")
    for t in sorted(set(our_toc) - set(et_toc)):
        print(" ", sum(x[1]["rows"] for x in our_toc[t]), t[:90])

    # form aggregates
    print("\n=== By form code ===")
    for code in ("0420409", "0420414", "0420431", "0420459", "0420415"):
        def agg(store):
            r = n = tx = ok = 0
            names = []
            for sn, s in store.items():
                if code in sn or code in s["toc"]:
                    r += s["rows"]
                    n += s["num"]
                    tx += s["text_num"]
                    ok += s["raw_okato"]
                    if s["rows"] > 0:
                        names.append(f"{sn}:{s['rows']}")
            return r, n, tx, ok, names

        er, en, etx, eok, enames = agg(et)
        or_, on, otx, ook, onames = agg(ours)
        print(
            f"{code}: rows {er}->{or_} num {en}->{on} textnum {etx}->{otx} okato {eok}->{ook}"
        )
        if enames[:5] or onames[:5]:
            print("  ET nonempty", enames[:8])
            print("  OUR nonempty", onames[:8])

    # spot 409 and 431 R4 headers/sample
    print("\n=== Spot sheets ===")
    for label, store in (("ETALON", et), ("OURS", ours)):
        for sn, s in store.items():
            if s["rows"] == 0:
                continue
            if "0420409" in sn and "Раздел 1" in sn:
                print(f"{label} 409 R1 {sn}: rows={s['rows']} cols={s['cols']} num={s['num']} text={s['text_num']}")
                print("  hdr", s["headers"][:8])
            if "0420431" in sn and "Раздел 4" in sn and s["rows"] > 100:
                print(f"{label} 431 R4 {sn}: rows={s['rows']} cols={s['cols']} num={s['num']} text={s['text_num']} okato={s['raw_okato']}")
                print("  hdr", s["headers"][:8])

    print("\nDONE")


if __name__ == "__main__":
    main()
