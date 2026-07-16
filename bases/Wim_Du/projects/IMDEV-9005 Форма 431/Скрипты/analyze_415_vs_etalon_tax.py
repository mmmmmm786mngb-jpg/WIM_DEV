#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Deep compare 415 _2 vs CBR etalon + taxonomy table mapping."""

import zipfile
import re
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ")
OURS = BASE / "XBRL_Orticon_taxonomy_июнь1.xlsx"
CBR = BASE / "401-414-415-0420437_июнь_2026_кроме 431.xlsx"
TAX = Path(r"C:\Users\Acer\AppData\Local\XBRLConverter\Taxonomies\20251230.zip")
XBRL_ZIP = BASE / "июнь_XBRL_1027739323600_ep_nso_purcb_m_q_y_10rd_ex_reestr_0420417_20260630.zip"


def dump_ours():
    print("=== OURS 0420415 Раздел 1. Операции с _2 ===")
    wb = load_workbook(OURS, read_only=True, data_only=True)
    ws = wb["0420415 Раздел 1. Операции с _2"]
    for r in range(1, 8):
        row = list(next(ws.iter_rows(min_row=r, max_row=r, values_only=True)))
        vals = [(i, str(c).replace("\n", " ")[:55]) for i, c in enumerate(row) if c]
        if vals:
            print(f"R{r}: {vals}")
    # value distribution per col
    cols = defaultdict(Counter)
    n = 0
    nonempty = Counter()
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row or row[0] is None:
            continue
        n += 1
        for i, c in enumerate(row):
            if c is None or str(c).strip() == "":
                continue
            nonempty[i] += 1
            if i in (4, 6, 8, 9):  # dim-like
                cols[i][str(c)[:60]] += 1
    print(f"rows={n}")
    print("nonempty per col:", dict(nonempty))
    for i, ctr in cols.items():
        print(f" col{i} unique={len(ctr)} top:", ctr.most_common(8))
    # how many rows have qty vs only id dims
    patterns = Counter()
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row or row[0] is None:
            continue
        has_qty = row[7] is not None or (len(row) > 5 and row[5] is not None)
        has_dim4 = len(row) > 4 and row[4]
        has_dim6 = len(row) > 6 and row[6]
        has_dim8 = len(row) > 8 and row[8]
        key = (
            f"qty={1 if has_qty else 0}",
            f"d4={1 if has_dim4 else 0}",
            f"d6={1 if has_dim6 else 0}",
            f"d8={1 if has_dim8 else 0}",
        )
        patterns[key] += 1
    print("row patterns:", patterns.most_common(20))
    wb.close()


def dump_cbr():
    print("\n=== CBR 0420415 Раздел 1 Операции с ц_3 (data sheet) ===")
    wb = load_workbook(CBR, data_only=True)
    ws = wb["0420415 Раздел 1 Операции с ц_3"]
    print(f"max_row={ws.max_row} max_col={ws.max_column}")
    for r in range(1, 13):
        vals = []
        for c in range(1, min(33, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if v:
                vals.append(f"{c}:{str(v).replace(chr(10), ' ')[:40]}")
        if vals:
            print(f"R{r}: {vals}")
    # TOC URI from R3
    print("URI:", ws.cell(3, 1).value)
    wb.close()


def tax_415_owned_securities():
    print("\n=== Taxonomy tables SR_0420415* ===")
    z = zipfile.ZipFile(TAX)
    labs = [n for n in z.namelist() if "0420415" in n and n.endswith("-lab.xml")]
    print("lab files:", len(labs))
    for n in sorted(labs):
        text = z.read(n).decode("utf-8", errors="replace")
        # table title
        titles = re.findall(r"<link:label[^>]*xml:lang=\"ru\"[^>]*>([^<]+)</link:label>", text)
        if not titles:
            titles = re.findall(r"<label[^>]*xml:lang=\"ru\"[^>]*>([^<]+)</label>", text)
        key = re.search(r"(SR_0420415[^/\\\\]+)-lab", n.replace("\\", "/"))
        k = key.group(1) if key else n
        # pick longest russian title
        title = max(titles, key=len) if titles else ""
        print(f"\n{k}")
        print(f"  title: {title[:120]}")
        # matching ours/CBR
        if "принадлежащ" in title.lower() or "праве собственности" in title.lower():
            print("  *** OWNED SECURITIES CANDIDATE ***")
            rend = n.replace("-lab.xml", "-rend.xml")
            if rend in z.namelist():
                rt = z.read(rend).decode("utf-8", errors="replace")
                dims = sorted(set(re.findall(r"([A-Za-z0-9_]+(?:Axis|Taxis))", rt)))
                print(f"  dims ({len(dims)}): {dims}")
                # breakdown concepts from hrefs
                concepts = re.findall(r"href=\"[^\"]*#([A-Za-z0-9_]+)\"", rt)
                concepts = [c for c in concepts if not c.endswith(("Axis", "Taxis", "Member", "Table", "LineItems", "Domain"))]
                # unique preserve order
                seen = []
                for c in concepts:
                    if c not in seen:
                        seen.append(c)
                print(f"  concepts ({len(seen)}): {seen[:40]}")


def match_uri():
    print("\n=== Sheet URI / TOC mapping ===")
    wb = load_workbook(OURS, read_only=True, data_only=True)
    for s in wb.sheetnames:
        if "415" in s and "Раздел 1" in s:
            ws = wb[s]
            r2 = r3 = None
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=4, max_col=1, values_only=True), 1):
                if i == 2:
                    r2 = row[0]
                if i == 3:
                    r3 = row[0]
            # count rows
            n = sum(1 for row in ws.iter_rows(min_row=7, max_col=1, values_only=True) if row and row[0])
            print(f"{s}: rows={n}")
            print(f"  toc={str(r2)[:80]}")
            print(f"  uri={str(r3)[:100]}")
    wb.close()

    wb = load_workbook(CBR, data_only=True)
    for s in wb.sheetnames:
        if "415" in s and "Раздел 1" in s:
            ws = wb[s]
            print(f"CBR {s}: R2={str(ws.cell(2,1).value)[:80]}")
            print(f"  uri={str(ws.cell(3,1).value)[:120]}")
            print(f"  maxR={ws.max_row}")
    wb.close()


if __name__ == "__main__":
    dump_ours()
    dump_cbr()
    match_uri()
    tax_415_owned_securities()
