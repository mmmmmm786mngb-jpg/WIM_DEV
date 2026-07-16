#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Analyze form 0420415 in June XBRL vs etalon Excel."""

import zipfile
import re
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ")
ZIP = BASE / "июнь_XBRL_1027739323600_ep_nso_purcb_m_q_y_10rd_ex_reestr_0420417_20260630.zip"
ETALON = BASE / "401-414-415-0420437_июнь_2026_кроме 431.xlsx"
OURS = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431"
    r"\Тестирование\ОШИБКИ_XBRL_Orticon_taxonomy_июнь.xlsx"
)


def analyze_etalon():
    print("=== ETALON Excel 415 sheets ===")
    wb = load_workbook(ETALON, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        if "415" not in sname:
            continue
        ws = wb[sname]
        # header-ish rows
        print(f"\n--- {sname} ---")
        for r in range(1, 16):
            try:
                row = list(next(ws.iter_rows(min_row=r, max_row=r, max_col=12, values_only=True)))
            except StopIteration:
                break
            vals = [str(c).replace("\n", " ")[:40] if c is not None else "" for c in row]
            if any(vals):
                print(f"  R{r}: {vals}")
        # count nonempty data rows after headers
        data_n = 0
        for row in ws.iter_rows(min_row=16, max_col=5, values_only=True):
            if any(c is not None and str(c).strip() for c in row):
                data_n += 1
            if data_n > 5000:
                break
        print(f"  data rows (from R16, capped): {data_n}")
    wb.close()


def analyze_ours():
    if not OURS.exists():
        print("\n=== OURS file missing ===")
        return
    print("\n=== OURS converter 415 sheets ===")
    wb = load_workbook(OURS, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        if "415" not in sname:
            continue
        ws = wb[sname]
        print(f"\n--- {sname} ---")
        for r in range(1, 12):
            try:
                row = list(next(ws.iter_rows(min_row=r, max_row=r, max_col=10, values_only=True)))
            except StopIteration:
                break
            vals = [str(c).replace("\n", " ")[:40] if c is not None else "" for c in row]
            if any(vals):
                print(f"  R{r}: {vals}")
        n = 0
        nonempty_cols = Counter()
        for row in ws.iter_rows(min_row=7, values_only=True):
            if row and row[0] is not None:
                n += 1
                for i, c in enumerate(row):
                    if c is not None and str(c).strip():
                        nonempty_cols[i] += 1
            if n > 3000:
                break
        print(f"  data rows: {n}, nonempty cols sample: {dict(list(nonempty_cols.items())[:15])}")
    wb.close()


def analyze_xbrl():
    print("\n=== XBRL 415 scan (stream) ===")
    with zipfile.ZipFile(ZIP) as zf:
        name = [n for n in zf.namelist() if n.lower().endswith(".xbrl")][0]
        print("entry:", name)
        # stream in chunks
        concepts_415 = Counter()
        contexts_with_415 = set()
        schema_ref = ""
        dim_members = Counter()
        sample_facts = []
        # also collect table keys from schemaRef / comments if any
        buf = ""
        fact_re = re.compile(
            r"<([A-Za-z0-9_-]+):([A-Za-z0-9_]+)[^>]*contextRef=\"([^\"]+)\"[^>]*>([^<]*)</"
        )
        # contexts mentioning 415 dimensions - hard; look for concept names with 415 or known ops
        # Better: find all concepts that appear; filter by name patterns related to operations
        known_prefixes = (
            "Operac",
            "Sdelk",
            "CZen",
            "Czen",
            "Vektor",
            "Veks",
            "Cifr",
            "Czif",
            "Pokup",
            "Proda",
            "Kol",
            "Summ",
            "Stoim",
        )
        # From taxonomy tables SR_0420415
        table_hint = re.compile(r"0420415|SR_0420415", re.I)

        with zf.open(name) as f:
            # read first 2MB for schemaRef
            head = f.read(2_000_000).decode("utf-8", errors="replace")
            m = re.search(r'schemaRef[^>]*href=\"([^\"]+)\"', head)
            if m:
                schema_ref = m.group(1)
            print("schemaRef:", schema_ref)
            print("0415 mentions in head:", len(table_hint.findall(head)))

            # reset and stream full for facts - file is 434MB, stream line by line
        with zf.open(name) as f:
            # decode streaming
            decoder = "utf-8"
            leftover = ""
            total_facts = 0
            facts_415ish = 0
            # concept name patterns for 415 form from taxonomy naming
            # We'll collect ALL unique concept local names and count those with Operacii / CZennye / Vekselya etc.
            concept_all = Counter()
            # Also look for typed members / dimensions often used in 415
            for chunk in iter(lambda: f.read(8 * 1024 * 1024), b""):
                text = leftover + chunk.decode("utf-8", errors="replace")
                # keep last incomplete line
                if "\n" in text:
                    parts = text.rsplit("\n", 1)
                    text, leftover = parts[0], parts[1]
                else:
                    leftover = text
                    continue
                for m in fact_re.finditer(text):
                    total_facts += 1
                    concept = m.group(2)
                    concept_all[concept] += 1
                    cref = m.group(3)
                    val = m.group(4)
                    cl = concept.lower()
                    # heuristic: 415 operations concepts
                    if (
                        "operac" in cl
                        or "sdelk" in cl
                        or "veksel" in cl
                        or "czifrov" in cl
                        or "cifr" in cl
                        or concept.startswith("Kolczb")
                        and "Oper" in concept
                    ):
                        facts_415ish += 1
                        concepts_415[concept] += 1
                        if len(sample_facts) < 15:
                            sample_facts.append((concept, cref, val[:40]))
                if total_facts and total_facts % 500000 < 1000:
                    pass
            if leftover:
                for m in fact_re.finditer(leftover):
                    total_facts += 1
                    concept = m.group(2)
                    concept_all[concept] += 1

        print(f"total facts scanned: {total_facts}")
        print(f"415-ish facts (heuristic): {facts_415ish}")
        print("top 415-ish concepts:")
        for c, n in concepts_415.most_common(30):
            print(f"  {n:6} {c}")

        # concepts containing 415 in name (unlikely)
        c415 = [(c, n) for c, n in concept_all.items() if "415" in c or "0415" in c]
        print("concepts with 415 in name:", c415[:20])

        # look for typical 415 measures from CBR naming
        keys = [
            "Pokupka",
            "Prodazha",
            "Vekselya",
            "CZennyeBumagi",
            "CifrovyePrava",
            "OperaciiS",
            "StoimOcz",
            "Kolczb",
            "ISIN",
            "Kod_valyuty",
        ]
        print("\nconcept name hits:")
        for k in keys:
            hits = [(c, n) for c, n in concept_all.items() if k.lower() in c.lower()]
            print(f"  {k}: {len(hits)} concepts, {sum(n for _,n in hits)} facts")
            for c, n in sorted(hits, key=lambda x: -x[1])[:8]:
                print(f"     {n:6} {c}")


if __name__ == "__main__":
    analyze_etalon()
    analyze_ours()
    analyze_xbrl()
