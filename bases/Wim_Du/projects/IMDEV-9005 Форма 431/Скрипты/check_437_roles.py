#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Check 437 role hrefs vs taxonomy labels and etalon."""

import re
import zipfile

from openpyxl import load_workbook

TAX = r"C:\Users\Acer\AppData\Local\XBRLConverter\Taxonomies\20251230.zip"
OURS = (
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ"
    r"\XBRL_Orticon_taxonomy_июнь_1_4_9.xlsx"
)
ETALON = (
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ"
    r"\401-414-415-0420437_июнь_2026_кроме 431.xlsx"
)


def local_from_href(value: str) -> str:
    text = str(value).strip()
    if "#" in text:
        text = text.split("#")[-1]
    if text.lower().startswith("mem-int_"):
        text = text[8:]
    return text


def load_mem_labels():
    z = zipfile.ZipFile(TAX)
    t = z.read("final_7_1/www.cbr.ru/xbrl/udr/dom/mem-int-label.xml").decode(
        "utf-8", "ignore"
    )
    labels = {}
    for m in re.finditer(
        r'xlink:label="label_([^"]+)"[^>]*xml:lang="ru"[^>]*>([^<]+)</link:label>',
        t,
    ):
        labels[m.group(1)] = m.group(2).strip()
    # also without requiring attribute order
    for m in re.finditer(
        r'<link:label[^>]*xlink:label="label_([^"]+)"[^>]*>([^<]+)</link:label>',
        t,
    ):
        if 'xml:lang="ru"' in m.group(0) or "xml:lang='ru'" in m.group(0):
            labels.setdefault(m.group(1), m.group(2).strip())
    return labels


def resolve(local: str, labels: dict) -> str:
    for cand in (local, local + "Member", local.replace("Member", "")):
        if cand in labels:
            return labels[cand]
    return ""


def main():
    labels = load_mem_labels()
    print("mem-int RU labels loaded:", len(labels))

    wb = load_workbook(OURS, read_only=True, data_only=True)
    ws = wb["0420437 Сведения о контрагентах"]
    roles = set()
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 8:
            continue
        v = row[9] if len(row) > 9 else None
        if v:
            roles.add(str(v))
    wb.close()

    print("unique role values in 1_4_9:", len(roles))
    for href in sorted(roles):
        local = local_from_href(href)
        ru = resolve(local, labels)
        print(" HREF:", href[:100])
        print(" LOCAL:", local, "-> RU:", ru[:120] if ru else "MISSING")

    wb = load_workbook(ETALON, data_only=True)
    ws = wb["0420437 Сведения о контрагентах"]
    etalon_roles = set()
    for r in range(10, ws.max_row + 1):
        v = ws.cell(r, 16).value
        if v:
            etalon_roles.add(str(v).strip())
    wb.close()
    print("etalon unique roles:", len(etalon_roles))
    for x in sorted(etalon_roles):
        print(" ETALON:", x[:140])


if __name__ == "__main__":
    main()
