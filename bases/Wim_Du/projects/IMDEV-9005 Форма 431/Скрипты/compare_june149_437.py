#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare june 1.4.9 vs CBR etalon; focus on 437 role column."""

from pathlib import Path
from openpyxl import load_workbook

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ")
OURS = BASE / "XBRL_Orticon_taxonomy_июнь_1_4_9.xlsx"
CBR = BASE / "401-414-415-0420437_июнь_2026_кроме 431.xlsx"


def find_sheets(wb, key):
    return [sn for sn in wb.sheetnames if key in sn]


def dump_437(path, label, use_full=False):
    print(f"\n=== {label} 437 ===")
    wb = load_workbook(path, data_only=True) if use_full else load_workbook(
        path, read_only=True, data_only=False
    )
    sheets = find_sheets(wb, "0420437")
    print("sheets:", sheets)
    for sn in sheets:
        ws = wb[sn]
        toc = ws.cell(2, 1).value
        gen = ws.cell(4, 1).value if not use_full else None
        print(f"\n-- {sn}")
        print(" TOC:", str(toc)[:100] if toc else None)
        if gen:
            print(" GEN:", str(gen)[:90])
        # headers: ours row 6; etalon multi-row
        if use_full:
            # find header row with Роль
            hdr_row = None
            for r in range(1, 15):
                for c in range(1, min(30, (ws.max_column or 1) + 1)):
                    v = ws.cell(r, c).value
                    if v and "рол" in str(v).lower():
                        hdr_row = r
                        print(f" role header at R{r}C{c}:", str(v).replace("\n", " ")[:80])
            # print row 6-12 headers
            for r in range(6, 13):
                vals = []
                for c in range(1, min(20, (ws.max_column or 1) + 1)):
                    v = ws.cell(r, c).value
                    if v:
                        vals.append(f"{c}:{str(v).replace(chr(10), ' ')[:40]}")
                if vals:
                    print(f" R{r}:", vals)
            # data samples last cols
            data_start = 13
            for r in range(1, 15):
                v = ws.cell(r, 1).value
                if v and ("Идентификатор" in str(v) or str(v).startswith("643_") or "_" in str(v)[:20]):
                    if "Идентификатор" in str(v):
                        data_start = r + 1
            print(" data_start guess", data_start, "max_row", ws.max_row, "max_col", ws.max_column)
            n = 0
            for r in range(data_start, min(ws.max_row, data_start + 15) + 1):
                rowvals = [ws.cell(r, c).value for c in range(1, (ws.max_column or 1) + 1)]
                if not any(x not in (None, "") for x in rowvals):
                    continue
                n += 1
                # show last 4 cols + first 3
                first = rowvals[:3]
                last = rowvals[-4:]
                print(f" data{n}: first={first} last={last}")
                if n >= 8:
                    break
            # count rows
            total = 0
            emails_in_last = 0
            roles_ok = 0
            last_col = ws.max_column
            for r in range(data_start, ws.max_row + 1):
                if any(ws.cell(r, c).value not in (None, "") for c in range(1, min(5, last_col) + 1)):
                    total += 1
                    v = ws.cell(r, last_col).value
                    if isinstance(v, str) and "@" in v:
                        emails_in_last += 1
                    elif v not in (None, ""):
                        roles_ok += 1
            print(f" rows={total} last_col={last_col} emails_in_last={emails_in_last} other_last={roles_ok}")
        else:
            # ours flat
            headers = []
            for c in range(1, 30):
                v = None
                # read header from row 6 via iter
                break
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
                if i == 2:
                    print(" TOC2:", str(row[0])[:100] if row[0] else None)
                if i == 4:
                    print(" GEN:", str(row[0])[:90] if row[0] else None)
                if i == 6:
                    headers = [(j + 1, str(v).replace("\n", " ")[:55]) for j, v in enumerate(row) if v]
                    print(" headers:")
                    for hc, hv in headers:
                        print(f"  H{hc}: {hv}")
            # data
            n = 0
            email_last = 0
            samples = []
            last_h = headers[-1][0] if headers else None
            for row in ws.iter_rows(min_row=7, values_only=True):
                if not any(c not in (None, "") for c in row):
                    continue
                n += 1
                last_v = row[last_h - 1] if last_h and last_h <= len(row) else None
                if isinstance(last_v, str) and "@" in last_v:
                    email_last += 1
                if n <= 8:
                    samples.append((row[0], row[1] if len(row) > 1 else None, last_v))
            print(f" rows={n} emails_in_last_col={email_last}")
            for s in samples:
                print("  sample", s)
    wb.close()


def check_415_frac(path):
    print("\n=== OURS 415 fractional qty ===")
    wb = load_workbook(path, read_only=True, data_only=False)
    for sn in wb.sheetnames:
        if "0420415" not in sn:
            continue
        ws = wb[sn]
        toc = ""
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=2, values_only=True), 2):
            toc = str(row[0] or "")
        if "1.3" not in toc:
            continue
        print("sheet", sn, toc[:70])
        frac = 0
        samples = []
        for row in ws.iter_rows(min_row=7, min_col=7, max_col=11, values_only=False):
            for cell in row:
                v = cell.value
                if isinstance(v, float) and abs(v - round(v)) > 1e-9:
                    frac += 1
                    if len(samples) < 5:
                        samples.append(v)
        print(" fractional cells in G-K:", frac, "samples", samples)
        break
    wb.close()


def main():
    print("OURS", OURS.exists(), OURS.name)
    print("CBR", CBR.exists(), CBR.name)
    dump_437(OURS, "OURS", use_full=False)
    dump_437(CBR, "ETALON", use_full=True)
    check_415_frac(OURS)


if __name__ == "__main__":
    main()
