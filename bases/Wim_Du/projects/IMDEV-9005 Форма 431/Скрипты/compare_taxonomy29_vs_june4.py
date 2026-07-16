#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare taxonomy29 (good baseline) vs june4 (many empty sheets)."""

from pathlib import Path
from openpyxl import load_workbook

OLD = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431"
    r"\ОРТИКОН\XBRL_Orticon_taxonomy29.xlsx"
)
NEW = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ"
    r"\XBRL_Orticon_taxonomy_июнь4.xlsx"
)


def sheet_stats(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = []
    gen = None
    for sn in wb.sheetnames:
        ws = wb[sn]
        toc = None
        uri = None
        g = None
        # read first few rows
        hdr = None
        data_rows = 0
        cols = 0
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
            vals = [c for c in row if c is not None and str(c).strip() != ""]
            if i == 2 and vals:
                toc = str(vals[0])[:90]
            if i == 3 and vals:
                uri = str(vals[0])[:80]
            if i == 4 and vals and "Generator" in str(vals[0]):
                g = str(vals[0])[:80]
                gen = gen or g
            if i == 6:
                hdr = sum(1 for c in row if c is not None and str(c).strip() != "")
                cols = hdr
        # count data rows (from row 7)
        for row in ws.iter_rows(min_row=7, values_only=True):
            if any(c is not None and str(c).strip() != "" for c in row):
                data_rows += 1
        rows.append(
            {
                "name": sn,
                "toc": toc or "",
                "rows": data_rows,
                "cols": cols or 0,
                "empty": data_rows == 0,
            }
        )
    wb.close()
    return gen, rows


def main():
    print("OLD", OLD.exists(), OLD)
    print("NEW", NEW.exists(), NEW)
    gen_old, old = sheet_stats(OLD)
    gen_new, new = sheet_stats(NEW)
    print("OLD gen:", gen_old)
    print("NEW gen:", gen_new)
    print("OLD sheets:", len(old), "nonempty:", sum(1 for x in old if not x["empty"]))
    print("NEW sheets:", len(new), "nonempty:", sum(1 for x in new if not x["empty"]))

    # map by sheet name and by toc
    old_by_name = {x["name"]: x for x in old}
    new_by_name = {x["name"]: x for x in new}

    print("\n=== NEW empty that OLD had data (by name) ===")
    for name, o in old_by_name.items():
        n = new_by_name.get(name)
        if o["rows"] > 0 and (n is None or n["rows"] == 0):
            print(f"  LOST {name}: old_rows={o['rows']} new={n['rows'] if n else 'MISSING'} toc={o['toc'][:70]}")

    print("\n=== NEW nonempty summary ===")
    for x in new:
        if not x["empty"]:
            print(f"  {x['rows']:5d} cols={x['cols']:2d} {x['name'][:40]} | {x['toc'][:60]}")

    print("\n=== OLD nonempty summary ===")
    for x in old:
        if not x["empty"]:
            print(f"  {x['rows']:5d} cols={x['cols']:2d} {x['name'][:40]} | {x['toc'][:60]}")

    print("\n=== ALL NEW sheets row counts ===")
    for x in new:
        mark = "EMPTY" if x["empty"] else "ok"
        print(f"  [{mark:5}] {x['rows']:5d} {x['name']}")

    # toc-based: old tocs with data missing in new
    print("\n=== TOC with data in OLD missing/empty in NEW ===")
    new_tocs = {}
    for x in new:
        key = x["toc"].strip().lower()
        new_tocs[key] = max(new_tocs.get(key, 0), x["rows"])
    for x in old:
        if x["rows"] <= 0:
            continue
        key = x["toc"].strip().lower()
        nr = new_tocs.get(key, -1)
        if nr <= 0:
            print(f"  TOC LOST rows_old={x['rows']} new={nr} | {x['toc'][:90]}")


if __name__ == "__main__":
    main()
