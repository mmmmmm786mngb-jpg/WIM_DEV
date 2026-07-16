#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Find measure columns still stored as text in taxonomy29_2."""

from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

PATH = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431"
    r"\ОРТИКОН\XBRL_Orticon_taxonomy29_2.xlsx"
)


def looks_like_number(s: str) -> bool:
    t = s.strip().replace(" ", "").replace("\u00a0", "")
    if not t or "_" in t:
        return False
    # dates
    if len(t) >= 8 and t[4:5] == "-" and t[7:8] == "-":
        return False
    # codes like 643-RUB
    if "-" in t and any(ch.isalpha() for ch in t):
        return False
    # percent / plain number with . or ,
    u = t.replace(",", ".")
    if u.count(".") > 1:
        return False
    body = u.replace(".", "").replace("-", "")
    return body.isdigit() and len(body) > 0


def main():
    wb = load_workbook(PATH, data_only=True)
    report = []
    for sn in wb.sheetnames:
        if sn == "TOC":
            continue
        ws = wb[sn]
        headers = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(6, c).value
            if h:
                headers[c] = str(h).replace("\n", " ")
        if not headers:
            continue
        stats = {
            c: {"num": 0, "textnum": 0, "samples": []}
            for c in headers
        }
        for r in range(7, ws.max_row + 1):
            empty = True
            for c in headers:
                v = ws.cell(r, c).value
                if v is not None and str(v).strip() != "":
                    empty = False
                    break
            if empty:
                continue
            for c in headers:
                v = ws.cell(r, c).value
                if v is None:
                    continue
                if isinstance(v, (int, float)):
                    stats[c]["num"] += 1
                elif isinstance(v, str) and looks_like_number(v):
                    stats[c]["textnum"] += 1
                    if len(stats[c]["samples"]) < 2:
                        stats[c]["samples"].append(v)
        for c, h in headers.items():
            st = stats[c]
            if st["textnum"] == 0:
                continue
            # skip pure id columns if header says идентификатор
            hl = h.lower()
            if "идентификатор" in hl and "сумм" not in hl:
                continue
            if "инн" in hl or "номер счета" in hl or "код " in hl[:10]:
                # still may be numeric codes - report if many
                pass
            report.append(
                (
                    st["textnum"],
                    st["num"],
                    sn,
                    c,
                    h[:60],
                    st["samples"],
                )
            )
    wb.close()
    report.sort(reverse=True)
    print("text-looking measure-ish columns (top 40):")
    for textn, num, sn, c, h, samples in report[:40]:
        print(f"  text={textn:5d} num={num:5d} {sn[:32]:32} H{c}: {h} | {samples}")
    print("total cols with text numbers:", len(report))


if __name__ == "__main__":
    main()
