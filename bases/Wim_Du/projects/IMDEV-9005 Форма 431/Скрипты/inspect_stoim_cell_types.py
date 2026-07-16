#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import zipfile
import re
from pathlib import Path
from openpyxl import load_workbook

p = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ"
    r"\XBRL_Orticon_taxonomy_июнь1.xlsx"
)

wb = load_workbook(p, read_only=True, data_only=False)
ws = wb["0420431 Раздел 4. Сведения о по"]
print("openpyxl data_type row7:")
for i, row in enumerate(ws.iter_rows(min_row=7, max_row=9)):
    for cell in row[:5]:
        if cell.value is not None:
            print(
                f"  {cell.coordinate} type={cell.data_type} "
                f"val={repr(cell.value)[:50]} number_format={cell.number_format}"
            )
wb.close()

with zipfile.ZipFile(p) as z:
    wb_xml = z.read("xl/workbook.xml").decode("utf-8")
    sheets = re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wb_xml)
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rid_to_target = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
    for name, rid in sheets:
        if "431" in name and "Раздел 4" in name and name.endswith("по"):
            target = "xl/" + rid_to_target[rid].lstrip("/")
            print("xml sheet", name, target)
            xml = z.read(target).decode("utf-8")
            rows = re.findall(r'<row[^>]*r="(\d+)"[^>]*>(.*?)</row>', xml, re.S)
            print("first row numbers", [r[0] for r in rows[:12]])
            for rnum, body in rows[5:9]:
                cells = re.findall(r'<c r="([A-Z]+)(\d+)"([^>]*)>(.*?)</c>', body, re.S)
                print("row", rnum)
                for ref, rn, attrs, inner in cells[:5]:
                    t = re.search(r't="([^"]+)"', attrs)
                    v = re.search(r"<v>([^<]*)</v>", inner)
                    print(
                        " ",
                        ref + rn,
                        "t=",
                        t.group(1) if t else "number",
                        "v=",
                        (v.group(1) if v else "")[:40],
                    )
            break
