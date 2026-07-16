#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Simulate 415 PR1.3 pivot on current flat export to estimate row/col counts."""

from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

OURS = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ"
    r"\XBRL_Orticon_taxonomy_июнь1.xlsx"
)
CBR = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ"
    r"\401-414-415-0420437_июнь_2026_кроме 431.xlsx"
)


def main():
    wb = load_workbook(OURS, read_only=True, data_only=True)
    ws = wb["0420415 Раздел 1. Операции с _2"]
    headers = [c for c in next(ws.iter_rows(min_row=6, max_row=6, values_only=True))]
    print("flat headers:", [str(h)[:40].replace("\n", " ") if h else None for h in headers])

    # columns: 0 Period, 1 Org, 2 stroka, 3 CB, then dim-like / measures mixed
    # From prior analysis: row key = Period + Org + stroka + CB
    rows_by_key = defaultdict(dict)
    flat_n = 0
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row or row[0] is None:
            continue
        flat_n += 1
        key = (row[0], row[1], row[2], row[3])
        # collect non-empty measure-ish cells
        for i, v in enumerate(row):
            if i < 4 or v is None or str(v).strip() == "":
                continue
            # skip pure dimension labels that look like axis members without numbers
            rows_by_key[key][i] = v
    wb.close()

    print("flat rows:", flat_n)
    print("pivoted unique keys:", len(rows_by_key))

    wb2 = load_workbook(CBR, data_only=True)
    ws2 = wb2["0420415 Раздел 1 Операции с ц_3"]
    etalon_n = 0
    for r in range(13, ws2.max_row + 1):
        if ws2.cell(r, 1).value is not None:
            etalon_n += 1
    print("etalon data rows:", etalon_n)
    print("diff pivot vs etalon:", len(rows_by_key) - etalon_n)
    wb2.close()


if __name__ == "__main__":
    main()
