#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Verify OKATO fix and numeric columns in June1 export."""

from openpyxl import load_workbook
from pathlib import Path
from collections import Counter
import re

P = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ"
    r"\XBRL_Orticon_taxonomy_июнь1.xlsx"
)


def classify_okato(v):
    s = str(v).strip()
    if re.match(r"^OKATO", s, re.I):
        return "raw_okato"
    if re.match(r"^\d{3}$", s):
        return "oksm_3digit"
    if re.match(r"^\d{2,5}\s*-", s):
        return "region_label"
    return "other"


def main():
    print("file:", P.name, "exists:", P.exists(), "MB:", round(P.stat().st_size / 1024 / 1024, 2))
    wb = load_workbook(P, read_only=True, data_only=True)

    # Generator
    ws = wb[wb.sheetnames[1]]
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=1, values_only=True):
        if row[0] and "Generator" in str(row[0]):
            print("Generator:", row[0])
    if "TOC" in wb.sheetnames:
        for row in wb["TOC"].iter_rows(min_row=1, max_row=5, max_col=2, values_only=True):
            if row and row[0] and "Generator" in str(row[0]):
                print("TOC Generator:", row[0], row[1] if len(row) > 1 else "")
            if row and len(row) > 1 and row[1] and "Generator" in str(row[0]):
                print("TOC:", row[0], row[1])

    # --- OKATO ---
    print("\n=== OKATO ===")
    raw_total = 0
    good_total = 0
    oksm_total = 0
    raw_vals = Counter()
    target_names = [
        "Херсон",
        "Донецк",
        "Луганск",
        "Запорож",
    ]
    found_targets = Counter()

    for sname in wb.sheetnames:
        if "431" not in sname and "415" not in sname:
            continue
        ws = wb[sname]
        okato_col = None
        hr = 6
        try:
            for r in range(4, 10):
                row = list(next(ws.iter_rows(min_row=r, max_row=r, values_only=True)))
                for i, h in enumerate(row):
                    if h and "ОКАТО" in str(h).upper():
                        okato_col = i
                        hr = r
                        break
                if okato_col is not None:
                    break
        except StopIteration:
            continue
        if okato_col is None:
            continue

        bad = 0
        good = 0
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            if not row or (row[0] is None and (len(row) <= okato_col or row[okato_col] is None)):
                continue
            if len(row) <= okato_col or row[okato_col] is None:
                continue
            v = str(row[okato_col]).strip()
            if not v:
                continue
            cl = classify_okato(v)
            if cl == "raw_okato":
                bad += 1
                raw_vals[v] += 1
            elif cl == "oksm_3digit":
                oksm_total += 1
            elif cl == "region_label":
                good += 1
                for t in target_names:
                    if t.lower() in v.lower():
                        found_targets[v] += 1
        if good or bad:
            print(f"  {sname[:48]:48} good={good:6} raw={bad:4}")
            good_total += good
            raw_total += bad

    print(f"TOTAL good region labels: {good_total}")
    print(f"TOTAL raw OKATO*: {raw_total}")
    print(f"TOTAL oksm 3digit only: {oksm_total}")
    if raw_vals:
        print("RAW leftovers:")
        for v, c in raw_vals.most_common(20):
            print(f"  {c:5} {v}")
    else:
        print("RAW leftovers: NONE (OK)")
    print("New regions samples:")
    for v, c in sorted(found_targets.items(), key=lambda x: -x[1])[:10]:
        print(f"  {c:5} {v}")

    # --- Numeric columns ---
    print("\n=== NUMERIC COLUMNS ===")
    checks = [
        ("0420431 Раздел 1. Сведения об_1", ["колич"]),
        ("0420431 Раздел 4. Сведения о по", ["стоим", "колич", "оцен"]),
        ("0420409 Раздел 1. Сведения о ба", ["остат", "дско", "сумм"]),
    ]
    # find actual sheet names
    for hint, keywords in [
        ("об_1", ["колич"]),
        ("Раздел 4", ["стоим", "колич"]),
        ("0409 Раздел 1", ["dsko", "остат", "сумм", "оборот"]),
    ]:
        sname = None
        for s in wb.sheetnames:
            if hint in s:
                sname = s
                break
        if not sname:
            print(f" sheet with '{hint}' not found")
            continue
        ws = wb[sname]
        headers = list(next(ws.iter_rows(min_row=6, max_row=6, values_only=True)))
        print(f"\n--- {sname} ---")
        for i, h in enumerate(headers):
            if not h:
                continue
            hs = str(h).replace("\n", " ")
            hl = hs.lower()
            if not any(k in hl for k in keywords) and not any(
                k in hl for k in ["колич", "стоим", "оцен", "сумм"]
            ):
                # also DSKO by checking sample later via known money headers
                if "количество" not in hl and "стоим" not in hl and "оцен" not in hl:
                    continue
            types = Counter()
            samples = []
            for row in ws.iter_rows(min_row=7, max_row=300, values_only=True):
                if len(row) <= i or row[i] is None:
                    continue
                types[type(row[i]).__name__] += 1
                if len(samples) < 3:
                    samples.append(repr(row[i])[:40])
            num = types.get("int", 0) + types.get("float", 0)
            text = types.get("str", 0)
            status = "OK" if num > 0 and text == 0 else ("MIXED" if num and text else ("TEXT" if text else "EMPTY"))
            print(f"  [{status}] col{i}: {hs[:55]}")
            print(f"         types={dict(types)} samples={samples}")

    # bank sheet DSKO if present
    for s in wb.sheetnames:
        if "0409" in s and "Раздел 1" in s:
            ws = wb[s]
            headers = list(next(ws.iter_rows(min_row=6, max_row=6, values_only=True)))
            print(f"\n--- {s} (all nonempty typed cols sample) ---")
            for i, h in enumerate(headers):
                if not h:
                    continue
                types = Counter()
                samples = []
                for row in ws.iter_rows(min_row=7, max_row=50, values_only=True):
                    if len(row) <= i or row[i] is None:
                        continue
                    types[type(row[i]).__name__] += 1
                    if len(samples) < 2:
                        samples.append(repr(row[i])[:35])
                if not types:
                    continue
                hs = str(h).replace("\n", " ")[:50]
                num = types.get("int", 0) + types.get("float", 0)
                text = types.get("str", 0)
                # only print if looks numeric-ish header or has numbers
                if num or any(x in hs.lower() for x in ["сумм", "остат", "оборот", "дско"]):
                    st = "OK" if num and not text else ("TEXT" if text and not num else "MIX")
                    print(f"  [{st}] {i}: {hs} {dict(types)} {samples}")
            break

    wb.close()
    print("\n=== VERDICT ===")
    okato_ok = raw_total == 0
    print(f"OKATO raw members gone: {'YES' if okato_ok else 'NO'}")


if __name__ == "__main__":
    main()
