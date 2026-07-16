#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Analyze OKATO issues in user complaint Excel file."""

from openpyxl import load_workbook
from pathlib import Path
from collections import Counter, defaultdict
import re

P = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431\Тестирование\ОШИБКИ_XBRL_Orticon_taxonomy_июнь.xlsx")

def classify(v):
    s = str(v).strip()
    if re.match(r"^OKATO", s, re.I):
        return "raw_okato_member"
    if re.match(r"^\d{3}$", s):
        return "oksm_3digit"
    if re.match(r"^\d{2}\s*-", s) or re.match(r"^\d{5}\s*-", s):
        return "region_label"
    if re.match(r"^\d{2,5}$", s):
        return "code_only"
    return "other"

wb = load_workbook(P, read_only=True, data_only=True)

# TOC / generator
ws = wb["TOC"]
print("=== TOC / meta ===")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, max_col=3, values_only=True), 1):
    vals = [str(c) if c is not None else "" for c in row]
    line = " | ".join(vals).strip(" |")
    if line.strip():
        print(f"{i}: {line[:120]}")

# Scan all sheets for OKATO columns and bad values
print("\n=== OKATO columns scan ===")
by_sheet = []
all_raw = Counter()
all_oksm = Counter()
all_good_unique = set()
all_bad_unique = set()

for sname in wb.sheetnames:
    ws = wb[sname]
    # find header with OKATO
    okato_cols = []
    header_row = None
    try:
        for r in range(1, 15):
            row = list(next(ws.iter_rows(min_row=r, max_row=r, values_only=True)))
            for i, h in enumerate(row):
                if h and ("ОКАТО" in str(h).upper() or "OKATO" in str(h).upper()):
                    okato_cols.append((r, i, str(h).replace("\n", " ")[:70]))
                    header_row = r
    except StopIteration:
        pass

    if not okato_cols:
        continue

    # use first okato col
    hr, col, hdr = okato_cols[0]
    vals = []
    classes = Counter()
    try:
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            if row is None:
                continue
            # stop heuristic
            if (row[0] is None or str(row[0]).strip() == "") and (len(row) <= col or row[col] is None):
                # allow empty first cols sometimes
                if all(c is None or str(c).strip() == "" for c in row[: min(5, len(row))]):
                    continue
            if len(row) <= col or row[col] is None:
                continue
            v = str(row[col]).strip()
            if not v:
                continue
            vals.append(v)
            cl = classify(v)
            classes[cl] += 1
            if cl == "raw_okato_member":
                all_raw[v] += 1
                all_bad_unique.add(v)
            elif cl == "oksm_3digit":
                all_oksm[v] += 1
                all_bad_unique.add(v)
            elif cl == "region_label":
                all_good_unique.add(v)
            elif cl == "other" and ("OKATO" in v.upper() or "Member" in v):
                all_raw[v] += 1
                all_bad_unique.add(v)
    except Exception as e:
        print("ERR", sname, e)
        continue

    bad_n = classes.get("raw_okato_member", 0) + classes.get("oksm_3digit", 0)
    if vals:
        by_sheet.append((sname, len(vals), len(set(vals)), dict(classes), bad_n, hdr))

wb.close()

print(f"sheets with OKATO column: {len(by_sheet)}")
for sname, n, u, cls, bad_n, hdr in sorted(by_sheet, key=lambda x: -x[4]):
    mark = " BAD" if bad_n else ""
    print(f"  {sname[:48]:48} filled={n:5} uniq={u:3} bad={bad_n:4}{mark}  {cls}")

print("\n=== RAW OKATO members (problem) ===")
print(f"unique raw: {len(all_raw)}  total cells: {sum(all_raw.values())}")
for v, c in all_raw.most_common(50):
    print(f"  {c:5}  {v}")

print("\n=== OKSM 3-digit ===")
for v, c in all_oksm.most_common(20):
    print(f"  {c:5}  {v}")

print("\n=== SUMMARY ===")
print(f"good region labels unique: {len(all_good_unique)}")
print(f"bad unique values: {len(all_bad_unique)}")
print(f"bad cells total: {sum(all_raw.values()) + sum(all_oksm.values())}")
