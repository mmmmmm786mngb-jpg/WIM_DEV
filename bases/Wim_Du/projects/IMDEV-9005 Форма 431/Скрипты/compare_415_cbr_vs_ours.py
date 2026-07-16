#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare form 415: CBR official Excel vs our Orticon converter."""

from openpyxl import load_workbook
from pathlib import Path
from collections import Counter
import re

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ")
CBR = BASE / "401-414-415-0420437_июнь_2026_кроме 431.xlsx"
OURS = BASE / "XBRL_Orticon_taxonomy_июнь.xlsx"


def sheet_profile_full(ws, title):
    """For small CBR workbook (non-read-only)."""
    mr = ws.max_row or 0
    mc = ws.max_column or 0
    # detect header block: rows 1-15
    headers = []
    for r in range(1, min(15, mr) + 1):
        vals = []
        for c in range(1, min(mc, 25) + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                vals.append((c, str(v).replace("\n", " ")[:50]))
        if vals:
            headers.append((r, vals[:8]))
    # data rows: from first row after 'Идентификатор' header or from row 13
    start = 13
    for r, vals in headers:
        texts = " ".join(v for _, v in vals)
        if "Идентификатор" in texts and r >= 9:
            start = r + 1
            break
    data_rows = 0
    numeric_cells = 0
    text_cells = 0
    empty_axis = False
    for r, vals in headers:
        for _, v in vals:
            if "EMPTY_AXIS" in v:
                empty_axis = True
    samples = []
    for r in range(start, mr + 1):
        rowvals = [ws.cell(r, c).value for c in range(1, mc + 1)]
        if not any(v not in (None, "") for v in rowvals):
            continue
        data_rows += 1
        for v in rowvals:
            if v is None or v == "":
                continue
            if isinstance(v, (int, float)):
                numeric_cells += 1
            else:
                text_cells += 1
        if len(samples) < 2:
            samples.append([str(v)[:28] if v is not None else "" for v in rowvals[:8]])
    return {
        "title": title,
        "max_row": mr,
        "max_col": mc,
        "data_rows": data_rows,
        "empty_axis": empty_axis,
        "num": numeric_cells,
        "txt": text_cells,
        "headers": headers[:6],
        "samples": samples,
        "layout": "pivot" if any("Наименование показателя" in str(ws.cell(6, 1).value or "") for _ in [0]) else "?",
    }


def sheet_profile_ours(ws, title):
    """Flat layout from our converter."""
    # row 6 = headers typically
    headers = []
    try:
        row6 = list(next(ws.iter_rows(min_row=6, max_row=6, values_only=True)))
        headers = [str(h).replace("\n", " ")[:45] for h in row6 if h]
    except StopIteration:
        pass
    n = 0
    types = Counter()
    sample = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        n += 1
        for c in row:
            if c is None:
                continue
            types[type(c).__name__] += 1
        if len(sample) < 2:
            sample.append([str(c)[:28] if c is not None else "" for c in row[:8]])
    return {
        "title": title,
        "data_rows": n,
        "headers": headers[:12],
        "types": dict(types),
        "samples": sample,
    }


def main():
    print("=== CBR official (pivot) 415 ===")
    wb = load_workbook(CBR, read_only=False, data_only=True)
    cbr_profiles = []
    for sname in wb.sheetnames:
        if "415" not in sname:
            continue
        p = sheet_profile_full(wb[sname], sname)
        cbr_profiles.append(p)
        flag = "EMPTY" if p["data_rows"] == 0 else "DATA"
        ea = " [EMPTY_AXIS]" if p["empty_axis"] else ""
        print(
            f"  [{flag}]{ea} {sname[:42]:42} "
            f"R={p['max_row']:3} C={p['max_col']:3} data={p['data_rows']:4} "
            f"num={p['num']} txt={p['txt']}"
        )
        if p["samples"]:
            print(f"         sample: {p['samples'][0]}")
        # show multirow header gist
        if p["data_rows"] > 0 or True:
            for r, vals in p["headers"]:
                if r >= 6:
                    print(f"         H{r}: {vals[:5]}")
    wb.close()

    print("\n=== OURS flat 415 ===")
    wb = load_workbook(OURS, read_only=True, data_only=True)
    ours_profiles = []
    for sname in wb.sheetnames:
        if "415" not in sname:
            continue
        p = sheet_profile_ours(wb[sname], sname)
        ours_profiles.append(p)
        flag = "EMPTY" if p["data_rows"] == 0 else "DATA"
        print(f"  [{flag}] {sname[:48]:48} data={p['data_rows']:5} types={p['types']}")
        if p["headers"]:
            print(f"         cols: {p['headers'][:6]}")
        if p["samples"]:
            print(f"         sample: {p['samples'][0]}")
    wb.close()

    print("\n=== SUMMARY ===")
    cbr_data = sum(1 for p in cbr_profiles if p["data_rows"] > 0)
    cbr_empty = sum(1 for p in cbr_profiles if p["data_rows"] == 0)
    ours_data = sum(1 for p in ours_profiles if p["data_rows"] > 0)
    ours_empty = sum(1 for p in ours_profiles if p["data_rows"] == 0)
    print(f"CBR sheets: {len(cbr_profiles)} (with data {cbr_data}, empty {cbr_empty})")
    print(f"OURS sheets: {len(ours_profiles)} (with data {ours_data}, empty {ours_empty})")
    print(f"CBR total data rows: {sum(p['data_rows'] for p in cbr_profiles)}")
    print(f"OURS total data rows: {sum(p['data_rows'] for p in ours_profiles)}")


if __name__ == "__main__":
    main()
