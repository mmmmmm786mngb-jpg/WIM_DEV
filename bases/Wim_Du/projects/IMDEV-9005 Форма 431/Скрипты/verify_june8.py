#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Verify XBRL_Orticon_taxonomy_июнь8.xlsx: version, numbers, 415, OKATO."""

from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ")
OURS = BASE / "XBRL_Orticon_taxonomy_июнь8.xlsx"
CBR = BASE / "401-414-415-0420437_июнь_2026_кроме 431.xlsx"


def looks_like_number(s: str) -> bool:
    t = s.strip().replace(" ", "").replace("\u00a0", "")
    if not t or "_" in t or ":" in t:
        return False
    if len(t) >= 8 and t[4:5] == "-" and t[7:8] == "-":
        return False
    if "-" in t and any(ch.isalpha() for ch in t):
        return False
    u = t.replace(",", ".")
    if u.count(".") > 1:
        return False
    body = u.replace(".", "").replace("-", "")
    return body.isdigit() and len(body) > 0


def is_id_header(h: str) -> bool:
    hl = h.lower()
    keys = (
        "идентификатор",
        "инн",
        "огрн",
        "номер счета",
        "код страны",
        "код по окато",
        "код территории",
        "период",
        "bic",
        "бик",
    )
    return any(k in hl for k in keys)


def analyze(path):
    wb = load_workbook(path, data_only=False)
    gen = None
    sheets_info = []
    measure_text = []
    measure_ok = []
    raw_okato = 0
    total_sheets = len(wb.sheetnames)
    nonempty = 0

    for sn in wb.sheetnames:
        ws = wb[sn]
        toc = str(ws.cell(2, 1).value or "")
        g = str(ws.cell(4, 1).value or "")
        if "Generator" in g:
            gen = g
        headers = {}
        for c in range(1, (ws.max_column or 1) + 1):
            h = ws.cell(6, c).value
            if h:
                headers[c] = str(h).replace("\n", " ")
        rows = 0
        for r in range(7, (ws.max_row or 6) + 1):
            if not any(
                ws.cell(r, c).value not in (None, "")
                for c in range(1, min(12, (ws.max_column or 1) + 1))
            ):
                continue
            rows += 1
            for c, h in headers.items():
                cell = ws.cell(r, c)
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith("OKATO") and any(
                    ch.isalpha() for ch in v[5:]
                ):
                    raw_okato += 1
                if is_id_header(h):
                    continue
                hl = h.lower()
                measure_like = any(
                    k in hl
                    for k in (
                        "стоим",
                        "сумм",
                        "количеств",
                        "остат",
                        "оборот",
                        "поступл",
                        "изъят",
                        "доходн",
                        "номинальн",
                        "процент",
                        "горизонт",
                        "оценк",
                        "всего",
                        "забаланс",
                        "балансов",
                    )
                )
                if not measure_like and not (
                    isinstance(v, str) and looks_like_number(v) and ("." in v or "," in v)
                ):
                    continue
                if isinstance(v, (int, float)) or cell.data_type == "n":
                    measure_ok.append((sn, h[:40], v, cell.data_type))
                elif isinstance(v, str) and looks_like_number(v):
                    measure_text.append((sn, h[:40], v, cell.data_type))
        if rows > 0:
            nonempty += 1
        sheets_info.append((sn, rows, toc[:70]))

    # 415 1.3
    s415 = None
    for sn, rows, toc in sheets_info:
        if "0420415" in sn or "0420415" in toc:
            if "1.3" in toc:
                s415 = (sn, rows, toc)
                break

    wb.close()
    return {
        "gen": gen,
        "total": total_sheets,
        "nonempty": nonempty,
        "raw_okato": raw_okato,
        "measure_ok_n": len(measure_ok),
        "measure_text_n": len(measure_text),
        "measure_text_top": measure_text[:25],
        "measure_ok_sample": measure_ok[:8],
        "s415": s415,
        "sheets": sheets_info,
    }


def main():
    print("FILE", OURS.exists(), OURS.name, "MB", round(OURS.stat().st_size / 1e6, 2))
    a = analyze(OURS)
    print("GEN:", a["gen"])
    print("sheets:", a["total"], "nonempty:", a["nonempty"], "empty:", a["total"] - a["nonempty"])
    print("raw OKATO:", a["raw_okato"])
    print("measure cells number:", a["measure_ok_n"], "still text:", a["measure_text_n"])
    print("415 1.3:", a["s415"])
    if a["measure_text_top"]:
        print("\nStill TEXT measures (sample):")
        # group by sheet+header
        g = defaultdict(int)
        samples = {}
        for sn, h, v, dt in a["measure_text_top"]:
            g[(sn, h)] += 1
            samples.setdefault((sn, h), v)
        # recount all text grouped
        g2 = defaultdict(int)
        samples2 = {}
        for sn, h, v, dt in [
            x
            for x in a["measure_text_top"]
        ]:
            pass
    # full group from re-scan lighter: print top from list
    from collections import Counter

    ctr = Counter((sn, h) for sn, h, v, dt in [
        # need full list - re-get from attribute
    ])
    # use measure_text which is only first collected - regenerate properly below

    print("\nOK number samples:")
    for x in a["measure_ok_sample"]:
        print(" ", x)

    # detailed text measures full pass on key sheets
    print("\n=== Key sheets cell types ===")
    wb = load_workbook(OURS, data_only=False)
    for sn in wb.sheetnames:
        if not any(k in sn for k in ("0420431 Раздел 4", "0420409 Раздел 1", "0420415 Раздел 1. Операции с _2", "0420414")):
            continue
        if "0420431 Раздел 4. Сведения о по" not in sn and "0420409 Раздел 1. Сведения о ба" not in sn and "0420415" not in sn and "0420414" not in sn:
            if "0420431 Раздел 4" not in sn:
                continue
        ws = wb[sn]
        print("--", sn, "rows~", ws.max_row)
        headers = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(6, c).value
            if h:
                headers[c] = str(h).replace("\n", " ")[:45]
        stats = {c: {"n": 0, "s_num": 0, "sample_n": None, "sample_s": None} for c in headers}
        for r in range(7, min(ws.max_row, 200) + 1):
            for c in headers:
                cell = ws.cell(r, c)
                v = cell.value
                if v is None:
                    continue
                if cell.data_type == "n" or isinstance(v, (int, float)):
                    stats[c]["n"] += 1
                    if stats[c]["sample_n"] is None:
                        stats[c]["sample_n"] = v
                elif isinstance(v, str) and looks_like_number(v):
                    stats[c]["s_num"] += 1
                    if stats[c]["sample_s"] is None:
                        stats[c]["sample_s"] = v
        for c, h in headers.items():
            st = stats[c]
            if st["n"] == 0 and st["s_num"] == 0:
                continue
            if is_id_header(h) and st["n"] == 0:
                continue
            mark = "OK" if st["s_num"] == 0 and st["n"] > 0 else ("TEXT" if st["s_num"] > 0 else "?")
            if mark == "TEXT" or (not is_id_header(h) and (st["n"] or st["s_num"])):
                if is_id_header(h):
                    continue
                print(
                    f"  [{mark}] H{c} n={st['n']} text={st['s_num']} {h} | n={st['sample_n']} s={st['sample_s']}"
                )

    # 415 full row count
    for sn in wb.sheetnames:
        ws = wb[sn]
        toc = str(ws.cell(2, 1).value or "")
        if "Подраздел 1.3" in toc or "1.3." in toc and "0420415" in toc:
            n = 0
            for r in range(7, ws.max_row + 1):
                if any(ws.cell(r, c).value not in (None, "") for c in range(1, 5)):
                    n += 1
            print("\n415 PR1.3 sheet", sn, "data rows", n, "gen", str(ws.cell(4, 1).value)[:80])
            # headers
            hdr = [str(ws.cell(6, c).value).replace("\n", " ")[:35] if ws.cell(6, c).value else None for c in range(1, 15)]
            print(" headers", hdr)
            # first data row types
            for c in range(1, 15):
                cell = ws.cell(7, c)
                if cell.value is not None:
                    print("  c", c, type(cell.value).__name__, cell.data_type, repr(cell.value)[:50])
            break

    # etalon 415 rows
    if CBR.exists():
        wb2 = load_workbook(CBR, data_only=True)
        ws2 = wb2["0420415 Раздел 1 Операции с ц_3"]
        et = sum(1 for r in range(13, ws2.max_row + 1) if ws2.cell(r, 2).value is not None or ws2.cell(r, 3).value is not None)
        print("etalon 415 rows", et)
        wb2.close()
    wb.close()
    print("DONE")


if __name__ == "__main__":
    main()
