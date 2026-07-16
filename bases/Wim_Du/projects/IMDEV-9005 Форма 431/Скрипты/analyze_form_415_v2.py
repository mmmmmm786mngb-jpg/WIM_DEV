#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""415: etalon (full load) + taxonomy tables + XBRL dim match."""

import zipfile
import re
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ")
ZIP = BASE / "июнь_XBRL_1027739323600_ep_nso_purcb_m_q_y_10rd_ex_reestr_0420417_20260630.zip"
ETALON = BASE / "401-414-415-0420437_июнь_2026_кроме 431.xlsx"
OURS = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431"
    r"\Тестирование\ОШИБКИ_XBRL_Orticon_taxonomy_июнь.xlsx"
)
TAX = Path(r"C:\Users\Acer\AppData\Local\XBRLConverter\Taxonomies\20251230.zip")


def etalon_415():
    print("=== ETALON 415 (full workbook) ===")
    wb = load_workbook(ETALON, read_only=False, data_only=True)
    for sname in wb.sheetnames:
        if "415" not in sname:
            continue
        ws = wb[sname]
        mr, mc = ws.max_row or 0, ws.max_column or 0
        # count rows with any value below row 12
        data = 0
        for r in range(12, mr + 1):
            if any(ws.cell(r, c).value not in (None, "") for c in range(1, min(mc, 20) + 1)):
                data += 1
        # TOC description from row2
        title = ws.cell(2, 1).value
        print(f"  {sname[:42]:42} maxR={mr:4} maxC={mc:3} dataFromR12={data:5} title={str(title)[:55]}")
        if data > 0:
            # sample first data row
            for r in range(12, mr + 1):
                vals = [ws.cell(r, c).value for c in range(1, min(8, mc) + 1)]
                if any(v not in (None, "") for v in vals):
                    print(f"    first data R{r}: {[str(v)[:28] if v is not None else '' for v in vals]}")
                    break
    wb.close()


def ours_415():
    print("\n=== OURS 415 ===")
    wb = load_workbook(OURS, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        if "415" not in sname:
            continue
        ws = wb[sname]
        n = 0
        for row in ws.iter_rows(min_row=7, max_col=4, values_only=True):
            if row and any(c is not None for c in row):
                n += 1
        print(f"  {sname[:48]:48} rows>~{n}")
    wb.close()


def tax_415():
    print("\n=== Taxonomy SR_0420415 ===")
    z = zipfile.ZipFile(TAX)
    keys = sorted(
        {
            m.group(1)
            for n in z.namelist()
            if (m := re.search(r"(SR_0420415_[A-Za-z0-9_]+)", n.replace("\\", "/")))
        }
    )
    # clean trailing -lab etc
    keys2 = []
    for k in keys:
        k2 = re.sub(r"-(lab|rend|def|pre)$", "", k)
        keys2.append(k2)
    keys2 = sorted(set(keys2))
    print("tables:", len(keys2))
    for k in keys2:
        print(" ", k)

    # dims per table from rend
    all_dims = Counter()
    for k in keys2:
        rends = [n for n in z.namelist() if k in n and n.endswith("-rend.xml")]
        if not rends:
            continue
        text = z.read(rends[0]).decode("utf-8", errors="replace")
        dims = re.findall(r"dimension=\"[^\"]*?([A-Za-z0-9_]+(?:Axis|Taxis))\"", text)
        if not dims:
            dims = re.findall(r"([A-Za-z0-9_]+(?:Axis|Taxis))", text)
        dims = sorted(set(dims))
        for d in dims:
            all_dims[d] += 1
        print(f"  {k}: dims={dims}")
    print("\nmost common 415 dims:")
    for d, n in all_dims.most_common(20):
        print(f"  {n:2} {d}")
    return list(all_dims.keys())


def xbrl_scan_dims(dim_names):
    print("\n=== XBRL contexts with 415 dims ===")
    # use unique distinctive dims
    needles = [
        d
        for d in dim_names
        if any(
            x in d.lower()
            for x in (
                "tip_oper",
                "nominal",
                "schet_tipa",
                "vid_scheta",
                "veksel",
                "cifrov",
                "czifrov",
                "depozitarn",
                "obremenen",
                "prinadl",
            )
        )
    ]
    if not needles:
        needles = dim_names[:15]
    print("needles:", needles[:20])

    with zipfile.ZipFile(ZIP) as zf:
        name = [n for n in zf.namelist() if n.lower().endswith(".xbrl")][0]
        ctx_hit = set()
        dim_hits = Counter()
        samples = []
        with zf.open(name) as f:
            buf = ""
            in_ctx = False
            ctx_id = ""
            body_parts = []
            for chunk in iter(lambda: f.read(12 * 1024 * 1024), b""):
                text = buf + chunk.decode("utf-8", errors="replace")
                while True:
                    if not in_ctx:
                        m = re.search(r"<[^>\s]*:?context\b[^>]*\bid=\"([^\"]+)\"", text)
                        if not m:
                            buf = text[-3000:]
                            break
                        in_ctx = True
                        ctx_id = m.group(1)
                        text = text[m.end() :]
                        body_parts = []
                    m2 = re.search(r"</[^>\s]*:?context>", text)
                    if not m2:
                        body_parts.append(text)
                        buf = ""
                        break
                    body_parts.append(text[: m2.start()])
                    body = "".join(body_parts)
                    hit = [n for n in needles if n in body]
                    if hit:
                        ctx_hit.add(ctx_id)
                        for h in hit:
                            dim_hits[h] += 1
                        if len(samples) < 6:
                            members = re.findall(
                                r"dimension=\"[^\"]*([A-Za-z0-9_]+(?:Axis|Taxis))\"[^>]*>([^<]{0,80})",
                                body,
                            )
                            samples.append((ctx_id, members[:10]))
                    in_ctx = False
                    text = text[m2.end() :]
                    body_parts = []

        print(f"matching contexts: {len(ctx_hit)}")
        for d, n in dim_hits.most_common(15):
            print(f"  dim hits {n:6} {d}")
        for s in samples:
            print(" sample", s[0], s[1])

        if not ctx_hit:
            return

        fact_re = re.compile(
            r"<([A-Za-z0-9_-]+):([A-Za-z0-9_]+)[^>]*contextRef=\"([^\"]+)\"[^>]*>"
        )
        facts = Counter()
        n = 0
        with zf.open(name) as f:
            leftover = ""
            for chunk in iter(lambda: f.read(12 * 1024 * 1024), b""):
                text = leftover + chunk.decode("utf-8", errors="replace")
                if "\n" in text:
                    text, leftover = text.rsplit("\n", 1)
                else:
                    leftover = text
                    continue
                for m in fact_re.finditer(text):
                    if m.group(3) in ctx_hit:
                        n += 1
                        facts[m.group(2)] += 1
        print(f"facts in 415-like contexts: {n}")
        for c, cnt in facts.most_common(30):
            print(f"  {cnt:6} {c}")


if __name__ == "__main__":
    etalon_415()
    ours_415()
    dims = tax_415()
    xbrl_scan_dims(dims)
