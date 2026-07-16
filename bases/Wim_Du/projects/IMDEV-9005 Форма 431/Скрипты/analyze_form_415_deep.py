#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Deep dive: 415 tables in taxonomy + XBRL dims + etalon/ours row counts."""

import zipfile
import re
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ")
ZIP = BASE / "июнь_XBRL_1027739323600_ep_nso_purcb_m_q_y_10rd_ex_reestr_0420417_20260630.zip"
ETALON = BASE / "401-414-415-0420437_июнь_2026_кроме 431.xlsx"
OURS = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431"
    r"\Тестирование\ОШИБКИ_XBRL_Orticon_taxonomy_июнь.xlsx"
)
TAX = Path(r"C:\Users\Acer\AppData\Local\XBRLConverter\Taxonomies\20251230.zip")


def count_sheet_rows(path, label):
    print(f"\n=== {label} 415 row counts ===")
    wb = load_workbook(path, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        if "415" not in sname:
            continue
        ws = wb[sname]
        # find first row that looks like data (has value in col A that is not TOC/header keyword)
        rows = 0
        header_hits = 0
        sample = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            a = str(row[0]).strip() if row[0] is not None else ""
            joined = " ".join(str(c) for c in row[:6] if c)
            if a in ("TOC", "Наименование показателя", "Generator") or a.startswith("http"):
                header_hits += 1
                continue
            if a.startswith("0420415") or a.startswith("T=") or a.startswith("Идентификатор"):
                header_hits += 1
                continue
            if re.match(r"^\d+$", a) and i < 15:
                header_hits += 1
                continue
            # data-ish
            rows += 1
            if len(sample) < 2:
                sample.append((i, [str(c)[:30] if c is not None else "" for c in row[:6]]))
        print(f"  {sname[:48]:48} data~{rows:5} hdr~{header_hits} sample={sample[:1]}")
    wb.close()


def taxonomy_415_tables():
    print("\n=== Taxonomy SR_0420415 tables ===")
    if not TAX.exists():
        print("tax missing")
        return []
    z = zipfile.ZipFile(TAX)
    tabs = [n for n in z.namelist() if "0420415" in n and n.endswith(("-lab.xml", "-rend.xml", ".xsd"))]
    print(f"files with 0420415: {len(tabs)}")
    tables = sorted({re.search(r"(SR_0420415[^/\\\\]+)", n.replace("\\", "/")) for n in tabs})
    keys = []
    for n in sorted(set(tabs)):
        if "-lab.xml" in n:
            print(" ", n.split("0420415")[-1][:80] if False else n[-90:])
            m = re.search(r"(SR_0420415[A-Za-z0-9_]*)", n)
            if m:
                keys.append(m.group(1))
    keys = sorted(set(keys))
    print("table keys:", len(keys))
    for k in keys:
        print(" ", k)

    # read one rend for dims/concepts
    for k in keys[:3]:
        rends = [n for n in z.namelist() if k in n and n.endswith("-rend.xml")]
        if not rends:
            continue
        text = z.read(rends[0]).decode("utf-8", errors="replace")
        dims = sorted(set(re.findall(r"dimension=\"[^\"]*#([A-Za-z0-9_]+)\"", text)))
        if not dims:
            dims = sorted(set(re.findall(r">([A-Za-z0-9_]+(?:Axis|Taxis))<", text)))
        concepts = sorted(set(re.findall(r"href=\"[^\"]*#([A-Za-z0-9_]+)\"", text)))
        concepts = [c for c in concepts if not c.endswith(("Axis", "Taxis", "Member", "Table", "LineItems"))]
        print(f"\n  {k}: dims={dims[:12]}")
        print(f"    concepts sample: {concepts[:15]}")
    return keys


def xbrl_415_by_typed_ids():
    """Find facts in contexts that look like 415 depositary ops (ID_stroki + security id patterns)."""
    print("\n=== XBRL: contexts/facts for 415-like typed axes ===")
    with zipfile.ZipFile(ZIP) as zf:
        name = [n for n in zf.namelist() if n.lower().endswith(".xbrl")][0]
        # First pass: collect context ids that have dimensions used by 415
        # From taxonomy often: Tip_operaciiAxis, Priznak_nominalnogo_derzhatelya, etc.
        dim_needles = [
            "Tip_operacii",
            "Priznak_nominal",
            "Schet_tipa_S",
            "Vid_scheta_depo",
            "Nominalnyj_derzhatel",
            "Vyshestoyashhij_depozitarij",
            "Cifrov",
            "Veksel",
            "Depozitarn",
        ]
        # Stream: for each context block, if any needle in it, mark context id
        ctx_415 = set()
        fact_by_concept = Counter()
        sample_ctx = []
        with zf.open(name) as f:
            buf = ""
            in_ctx = False
            ctx_id = ""
            ctx_body = []
            for chunk in iter(lambda: f.read(8 * 1024 * 1024), b""):
                text = buf + chunk.decode("utf-8", errors="replace")
                # process complete contexts
                while True:
                    if not in_ctx:
                        m = re.search(r"<[^>]*:context[^>]*\bid=\"([^\"]+)\"[^>]*>", text)
                        if not m:
                            # keep tail
                            buf = text[-2000:]
                            break
                        in_ctx = True
                        ctx_id = m.group(1)
                        text = text[m.end() :]
                        ctx_body = []
                    # find end
                    m2 = re.search(r"</[^>]*:context>", text)
                    if not m2:
                        ctx_body.append(text)
                        buf = ""
                        break
                    ctx_body.append(text[: m2.start()])
                    body = "".join(ctx_body)
                    if any(n.lower() in body.lower() for n in dim_needles):
                        ctx_415.add(ctx_id)
                        if len(sample_ctx) < 5:
                            dims = re.findall(
                                r"dimension=\"[^\"]*#?([A-Za-z0-9_]+)\"[^>]*>([^<]+)<",
                                body,
                            )
                            if not dims:
                                dims = re.findall(
                                    r"<[^>]*:explicitMember[^>]*dimension=\"([^\"]+)\"[^>]*>([^<]+)<",
                                    body,
                                )
                            sample_ctx.append((ctx_id, dims[:8]))
                    in_ctx = False
                    text = text[m2.end() :]
                    ctx_body = []

        print(f"contexts matching 415-like dims: {len(ctx_415)}")
        for s in sample_ctx:
            print(" sample ctx", s[0], "dims:", s[1])

        if not ctx_415:
            return

        # second pass: count facts in those contexts
        fact_re = re.compile(
            r"<([A-Za-z0-9_-]+):([A-Za-z0-9_]+)[^>]*contextRef=\"([^\"]+)\"[^>]*>([^<]*)</"
        )
        with zf.open(name) as f:
            leftover = ""
            n_facts = 0
            for chunk in iter(lambda: f.read(8 * 1024 * 1024), b""):
                text = leftover + chunk.decode("utf-8", errors="replace")
                if "\n" in text:
                    text, leftover = text.rsplit("\n", 1)
                else:
                    leftover = text
                    continue
                for m in fact_re.finditer(text):
                    if m.group(3) in ctx_415:
                        n_facts += 1
                        fact_by_concept[m.group(2)] += 1
            if leftover:
                for m in fact_re.finditer(leftover):
                    if m.group(3) in ctx_415:
                        n_facts += 1
                        fact_by_concept[m.group(2)] += 1
        print(f"facts in those contexts: {n_facts}")
        for c, n in fact_by_concept.most_common(25):
            print(f"  {n:6} {c}")


def compare_structure():
    print("\n=== Structure contrast (etalon pivot vs ours flat) ===")
    wb_e = load_workbook(ETALON, read_only=True, data_only=True)
    wb_o = load_workbook(OURS, read_only=True, data_only=True)
    # pick etalon sheet with data (ц_3)
    e_name = "0420415 Раздел 1 Операции с ц_3"
    o_name = "0420415 Раздел 1. Операции с _2"
    print("ETALON sheet:", e_name)
    ws = wb_e[e_name]
    for r in range(11, 18):
        row = list(next(ws.iter_rows(min_row=r, max_row=r, max_col=8, values_only=True)))
        print(" ", r, [str(c)[:35] if c is not None else "" for c in row])
    print("OURS sheet:", o_name)
    ws = wb_o[o_name]
    for r in range(5, 12):
        row = list(next(ws.iter_rows(min_row=r, max_row=r, max_col=10, values_only=True)))
        print(" ", r, [str(c)[:35] if c is not None else "" for c in row])
    wb_e.close()
    wb_o.close()


if __name__ == "__main__":
    count_sheet_rows(ETALON, "ETALON")
    count_sheet_rows(OURS, "OURS")
    taxonomy_415_tables()
    compare_structure()
    xbrl_415_by_typed_ids()
