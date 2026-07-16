#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare taxonomy29 (v1.3.29) vs taxonomy29_2 (v1.4.6) same ORTICON package."""

from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

DIR = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\IMDEV-9005 Форма 431\ОРТИКОН")
OLD = DIR / "XBRL_Orticon_taxonomy29.xlsx"
NEW = DIR / "XBRL_Orticon_taxonomy29_2.xlsx"


def analyze(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    gen = None
    sheets = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        toc = ""
        rows = 0
        cols = 0
        headers = []
        raw_okato = 0
        num_cells = 0
        text_num = 0
        sample_vals = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
            vals = [c for c in row if c is not None and str(c).strip() != ""]
            if i == 2 and vals:
                toc = str(vals[0])[:100]
            if i == 4 and vals and "Generator" in str(vals[0]):
                gen = str(vals[0])
            if i == 6:
                headers = [str(c).replace("\n", " ")[:40] if c else "" for c in row if c is not None]
                cols = len(headers)
        for row in ws.iter_rows(min_row=7, values_only=True):
            if not any(c is not None and str(c).strip() != "" for c in row):
                continue
            rows += 1
            for c in row:
                if c is None:
                    continue
                s = str(c)
                if s.startswith("OKATO") and any(ch.isalpha() for ch in s[5:]):
                    raw_okato += 1
                if isinstance(c, (int, float)):
                    num_cells += 1
                elif isinstance(c, str):
                    t = s.replace(".", "").replace(",", "").replace("-", "")
                    if t.isdigit() and (("." in s) or ("," in s) or s.isdigit()):
                        # skip pure id-like long codes with underscores elsewhere
                        if "_" in s:
                            continue
                        text_num += 1
            if rows <= 1 and sn != "TOC":
                sample_vals = [str(c)[:30] if c is not None else None for c in row[:8]]
        sheets[sn] = {
            "toc": toc,
            "rows": rows,
            "cols": cols,
            "headers": headers,
            "raw_okato": raw_okato,
            "num": num_cells,
            "text_num": text_num,
            "sample": sample_vals,
        }
    wb.close()
    return gen, sheets


def norm_toc(t):
    return " ".join(str(t).lower().split())


def main():
    print("OLD", OLD.exists(), OLD.name, "MB", round(OLD.stat().st_size / 1e6, 2))
    print("NEW", NEW.exists(), NEW.name, "MB", round(NEW.stat().st_size / 1e6, 2))
    gen_o, old = analyze(OLD)
    gen_n, new = analyze(NEW)
    print("OLD gen:", gen_o)
    print("NEW gen:", gen_n)
    print(
        "OLD sheets",
        len(old),
        "nonempty",
        sum(1 for s in old.values() if s["rows"] > 0),
        "rows_total",
        sum(s["rows"] for s in old.values()),
    )
    print(
        "NEW sheets",
        len(new),
        "nonempty",
        sum(1 for s in new.values() if s["rows"] > 0),
        "rows_total",
        sum(s["rows"] for s in new.values()),
    )

    # by sheet name
    all_names = sorted(set(old) | set(new))
    print("\n=== By sheet name: row deltas ===")
    only_old = []
    only_new = []
    changed = []
    same = []
    for sn in all_names:
        o = old.get(sn)
        n = new.get(sn)
        if o and not n:
            only_old.append((sn, o["rows"], o["toc"][:60]))
        elif n and not o:
            only_new.append((sn, n["rows"], n["toc"][:60]))
        else:
            dr = n["rows"] - o["rows"]
            if dr != 0 or o["cols"] != n["cols"]:
                changed.append((sn, o["rows"], n["rows"], o["cols"], n["cols"], o["toc"][:50]))
            else:
                same.append(sn)

    print("same name+rows+cols:", len(same))
    print("changed:", len(changed))
    for sn, ro, rn, co, cn, toc in changed:
        print(f"  {sn}: rows {ro}->{rn} (d={rn-ro}) cols {co}->{cn} | {toc}")
    if only_old:
        print("only OLD:")
        for x in only_old:
            print(" ", x)
    if only_new:
        print("only NEW:")
        for x in only_new:
            print(" ", x)

    # by TOC for nonempty
    print("\n=== Nonempty by TOC match ===")
    old_toc = defaultdict(list)
    new_toc = defaultdict(list)
    for sn, s in old.items():
        if s["rows"] > 0 and sn != "TOC":
            old_toc[norm_toc(s["toc"])].append((sn, s))
    for sn, s in new.items():
        if s["rows"] > 0 and sn != "TOC":
            new_toc[norm_toc(s["toc"])].append((sn, s))

    all_tocs = sorted(set(old_toc) | set(new_toc))
    for t in all_tocs:
        olist = old_toc.get(t, [])
        nlist = new_toc.get(t, [])
        orows = sum(x[1]["rows"] for x in olist)
        nrows = sum(x[1]["rows"] for x in nlist)
        ook = sum(x[1]["raw_okato"] for x in olist)
        nok = sum(x[1]["raw_okato"] for x in nlist)
        onum = sum(x[1]["num"] for x in olist)
        nnum = sum(x[1]["num"] for x in nlist)
        otx = sum(x[1]["text_num"] for x in olist)
        ntx = sum(x[1]["text_num"] for x in nlist)
        mark = "OK" if orows == nrows else "DIFF"
        short = t[:75] if t else "(empty toc)"
        print(
            f"[{mark}] rows {orows}->{nrows} okato_raw {ook}->{nok} "
            f"num {onum}->{nnum} textnum {otx}->{ntx} | {short}"
        )
        if orows != nrows or ook != nok:
            print("   OLD sheets:", [(x[0], x[1]["rows"]) for x in olist])
            print("   NEW sheets:", [(x[0], x[1]["rows"]) for x in nlist])

    # highlight key forms
    print("\n=== Key forms (409/431/459) ===")
    for key in ("0420409", "0420431", "0420459", "0420414", "0420415"):
        print(f"-- {key}")
        for label, store in (("OLD", old), ("NEW", new)):
            for sn, s in store.items():
                if key in sn or key in s["toc"]:
                    if s["rows"] == 0 and key not in sn:
                        continue
                    print(
                        f"  {label} {sn}: rows={s['rows']} cols={s['cols']} "
                        f"okato={s['raw_okato']} num={s['num']} textnum={s['text_num']}"
                    )

    # OKATO samples from 431 R4 if present
    print("\n=== Sample first data row for matching nonempty sheets (first 5 DIFF or 409) ===")
    shown = 0
    for sn in all_names:
        o = old.get(sn)
        n = new.get(sn)
        if not o or not n:
            continue
        if o["rows"] == 0 and n["rows"] == 0:
            continue
        if "0420409" in sn or o["rows"] != n["rows"] or shown < 3:
            print(f"{sn}:")
            print("  OLD sample", o["sample"])
            print("  NEW sample", n["sample"])
            print("  OLD hdr", o["headers"][:8])
            print("  NEW hdr", n["headers"][:8])
            shown += 1
            if shown >= 6:
                break

    print("\nDONE")


if __name__ == "__main__":
    main()
