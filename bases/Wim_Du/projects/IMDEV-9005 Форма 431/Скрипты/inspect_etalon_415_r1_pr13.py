#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Inspect CBR etalon sheet for 415 R1 PR1.3."""

from pathlib import Path
from openpyxl import load_workbook

CBR = Path(
    r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ"
    r"\401-414-415-0420437_июнь_2026_кроме 431.xlsx"
)


def main():
    wb = load_workbook(CBR, data_only=True)
    for sn in wb.sheetnames:
        if "0420415" not in sn:
            continue
        ws = wb[sn]
        print("====", sn, "max_row=", ws.max_row, "max_col=", ws.max_column)
        toc = ws.cell(2, 1).value
        uri = ws.cell(3, 1).value
        print(" TOC:", str(toc)[:100] if toc else None)
        print(" URI:", str(uri)[:120] if uri else None)
        if ws.max_row < 5:
            continue
        for r in range(1, 14):
            vals = []
            for c in range(1, min(30, ws.max_column) + 1):
                v = ws.cell(r, c).value
                if v is not None and str(v).strip() != "":
                    vals.append(f"{c}:{str(v).replace(chr(10), ' ')[:35]}")
            if vals:
                print(f" R{r}: {vals}")
        # find first data row with ID-like
        found = 0
        for r in range(10, min(ws.max_row, 50) + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, 6)]
            s = " ".join("" if v is None else str(v) for v in row_vals)
            if "USL" in s or "643_" in s or any(
                isinstance(v, str) and ("_" in v) and len(v) > 8 for v in row_vals
            ):
                print(" DATA", r, [str(v)[:40] if v is not None else None for v in row_vals])
                found += 1
                if found >= 3:
                    break
        # count nonempty rows
        n = 0
        for r in range(1, ws.max_row + 1):
            if any(ws.cell(r, c).value not in (None, "") for c in range(1, min(10, ws.max_column) + 1)):
                n += 1
        print(" nonempty rows (col1-10):", n)
    wb.close()


if __name__ == "__main__":
    main()
