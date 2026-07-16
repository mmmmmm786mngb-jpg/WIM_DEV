#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Verify XBRL_Orticon_taxonomy_июнь3.xlsx vs CBR etalon (415 pivot, numbers, OKATO)."""

from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(r"c:\1c\Cursor_1c\WIM_DEV\bases\Wim_Du\projects\9005_Ортикон\СУПЕРТЕСТ")
OURS = BASE / "XBRL_Orticon_taxonomy_июнь3.xlsx"
CBR = BASE / "401-414-415-0420437_июнь_2026_кроме 431.xlsx"


def sheet_by_toc_substr(wb, substr):
    for sn in wb.sheetnames:
        ws = wb[sn]
        toc = ws.cell(2, 1).value
        if toc and substr in str(toc):
            return sn, ws, str(toc)
    return None, None, None


def find_ours_415_13(wb):
    for sn in wb.sheetnames:
        if "0420415" not in sn:
            continue
        ws = wb[sn]
        toc = str(ws.cell(2, 1).value or "")
        gen = str(ws.cell(4, 1).value or "")
        if "1.3" in toc or "Подраздел 1.3" in toc:
            return sn, ws, toc, gen
    return None, None, None, None


def main():
    if not OURS.exists():
        print("MISSING", OURS)
        return

    print("=== FILE", OURS.name, "size_MB=", round(OURS.stat().st_size / 1e6, 2))
    wb = load_workbook(OURS, data_only=True)

    # generator from first 415 or any data sheet
    sn, ws, toc, gen = find_ours_415_13(wb)
    print("415_1.3 sheet:", sn)
    print(" TOC:", (toc or "")[:120])
    print(" Generator:", gen[:100] if gen else None)
    print(" URI row3:", str(ws.cell(3, 1).value or "")[:100] if ws else None)

    if ws is None:
        print("ERROR: 415 subsection 1.3 sheet not found")
        print("415 sheets:", [s for s in wb.sheetnames if "0420415" in s])
        wb.close()
        return

    # headers
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(6, c).value
        if v is not None:
            headers.append((c, str(v).replace("\n", " ")[:55]))
    print("headers count:", len(headers))
    for c, h in headers:
        print(f"  H{c}: {h}")

    # data rows
    n_rows = 0
    raw_okato = 0
    text_numbers = 0
    real_numbers = 0
    empty_measure_only = 0
    axis_like_in_cells = Counter()
    sample_rows = []
    ids_cb, ids_org, ids_st = set(), set(), set()

    # detect column roles by header text
    col_period = col_org = col_st = col_cb = None
    for c, h in headers:
        hl = h.lower()
        if "период" in hl:
            col_period = c
        elif "идентификатор организации" in hl:
            col_org = c
        elif "идентификатор строки" in hl:
            col_st = c
        elif "идентификатор ценной бумаги" in hl:
            col_cb = c

    print("id cols:", {"period": col_period, "org": col_org, "stroka": col_st, "cb": col_cb})

    for r in range(7, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        n_rows += 1
        if col_cb:
            v = ws.cell(r, col_cb).value
            if v is not None:
                ids_cb.add(str(v))
        if col_org:
            v = ws.cell(r, col_org).value
            if v is not None:
                ids_org.add(str(v))
        if col_st:
            v = ws.cell(r, col_st).value
            if v is not None:
                ids_st.add(str(v))

        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v)
            if s.startswith("OKATO") and not s[:5].replace("OKATO", "").isdigit():
                # raw member like OKATO74000...
                if "OKATO" in s and any(ch.isalpha() for ch in s[5:]):
                    raw_okato += 1
            if isinstance(v, (int, float)):
                real_numbers += 1
            elif isinstance(v, str) and s.replace(".", "", 1).replace(",", "", 1).replace("-", "", 1).isdigit():
                # look like number stored as text
                if "." in s or "," in s or s.isdigit():
                    # only count measure-ish columns (not IDs)
                    if col_cb and c == col_cb:
                        continue
                    if col_org and c == col_org:
                        continue
                    if col_st and c == col_st:
                        continue
                    if col_period and c == col_period:
                        continue
                    text_numbers += 1

        # detect leftover axis labels in data (old flat pattern)
        for c, h in headers:
            if c in (col_period, col_org, col_st, col_cb):
                continue
            v = ws.cell(r, c).value
            if isinstance(v, str) and (
                "на счетах доверительного" in v.lower()
                or "по типу" in v.lower()
                or v.endswith("Member")
            ):
                axis_like_in_cells[v[:50]] += 1

        if n_rows <= 3:
            sample_rows.append([ws.cell(r, c).value for c in range(1, min(12, ws.max_column) + 1)])

    print("data rows:", n_rows)
    print("unique cb/org/stroka:", len(ids_cb), len(ids_org), len(ids_st))
    print("numeric cells: number=", real_numbers, "text-looking=", text_numbers)
    print("raw OKATO-like cells:", raw_okato)
    print("axis-like string values in measure cols (top):", axis_like_in_cells.most_common(8))
    print("sample rows:")
    for i, row in enumerate(sample_rows, 1):
        print(" ", i, [str(x)[:35] if x is not None else None for x in row])

    # compare etalon
    print("\n=== ETALON ц_3 ===")
    wb2 = load_workbook(CBR, data_only=True)
    ws2 = wb2["0420415 Раздел 1 Операции с ц_3"]
    et_n = 0
    et_cb, et_org, et_st = set(), set(), set()
    for r in range(13, ws2.max_row + 1):
        a, b, c = ws2.cell(r, 1).value, ws2.cell(r, 2).value, ws2.cell(r, 3).value
        if a is None and b is None and c is None:
            continue
        et_n += 1
        if a is not None:
            et_cb.add(str(a))
        if b is not None:
            et_org.add(str(b))
        if c is not None:
            et_st.add(str(c))
    print("etalon rows:", et_n, "unique cb/org/stroka:", len(et_cb), len(et_org), len(et_st))
    print("row delta ours-etalon:", n_rows - et_n)
    print("cb overlap:", len(ids_cb & et_cb), "/", len(et_cb))
    print("org overlap:", len(ids_org & et_org), "/", len(et_org))
    print("stroka overlap:", len(ids_st & et_st), "/", len(et_st))

    # spot-check one common security totals
    # pick first shared CB
    common = sorted(ids_cb & et_cb)
    if common:
        cb = common[0]
        print("\nspot CB:", cb)
        # ours rows
        for r in range(7, ws.max_row + 1):
            if col_cb and str(ws.cell(r, col_cb).value) == cb:
                print(
                    " ours r",
                    r,
                    "org=",
                    ws.cell(r, col_org).value if col_org else None,
                    "st=",
                    ws.cell(r, col_st).value if col_st else None,
                )
                measures = []
                for c, h in headers:
                    if c in (col_period, col_org, col_st, col_cb):
                        continue
                    v = ws.cell(r, c).value
                    if v is not None and str(v).strip() != "":
                        measures.append((h[:40], v, type(v).__name__))
                print("  measures:", measures[:12])
                break
        for r in range(13, ws2.max_row + 1):
            if str(ws2.cell(r, 1).value) == cb:
                print(
                    " etalon r",
                    r,
                    "org=",
                    ws2.cell(r, 2).value,
                    "st=",
                    ws2.cell(r, 3).value,
                    "c4=",
                    ws2.cell(r, 4).value,
                    "c5=",
                    ws2.cell(r, 5).value,
                    "c13=",
                    ws2.cell(r, 13).value,
                    "c16=",
                    str(ws2.cell(r, 16).value)[:40] if ws2.cell(r, 16).value else None,
                    "c28=",
                    ws2.cell(r, 28).value,
                )
                break

    # other 415 sheets row counts
    print("\n=== all 0420415 sheets (ours) ===")
    for sn2 in wb.sheetnames:
        if "0420415" not in sn2:
            continue
        w = wb[sn2]
        toc2 = str(w.cell(2, 1).value or "")[:80]
        nr = 0
        for r in range(7, w.max_row + 1):
            if any(w.cell(r, c).value not in (None, "") for c in range(1, min(8, w.max_column) + 1)):
                nr += 1
        print(f" {sn2}: rows={nr} toc={toc2}")

    wb.close()
    wb2.close()
    print("\nDONE")


if __name__ == "__main__":
    main()
