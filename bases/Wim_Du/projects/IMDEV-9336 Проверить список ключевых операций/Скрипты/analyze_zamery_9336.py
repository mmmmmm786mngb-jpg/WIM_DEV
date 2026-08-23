#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Analiz zamery Excel dlya IMDEV-9336."""

import json
import re
import statistics
from collections import defaultdict
from pathlib import Path

import openpyxl

THRESHOLD_SEC = 0.5 * 3600  # 1800 sec - plan from IMDEV-8927

PROJECT = Path(__file__).resolve().parents[1]
FILES = {
    "broker": PROJECT / "Замеры_СверкаСделокСОтчетомБрокера.xlsx",
    "cb": PROJECT / "Замеры_СверкаЦБ.xlsx",
}


def parse_comment(raw):
    if not raw:
        return {}
    text = str(raw)
    match = re.search(r'"ДопИнф":"([^"]*)"', text)
    if not match:
        return {}
    info = match.group(1).replace("\xa0", " ")
    result = {}
    for part in info.split(";"):
        part = part.strip()
        if "=" in part:
            key, value = part.split("=", 1)
            result[key.strip()] = value.strip()
    return result


def load_rows(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    rows = []
    for row_idx in range(2, sheet.max_row + 1):
        rows.append(
            {headers[col - 1]: sheet.cell(row_idx, col).value for col in range(1, sheet.max_column + 1)}
        )
    return rows


def calc_stats(values):
    if not values:
        return {}
    values = sorted(values)
    count = len(values)

    def percentile(p):
        index = int(round((count - 1) * p))
        return values[index]

    return {
        "n": count,
        "min": round(min(values), 2),
        "max": round(max(values), 2),
        "avg": round(sum(values) / count, 2),
        "median": round(statistics.median(values), 2),
        "p95": round(percentile(0.95), 2),
    }


def normalize_op(name):
    return re.sub(r"\s+", " ", str(name or "").lower().strip())


def analyze_dataset(label, rows):
    by_op = defaultdict(list)
    dates = []
    fail_count = 0
    warn_count = 0
    deals = []

    for row in rows:
        op = normalize_op(row.get("Ключевая операция"))
        duration = float(row.get("Время выполнения в секундах") or 0)
        by_op[op].append(duration)

        dt = row.get("Дата записи локальная") or row.get("Дата записи")
        if dt:
            dates.append(str(dt))

        comment = parse_comment(row.get("Комментарий"))
        if comment.get("FAIL") == "Да":
            fail_count += 1
        if comment.get("WARN") == "Да":
            warn_count += 1
        if "СделокИзSQL" in comment:
            deals.append(int(re.sub(r"\s", "", comment["СделокИзSQL"])))

    all_durations = [float(row.get("Время выполнения в секундах") or 0) for row in rows]
    slowest = sorted(rows, key=lambda item: float(item.get("Время выполнения в секундах") or 0), reverse=True)[:5]

    result = {
        "label": label,
        "records": len(rows),
        "period_from": min(dates) if dates else "",
        "period_to": max(dates) if dates else "",
        "fail_rows": fail_count,
        "warn_rows": warn_count,
        "deals_sql": {
            "min": min(deals) if deals else 0,
            "max": max(deals) if deals else 0,
            "avg": round(sum(deals) / len(deals), 0) if deals else 0,
        },
        "all_stats": calc_stats(all_durations),
        "over_threshold": sum(1 for value in all_durations if value > THRESHOLD_SEC),
        "by_operation": {},
        "slowest": [],
    }

    for op, durations in sorted(by_op.items(), key=lambda item: -len(item[1])):
        result["by_operation"][op] = {
            "stats": calc_stats(durations),
            "over_threshold": sum(1 for value in durations if value > THRESHOLD_SEC),
        }

    for row in slowest:
        comment = parse_comment(row.get("Комментарий"))
        result["slowest"].append(
            {
                "date": str(row.get("Дата записи локальная") or row.get("Дата записи") or ""),
                "seconds": round(float(row.get("Время выполнения в секундах") or 0), 1),
                "operation": normalize_op(row.get("Ключевая операция")),
                "deals_sql": comment.get("СделокИзSQL", ""),
            }
        )

    return result


def main():
    output = {}
    for key, path in FILES.items():
        rows = load_rows(path)
        output[key] = analyze_dataset(key, rows)

    report_path = PROJECT / "Тестирование" / "reports" / "zamery_9336_analysis.json"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as handle:
        json.dump(output, handle, ensure_ascii=False, indent=2)

    print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
