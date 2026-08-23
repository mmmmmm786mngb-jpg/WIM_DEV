#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Analiz zamery Excel dlya IMDEV-9336 s uchetom vesa."""

import json
import re
import statistics
from collections import defaultdict
from pathlib import Path

import openpyxl

THRESHOLD_SEC = 0.5 * 3600
PROJECT = Path(__file__).resolve().parents[1]
FILES = {
    "broker": PROJECT / "Замеры_СверкаСделокСОтчетомБрокера.xlsx",
    "cb": PROJECT / "Замеры_СверкаЦБ.xlsx",
}


def parse_comment(raw):
    if not raw:
        return {}
    text = str(raw)
    match = re.search(r'"ДопИнф":"([^"]*)"', text)
    if not match:
        return {}
    info = match.group(1).replace("\xa0", " ")
    result = {}
    for part in info.split(";"):
        part = part.strip()
        if "=" in part:
            key, value = part.split("=", 1)
            result[key.strip()] = value.strip()
    return result


def load_rows(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    rows = []
    for row_idx in range(2, sheet.max_row + 1):
        rows.append(
            {headers[col - 1]: sheet.cell(row_idx, col).value for col in range(1, sheet.max_column + 1)}
        )
    return rows


def normalize_op(name):
    return re.sub(r"\s+", " ", str(name or "").lower().strip())


def calc_stats(values):
    if not values:
        return {}
    values = sorted(values)
    count = len(values)

    def percentile(p):
        index = int(round((count - 1) * p))
        return values[index]

    return {
        "n": count,
        "min": round(min(values), 2),
        "max": round(max(values), 2),
        "avg": round(sum(values) / count, 2),
        "median": round(statistics.median(values), 2),
        "p95": round(percentile(0.95), 2),
    }


def analyze(label, rows):
    by_op = defaultdict(list)
    for row in rows:
        op = normalize_op(row.get("Ключевая операция"))
        duration = float(row.get("Время выполнения в секундах") or 0)
        weight = float(row.get("Вес замера") or 0)
        if weight <= 0:
            weight = 1.0
        comment = parse_comment(row.get("Комментарий"))
        deals_sql = 0
        if "СделокИзSQL" in comment:
            deals_sql = int(re.sub(r"\s", "", comment["СделокИзSQL"]) or "0")
        sec_per_unit = duration / weight
        by_op[op].append(
            {
                "duration": duration,
                "weight": weight,
                "sec_per_unit": sec_per_unit,
                "units_per_sec": weight / duration if duration > 0 else 0,
                "deals_sql": deals_sql,
                "date": str(row.get("Дата записи локальная") or row.get("Дата записи") or ""),
                "fail": comment.get("FAIL") == "Да",
                "warn": comment.get("WARN") == "Да",
            }
        )

    result = {"label": label, "operations": {}}
    for op, items in sorted(by_op.items(), key=lambda item: -len(item[1])):
        durations = [item["duration"] for item in items]
        weights = [item["weight"] for item in items]
        sec_per = [item["sec_per_unit"] for item in items]
        ups = [item["units_per_sec"] for item in items]
        deals = [item["deals_sql"] for item in items if item["deals_sql"] > 0]
        total_weight = sum(weights)
        total_duration = sum(durations)
        weighted_avg_sec_per = total_duration / total_weight if total_weight else 0
        result["operations"][op] = {
            "n": len(items),
            "duration": calc_stats(durations),
            "weight": calc_stats(weights),
            "sec_per_unit": calc_stats(sec_per),
            "units_per_sec": calc_stats(ups),
            "weighted_avg_sec_per_unit": round(weighted_avg_sec_per, 4),
            "weighted_avg_units_per_sec": round(total_weight / total_duration, 2) if total_duration else 0,
            "deals_sql": calc_stats(deals) if deals else {},
            "fail": sum(1 for item in items if item["fail"]),
            "warn": sum(1 for item in items if item["warn"]),
            "over_threshold": sum(1 for item in items if item["duration"] > THRESHOLD_SEC),
            "slowest": sorted(items, key=lambda item: item["duration"], reverse=True)[:3],
            "worst_per_unit": sorted(items, key=lambda item: item["sec_per_unit"], reverse=True)[:3],
        }
    return result


def main():
    output = {}
    for key, path in FILES.items():
        output[key] = analyze(key, load_rows(path))
    report_path = PROJECT / "Тестирование" / "reports" / "zamery_9336_weight_analysis.json"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as handle:
        json.dump(output, handle, ensure_ascii=False, indent=2, default=str)
    print(json.dumps(output, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
