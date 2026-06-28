#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Compare Regress2105 XLSX exports (bylo vs stalo) by business multiset.
Ignores row order and UUID/link columns that change between runs.
"""

import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

REG_DIR = Path(__file__).resolve().parent / "Регресс2105"

# --- normalize helpers ---

DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")


def norm_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    text = str(value).strip()
    if " 0:00:00" in text:
        text = text.split(" 0:00:00")[0].strip()
    return text


def norm_date(value) -> str:
    text = norm_str(value)
    match = DATE_RE.search(text)
    if match:
        return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
    return text


def norm_amount(value) -> str:
    text = norm_str(value).replace(" ", "").replace("\u00a0", "")
    if not text:
        return ""
    return text.replace(".", ",") if "," not in text and "." in text else text


def norm_account(value) -> str:
    return norm_str(value).replace(" ", "")


def read_sheet(path: Path) -> tuple[list[str], list[dict]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [norm_str(c) for c in next(rows_iter)]
    data = []
    for row in rows_iter:
        if not any(c is not None and norm_str(c) for c in row):
            continue
        item = {header[i]: row[i] if i < len(row) else None for i in range(len(header))}
        data.append(item)
    wb.close()
    return header, data


def compare_multiset(cnt_bylo: Counter, cnt_stalo: Counter) -> dict:
    diff = (cnt_bylo - cnt_stalo) + (cnt_stalo - cnt_bylo)
    return {
        "identical": len(diff) == 0,
        "diff_types": len(diff),
        "only_bylo": sum((cnt_bylo - cnt_stalo).values()),
        "only_stalo": sum((cnt_stalo - cnt_bylo).values()),
        "shared_types": len(cnt_bylo & cnt_stalo),
        "top_diffs": [
            {"delta": delta, "key": key}
            for key, delta in sorted(diff.items(), key=lambda x: -abs(x[1]))[:15]
        ],
    }


# --- Vypiski ---

VYPISKI_BUSINESS = (
    "Дата создания",
    "Банковский счет",
    "Номер счета",
    "Начальный остаток",
    "Всего поступило",
    "Всего списано",
    "Конечный остаток",
    "Договор ДУ",
    "Загружать",
    "Загружена",
)


def vypiski_key(row: dict) -> tuple:
    return (
        norm_date(row.get("Дата создания")),
        norm_str(row.get("Банковский счет")),
        norm_account(row.get("Номер счета")),
        norm_amount(row.get("Начальный остаток")),
        norm_amount(row.get("Всего поступило")),
        norm_amount(row.get("Всего списано")),
        norm_amount(row.get("Конечный остаток")),
        norm_str(row.get("Договор ДУ")) or "(empty)",
        norm_str(row.get("Загружать")),
        norm_str(row.get("Загружена")),
    )


def compare_vypiski_xlsx(path_bylo: Path, path_stalo: Path) -> dict:
    _, rows_b = read_sheet(path_bylo)
    _, rows_s = read_sheet(path_stalo)
    biz_b = Counter(vypiski_key(r) for r in rows_b)
    biz_s = Counter(vypiski_key(r) for r in rows_s)
    return {
        "rows": (len(rows_b), len(rows_s)),
        "business": compare_multiset(biz_b, biz_s),
    }


# --- PP ---

PP_CORE = (
    "Дата",
    "Номер",
    "Сумма",
    "Плательщик счет",
    "Получатель счет",
    "Назначение платежа",
)

PP_META = ("Ключ выписки", "Ключ документа", "Загружен", "Лог", "Документ", "Документ ссылка1", "Документ ссылка2")


def pp_core_key(row: dict) -> tuple:
    return (
        norm_date(row.get("Дата")),
        norm_str(row.get("Номер")),
        norm_amount(row.get("Сумма")),
        norm_account(row.get("Плательщик счет")),
        norm_account(row.get("Получатель счет")),
        norm_str(row.get("Назначение платежа")),
    )


def pp_dn_key(row: dict) -> tuple:
    return (norm_date(row.get("Дата")), norm_str(row.get("Номер")))


def compare_pp_xlsx(path_bylo: Path, path_stalo: Path) -> dict:
    _, rows_b = read_sheet(path_bylo)
    _, rows_s = read_sheet(path_stalo)

    pp_b = [r for r in rows_b if norm_str(r.get("Операция")) == "Платежное поручение"]
    pp_s = [r for r in rows_s if norm_str(r.get("Операция")) == "Платежное поручение"]

    core_b = Counter(pp_core_key(r) for r in pp_b)
    core_s = Counter(pp_core_key(r) for r in pp_s)

    # map by (date, number) for field-level diff on shared payments
    map_b = {pp_dn_key(r): r for r in pp_b}
    map_s = {pp_dn_key(r): r for r in pp_s}
    keys_b, keys_s = set(map_b), set(map_s)

    loaded_diffs = []
    meta_diffs = []
    for key in sorted(keys_b & keys_s):
        rb, rs = map_b[key], map_s[key]
        if norm_str(rb.get("Загружен")) != norm_str(rs.get("Загружен")):
            loaded_diffs.append({"key": list(key), "bylo": norm_str(rb.get("Загружен")), "stalo": norm_str(rs.get("Загружен"))})
        for col in PP_META:
            if norm_str(rb.get(col)) != norm_str(rs.get(col)):
                meta_diffs.append({"key": list(key), "column": col, "bylo": norm_str(rb.get(col))[:80], "stalo": norm_str(rs.get(col))[:80]})
                break  # one sample per payment for meta

    probe = ("2026-05-21", "485038")
    pb = map_b.get(probe) or next((r for r in pp_b if norm_str(r.get("Номер")) == "485038"), None)
    ps = map_s.get(probe) or next((r for r in pp_s if norm_str(r.get("Номер")) == "485038"), None)

    return {
        "all_rows": (len(rows_b), len(rows_s)),
        "pp_rows": (len(pp_b), len(pp_s)),
        "only_bylo_dn": sorted(keys_b - keys_s)[:20],
        "only_stalo_dn": sorted(keys_s - keys_b)[:20],
        "only_bylo_count": len(keys_b - keys_s),
        "only_stalo_count": len(keys_s - keys_b),
        "core_multiset": compare_multiset(core_b, core_s),
        "loaded_diff_count": len(loaded_diffs),
        "loaded_diff_samples": loaded_diffs[:10],
        "meta_diff_count": len(meta_diffs),
        "meta_diff_samples": meta_diffs[:10],
        "probe_485038": {
            "in_bylo": pb is not None,
            "in_stalo": ps is not None,
            "core_equal": pp_core_key(pb) == pp_core_key(ps) if pb and ps else None,
            "bylo": {c: norm_str((pb or {}).get(c))[:60] for c in PP_CORE} if pb else None,
            "stalo": {c: norm_str((ps or {}).get(c))[:60] for c in PP_CORE} if ps else None,
        },
    }


def safe_print(text: str) -> None:
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", errors="replace").decode("ascii"))


def main() -> int:
    report = {
        "format": "xlsx",
        "folder": str(REG_DIR),
        "vypiski": compare_vypiski_xlsx(
            REG_DIR / "2105_2105_ВЫПИСКИ_было.xlsx",
            REG_DIR / "2105_2105_ВЫПИСКИ_стало.xlsx",
        ),
        "pp": compare_pp_xlsx(
            REG_DIR / "2105_2105_ПП_было.xlsx",
            REG_DIR / "2105_2105_ПП_стало.xlsx",
        ),
    }
    out = REG_DIR / "2105_xlsx_compare_report.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    safe_print("=== REGRESS2105 XLSX COMPARE ===")
    safe_print("")
    v = report["vypiski"]
    safe_print("1. VYPISKI")
    safe_print(f"   Rows: {v['rows'][0]} / {v['rows'][1]}")
    b = v["business"]
    safe_print(f"   Business multiset identical: {b['identical']}")
    safe_print(f"   diff_types={b['diff_types']} only_b={b['only_bylo']} only_s={b['only_stalo']}")
    if b["top_diffs"]:
        for d in b["top_diffs"][:5]:
            safe_print(f"     delta={d['delta']:+d} {d['key']}")
    safe_print("")

    p = report["pp"]
    safe_print("2. PP (Платежное поручение)")
    safe_print(f"   All rows: {p['all_rows'][0]} / {p['all_rows'][1]}")
    safe_print(f"   PP rows: {p['pp_rows'][0]} / {p['pp_rows'][1]}")
    c = p["core_multiset"]
    safe_print(f"   Core multiset identical: {c['identical']}")
    safe_print(f"   only BYLO (Date,Number): {p['only_bylo_count']}")
    safe_print(f"   only STALO (Date,Number): {p['only_stalo_count']}")
    safe_print(f"   Zагружен diffs (same payment): {p['loaded_diff_count']}")
    safe_print(f"   Meta/UUID diffs (same payment): {p['meta_diff_count']}")
    safe_print(f"   485038: {p['probe_485038']}")
    if p["only_bylo_dn"]:
        safe_print(f"   sample only BYLO: {p['only_bylo_dn'][:5]}")
    if p["only_stalo_dn"]:
        safe_print(f"   sample only STALO: {p['only_stalo_dn'][:5]}")
    safe_print("")
    safe_print(f"Saved: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
