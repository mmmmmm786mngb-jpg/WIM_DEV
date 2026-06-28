#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Deep analysis of vypiski diffs between 1805_3105 bylo/stalo XLSX:
- 17 empty ERS on 2026-05-21
- Flag changes for DU 9957 and DU 10076
"""

import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

REG_DIR = Path(__file__).resolve().parent / "Регресс"
DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")

FLAG_COLS = ("Загружать", "Загружена")


def norm_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    text = str(value).strip()
    if " 0:00:00" in text:
        text = text.split(" 0:00:00")[0].strip()
    return text


def norm_date(value) -> str:
    text = norm_str(value)
    match = DATE_RE.search(text)
    if match:
        return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
    return text


def norm_amount(value) -> str:
    text = norm_str(value).replace(" ", "").replace("\u00a0", "")
    if not text:
        return ""
    return text.replace(".", ",") if "," not in text and "." in text else text


def norm_account(value) -> str:
    return norm_str(value).replace(" ", "")


def read_sheet(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [norm_str(c) for c in next(rows_iter)]
    data = []
    for row in rows_iter:
        if not any(c is not None and norm_str(c) for c in row):
            continue
        data.append({header[i]: row[i] if i < len(row) else None for i in range(len(header))})
    wb.close()
    return data


def vypiski_key(row: dict) -> tuple:
    return (
        norm_date(row.get("Дата создания")),
        norm_str(row.get("Банковский счет")),
        norm_account(row.get("Номер счета")),
        norm_amount(row.get("Начальный остаток")),
        norm_amount(row.get("Всего поступило")),
        norm_amount(row.get("Всего списано")),
        norm_amount(row.get("Конечный остаток")),
        norm_str(row.get("Договор ДУ")) or "(empty)",
        norm_str(row.get("Загружать")),
        norm_str(row.get("Загружена")),
    )


def row_summary(row: dict) -> dict:
    return {
        "data": norm_date(row.get("Дата создания")),
        "bank": norm_str(row.get("Банковский счет")),
        "account": norm_account(row.get("Номер счета")),
        "dogovor": norm_str(row.get("Договор ДУ")) or "(empty)",
        "nach": norm_amount(row.get("Начальный остаток")),
        "post": norm_amount(row.get("Всего поступило")),
        "spis": norm_amount(row.get("Всего списано")),
        "kon": norm_amount(row.get("Конечный остаток")),
        "zagruzhat": norm_str(row.get("Загружать")),
        "zagruzhena": norm_str(row.get("Загружена")),
    }


def analyze_empty_ers_2105(rows_bylo: list[dict], rows_stalo: list[dict]) -> dict:
    target_date = "2026-05-21"
    ers_bank = "р/с_ВТБ_ДС_RUR_ЕРС"
    ers_account = "40701810000030000413"

    def is_target(r: dict) -> bool:
        return (
            norm_date(r.get("Дата создания")) == target_date
            and norm_str(r.get("Банковский счет")) == ers_bank
            and norm_account(r.get("Номер счета")) == ers_account
            and not norm_str(r.get("Договор ДУ"))
        )

    bylo_ers = [r for r in rows_bylo if is_target(r)]
    stalo_ers = [r for r in rows_stalo if is_target(r)]

    bylo_keys = Counter(vypiski_key(r) for r in bylo_ers)
    stalo_keys = Counter(vypiski_key(r) for r in stalo_ers)
    only_bylo_keys = bylo_keys - stalo_keys

    only_bylo_rows = []
    for r in bylo_ers:
        if only_bylo_keys[vypiski_key(r)] > 0:
            only_bylo_rows.append(row_summary(r))

    # same date/account but WITH dogovor in stalo
    stalo_with_du = [
        row_summary(r)
        for r in rows_stalo
        if norm_date(r.get("Дата создания")) == target_date
        and norm_str(r.get("Банковский счет")) == ers_bank
        and norm_account(r.get("Номер счета")) == ers_account
        and norm_str(r.get("Договор ДУ"))
    ]

    bylo_with_du_same = [
        row_summary(r)
        for r in rows_bylo
        if norm_date(r.get("Дата создания")) == target_date
        and norm_str(r.get("Банковский счет")) == ers_bank
        and norm_account(r.get("Номер счета")) == ers_account
        and norm_str(r.get("Договор ДУ"))
    ]

    # unique amounts in only-bylo empty ERS
    amount_groups = Counter(
        (s["nach"], s["post"], s["spis"], s["kon"], s["zagruzhat"], s["zagruzhena"])
        for s in only_bylo_rows
    )

    return {
        "date": target_date,
        "bank": ers_bank,
        "account": ers_account,
        "counts": {
            "bylo_empty_ers": len(bylo_ers),
            "stalo_empty_ers": len(stalo_ers),
            "only_bylo_empty": len(only_bylo_rows),
            "bylo_with_dogovor": len(bylo_with_du_same),
            "stalo_with_dogovor": len(stalo_with_du),
        },
        "only_bylo_amount_groups": [
            {"count": cnt, "amounts": list(key)}
            for key, cnt in amount_groups.most_common()
        ],
        "only_bylo_samples": only_bylo_rows[:5],
        "stalo_with_dogovor_samples": stalo_with_du[:10],
        "bylo_with_dogovor_dogovors": sorted({s["dogovor"] for s in bylo_with_du_same}),
        "stalo_with_dogovor_dogovors": sorted({s["dogovor"] for s in stalo_with_du}),
    }


def analyze_dogovor_flags(
    rows_bylo: list[dict], rows_stalo: list[dict], dogovor_substr: str, date_filter: str | None = None
) -> dict:
    def matches(r: dict) -> bool:
        if dogovor_substr not in norm_str(r.get("Договор ДУ")):
            return False
        if date_filter and norm_date(r.get("Дата создания")) != date_filter:
            return False
        return True

    bylo = [r for r in rows_bylo if matches(r)]
    stalo = [r for r in rows_stalo if matches(r)]

    bylo_flags = Counter(
        (norm_date(r.get("Дата создания")), norm_str(r.get("Загружать")), norm_str(r.get("Загружена")))
        for r in bylo
    )
    stalo_flags = Counter(
        (norm_date(r.get("Дата создания")), norm_str(r.get("Загружать")), norm_str(r.get("Загружена")))
        for r in stalo
    )

    bylo_biz = Counter(vypiski_key(r) for r in bylo)
    stalo_biz = Counter(vypiski_key(r) for r in stalo)

    only_bylo = bylo_biz - stalo_biz
    only_stalo = stalo_biz - bylo_biz
    shared = bylo_biz & stalo_biz

    def key_to_row(key: tuple) -> dict:
        return {
            "data": key[0],
            "bank": key[1],
            "account": key[2],
            "nach": key[3],
            "post": key[4],
            "spis": key[5],
            "kon": key[6],
            "dogovor": key[7],
            "zagruzhat": key[8],
            "zagruzhena": key[9],
        }

    return {
        "dogovor_filter": dogovor_substr,
        "date_filter": date_filter,
        "counts": {"bylo_rows": len(bylo), "stalo_rows": len(stalo)},
        "flag_multiset_bylo": {f"{d}|{z}|{zh}": c for (d, z, zh), c in bylo_flags.items()},
        "flag_multiset_stalo": {f"{d}|{z}|{zh}": c for (d, z, zh), c in stalo_flags.items()},
        "only_bylo_business": [
            {"delta": only_bylo[k], "row": key_to_row(k)} for k in sorted(only_bylo, key=lambda x: x[0])
        ],
        "only_stalo_business": [
            {"delta": only_stalo[k], "row": key_to_row(k)} for k in sorted(only_stalo, key=lambda x: x[0])
        ],
        "shared_business_types": len(shared),
        "amounts_identical_on_shared": all(
            k[:7] in {s[:7] for s in shared} for k in only_bylo
        ) and all(k[:7] in {s[:7] for s in shared} for k in only_stalo),
    }


def safe_print(text: str) -> None:
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", errors="replace").decode("ascii"))


def main() -> int:
    path_bylo = REG_DIR / "1805_3105_ВЫПИСКИ_было.xlsx"
    path_stalo = REG_DIR / "1805_3105_ВЫПИСКИ_стало.xlsx"

    rows_bylo = read_sheet(path_bylo)
    rows_stalo = read_sheet(path_stalo)

    report = {
        "empty_ers_2105": analyze_empty_ers_2105(rows_bylo, rows_stalo),
        "du_9957": analyze_dogovor_flags(rows_bylo, rows_stalo, "9957", "2026-05-18"),
        "du_10076": analyze_dogovor_flags(rows_bylo, rows_stalo, "10076", "2026-05-28"),
    }

    out = REG_DIR / "1805_3105_vypiski_diff_analysis.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    e = report["empty_ers_2105"]
    safe_print("=== EMPTY ERS 2026-05-21 ===")
    safe_print(f"bylo empty ERS: {e['counts']['bylo_empty_ers']}")
    safe_print(f"stalo empty ERS: {e['counts']['stalo_empty_ers']}")
    safe_print(f"only in BYLO (empty): {e['counts']['only_bylo_empty']}")
    safe_print(f"bylo with dogovor same date/account: {e['counts']['bylo_with_dogovor']}")
    safe_print(f"stalo with dogovor same date/account: {e['counts']['stalo_with_dogovor']}")
    safe_print("amount groups only BYLO:")
    for g in e["only_bylo_amount_groups"]:
        safe_print(f"  x{g['count']}: nach={g['amounts'][0]} post={g['amounts'][1]} spis={g['amounts'][2]} kon={g['amounts'][3]} z={g['amounts'][4]} zh={g['amounts'][5]}")
    safe_print("")

    for name, key in [("DU 9957 18.05", "du_9957"), ("DU 10076 28.05", "du_10076")]:
        d = report[key]
        safe_print(f"=== {name} ===")
        safe_print(f"rows: {d['counts']['bylo_rows']} / {d['counts']['stalo_rows']}")
        safe_print(f"flags BYLO: {d['flag_multiset_bylo']}")
        safe_print(f"flags STALO: {d['flag_multiset_stalo']}")
        safe_print("only BYLO business:")
        for item in d["only_bylo_business"]:
            r = item["row"]
            safe_print(f"  +{item['delta']} z={r['zagruzhat']} zh={r['zagruzhena']} amounts={r['nach']}/{r['post']}/{r['spis']}/{r['kon']}")
        safe_print("only STALO business:")
        for item in d["only_stalo_business"]:
            r = item["row"]
            safe_print(f"  +{item['delta']} z={r['zagruzhat']} zh={r['zagruzhena']} amounts={r['nach']}/{r['post']}/{r['spis']}/{r['kon']}")
        safe_print("")

    safe_print(f"Saved: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
