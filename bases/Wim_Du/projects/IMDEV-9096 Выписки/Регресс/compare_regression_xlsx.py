#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Compare regression XLSX exports (bylo vs stalo) by business multiset.
Ignores row order and UUID/link columns that change between runs.

Usage:
  python compare_regression_xlsx.py <folder> <prefix>

Example:
  python compare_regression_xlsx.py "Регресс2105" "2105_2105"
  python compare_regression_xlsx.py "Регресс" "1805_3105"
"""

import argparse
import json
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent

DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")

PP_CORE = (
    "Дата",
    "Номер",
    "Сумма",
    "Плательщик счет",
    "Получатель счет",
    "Назначение платежа",
)

PP_META = (
    "Ключ выписки",
    "Ключ документа",
    "Загружен",
    "Лог",
    "Документ",
    "Документ ссылка1",
    "Документ ссылка2",
)


def norm_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    text = str(value).strip()
    if " 0:00:00" in text:
        text = text.split(" 0:00:00")[0].strip()
    return text


def norm_date(value) -> str:
    text = norm_str(value)
    match = DATE_RE.search(text)
    if match:
        return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
    return text


def norm_amount(value) -> str:
    text = norm_str(value).replace(" ", "").replace("\u00a0", "")
    if not text:
        return ""
    return text.replace(".", ",") if "," not in text and "." in text else text


def norm_account(value) -> str:
    return norm_str(value).replace(" ", "")


def read_sheet(path: Path) -> tuple[list[str], list[dict]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [norm_str(c) for c in next(rows_iter)]
    data = []
    for row in rows_iter:
        if not any(c is not None and norm_str(c) for c in row):
            continue
        item = {header[i]: row[i] if i < len(row) else None for i in range(len(header))}
        data.append(item)
    wb.close()
    return header, data


def compare_multiset(cnt_bylo: Counter, cnt_stalo: Counter) -> dict:
    diff = (cnt_bylo - cnt_stalo) + (cnt_stalo - cnt_bylo)
    return {
        "identical": len(diff) == 0,
        "diff_types": len(diff),
        "only_bylo": sum((cnt_bylo - cnt_stalo).values()),
        "only_stalo": sum((cnt_stalo - cnt_bylo).values()),
        "shared_types": len(cnt_bylo & cnt_stalo),
        "top_diffs": [
            {"delta": delta, "key": key}
            for key, delta in sorted(diff.items(), key=lambda x: -abs(x[1]))[:20]
        ],
    }


def vypiski_key(row: dict) -> tuple:
    return (
        norm_date(row.get("Дата создания")),
        norm_str(row.get("Банковский счет")),
        norm_account(row.get("Номер счета")),
        norm_amount(row.get("Начальный остаток")),
        norm_amount(row.get("Всего поступило")),
        norm_amount(row.get("Всего списано")),
        norm_amount(row.get("Конечный остаток")),
        norm_str(row.get("Договор ДУ")) or "(empty)",
        norm_str(row.get("Загружать")),
        norm_str(row.get("Загружена")),
    )


def vypiski_alt_key(row: dict) -> tuple:
    return (
        norm_date(row.get("Дата создания")),
        norm_str(row.get("Банковский счет")),
        norm_str(row.get("Договор ДУ")) or "(empty)",
    )


def compare_vypiski_xlsx(path_bylo: Path, path_stalo: Path) -> dict:
    _, rows_b = read_sheet(path_bylo)
    _, rows_s = read_sheet(path_stalo)
    biz_b = Counter(vypiski_key(r) for r in rows_b)
    biz_s = Counter(vypiski_key(r) for r in rows_s)
    alt_b = Counter(vypiski_alt_key(r) for r in rows_b)
    alt_s = Counter(vypiski_alt_key(r) for r in rows_s)
    return {
        "rows": (len(rows_b), len(rows_s)),
        "row_delta": len(rows_b) - len(rows_s),
        "business": compare_multiset(biz_b, biz_s),
        "alt_dogovor": compare_multiset(alt_b, alt_s),
    }


def pp_core_key(row: dict) -> tuple:
    return (
        norm_date(row.get("Дата")),
        norm_str(row.get("Номер")),
        norm_amount(row.get("Сумма")),
        norm_account(row.get("Плательщик счет")),
        norm_account(row.get("Получатель счет")),
        norm_str(row.get("Назначение платежа")),
    )


def pp_dn_key(row: dict) -> tuple:
    return (norm_date(row.get("Дата")), norm_str(row.get("Номер")))


def compare_pp_xlsx(path_bylo: Path, path_stalo: Path) -> dict:
    _, rows_b = read_sheet(path_bylo)
    _, rows_s = read_sheet(path_stalo)

    pp_b = [r for r in rows_b if norm_str(r.get("Операция")) == "Платежное поручение"]
    pp_s = [r for r in rows_s if norm_str(r.get("Операция")) == "Платежное поручение"]

    core_b = Counter(pp_core_key(r) for r in pp_b)
    core_s = Counter(pp_core_key(r) for r in pp_s)

    map_b = {pp_dn_key(r): r for r in pp_b}
    map_s = {pp_dn_key(r): r for r in pp_s}
    keys_b, keys_s = set(map_b), set(map_s)

    loaded_diffs = []
    meta_diffs = []
    meta_by_column = Counter()
    for key in sorted(keys_b & keys_s):
        rb, rs = map_b[key], map_s[key]
        if norm_str(rb.get("Загружен")) != norm_str(rs.get("Загружен")):
            loaded_diffs.append(
                {
                    "key": list(key),
                    "bylo": norm_str(rb.get("Загружен")),
                    "stalo": norm_str(rs.get("Загружен")),
                }
            )
        for col in PP_META:
            if norm_str(rb.get(col)) != norm_str(rs.get(col)):
                meta_by_column[col] += 1
                if len(meta_diffs) < 15:
                    meta_diffs.append(
                        {
                            "key": list(key),
                            "column": col,
                            "bylo": norm_str(rb.get(col))[:80],
                            "stalo": norm_str(rs.get(col))[:80],
                        }
                    )

    return {
        "all_rows": (len(rows_b), len(rows_s)),
        "pp_rows": (len(pp_b), len(pp_s)),
        "only_bylo_dn": [list(k) for k in sorted(keys_b - keys_s)[:30]],
        "only_stalo_dn": [list(k) for k in sorted(keys_s - keys_b)[:30]],
        "only_bylo_count": len(keys_b - keys_s),
        "only_stalo_count": len(keys_s - keys_b),
        "shared_dn_count": len(keys_b & keys_s),
        "core_multiset": compare_multiset(core_b, core_s),
        "loaded_diff_count": len(loaded_diffs),
        "loaded_diff_samples": loaded_diffs[:15],
        "meta_diff_count": sum(meta_by_column.values()),
        "meta_diff_by_column": dict(meta_by_column),
        "meta_diff_samples": meta_diffs,
    }


def safe_print(text: str) -> None:
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", errors="replace").decode("ascii"))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("folder", help="Subfolder under Regress/, e.g. Regress or Regress2105")
    parser.add_argument("prefix", help="File prefix, e.g. 1805_3105 or 2105_2105")
    args = parser.parse_args()

    reg_dir = SCRIPT_DIR / args.folder
    prefix = args.prefix

    v_bylo = reg_dir / f"{prefix}_ВЫПИСКИ_было.xlsx"
    v_stalo = reg_dir / f"{prefix}_ВЫПИСКИ_стало.xlsx"
    p_bylo = reg_dir / f"{prefix}_ПП_было.xlsx"
    p_stalo = reg_dir / f"{prefix}_ПП_стало.xlsx"

    for path in (v_bylo, v_stalo, p_bylo, p_stalo):
        if not path.exists():
            safe_print(f"ERROR: file not found: {path}")
            return 1

    report = {
        "format": "xlsx",
        "period": prefix,
        "folder": str(reg_dir),
        "files": {
            "vypiski": (v_bylo.name, v_stalo.name),
            "pp": (p_bylo.name, p_stalo.name),
        },
        "vypiski": compare_vypiski_xlsx(v_bylo, v_stalo),
        "pp": compare_pp_xlsx(p_bylo, p_stalo),
    }
    out = reg_dir / f"{prefix}_xlsx_compare_report.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    safe_print(f"=== XLSX COMPARE {prefix} ({args.folder}) ===")
    safe_print("")

    v = report["vypiski"]
    safe_print("1. VYPISKI")
    safe_print(f"   Rows: {v['rows'][0]} / {v['rows'][1]} (delta {v['row_delta']})")
    b = v["business"]
    safe_print(f"   Business multiset identical: {b['identical']}")
    safe_print(
        f"   Business: diff_types={b['diff_types']} only_b={b['only_bylo']} only_s={b['only_stalo']}"
    )
    a = v["alt_dogovor"]
    safe_print(
        f"   Alt (Date,Bank,Dogovor): identical={a['identical']} "
        f"diff_types={a['diff_types']} only_b={a['only_bylo']} only_s={a['only_stalo']}"
    )
    if b["top_diffs"]:
        safe_print("   Top business diffs:")
        for d in b["top_diffs"][:10]:
            safe_print(f"     delta={d['delta']:+d} {d['key']}")
    safe_print("")

    p = report["pp"]
    safe_print("2. PP (Платежное поручение)")
    safe_print(f"   All rows: {p['all_rows'][0]} / {p['all_rows'][1]}")
    safe_print(f"   PP rows: {p['pp_rows'][0]} / {p['pp_rows'][1]}")
    c = p["core_multiset"]
    safe_print(f"   Core multiset identical: {c['identical']}")
    safe_print(
        f"   Core: diff_types={c['diff_types']} only_b={c['only_bylo']} only_s={c['only_stalo']}"
    )
    safe_print(f"   Shared (Date,Number): {p['shared_dn_count']}")
    safe_print(f"   Only BYLO (Date,Number): {p['only_bylo_count']}")
    safe_print(f"   Only STALO (Date,Number): {p['only_stalo_count']}")
    if p["only_bylo_dn"]:
        safe_print(f"   Sample only BYLO: {p['only_bylo_dn'][:8]}")
    if p["only_stalo_dn"]:
        safe_print(f"   Sample only STALO: {p['only_stalo_dn'][:8]}")
    if c["top_diffs"]:
        safe_print("   Top core diffs:")
        for d in c["top_diffs"][:8]:
            safe_print(f"     delta={d['delta']:+d} date={d['key'][1]} num={d['key'][2]} sum={d['key'][3]}")
    safe_print(f"   Zагружен diffs (same DN): {p['loaded_diff_count']}")
    safe_print(f"   Meta/UUID field diffs (same DN): {p['meta_diff_count']}")
    if p["meta_diff_by_column"]:
        safe_print(f"   Meta by column: {p['meta_diff_by_column']}")
    safe_print("")
    safe_print(f"Saved: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
