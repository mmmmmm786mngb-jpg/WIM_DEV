#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Inspect XLSX regression export structure."""

from openpyxl import load_workbook
from pathlib import Path

REG = Path(__file__).resolve().parent / "Регресс2105"

for name in sorted(REG.glob("*.xlsx")):
    wb = load_workbook(name, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row)
        if i >= 2:
            break
    print("===", name.name, "sheet:", ws.title, "===")
    print("header:", [str(c)[:40] if c is not None else None for c in rows[0]])
    if len(rows) > 1:
        print("row1:", [str(c)[:40] if c is not None else None for c in rows[1]])
    # count rows
    n = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True) if any(c is not None and str(c).strip() for c in _))
    print("data rows (approx):", n)
    wb.close()
    print()
