#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Чтение содержимого Excel-файла с ключевыми операциями
и сохранение в JSON для дальнейшего анализа.
"""

import sys
import os
import json

import openpyxl

EXCEL_PATH = os.path.join(
    r'C:\1c\Cursor_1c\WIM_DEV',
    r'bases\Wim_Du\projects',
    'IMDEV-8927 Проанализировать все места вызова программного кода для ключевых операций ДУ',
    'Ключевые операции 15.xlsx'
)


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Sheet1']

    print('Dimensions:', ws.dimensions)
    print('Max row:', ws.max_row, 'Max col:', ws.max_column)

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(v) if v is not None else None for v in row])

    out_path = os.path.join(os.path.dirname(__file__), 'excel_data.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print('Saved to:', out_path)
    print('Total rows:', len(rows))
    print('First 3 rows:')
    for i, r in enumerate(rows[:3]):
        print(f'  Row {i}:', r)


if __name__ == '__main__':
    main()
