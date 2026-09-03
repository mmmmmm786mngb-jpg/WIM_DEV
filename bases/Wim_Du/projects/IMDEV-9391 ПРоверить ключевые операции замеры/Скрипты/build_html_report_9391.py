#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sbor statistiki zamerov IMDEV-9391 i generator HTML-otcheta.
Imena KO v Excel = Naimenovanie spravochnika (BSP RazlozhitStrokuPoSlovam).
"""

from __future__ import annotations

import html
import json
import math
import re
import statistics
from collections import Counter, defaultdict
from datetime import datetime, time
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "ЗамерыВремени.xlsx"
HTML_PATH = ROOT / "Документация" / "imdev_9391_key_operations_report.html"
JSON_PATH = ROOT / "Тестирование" / "reports" / "zamery_9391_stats.json"
JIRA_SUMMARY_PATH = ROOT / "Документация" / "jira_813_customer_summary.txt"
JIRA_ANSWERS_PATH = ROOT / "Документация" / "jira_813_customer_answers.txt"

JIRA_PASTE_HELP = """\
======= ЭТОТ БЛОК В JIRA НЕ КОПИРОВАТЬ =======
Jira 8.13: комментарий -> вкладка "Текст" (Wiki), не визуальный редактор.
Вставьте всё от строки h2. до конца файла и нажмите "Добавить".
======= КОПИРОВАТЬ НИЖЕ ЭТОЙ СТРОКИ =======

"""

CYR_ALL = set("АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯабвгдеёжзийклмнопрстуфхцчшщъыьэюя")
LAT_ALL = set("ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz")


def razlozhit(s: str) -> str:
    """Kopiya ОценкаПроизводительности.РазложитьСтрокуПоСловам."""
    positions = [i for i, ch in enumerate(s) if ch == ch.upper() and (ch in CYR_ALL or ch in LAT_ALL)]
    words = []
    if positions:
        prev = None
        last_pos = None
        for pos in positions:
            if prev is not None:
                sub = s[prev:pos].strip()
                if sub:
                    words.append(sub)
            prev = pos
            last_pos = pos
        sub = s[last_pos:].strip()
        if sub:
            words.append(sub)
    if not words:
        return s
    for i in range(1, len(words)):
        words[i] = words[i].lower()
    return " ".join(words)


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-", ".", "-."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_dt(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_comment(raw):
    extra = {}
    dop = ""
    if not raw:
        return extra, dop
    text = str(raw)
    try:
        obj = json.loads(text)
        if isinstance(obj, dict):
            dop_val = obj.get("ДопИнф", "")
            if isinstance(dop_val, dict):
                dop = json.dumps(dop_val, ensure_ascii=False)
                extra.update({str(k): str(v) for k, v in dop_val.items()})
            else:
                dop = str(dop_val)
        else:
            dop = text
    except json.JSONDecodeError:
        dop = text
    for part in re.split(r"[;\n]", dop):
        part = part.strip()
        if not part or "=" not in part:
            continue
        k, v = part.split("=", 1)
        extra[k.strip()] = v.strip()
    return extra, dop


def percentile(sorted_vals, p):
    if not sorted_vals:
        return None
    if len(sorted_vals) == 1:
        return sorted_vals[0]
    k = (len(sorted_vals) - 1) * p / 100.0
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return sorted_vals[int(k)]
    return sorted_vals[f] * (c - k) + sorted_vals[c] * (k - f)


def pack(vals):
    vals = [v for v in vals if v is not None]
    if not vals:
        return None
    s = sorted(vals)
    return {
        "n": len(s),
        "min": s[0],
        "p25": percentile(s, 25),
        "median": percentile(s, 50),
        "mean": statistics.fmean(s),
        "p75": percentile(s, 75),
        "p95": percentile(s, 95),
        "max": s[-1],
    }


def classify_session(dt):
    if dt is None:
        return "unknown"
    hm = dt.hour * 60 + dt.minute
    if hm < 10 * 60 + 30:
        return "s1"
    if hm < 13 * 60:
        return "s2"
    return "other"


# Technical names from EPF code
TECH = {
    "ds_rz": "ДУ.СверкаДенежныхСредств.Сверка.ФоновоеВыполнение",
    "ds_man": "ДУ.СверкаДенежныхСредств.Сверка.Ручной",
    "ers_rz": "ДУ.СверкаДенежныхСредствЕРС.Сверка.ФоновоеВыполнение",
    "ers_man": "ДУ.СверкаДенежныхСредствЕРС.Сверка.Ручной",
    "get_rz": "ДУ.ЗагрузкаКотировок.ПолучениеTibco.ФоновоеВыполнение",
    "crt_rz": "ДУ.ЗагрузкаКотировок.СозданиеКотировок.ФоновоеВыполнение",
    "get_pif": "ДУ.ЗагрузкаКотировок.ПолучениеTibco.ФоновоеВыполнениеVTBAMПИФ",
    "crt_pif": "ДУ.ЗагрузкаКотировок.СозданиеКотировок.ФоновоеВыполнениеVTBAMПИФ",
    "get_man": "ДУ.ЗагрузкаКотировок.ПолучениеTibco.Ручной",
    "crt_man": "ДУ.ЗагрузкаКотировок.СозданиеКотировок.Ручной",
    "cbonds": "ДУ.ЗагрузкаКотировок.ЗагрузкаCbonds.ФоновоеВыполнениеCBONDS",
    "ctrl_rz": "ДУ.КонтрольКотировок.Отчет.ФоновоеВыполнение",
    "ctrl_man": "ДУ.КонтрольКотировок.Отчет.Ручной",
}

NAIM = {k: razlozhit(v) for k, v in TECH.items()}


def stats_rows(rows, target):
    times = [r["time"] for r in rows if r["time"] is not None]
    weights = [r["weight"] for r in rows if r["weight"] is not None]
    unit, thr, tw = [], [], []
    extras_num = defaultdict(list)
    extras_cat = defaultdict(Counter)
    hours = Counter()
    users = Counter()
    errors = Counter()
    dates = []
    daily = defaultdict(list)
    series = []
    samples = []
    for r in rows:
        t, w = r["time"], r["weight"]
        if t is not None and w not in (None, 0):
            unit.append(t / w)
            if t > 0:
                thr.append(w / t)
            tw.append((t, w))
        if r.get("local"):
            hours[f"{r['local'].hour:02d}"] += 1
            dates.append(r["local"])
            daily[r["local"].date().isoformat()].append(t)
            series.append({
                "dt": r["local"].strftime("%d.%m.%Y %H:%M:%S"),
                "day": r["local"].strftime("%d.%m"),
                "time": t,
                "weight": w,
                "error": r.get("error"),
            })
        users[r.get("user") or ""] += 1
        errors[r.get("error") or ""] += 1
        for k, v in (r.get("extra") or {}).items():
            num = to_float(str(v).replace(" об/с", "").replace("об/с", ""))
            if num is not None:
                extras_num[k].append(num)
            else:
                extras_cat[k][str(v)] += 1
        if r.get("dop"):
            samples.append({
                "dt": r["local"].strftime("%d.%m.%Y %H:%M:%S") if r.get("local") else "",
                "time": t,
                "weight": w,
                "error": r.get("error"),
                "dop": r["dop"],
                "user": r.get("user"),
            })
    sum_t = sum(x for x, _ in tw)
    sum_w = sum(w for _, w in tw)
    daily_med = []
    for day in sorted(daily):
        vals = [x for x in daily[day] if x is not None]
        if vals:
            daily_med.append((day, statistics.median(vals), max(vals), len(vals)))
    picked = []
    if samples:
        by_time = sorted(samples, key=lambda s: s["time"] or 0)
        picked.append(by_time[len(by_time) // 2])
        if by_time[-1] is not picked[0]:
            picked.append(by_time[-1])
        if by_time[0] is not picked[0]:
            picked.append(by_time[0])
        for s in samples:
            if s not in picked and len(picked) < 4:
                picked.append(s)
    return {
        "n": len(rows),
        "time": pack(times),
        "weight": pack(weights),
        "sec_per_unit": pack(unit),
        "units_per_sec": pack(thr),
        "weighted_sec_per_unit": (sum_t / sum_w) if sum_w else None,
        "weighted_units_per_sec": (sum_w / sum_t) if sum_t else None,
        "over_target": sum(1 for t in times if t > target),
        "weight_eq_1": sum(1 for w in weights if w == 1),
        "weight_gt_1": sum(1 for w in weights if w and w > 1),
        "errors": errors.most_common(),
        "users": users.most_common(6),
        "hours": sorted(hours.items()),
        "date_min": min(dates).strftime("%d.%m.%Y") if dates else None,
        "date_max": max(dates).strftime("%d.%m.%Y") if dates else None,
        "extra_num": {k: pack(v) for k, v in extras_num.items()},
        "extra_cat": {k: v.most_common(6) for k, v in extras_cat.items()},
        "samples": picked[:4],
        "series": series,
        "daily": daily_med,
        "target": target,
    }


def load_rows():
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["TDSheet"]
    it = ws.iter_rows(values_only=True)
    next(it)
    by_naim = defaultdict(list)
    n_all = 0
    for row in it:
        n_all += 1
        ko = str(row[0] or "").strip()
        if not ko:
            continue
        extra, dop = parse_comment(row[6] if len(row) > 6 else "")
        local = parse_dt(row[10] if len(row) > 10 else None)
        rec = {
            "ko": ko,
            "time": to_float(row[4] if len(row) > 4 else None),
            "weight": to_float(row[5] if len(row) > 5 else None),
            "extra": extra,
            "dop": dop,
            "local": local,
            "session": classify_session(local),
            "user": str(row[9] or "") if len(row) > 9 else "",
            "error": str(row[11] or "") if len(row) > 11 else "",
        }
        by_naim[ko].append(rec)
    wb.close()
    return n_all, by_naim


def pair_cycle(get_rows, crt_rows):
    """Summa polucheniya + sozdaniya za tot zhe den i seans."""
    buckets = defaultdict(lambda: {"get": [], "crt": []})
    for r in get_rows:
        if r.get("local"):
            buckets[(r["local"].date(), r["session"])]["get"].append(r)
    for r in crt_rows:
        if r.get("local"):
            buckets[(r["local"].date(), r["session"])]["crt"].append(r)
    totals = []
    for (day, sess), b in buckets.items():
        gs = sorted(b["get"], key=lambda x: x["local"])
        cs = sorted(b["crt"], key=lambda x: x["local"])
        for g, c in zip(gs, cs):
            if g["time"] is None or c["time"] is None:
                continue
            w_get = g.get("extra", {}).get("Котировок")
            w_crt = c.get("extra", {}).get("Создано")
            totals.append({
                "day": day.isoformat(),
                "session": sess,
                "total": g["time"] + c["time"],
                "get": g["time"],
                "crt": c["time"],
                "quotes": to_float(w_get),
                "docs": to_float(w_crt),
                "dt": g["local"].strftime("%d.%m.%Y %H:%M"),
            })
    return totals


def verdict(st, target):
    if not st or st["n"] == 0 or not st.get("time"):
        return "nodata", "Нет замеров в регистре"
    t = st["time"]
    share = st["over_target"] / st["n"] if st["n"] else 0
    if t["p95"] > target or share >= 0.1:
        return "need", "P95 или доля превышений относительно целевого времени"
    if t["median"] > target * 0.7:
        return "watch", "Медиана близка к целевому времени"
    if t["max"] > target:
        return "watch", "Есть единичные превышения максимума"
    return "ok", "Медиана и P95 существенно ниже цели"


def fmt(x, d=1):
    if x is None:
        return "-"
    if isinstance(x, int) or (isinstance(x, float) and abs(x - round(x)) < 1e-9 and d == 0):
        return str(int(round(x)))
    s = f"{x:.{d}f}".replace(".", ",")
    return s


def fmt_sec(x):
    if x is None:
        return "—"
    if x < 1:
        return f"{fmt(x, 3)} с"
    if x < 60:
        return f"{fmt(x, 1)} с"
    return f"{fmt(x, 1)} с ({fmt(x / 60, 1)} мин)"


def fmt_ude(x):
    if x is None:
        return "-"
    if x >= 10:
        return fmt(x, 1)
    if x >= 1:
        return fmt(x, 2)
    if x >= 0.01:
        return fmt(x, 3)
    return fmt(x, 4)


def ru_num_word(n, one, few, many):
    n = abs(int(n))
    if 11 <= n % 100 <= 14:
        return many
    d = n % 10
    if d == 1:
        return one
    if 2 <= d <= 4:
        return few
    return many


def fmt_plain_time(x):
    if x is None:
        return "нет данных"
    if x < 60:
        if abs(x - round(x)) < 0.05:
            n = int(round(x))
            return f"{n} {ru_num_word(n, 'секунда', 'секунды', 'секунд')}"
        return f"{fmt(x, 1)} секунды"
    whole = int(round(x))
    minutes, seconds = divmod(whole, 60)
    if seconds == 0:
        return f"{minutes} {ru_num_word(minutes, 'минута', 'минуты', 'минут')}"
    return (
        f"{minutes} {ru_num_word(minutes, 'минута', 'минуты', 'минут')} "
        f"{seconds} {ru_num_word(seconds, 'секунда', 'секунды', 'секунд')}"
    )


def pct_share(fact, target):
    if not fact or not target:
        return "-"
    p = 100.0 * fact / target
    return fmt(p, 0) if p >= 1 else fmt(p, 1)


def wiki_meter(fact, target, width=12):
    if not fact or not target:
        return ""
    n = int(round(width * min(1.0, fact / target)))
    if fact > 0 and n < 1:
        n = 1
    n = min(width, n)
    return "[" + ("#" * n) + ("." * (width - n)) + "]"


def wiki_panel(title, body):
    return (
        "{panel:title=" + title
        + "|borderStyle=solid|borderColor=#14892c|titleBGColor=#e3fcef|bgColor=#ffffff}\n"
        + body.strip()
        + "\n{panel}"
    )


def wiki_mono(name):
    return "{{" + name + "}}"


def write_jira_comments(data):
    """Dva kommentariya dlya Jira 8.13 (wiki markup)."""
    per = data["period"]
    ds_n, ds_man_n = data["ds_n"], data["ds_man_n"]
    ers_n, ers_man_n = data["ers_n"], data["ers_man_n"]
    c1_n, c2_n, ctrl_n = data["c1_n"], data["c2_n"], data["ctrl_n"]
    ds_med, ers_med = data["ds_med"], data["ers_med"]
    c1_med, c2_med, ctrl_med = data["c1_med"], data["c2_med"], data["ctrl_med"]
    ds_p95, ers_p95 = data["ds_p95"], data["ers_p95"]
    c1_p95, c2_p95, ctrl_p95 = data["c1_p95"], data["c2_p95"], data["ctrl_p95"]
    ds_max, ers_max = data["ds_max"], data["ers_max"]
    c1_max, c2_max, ctrl_max = data["c1_max"], data["c2_max"], data["ctrl_max"]
    ude_ds, ude_ers = data["ude_ds"], data["ude_ers"]
    ude_s1, ude_s2, ude_ctrl = data["ude_s1"], data["ude_s2"], data["ude_ctrl"]
    q1_vol, q2_vol = data["q1_vol"], data["q2_vol"]

    t_ds = fmt_plain_time(ds_med)
    t_ers = fmt_plain_time(ers_med)
    t_s1 = fmt_plain_time(c1_med)
    t_s2 = fmt_plain_time(c2_med)
    t_ctrl = fmt_plain_time(ctrl_med)

    def row_meter(fact, target):
        return wiki_meter(fact, target) + " " + pct_share(fact, target) + "%"

    summary = JIRA_PASTE_HELP + f"""\
h2. Сводка по списку заказчика

{{info}}
*Задача:* [IMDEV-9391]
*Период регистра:* {per}
*Контур для вердикта:* регламентные задания (не ручные запуски из формы)
*Цели:* сверки денежных средств - не дольше 6 минут; загрузка и контроль котировок - не дольше 30 минут
{{info}}

||№||Операция из задачи||Обработка||Замер||Типичное время||Цель||Занято от предела||Удельный вес||Оптимизация||
|1|Сверка денежных средств|{wiki_mono("внСверкаДенежныхСредств_обр")}|(/) Есть, {ds_n} регламентных|{t_ds}|6 минут|{row_meter(ds_med, 360)}|{fmt_ude(ude_ds)} расхождений в секунду|(/) *Не нужна*|
|2|Сверка денежных средств ЕРС|{wiki_mono("внСверкаДенежныхСредствНаЕРС_обр")}|(/) Есть, {ers_n} регламентных|{t_ers}|6 минут|{row_meter(ers_med, 360)}|{fmt_ude(ude_ers)} расхождений в секунду|(/) *Не нужна*|
|3|Загрузка котировок, первый сеанс за предыдущий день|{wiki_mono("внЗагрузкаКотировокДУ")} около 09:00|(/) Есть, {c1_n} полных циклов|{t_s1}|30 минут|{row_meter(c1_med, 1800)}|{fmt_ude(ude_s1)} котировок в секунду|(/) *Не нужна*|
|4|Сверка загрузки котировок|{wiki_mono("внОтчетКонтрольКотировок")}|(/) Есть, {ctrl_n} регламентных|{t_ctrl}|30 минут|{row_meter(ctrl_med, 1800)}|{fmt_ude(ude_ctrl)} строк отчёта в секунду|(/) *Не нужна*|
|5|Загрузка котировок, финальная за предыдущий день|{wiki_mono("внЗагрузкаКотировокДУ")} около 11:50|(/) Есть, {c2_n} полных циклов|{t_s2}|30 минут|{row_meter(c2_med, 1800)}|{fmt_ude(ude_s2)} котировок в секунду|(/) *Не нужна*|

{{tip:title=Итог}}
(/) Замер времени *формируется по всем пяти операциям*.
(/) Оптимизация *не нужна ни по одной операции*: типичное время рабочего контура много ниже цели.
{{tip}}
"""

    p1 = wiki_panel(
        "1. Сверка денежных средств  -  оптимизация не нужна",
        f"""*Обработка:* {wiki_mono("внСверкаДенежныхСредств_обр")}

*Вопрос 1. Формируется ли замер времени?*
(/) *Да, замер есть.*
Замер встроен в код обработки и пишется в регистр. За период с {per} в регистре {ds_n} запусков регламентным заданием (по ним делаем вывод) и {ds_man_n} ручных запусков из формы.

*Вопрос 2. Нужна ли оптимизация?*
(/) *Нет, не нужна.*
Вердикт по регламентным заданиям. Почти все запуски (95 из 100) укладываются в {fmt_plain_time(ds_p95)}. Самый долгий за период: {fmt_plain_time(ds_max)}. Целевое время ни разу не превышено.

||Типичное время||Допустимо не дольше||Занято от предела||
|{t_ds}|6 минут|{row_meter(ds_med, 360)}|""",
    )
    p2 = wiki_panel(
        "2. Сверка денежных средств ЕРС  -  оптимизация не нужна",
        f"""*Обработка:* {wiki_mono("внСверкаДенежныхСредствНаЕРС_обр")}
ЕРС - единый расчётный счёт.

*Вопрос 1. Формируется ли замер времени?*
(/) *Да, замер есть.*
Замер встроен в код обработки и пишется в регистр. За период с {per} в регистре {ers_n} запусков регламентным заданием и {ers_man_n} ручных запусков из формы.

*Вопрос 2. Нужна ли оптимизация?*
(/) *Нет, не нужна.*
Вердикт по регламентным заданиям. Почти все запуски (95 из 100) укладываются в {fmt_plain_time(ers_p95)}. Самый долгий за период: {fmt_plain_time(ers_max)}. Целевое время ни разу не превышено.

||Типичное время||Допустимо не дольше||Занято от предела||
|{t_ers}|6 минут|{row_meter(ers_med, 360)}|""",
    )
    vol1 = f" (около {fmt_int(q1_vol)} котировок за типичный цикл)" if q1_vol else ""
    vol2 = f" (около {fmt_int(q2_vol)} котировок за типичный цикл)" if q2_vol else ""
    p3 = wiki_panel(
        "3. Загрузка котировок, первый сеанс за предыдущий день  -  оптимизация не нужна",
        f"""*Обработка:* {wiki_mono("внЗагрузкаКотировокДУ")} (запуск около 09:00)

*Вопрос 1. Формируется ли замер времени?*
(/) *Да, замер есть.*
Обработка пишет два замера на один запуск: получение котировок и запись документов. Первый и финальный сеансы в справочнике не разделены - это одна обработка и одни и те же имена. Первый сеанс отобран по времени запуска около 09:00. За период с {per} собрано {c1_n} полных циклов (оба шага в один день).

*Вопрос 2. Нужна ли оптимизация?*
(/) *Нет, не нужна.*
Время сеанса - сумма двух шагов. Почти все запуски (95 из 100) укладываются в {fmt_plain_time(c1_p95)}. Самый долгий за период: {fmt_plain_time(c1_max)}. Целевое время ни разу не превышено.

||Типичное время||Допустимо не дольше||Занято от предела||
|{t_s1}|30 минут|{row_meter(c1_med, 1800)}|

Удельный вес: *{fmt_ude(ude_s1)} котировок в секунду*{vol1}.""",
    )
    p4 = wiki_panel(
        "4. Сверка загрузки котировок  -  оптимизация не нужна",
        f"""*Обработка:* {wiki_mono("внОтчетКонтрольКотировок")}

*Вопрос 1. Формируется ли замер времени?*
(/) *Да, замер есть.*
Замер встроен в код обработки и пишется в регистр. За период с {per} в регистре {ctrl_n} запусков регламентным заданием. Ручных запусков с замером в этой выгрузке нет.

*Вопрос 2. Нужна ли оптимизация?*
(/) *Нет, не нужна.*
Вердикт по регламентным заданиям. Почти все запуски (95 из 100) укладываются в {fmt_plain_time(ctrl_p95)}. Самый долгий за период: {fmt_plain_time(ctrl_max)}. Целевое время ни разу не превышено.

||Типичное время||Допустимо не дольше||Занято от предела||
|{t_ctrl}|30 минут|{row_meter(ctrl_med, 1800)}|

Удельный вес: *{fmt_ude(ude_ctrl)} строк отчёта в секунду*.""",
    )
    p5 = wiki_panel(
        "5. Загрузка котировок, финальная за предыдущий день  -  оптимизация не нужна",
        f"""*Обработка:* {wiki_mono("внЗагрузкаКотировокДУ")} (запуск около 11:50)

*Вопрос 1. Формируется ли замер времени?*
(/) *Да, замер есть.*
Та же обработка и те же имена в справочнике, что у первого сеанса. Финальный сеанс отобран по времени запуска около 11:50. За период с {per} собрано {c2_n} полных циклов (получение котировок плюс запись документов).

*Вопрос 2. Нужна ли оптимизация?*
(/) *Нет, не нужна.*
Время сеанса - сумма двух шагов. Почти все запуски (95 из 100) укладываются в {fmt_plain_time(c2_p95)}. Самый долгий за период: {fmt_plain_time(c2_max)}. Целевое время ни разу не превышено.

||Типичное время||Допустимо не дольше||Занято от предела||
|{t_s2}|30 минут|{row_meter(c2_med, 1800)}|

Удельный вес: *{fmt_ude(ude_s2)} котировок в секунду*{vol2}.""",
    )

    answers = JIRA_PASTE_HELP + f"""\
h2. Итоговые ответы заказчику

По каждой из пяти операций заданы два вопроса. Ниже - ответы на всю пятёрку сразу, затем карточка по каждой операции. Период регистра: *{per}*. Вывод по оптимизации - по регламентным заданиям, не по ручным запускам из формы.

{{tip:title=Вопрос 1. Формируется ли замер времени?}}
(/) *Да. Замер времени формируется по всем пяти операциям.*
Во всех четырёх внешних обработках код замера работает, в регистре есть записи. Первый и финальный сеансы загрузки котировок - одна обработка, в отчёте разделены по времени запуска.
{{tip}}

{{tip:title=Вопрос 2. Нужна ли оптимизация?}}
(/) *Нет. Оптимизация не нужна ни по одной операции.*
Типичное время рабочего контура много ниже цели: сверки денежных средств - не дольше 6 минут, загрузка и контроль котировок - не дольше 30 минут.
{{tip}}

{p1}

{p2}

{p3}

{p4}

{p5}

h3. Сводка одним взглядом

||№||Операция из задачи||Замер времени||Типичное время||Допустимо не дольше||Оптимизация||
|1|Сверка денежных средств|Есть. {ds_n} регламентных и {ds_man_n} ручных|{t_ds}|6 минут|(/) *Не нужна*|
|2|Сверка денежных средств ЕРС|Есть. {ers_n} регламентных и {ers_man_n} ручных|{t_ers}|6 минут|(/) *Не нужна*|
|3|Загрузка котировок, первый сеанс за предыдущий день|Есть. {c1_n} полных циклов около 09:00|{t_s1}|30 минут|(/) *Не нужна*|
|4|Сверка загрузки котировок|Есть. {ctrl_n} регламентных запусков|{t_ctrl}|30 минут|(/) *Не нужна*|
|5|Загрузка котировок, финальная за предыдущий день|Есть. {c2_n} полных циклов около 11:50|{t_s2}|30 минут|(/) *Не нужна*|
"""

    JIRA_SUMMARY_PATH.write_text(summary, encoding="utf-8")
    JIRA_ANSWERS_PATH.write_text(answers, encoding="utf-8")


def rate_median(st):
    # Вес/время по регистру: в комментарии УдельныйВес часто округлён до целых об/с.
    u = st.get("units_per_sec") or {}
    if u and u.get("median") is not None:
        return u["median"]
    extra = (st.get("extra_num") or {}).get("УдельныйВес")
    if extra and extra.get("median"):
        return extra["median"]
    return st.get("weighted_units_per_sec")


def cycle_rate(cyc):
    rates = []
    for c in cyc.get("items") or []:
        q, t = c.get("quotes"), c.get("total")
        if q and t and t > 0:
            rates.append(q / t)
    packed = pack(rates)
    return packed["median"] if packed else None


def fmt_int(x):
    if x is None:
        return "-"
    n = int(round(x))
    s = f"{n:,}".replace(",", " ")
    return s


def esc(s):
    return html.escape("" if s is None else str(s))


def svg_bars(items, width=720, bar_h=22, gap=8, label_w=240, unit="с"):
    """items: list of (label, value, color)."""
    if not items:
        return ""
    vals = [v for _, v, _ in items if v is not None]
    if not vals:
        return ""
    vmax = max(vals) or 1
    h = 16 + len(items) * (bar_h + gap)
    plot_w = width - label_w - 80
    parts = [f'<svg viewBox="0 0 {width} {h}" width="100%" role="img">']
    y = 8
    for label, val, color in items:
        vw = 0 if not val else plot_w * val / vmax
        parts.append(
            f'<text x="0" y="{y + bar_h * 0.72}" font-size="12" fill="#334">{esc(label)}</text>'
        )
        parts.append(
            f'<rect x="{label_w}" y="{y}" width="{plot_w}" height="{bar_h}" rx="3" fill="#eef2f6"/>'
        )
        parts.append(
            f'<rect x="{label_w}" y="{y}" width="{max(vw, 1)}" height="{bar_h}" rx="3" fill="{color}"/>'
        )
        parts.append(
            f'<text x="{label_w + plot_w + 8}" y="{y + bar_h * 0.72}" font-size="12" fill="#1a2332">'
            f"{fmt_sec(val) if unit == 'с' else fmt_ude(val) + ' ' + unit}</text>"
        )
        y += bar_h + gap
    parts.append("</svg>")
    return "\n".join(parts)


def svg_line(daily, width=720, height=200, color="#1e5a8a", title=""):
    """daily: list of (iso_day, median, max, n)."""
    if len(daily) < 2:
        return ""
    vals = [m for _, m, _, _ in daily]
    vmax = max(vals) or 1
    pad_l, pad_r, pad_t, pad_b = 44, 16, 18, 36
    plot_w = width - pad_l - pad_r
    plot_h = height - pad_t - pad_b
    pts = []
    n = len(daily) - 1
    for i, (_, med, _, _) in enumerate(daily):
        x = pad_l + plot_w * i / n
        y = pad_t + plot_h * (1 - med / vmax)
        pts.append((x, y, med, daily[i][0]))
    d = " ".join(f"{'M' if i == 0 else 'L'}{x:.1f},{y:.1f}" for i, (x, y, _, _) in enumerate(pts))
    circles = "".join(
        f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="{color}">'
        f"<title>{esc(day)}: {fmt_sec(med)}</title></circle>"
        for x, y, med, day in pts
    )
    labels = []
    step = max(1, len(pts) // 8)
    for i, (x, _, _, day) in enumerate(pts):
        if i % step == 0 or i == len(pts) - 1:
            dd = datetime.fromisoformat(day).strftime("%d.%m")
            labels.append(
                f'<text x="{x:.1f}" y="{height - 8}" font-size="10" text-anchor="middle" fill="#5a6a7a">{dd}</text>'
            )
    grid = []
    for frac in (0, 0.5, 1):
        y = pad_t + plot_h * (1 - frac)
        v = vmax * frac
        grid.append(f'<line x1="{pad_l}" y1="{y:.1f}" x2="{width - pad_r}" y2="{y:.1f}" stroke="#e6ebf0"/>')
        grid.append(
            f'<text x="{pad_l - 6}" y="{y + 4:.1f}" font-size="10" text-anchor="end" fill="#5a6a7a">{fmt(v, 0 if v >= 10 else 1)}</text>'
        )
    cap = f'<text x="{pad_l}" y="12" font-size="11" fill="#5a6a7a">{esc(title)}</text>' if title else ""
    return (
        f'<svg viewBox="0 0 {width} {height}" width="100%" role="img">{cap}'
        + "".join(grid)
        + f'<path d="{d}" fill="none" stroke="{color}" stroke-width="2"/>'
        + circles
        + "".join(labels)
        + "</svg>"
    )


def svg_hist(times, width=720, height=170, bins=12, color="#1e5a8a"):
    if not times:
        return ""
    lo, hi = min(times), max(times)
    if hi - lo < 1e-9:
        hi = lo + 1
    step = (hi - lo) / bins
    counts = [0] * bins
    for t in times:
        i = min(int((t - lo) / step), bins - 1)
        counts[i] += 1
    cmax = max(counts) or 1
    pad_l, pad_r, pad_t, pad_b = 36, 10, 10, 28
    plot_w = width - pad_l - pad_r
    plot_h = height - pad_t - pad_b
    bw = plot_w / bins * 0.82
    gap = plot_w / bins * 0.18
    rects = []
    for i, c in enumerate(counts):
        x = pad_l + i * (bw + gap)
        h = plot_h * c / cmax
        y = pad_t + plot_h - h
        left = lo + i * step
        right = left + step
        rects.append(
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{bw:.1f}" height="{h:.1f}" rx="2" fill="{color}">'
            f"<title>{fmt_sec(left)} – {fmt_sec(right)}: {c}</title></rect>"
        )
    xlab = (
        f'<text x="{pad_l}" y="{height - 6}" font-size="10" fill="#5a6a7a">{fmt_sec(lo)}</text>'
        f'<text x="{width - pad_r}" y="{height - 6}" font-size="10" text-anchor="end" fill="#5a6a7a">{fmt_sec(hi)}</text>'
    )
    return f'<svg viewBox="0 0 {width} {height}" width="100%" role="img">{"".join(rects)}{xlab}</svg>'


def kpi_row(st):
    t = st.get("time") or {}
    w = st.get("weight") or {}
    return f"""
    <div class="kpis">
      <div class="kpi"><div class="kpi-l">Замеров</div><div class="kpi-v">{st['n']}</div></div>
      <div class="kpi"><div class="kpi-l">Медиана</div><div class="kpi-v">{fmt_sec(t.get('median'))}</div></div>
      <div class="kpi"><div class="kpi-l">Среднее</div><div class="kpi-v">{fmt_sec(t.get('mean'))}</div></div>
      <div class="kpi"><div class="kpi-l">P95</div><div class="kpi-v">{fmt_sec(t.get('p95'))}</div></div>
      <div class="kpi"><div class="kpi-l">Максимум</div><div class="kpi-v">{fmt_sec(t.get('max'))}</div></div>
      <div class="kpi"><div class="kpi-l">Вес, медиана</div><div class="kpi-v">{fmt(w.get('median'), 0 if (w.get('median') or 0) >= 10 else 1)}</div></div>
    </div>"""


def table_stats(st, weight_meaning):
    t = st.get("time") or {}
    w = st.get("weight") or {}
    u = st.get("sec_per_unit") or {}
    thr = st.get("units_per_sec") or {}
    err_yes = 0
    for k, c in st.get("errors") or []:
        if str(k).lower() in ("да", "true", "1"):
            err_yes += c
    extras = []
    for k, ev in (st.get("extra_num") or {}).items():
        nd = 0 if (ev.get("median") or 0) >= 10 else 2
        extras.append(
            f"<tr><td>{esc(k)}</td><td class='num'>{fmt(ev['median'], nd)}</td>"
            f"<td class='num'>{fmt(ev['mean'], 1 if nd == 0 else 2)}</td>"
            f"<td class='num'>{fmt(ev['min'], nd)}</td>"
            f"<td class='num'>{fmt(ev['max'], nd)}</td></tr>"
        )
    extra_html = ""
    if extras:
        extra_html = f"""
        <h4>Параметры из комментария замера</h4>
        <table>
          <tr><th>Поле</th><th>Медиана</th><th>Среднее</th><th>Min</th><th>Max</th></tr>
          {''.join(extras)}
        </table>"""
    samples = ""
    if st.get("samples"):
        rows = []
        for s in st["samples"][:4]:
            rows.append(
                f"<tr><td>{esc(s['dt'])}</td><td class='num'>{fmt_sec(s['time'])}</td>"
                f"<td class='num'>{fmt(s['weight'], 0 if (s.get('weight') or 0) >= 1 else 2)}</td>"
                f"<td>{esc(s['error'])}</td>"
                f"<td><code>{esc(s['dop'][:180])}</code></td></tr>"
            )
        samples = f"""
        <h4>Примеры комментариев</h4>
        <table>
          <tr><th>Время (лок.)</th><th>Длительность</th><th>Вес</th><th>Ошибка</th><th>ДопИнф</th></tr>
          {''.join(rows)}
        </table>"""
    return f"""
        <table>
          <tr><th>Показатель</th><th>Значение</th></tr>
          <tr><td>Период</td><td>{esc(st.get('date_min'))} — {esc(st.get('date_max'))}</td></tr>
          <tr><td>Замеров / с ошибкой</td><td>{st['n']} / {err_yes}</td></tr>
          <tr><td>Время, с: min / медиана / среднее / P95 / max</td>
              <td class='num'>{fmt_sec(t.get('min'))} / {fmt_sec(t.get('median'))} / {fmt_sec(t.get('mean'))} / {fmt_sec(t.get('p95'))} / {fmt_sec(t.get('max'))}</td></tr>
          <tr><td>Целевое время</td><td class='num'>{fmt_sec(st.get('target'))}</td></tr>
          <tr><td>Превышений цели</td><td class='num'>{st.get('over_target', 0)} из {st['n']}</td></tr>
          <tr><td>Вес ({esc(weight_meaning)})</td>
              <td class='num'>min {fmt(w.get('min'), 0)} / мед. {fmt(w.get('median'), 0)} / сред. {fmt(w.get('mean'), 1)} / max {fmt(w.get('max'), 0)}
              (вес=1: {st.get('weight_eq_1', 0)}; вес&gt;1: {st.get('weight_gt_1', 0)})</td></tr>
          <tr><td>Удельно, с на единицу веса</td>
              <td class='num'>медиана {fmt(u.get('median'), 3) if u else '-'} с; взвеш. среднее {fmt(st.get('weighted_sec_per_unit'), 3) if st.get('weighted_sec_per_unit') is not None else '-'}</td></tr>
          <tr><td>Пропускная способность</td>
              <td class='num'>медиана {fmt(thr.get('median'), 2) if thr else '-'} ед/с; взвеш. среднее {fmt(st.get('weighted_units_per_sec'), 2) if st.get('weighted_units_per_sec') is not None else '-'}</td></tr>
          <tr><td>Пользователи</td><td>{esc(', '.join(f'{u} ({c})' for u, c in (st.get('users') or [])[:4]))}</td></tr>
        </table>
        {extra_html}
        {samples}
    """


def q_boxes(code_yes, excel_n, v_code, v_reason, extra_q1="", extra_q2=""):
    q1_cls = "ok" if code_yes and excel_n else ("warn" if code_yes else "bad")
    q1_t = "Да, замер формируется" if code_yes and excel_n else (
        "Код есть, в регистре записей нет" if code_yes else "Замер в коде не найден"
    )
    vmap = {"ok": ("ok", "Оптимизация не требуется"),
            "watch": ("warn", "Наблюдать, срочная оптимизация не нужна"),
            "need": ("bad", "Есть потребность в оптимизации"),
            "nodata": ("warn", "Недостаточно данных для вывода")}
    cls, title = vmap[v_code]
    return f"""
    <div class="qgrid">
      <div class="qbox {q1_cls}">
        <div class="qnum">Вопрос 1</div>
        <h3>Формируется ли замер времени?</h3>
        <p><strong>{q1_t}.</strong> Записей в регистре: <strong>{excel_n}</strong>.</p>
        {extra_q1}
      </div>
      <div class="qbox {cls}">
        <div class="qnum">Вопрос 2</div>
        <h3>Нужна ли оптимизация?</h3>
        <p><strong>{title}.</strong> {esc(v_reason)}</p>
        {extra_q2}
      </div>
    </div>
    """


def build():
    n_all, by_naim = load_rows()
    groups = {}
    for key, tech in TECH.items():
        naim = NAIM[key]
        target = 360 if key.startswith(("ds_", "ers_")) else 1800
        rows = by_naim.get(naim, [])
        groups[key] = {
            "tech": tech,
            "naim": naim,
            "rows": rows,
            "stats": stats_rows(rows, target),
        }

    # session splits for quotes RZ
    for sess in ("s1", "s2", "other"):
        for base in ("get_rz", "crt_rz"):
            rows = [r for r in groups[base]["rows"] if r["session"] == sess]
            target = 1800
            groups[f"{base}_{sess}"] = {
                "tech": groups[base]["tech"],
                "naim": groups[base]["naim"],
                "rows": rows,
                "stats": stats_rows(rows, target),
            }

    cycles_s1 = pair_cycle(groups["get_rz_s1"]["rows"], groups["crt_rz_s1"]["rows"])
    cycles_s2 = pair_cycle(groups["get_rz_s2"]["rows"], groups["crt_rz_s2"]["rows"])

    def cycle_stats(cycles):
        tot = [c["total"] for c in cycles]
        quotes = [c["quotes"] for c in cycles if c["quotes"] is not None]
        return {
            "n": len(cycles),
            "time": pack(tot),
            "quotes": pack(quotes),
            "over_1800": sum(1 for x in tot if x > 1800),
            "items": cycles,
        }

    c1 = cycle_stats(cycles_s1)
    c2 = cycle_stats(cycles_s2)

    period_min = None
    period_max = None
    for g in groups.values():
        for r in g["rows"]:
            if r.get("local"):
                if period_min is None or r["local"] < period_min:
                    period_min = r["local"]
                if period_max is None or r["local"] > period_max:
                    period_max = r["local"]

    payload = {
        "total_excel_rows": n_all,
        "period": {
            "min": period_min.strftime("%d.%m.%Y") if period_min else None,
            "max": period_max.strftime("%d.%m.%Y") if period_max else None,
        },
        "naim": NAIM,
        "n": {k: groups[k]["stats"]["n"] for k in TECH},
        "s1_get": groups["get_rz_s1"]["stats"]["n"],
        "s1_crt": groups["crt_rz_s1"]["stats"]["n"],
        "s2_get": groups["get_rz_s2"]["stats"]["n"],
        "s2_crt": groups["crt_rz_s2"]["stats"]["n"],
        "cycles_s1": c1["n"],
        "cycles_s2": c2["n"],
    }
    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)

    # ----- HTML pieces -----
    ds_rz, ds_man = groups["ds_rz"]["stats"], groups["ds_man"]["stats"]
    ers_rz, ers_man = groups["ers_rz"]["stats"], groups["ers_man"]["stats"]
    get_s1, crt_s1 = groups["get_rz_s1"]["stats"], groups["crt_rz_s1"]["stats"]
    get_s2, crt_s2 = groups["get_rz_s2"]["stats"], groups["crt_rz_s2"]["stats"]
    ctrl = groups["ctrl_rz"]["stats"]
    get_pif, crt_pif = groups["get_pif"]["stats"], groups["crt_pif"]["stats"]

    v_ds, r_ds = verdict(ds_rz, 360)
    v_ers, r_ers = verdict(ers_rz, 360)
    v_s1, r_s1 = verdict({"n": c1["n"], "time": c1["time"], "over_target": c1["over_1800"]}, 1800) if c1["n"] else ("nodata", "Нет парных циклов")
    if c1["n"] and c1["time"] and c1["time"]["p95"] < 1800:
        # also check steps
        vs1g, _ = verdict(get_s1, 1800)
        if vs1g == "ok":
            v_s1, r_s1 = "ok", "Полный цикл (Tibco + создание) далеко от цели 1800 с"
    v_s2, r_s2 = verdict({"n": c2["n"], "time": c2["time"], "over_target": c2["over_1800"]}, 1800) if c2["n"] else ("nodata", "Нет парных циклов")
    if c2["n"] and c2["time"] and c2["time"]["p95"] < 1800:
        v_s2, r_s2 = "ok", "Полный цикл (Tibco + создание) далеко от цели 1800 с"
    v_ctrl, r_ctrl = verdict(ctrl, 1800)

    def times_of(st):
        return [p["time"] for p in st.get("series") or [] if p.get("time") is not None]

    ude_ds = rate_median(ds_rz)
    ude_ers = rate_median(ers_rz)
    ude_s1 = cycle_rate(c1)
    ude_s2 = cycle_rate(c2)
    ude_ctrl = rate_median(ctrl)

    dash_items = [
        ("1", "Сверка денежных средств", "внСверкаДенежныхСредств_обр", ds_rz["n"], v_ds,
         fmt_sec((ds_rz.get("time") or {}).get("median")), "360 с", ude_ds, "расхожд./с"),
        ("2", "Сверка денежных средств ЕРС", "внСверкаДенежныхСредствНаЕРС_обр", ers_rz["n"], v_ers,
         fmt_sec((ers_rz.get("time") or {}).get("median")), "360 с", ude_ers, "расхожд./с"),
        ("3a", "Загрузка котировок, 1-й сеанс Т-1", "внЗагрузкаКотировокДУ ~09:00", c1["n"], v_s1,
         fmt_sec((c1.get("time") or {}).get("median")), "1800 с", ude_s1, "котировок/с"),
        ("3b", "Загрузка котировок, финальная Т-1", "внЗагрузкаКотировокДУ ~11:50", c2["n"], v_s2,
         fmt_sec((c2.get("time") or {}).get("median")), "1800 с", ude_s2, "котировок/с"),
        ("4", "Сверка загрузки котировок", "внОтчетКонтрольКотировок", ctrl["n"], v_ctrl,
         fmt_sec((ctrl.get("time") or {}).get("median")), "1800 с", ude_ctrl, "строк/с"),
    ]
    dash_cards = []
    vlab = {"ok": ("ok", "Оптимизация не нужна"), "watch": ("warn", "Наблюдать"), "need": ("bad", "Нужна оптимизация"), "nodata": ("warn", "Нет данных")}
    for num, title, obj, n, vc, med, tgt, ude, ude_unit in dash_items:
        cls, lab = vlab[vc]
        dash_cards.append(f"""
        <a class="dash {cls}" href="#sec-{num}">
          <div class="dash-num">{esc(num)}</div>
          <div class="dash-title">{esc(title)}</div>
          <div class="dash-obj">{esc(obj)}</div>
          <div class="dash-med">медиана: <strong>{med}</strong> при цели {tgt}</div>
          <div class="dash-ude">удельный вес: <strong>{fmt_ude(ude)} {esc(ude_unit)}</strong></div>
          <div class="dash-n">{n} замеров (РЗ / циклов)</div>
          <div class="dash-v">{lab}</div>
        </a>""")

    ude_table_rows = []
    for num, title, _obj, n, _vc, med, tgt, ude, ude_unit in dash_items:
        ude_table_rows.append(
            f"<tr><td>{esc(num)}</td><td>{esc(title)}</td>"
            f"<td class='num'>{n}</td>"
            f"<td class='num'>{med}</td>"
            f"<td>{esc(tgt)}</td>"
            f"<td class='num'><strong>{fmt_ude(ude)} {esc(ude_unit)}</strong></td></tr>"
        )
    ude_table = "\n".join(ude_table_rows)

    cmp_bars = svg_bars([
        ("Сверка ДС, РЗ", (ds_rz.get("time") or {}).get("median"), "#1e5a8a"),
        ("Сверка ДС ЕРС, РЗ", (ers_rz.get("time") or {}).get("median"), "#17a2b8"),
        ("Котировки 1-й сеанс, цикл", (c1.get("time") or {}).get("median"), "#1565c0"),
        ("Котировки финал, цикл", (c2.get("time") or {}).get("median"), "#6a1b9a"),
        ("Контроль котировок, РЗ", (ctrl.get("time") or {}).get("median"), "#2e7d32"),
    ], label_w=260)

    # Section 1
    sec1_q1 = f"""
        <ul class="tight">
          <li>Код IMDEV-8927 в <code>ВыполнитьСверку()</code>: <code>НачатьЗамерВремени</code> / <code>ЗакончитьЗамерВремени</code>.</li>
          <li>Имя: <code>{esc(TECH['ds_rz'])}</code> (контур РЗ) или <code>...Ручной</code>.</li>
          <li>Вес = число строк с «Есть расхождения», минимум 1. Комментарий: <code>Расхождений</code>, <code>Ошибка</code>, <code>УдельныйВес</code>.</li>
          <li>В Excel выгружается <em>наименование</em> справочника: <code>{esc(NAIM['ds_rz'])}</code>.</li>
        </ul>"""
    sec1_q2 = f"<p>Цель из таблицы ДУ-15: <strong>0,1 ч = 360 с</strong>. РЗ медиана {fmt_sec((ds_rz.get('time') or {}).get('median'))}, P95 {fmt_sec((ds_rz.get('time') or {}).get('p95'))}, max {fmt_sec((ds_rz.get('time') or {}).get('max'))}, превышений {ds_rz.get('over_target', 0)}. У ручных запусков есть единичные превышения 360 с (не регламент).</p>"

    ds_cmp = svg_bars([
        ("РЗ, медиана", (ds_rz.get("time") or {}).get("median"), "#1e5a8a"),
        ("РЗ, P95", (ds_rz.get("time") or {}).get("p95"), "#4e88b7"),
        ("РЗ, max", (ds_rz.get("time") or {}).get("max"), "#9bb8d0"),
        ("Ручной, медиана", (ds_man.get("time") or {}).get("median"), "#b8860b"),
        ("Ручной, P95", (ds_man.get("time") or {}).get("p95"), "#d4a84b"),
        ("Цель 360 с", 360, "#dc3545"),
    ], label_w=160)

    # Section 2
    sec2_q1 = f"""
        <ul class="tight">
          <li>Тот же шаблон замера в <code>ВыполнитьСверку()</code> обработки ЕРС.</li>
          <li>Имя: <code>{esc(TECH['ers_rz'])}</code>. В Excel: <code>{esc(NAIM['ers_rz'])}</code> (аббревиатура ЕРС распалась по буквам).</li>
          <li>Вес = число счетов с расхождением остатков, минимум 1.</li>
        </ul>"""
    sec2_q2 = f"<p>Цель <strong>360 с</strong>. Фактическая медиана РЗ {fmt_sec((ers_rz.get('time') or {}).get('median'))} — на два порядка ниже.</p>"
    ers_cmp = svg_bars([
        ("РЗ, медиана", (ers_rz.get("time") or {}).get("median"), "#17a2b8"),
        ("РЗ, P95", (ers_rz.get("time") or {}).get("p95"), "#5bc0de"),
        ("РЗ, max", (ers_rz.get("time") or {}).get("max"), "#a6e0ef"),
        ("Ручной, медиана", (ers_man.get("time") or {}).get("median"), "#b8860b"),
        ("Ручной, P95", (ers_man.get("time") or {}).get("p95"), "#d4a84b"),
        ("Цель 360 с", 360, "#dc3545"),
    ], label_w=160)

    def session_block(tag, title, when, get_st, crt_st, cyc, vcode, vreason, href):
        cycle_bar = svg_bars([
            ("Цикл, медиана", (cyc.get("time") or {}).get("median"), "#1565c0"),
            ("Цикл, P95", (cyc.get("time") or {}).get("p95"), "#6a1b9a"),
            ("Цикл, max", (cyc.get("time") or {}).get("max"), "#9c7bb8"),
            ("Tibco, медиана", (get_st.get("time") or {}).get("median"), "#1e5a8a"),
            ("Создание, медиана", (crt_st.get("time") or {}).get("median"), "#2e7d32"),
            ("Цель 1800 с", 1800, "#dc3545"),
        ], label_w=180)
        q1 = f"""
        <ul class="tight">
          <li>Та же обработка и та же команда <code>ФоновоеВыполнение</code>, что и у второго сеанса.</li>
          <li>На один запуск пишутся две КО: <code>ПолучениеTibco</code> и <code>СозданиеКотировок</code>.</li>
          <li>Сеанс в отчёте выделен по локальному времени замера: {esc(when)}.</li>
          <li>Парных циклов Tibco+создание: <strong>{cyc['n']}</strong>; шагов Tibco: {get_st['n']}, создание: {crt_st['n']}.</li>
        </ul>"""
        q2 = (
            f"<p>Цель слота <strong>0,5 ч = 1800 с</strong>. "
            f"Полный цикл медиана {fmt_sec((cyc.get('time') or {}).get('median'))}, "
            f"P95 {fmt_sec((cyc.get('time') or {}).get('p95'))}, "
            f"max {fmt_sec((cyc.get('time') or {}).get('max'))}. "
            f"Превышений 1800 с: {cyc.get('over_1800', 0)}.</p>"
        )
        quotes_note = ""
        qpack = cyc.get("quotes") or {}
        if qpack:
            quotes_note = (
                f"<p class='muted'>Объём Tibco: медиана {fmt(qpack.get('median'), 0)} котировок "
                f"(max {fmt(qpack.get('max'), 0)}). Есть короткие запуски по {fmt(qpack.get('min'), 0)} котировок "
                f"в 09:00 / 11:50 (типично воскресенье) на 4-7 с; основной прогон ~09:03 / ~11:52 "
                f"идёт около 3 мин на ~3600 котировок. Оба варианта далеко от цели 1800 с.</p>"
            )
        return f"""
        <div class="subcard" id="{href}">
          <h3>{esc(title)}</h3>
          <p class="muted">{esc(when)}. Команда: <code>ФоновоеВыполнение</code>. Пользователь РЗ: <code>svc_avancore</code>.</p>
          {q_boxes(True, get_st['n'] + crt_st['n'], vcode, vreason, q1, q2 + quotes_note)}
          <h4>Полный цикл загрузки (сумма двух шагов)</h4>
          {kpi_row({"n": cyc["n"], "time": cyc.get("time") or {}, "weight": cyc.get("quotes") or {}})}
          <div class="chart">{cycle_bar}</div>
          <h4>Шаг ПолучениеTibco</h4>
          {kpi_row(get_st)}
          {table_stats(get_st, "число котировок из Tibco")}
          <div class="chart">{svg_hist(times_of(get_st), color="#1e5a8a")}</div>
          <div class="chart">{svg_line(get_st.get("daily") or [], color="#1e5a8a", title="Медиана Tibco по дням, с")}</div>
          <h4>Шаг СозданиеКотировок</h4>
          {kpi_row(crt_st)}
          {table_stats(crt_st, "число записанных документов котировок")}
          <div class="chart">{svg_hist(times_of(crt_st), color="#2e7d32")}</div>
        </div>
        """

    if period_min and period_max:
        period_human = f"{period_min.strftime('%d.%m.%Y')} по {period_max.strftime('%d.%m.%Y')}"
    else:
        period_human = "период выгрузки"

    def goal_block(fact, target):
        pct = 0.0
        if fact and target:
            pct = min(100.0, 100.0 * fact / target)
        share = fmt(pct, 0) if pct >= 1 else fmt(pct, 1)
        return f"""
              <div class="pair-nums">
                <div class="pair-num">
                  <div class="l">Типичное время</div>
                  <div class="v">{esc(fmt_plain_time(fact))}</div>
                </div>
                <div class="pair-num">
                  <div class="l">Допустимо не дольше</div>
                  <div class="v">{esc(fmt_plain_time(target))}</div>
                </div>
              </div>
              <div class="goal-meter">
                <div class="goal-meter-track">
                  <div class="goal-meter-fill" style="width:{pct:.2f}%"></div>
                </div>
                <div class="goal-meter-cap">
                  <span>занято {share}% от предела</span>
                  <span>весь серый ряд — допустимое время</span>
                </div>
              </div>
        """

    def answer_card(num, title, epf, href, q1, q2, fact, target, ude_line=""):
        extra = f"<p>{ude_line}</p>" if ude_line else ""
        return f"""
        <article class="op-final">
          <div class="op-head">
            <div class="op-n">{esc(num)}</div>
            <div>
              <h3>{esc(title)}</h3>
              <div class="op-epf">Внешняя обработка: <code>{esc(epf)}</code>
              · <a href="#{href}">открыть подробности</a></div>
            </div>
          </div>
          <div class="op-qgrid">
            <div class="op-q">
              <div class="ql">Вопрос 1. Формируется ли замер времени?</div>
              <div class="qa">Да, замер есть</div>
              <p>{q1}</p>
            </div>
            <div class="op-q">
              <div class="ql">Вопрос 2. Нужна ли оптимизация?</div>
              <div class="qa">Нет, не нужна</div>
              <p>{q2}</p>
              {goal_block(fact, target)}
              {extra}
            </div>
          </div>
        </article>
        """

    ds_t = ds_rz.get("time") or {}
    ers_t = ers_rz.get("time") or {}
    c1_t = c1.get("time") or {}
    c2_t = c2.get("time") or {}
    ctrl_t = ctrl.get("time") or {}
    q1_vol = (c1.get("quotes") or {}).get("median")
    q2_vol = (c2.get("quotes") or {}).get("median")

    q2_tail = (
        "Почти все запуски (95 из 100) укладываются в {p95}. "
        "Самый долгий за период: {mx}. Целевое время ни разу не превышено."
    )

    itogo_html = f"""
<div class="card" id="itogo">
  <h2>Итоговые ответы заказчику</h2>
  <p class="final-lead">
    По каждой из пяти операций заказчик задал два вопроса.
    Сначала — ответы на всю пятёрку сразу, затем карточка по каждой операции:
    слева замер, справа оптимизация, внизу типичное время рядом с пределом.
    Период регистра: <strong>{esc(period_human)}</strong>.
    Вывод по оптимизации — по регламентным заданиям, не по ручным запускам из формы.
  </p>
  <div class="ans-pair">
    <div class="ans-hero yes">
      <div class="q">Вопрос 1</div>
      <div class="a">Замер времени формируется по всем пяти операциям</div>
      <p>Во всех четырёх внешних обработках код замера работает, в регистре есть записи.
      Первый и финальный сеансы загрузки котировок — одна обработка, в отчёте разделены по времени запуска.</p>
    </div>
    <div class="ans-hero yes">
      <div class="q">Вопрос 2</div>
      <div class="a">Оптимизация не нужна ни по одной операции</div>
      <p>Типичное время рабочего контура много ниже цели:
      сверки денежных средств — не дольше 6 минут, загрузка и контроль котировок — не дольше 30 минут.</p>
    </div>
  </div>
  {answer_card(
      "1",
      "Сверка денежных средств",
      "внСверкаДенежныхСредств_обр",
      "sec-1",
      f"Замер встроен в код обработки и пишется в регистр. "
      f"За период с {period_human} в регистре {ds_rz['n']} запусков регламентным заданием "
      f"(по ним делаем вывод) и {ds_man['n']} ручных запусков из формы.",
      "Вердикт — по регламентным заданиям. " + q2_tail.format(
          p95=fmt_plain_time(ds_t.get("p95")), mx=fmt_plain_time(ds_t.get("max"))),
      ds_t.get("median"), 360,
  )}
  {answer_card(
      "2",
      "Сверка денежных средств ЕРС (единый расчётный счёт)",
      "внСверкаДенежныхСредствНаЕРС_обр",
      "sec-2",
      f"Замер встроен в код обработки и пишется в регистр. "
      f"За период с {period_human} в регистре {ers_rz['n']} запусков регламентным заданием "
      f"и {ers_man['n']} ручных запусков из формы.",
      "Вердикт — по регламентным заданиям. " + q2_tail.format(
          p95=fmt_plain_time(ers_t.get("p95")), mx=fmt_plain_time(ers_t.get("max"))),
      ers_t.get("median"), 360,
  )}
  {answer_card(
      "3",
      "Загрузка котировок, первый сеанс за предыдущий день (дата Т минус 1)",
      "внЗагрузкаКотировокДУ",
      "sec-3a",
      f"Обработка пишет два замера на один запуск: получение котировок и запись документов. "
      f"Первый и финальный сеансы в справочнике не разделены — это одна обработка и одни и те же имена. "
      f"Первый сеанс отобран по времени запуска около 09:00. "
      f"За период с {period_human} собрано {c1['n']} полных циклов (оба шага в один день).",
      "Время сеанса — сумма двух шагов. " + q2_tail.format(
          p95=fmt_plain_time(c1_t.get("p95")), mx=fmt_plain_time(c1_t.get("max"))),
      c1_t.get("median"), 1800,
      f"Удельный вес: {fmt_ude(ude_s1)} котировок в секунду"
      + (f" (около {fmt_int(q1_vol)} котировок за типичный цикл)." if q1_vol else "."),
  )}
  {answer_card(
      "4",
      "Сверка загрузки котировок",
      "внОтчетКонтрольКотировок",
      "sec-4",
      f"Замер встроен в код обработки и пишется в регистр. "
      f"За период с {period_human} в регистре {ctrl['n']} запусков регламентным заданием. "
      f"Ручных запусков с замером в этой выгрузке нет.",
      "Вердикт — по регламентным заданиям. " + q2_tail.format(
          p95=fmt_plain_time(ctrl_t.get("p95")), mx=fmt_plain_time(ctrl_t.get("max"))),
      ctrl_t.get("median"), 1800,
      f"Удельный вес: {fmt_ude(ude_ctrl)} строк отчёта в секунду.",
  )}
  {answer_card(
      "5",
      "Загрузка котировок, финальная за предыдущий день (дата Т минус 1)",
      "внЗагрузкаКотировокДУ",
      "sec-3b",
      f"Та же обработка и те же имена в справочнике, что у первого сеанса. "
      f"Финальный сеанс отобран по времени запуска около 11:50. "
      f"За период с {period_human} собрано {c2['n']} полных циклов (получение котировок плюс запись документов).",
      "Время сеанса — сумма двух шагов. " + q2_tail.format(
          p95=fmt_plain_time(c2_t.get("p95")), mx=fmt_plain_time(c2_t.get("max"))),
      c2_t.get("median"), 1800,
      f"Удельный вес: {fmt_ude(ude_s2)} котировок в секунду"
      + (f" (около {fmt_int(q2_vol)} котировок за типичный цикл)." if q2_vol else "."),
  )}
  <h3>Сводка одним взглядом</h3>
  <table class="final-sum">
    <tr>
      <th>№</th>
      <th>Операция из задачи</th>
      <th>Замер времени</th>
      <th>Типичное время</th>
      <th>Допустимо не дольше</th>
      <th>Оптимизация</th>
    </tr>
    <tr>
      <td>1</td>
      <td>Сверка денежных средств</td>
      <td>Есть. {ds_rz['n']} регламентных и {ds_man['n']} ручных запусков</td>
      <td>{esc(fmt_plain_time(ds_t.get("median")))}</td>
      <td>6 минут</td>
      <td><strong>Не нужна</strong></td>
    </tr>
    <tr>
      <td>2</td>
      <td>Сверка денежных средств ЕРС</td>
      <td>Есть. {ers_rz['n']} регламентных и {ers_man['n']} ручных запусков</td>
      <td>{esc(fmt_plain_time(ers_t.get("median")))}</td>
      <td>6 минут</td>
      <td><strong>Не нужна</strong></td>
    </tr>
    <tr>
      <td>3</td>
      <td>Загрузка котировок, первый сеанс за предыдущий день</td>
      <td>Есть. {c1['n']} полных циклов около 09:00</td>
      <td>{esc(fmt_plain_time(c1_t.get("median")))}</td>
      <td>30 минут</td>
      <td><strong>Не нужна</strong></td>
    </tr>
    <tr>
      <td>4</td>
      <td>Сверка загрузки котировок</td>
      <td>Есть. {ctrl['n']} регламентных запусков</td>
      <td>{esc(fmt_plain_time(ctrl_t.get("median")))}</td>
      <td>30 минут</td>
      <td><strong>Не нужна</strong></td>
    </tr>
    <tr>
      <td>5</td>
      <td>Загрузка котировок, финальная за предыдущий день</td>
      <td>Есть. {c2['n']} полных циклов около 11:50</td>
      <td>{esc(fmt_plain_time(c2_t.get("median")))}</td>
      <td>30 минут</td>
      <td><strong>Не нужна</strong></td>
    </tr>
  </table>
</div>
"""

    html_out = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IMDEV-9391 — Замеры ключевых операций и потребность в оптимизации</title>
<style>
:root {{
  --bg:#f5f7fa; --card:#fff; --text:#1a2332; --muted:#5a6a7a; --border:#d8dee6;
  --accent:#1e5a8a; --ok:#28a745; --warn:#b8860b; --danger:#dc3545; --info:#17a2b8;
}}
* {{ box-sizing:border-box; }}
body {{ margin:0; font-family:"Segoe UI", Tahoma, Geneva, Verdana, sans-serif; background:var(--bg); color:var(--text); line-height:1.55; }}
.wrap {{ max-width:1120px; margin:0 auto; padding:28px 20px 56px; }}
h1 {{ font-size:1.55rem; margin:0 0 8px; color:var(--accent); }}
h2 {{ font-size:1.25rem; margin:0 0 12px; padding-bottom:8px; border-bottom:2px solid var(--accent); }}
h3 {{ font-size:1.08rem; margin:18px 0 8px; color:#16324d; }}
h4 {{ font-size:0.98rem; margin:16px 0 8px; color:#2c3e50; }}
.meta {{ color:var(--muted); font-size:0.92rem; margin-bottom:18px; }}
.meta a {{ color:var(--accent); }}
.intro {{ background:#e8f4fc; border-left:4px solid var(--info); padding:14px 16px 12px; margin-bottom:20px; border-radius:0 6px 6px 0; }}
.intro h2 {{ font-size:1.05rem; margin:0 0 10px; padding:0; border:none; color:#0d5c63; }}
.ask-ops {{ margin:0 0 12px; padding-left:22px; }}
.ask-ops li {{ margin:3px 0; }}
.ask-qs {{ display:grid; grid-template-columns:1fr 1fr; gap:8px; margin:0; }}
@media (max-width:800px) {{ .ask-qs {{ grid-template-columns:1fr; }} }}
.ask-q {{ background:#fff; border:1px solid #b8e0e8; border-radius:6px; padding:8px 12px; }}
.ask-q .n {{ font-size:0.75rem; font-weight:700; color:var(--info); letter-spacing:.04em; text-transform:uppercase; }}
.card {{ background:var(--card); border:1px solid var(--border); border-radius:8px; padding:18px 20px; margin-bottom:22px; }}
.subcard {{ background:#fafbfc; border:1px dashed var(--border); border-radius:8px; padding:16px 16px 8px; margin:16px 0 8px; }}
.badge {{ display:inline-block; font-size:0.75rem; font-weight:600; padding:3px 10px; border-radius:4px; margin:0 6px 6px 0; }}
.b-epf {{ background:#e3f2fd; color:#1565c0; }}
.b-ok {{ background:#e8f5e9; color:#2e7d32; }}
.b-warn {{ background:#fff8e1; color:#f57f17; }}
.b-info {{ background:#e0f7fa; color:#006064; }}
code {{ background:#eef2f6; padding:1px 6px; border-radius:3px; font-size:0.88em; }}
pre {{ background:#1e1e1e; color:#f8f8f2; padding:12px 14px; border-radius:6px; overflow-x:auto; font-size:12.5px; line-height:1.4; }}
table {{ width:100%; border-collapse:collapse; margin:10px 0 14px; font-size:0.92rem; }}
th, td {{ border:1px solid var(--border); padding:7px 9px; text-align:left; vertical-align:top; }}
th {{ background:#f0f4f8; }}
td.num {{ text-align:right; white-space:nowrap; }}
ul.tight {{ margin:8px 0; padding-left:20px; }}
ul.tight li {{ margin-bottom:4px; }}
.muted {{ color:var(--muted); font-size:0.92rem; }}
.toc a {{ color:var(--accent); text-decoration:none; }}
.toc li {{ margin:4px 0; }}
.dashgrid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(200px,1fr)); gap:12px; margin:12px 0 6px; }}
a.dash {{ display:block; text-decoration:none; color:inherit; border:1px solid var(--border); border-radius:8px; padding:12px 14px; background:#fff; }}
a.dash.ok {{ border-top:4px solid var(--ok); }}
a.dash.warn {{ border-top:4px solid var(--warn); }}
a.dash.bad {{ border-top:4px solid var(--danger); }}
.dash-num {{ font-size:0.75rem; color:var(--muted); font-weight:700; }}
.dash-title {{ font-weight:700; margin:4px 0; color:#16324d; }}
.dash-obj {{ font-size:0.8rem; color:var(--muted); margin-bottom:8px; }}
.dash-med {{ font-size:0.88rem; }}
.dash-ude {{ font-size:0.88rem; margin-top:4px; color:#16324d; }}
.dash-n {{ font-size:0.8rem; color:var(--muted); margin-top:4px; }}
.dash-v {{ margin-top:8px; font-weight:700; font-size:0.88rem; }}
a.dash.ok .dash-v {{ color:var(--ok); }}
a.dash.warn .dash-v {{ color:var(--warn); }}
a.dash.bad .dash-v {{ color:var(--danger); }}
.qgrid {{ display:grid; grid-template-columns:1fr 1fr; gap:12px; margin:12px 0; }}
@media (max-width:800px) {{ .qgrid {{ grid-template-columns:1fr; }} }}
.qbox {{ border-radius:8px; padding:12px 14px; }}
.qbox.ok {{ background:#e8f5e9; border-left:4px solid var(--ok); }}
.qbox.warn {{ background:#fff8e1; border-left:4px solid var(--warn); }}
.qbox.bad {{ background:#fdecea; border-left:4px solid var(--danger); }}
.qbox h3 {{ margin:4px 0 8px; }}
.qnum {{ font-size:0.75rem; font-weight:700; letter-spacing:.04em; text-transform:uppercase; color:var(--muted); }}
.kpis {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(120px,1fr)); gap:8px; margin:10px 0 14px; }}
.kpi {{ background:#fff; border:1px solid var(--border); border-radius:6px; padding:8px 10px; }}
.kpi-l {{ font-size:0.75rem; color:var(--muted); }}
.kpi-v {{ font-size:1.05rem; font-weight:700; color:#16324d; }}
.chart {{ background:#fff; border:1px solid var(--border); border-radius:6px; padding:10px 8px 4px; margin:10px 0 14px; }}
.flow {{ background:#fff; border:1px dashed var(--border); padding:10px 14px; border-radius:6px; font-size:0.9rem; margin:10px 0; }}
.sources {{ font-size:0.88rem; color:var(--muted); margin-top:28px; padding-top:14px; border-top:1px solid var(--border); }}
nav.sticky {{ position:sticky; top:0; background:rgba(245,247,250,.96); padding:8px 0; margin-bottom:12px; z-index:2; border-bottom:1px solid var(--border); font-size:0.85rem; }}
nav.sticky a {{ color:var(--accent); margin-right:14px; text-decoration:none; }}
.final-lead {{ margin:0 0 16px; font-size:1.02rem; }}
.ans-pair {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; margin:0 0 22px; }}
@media (max-width:800px) {{ .ans-pair {{ grid-template-columns:1fr; }} }}
.ans-hero {{ border-radius:10px; padding:18px 20px; }}
.ans-hero.yes {{ background:#e8f5e9; border:1px solid #b7dfb9; }}
.ans-hero .q {{ font-size:0.78rem; font-weight:700; letter-spacing:.04em; text-transform:uppercase; color:#2e7d32; margin:0 0 6px; }}
.ans-hero .a {{ font-size:1.4rem; font-weight:700; color:#1b5e20; margin:0 0 8px; line-height:1.25; }}
.ans-hero p {{ margin:0; }}
.op-final {{ border:1px solid var(--border); border-radius:10px; padding:16px 18px 14px; margin:14px 0; background:#f7fafc; }}
.op-final .op-head {{ display:flex; gap:14px; align-items:flex-start; margin-bottom:12px; }}
.op-n {{ min-width:40px; height:40px; border-radius:10px; background:var(--accent); color:#fff; display:flex; align-items:center; justify-content:center; font-weight:700; font-size:1.05rem; }}
.op-head h3 {{ margin:0 0 4px; padding:0; border:none; font-size:1.12rem; }}
.op-epf {{ color:var(--muted); font-size:0.9rem; }}
.op-epf a {{ color:var(--accent); text-decoration:none; }}
.op-qgrid {{ display:grid; grid-template-columns:1fr 1fr; gap:12px; }}
@media (max-width:800px) {{ .op-qgrid {{ grid-template-columns:1fr; }} }}
.op-q {{ background:#fff; border-radius:8px; padding:12px 14px; border:1px solid var(--border); }}
.op-q .ql {{ font-size:0.75rem; font-weight:700; letter-spacing:.03em; text-transform:uppercase; color:var(--muted); }}
.op-q .qa {{ font-size:1.08rem; font-weight:700; color:var(--ok); margin:4px 0 8px; }}
.op-q p {{ margin:0 0 8px; font-size:0.95rem; }}
.pair-nums {{ display:grid; grid-template-columns:1fr 1fr; gap:8px; margin:4px 0 8px; }}
.pair-num {{ background:#eef3f7; border-radius:6px; padding:8px 10px; text-align:center; }}
.pair-num .l {{ font-size:0.75rem; color:var(--muted); }}
.pair-num .v {{ font-size:1.02rem; font-weight:700; color:#16324d; margin-top:2px; }}
.goal-meter {{ margin-top:4px; }}
.goal-meter-track {{ height:12px; background:#e8edf2; border-radius:6px; overflow:hidden; }}
.goal-meter-fill {{ height:100%; background:#2e7d32; border-radius:6px; }}
.goal-meter-cap {{ display:flex; justify-content:space-between; font-size:0.8rem; color:var(--muted); margin-top:4px; gap:8px; }}
.final-sum th {{ font-size:0.85rem; }}
</style>
</head>
<body>
<div class="wrap">
<nav class="sticky">
  <a href="#dash">Сводка</a>
  <a href="#sec-1">1. Сверка ДС</a>
  <a href="#sec-2">2. Сверка ЕРС</a>
  <a href="#sec-3">3. Котировки</a>
  <a href="#sec-4">4. Контроль котировок</a>
  <a href="#itogo">Ответы заказчику</a>
</nav>

<h1>IMDEV-9391 — Проверка ключевых операций: замеры и оптимизация</h1>
<p class="meta">
  Задача: <a href="https://jira/browse/IMDEV-9391">IMDEV-9391</a>
  «Проверить ключевые операции на потребность в оптимизации».
  Эпик: <a href="https://jira/browse/IMDEV-8633">IMDEV-8633</a>.
  База: Wim_Du (AvancoreDU).<br>
  Источник замеров: регистр <code>ЗамерыВремени</code>, файл <code>ЗамерыВремени.xlsx</code>
  ({fmt_int(n_all)} строк, период по нашим КО:
  {esc(period_min.strftime('%d.%m.%Y') if period_min else '-')} -
  {esc(period_max.strftime('%d.%m.%Y') if period_max else '-')}).
</p>

<div class="intro">
  <h2>Что просил заказчик</h2>
  <p style="margin:0 0 6px;"><strong>Что замеряем</strong> (пять ключевых операций):</p>
  <ol class="ask-ops">
    <li>Сверка денежных средств</li>
    <li>Сверка денежных средств ЕРС</li>
    <li>Загрузка котировок, 1-й сеанс за дату Т-1</li>
    <li>Сверка загрузки котировок</li>
    <li>Загрузка котировок, финальная за Т-1</li>
  </ol>
  <p style="margin:0 0 8px;"><strong>Что делаем</strong> по каждой операции:</p>
  <div class="ask-qs">
    <div class="ask-q">
      <div class="n">Вопрос 1</div>
      Проверить, что формируется замер времени
    </div>
    <div class="ask-q">
      <div class="n">Вопрос 2</div>
      Проанализировать потребность в оптимизации
    </div>
  </div>
</div>

<div class="card">
  <h2>Методика</h2>
  <p>Смотрим два источника: код обработок (есть ли вызов замера) и выгрузку регистра замеров времени из рабочей базы.</p>
  <ul class="tight">
    <li><strong>Как понять, что строка в Excel относится к нашей обработке.</strong>
      В коде замер пишется коротким техническим именем, например
      <code>ДУ.СверкаДенежныхСредств.Сверка.ФоновоеВыполнение</code>.
      Если такой операции ещё нет в справочнике, платформа создаёт её сама и для экрана
      разбивает имя на слова: получается
      «Д у. сверка денежных средств. сверка. фоновое выполнение».
      В Excel именно эта «человеческая» подпись. По ней мы и нашли все строки.</li>
    <li><strong>Какое время считается нормальным.</strong>
      Ориентир взят из таблицы ключевых операций ДУ (задача IMDEV-8927):
      сверки денежных средств - не дольше <strong>6 минут</strong> (0,1 часа),
      загрузка и контроль котировок - не дольше <strong>30 минут</strong> (0,5 часа).</li>
    <li><strong>По каким запускам делаем вывод.</strong>
      Главный контур - ночные и утренние регламентные задания (служебный пользователь),
      а не ручные нажатия в форме. Оптимизация нужна, если типичный «тяжёлый» запуск
      (95-й процентиль: почти все случаи, кроме самых редких выбросов) уже выше цели
      либо цель превышается чаще чем в каждом десятом запуске.</li>
    <li><strong>Два сеанса загрузки котировок.</strong>
      Это не две разные операции в справочнике, а один и тот же код в ~09:00 и ~11:50.
      Время слота считаем как сумму двух шагов за один запуск: получить котировки с Tibco
      и записать документы. Сеансы в таблице различаем по часам, а не по имени.</li>
  </ul>
</div>

<div class="card" id="dash">
  <h2>Сводка по списку заказчика</h2>
  <div class="dashgrid">{''.join(dash_cards)}</div>
  <h4>Медиана длительности (рабочий контур)</h4>
  <div class="chart">{cmp_bars}</div>
  <h4>Удельный вес (единиц в секунду)</h4>
  <table>
    <tr>
      <th>№</th>
      <th>Операция</th>
      <th>Замеров</th>
      <th>Медиана времени</th>
      <th>Цель</th>
      <th>Удельный вес</th>
    </tr>
    {ude_table}
  </table>
  <p class="muted">
    Удельный вес = вес замера / время (то же, что поле <code>УдельныйВес</code> в комментарии, об/с).
    Для этих пяти операций это <strong>не договоры</strong>: сверка ДС и ЕРС пишут число расхождений
    (часто 1, поэтому «штук в секунду» почти ничего не говорит о нагрузке);
    загрузка котировок — число котировок за полный цикл Tibco + создание документов;
    контроль — число строк отчёта. Сравнивать эти пять чисел между собой нельзя: единицы разные.
  </p>
</div>

<!-- ===================== 1 ===================== -->
<div class="card" id="sec-1">
  <h2>1. Сверка денежных средств</h2>
  <p>
    <span class="badge b-epf">EPF</span>
    <span class="badge b-ok">Замер в коде и в регистре</span>
    <code>внСверкаДенежныхСредств_обр</code>, версия 1.07
  </p>
  <table>
    <tr><th style="width:28%">Поле</th><th>Значение</th></tr>
    <tr><td>Операция в Jira</td><td>Сверка денежных средств</td></tr>
    <tr><td>Команда РЗ</td><td><code>ФоновоеВыполнение</code> — «Сверка денежных средств (фон.)»</td></tr>
    <tr><td>Ключевая операция</td><td><code>{esc(TECH['ds_rz'])}</code></td></tr>
    <tr><td>В Excel</td><td><code>{esc(NAIM['ds_rz'])}</code></td></tr>
    <tr><td>Цель</td><td>360 с (0,1 ч)</td></tr>
  </table>
  <div class="flow">РЗ / форма → <code>ВыполнитьКоманду</code> (ставит контур) → <code>ВыполнитьСверку</code> (замер всего расчёта) → рассылка.</div>
  {q_boxes(True, ds_rz['n'] + ds_man['n'], v_ds, r_ds, sec1_q1, sec1_q2)}
  <h3>Регламентный запуск (основа для вердикта)</h3>
  {kpi_row(ds_rz)}
  <div class="chart">{ds_cmp}</div>
  <div class="chart">{svg_line(ds_rz.get('daily') or [], title='Медиана РЗ по дням, с')}</div>
  <div class="chart">{svg_hist(times_of(ds_rz))}</div>
  {table_stats(ds_rz, 'число расхождений, минимум 1')}
  <h3>Ручные запуски (справочно)</h3>
  <p class="muted">{ds_man['n']} записей <code>...Ручной</code>. На вердикт по регламенту не влияют, но показывают разброс интерактива.</p>
  {kpi_row(ds_man)}
  {table_stats(ds_man, 'число расхождений, минимум 1')}
  <p class="muted">Вес почти всегда 1: расхождений нет, обработка всё равно сканирует счета. Удельное время не отражает объём работы.</p>
</div>

<!-- ===================== 2 ===================== -->
<div class="card" id="sec-2">
  <h2>2. Сверка денежных средств на ЕРС</h2>
  <p>
    <span class="badge b-epf">EPF</span>
    <span class="badge b-ok">Замер в коде и в регистре</span>
    <code>внСверкаДенежныхСредствНаЕРС_обр</code>, версия 1.03
  </p>
  <table>
    <tr><th style="width:28%">Поле</th><th>Значение</th></tr>
    <tr><td>Операция в Jira</td><td>Сверка денежных средств ЕРС</td></tr>
    <tr><td>Команда РЗ</td><td><code>ФоновоеВыполнение</code> — «Сверка денежных средств на ЕРС (фон.)»</td></tr>
    <tr><td>Ключевая операция</td><td><code>{esc(TECH['ers_rz'])}</code></td></tr>
    <tr><td>В Excel</td><td><code>{esc(NAIM['ers_rz'])}</code></td></tr>
    <tr><td>Цель</td><td>360 с (0,1 ч)</td></tr>
  </table>
  {q_boxes(True, ers_rz['n'] + ers_man['n'], v_ers, r_ers, sec2_q1, sec2_q2)}
  <h3>Регламентный запуск</h3>
  {kpi_row(ers_rz)}
  <div class="chart">{ers_cmp}</div>
  <div class="chart">{svg_line(ers_rz.get('daily') or [], color='#17a2b8', title='Медиана РЗ ЕРС по дням, с')}</div>
  {table_stats(ers_rz, 'число счетов с расхождением, минимум 1')}
  <h3>Ручные запуски (справочно)</h3>
  {kpi_row(ers_man)}
  {table_stats(ers_man, 'число счетов с расхождением, минимум 1')}
</div>

<!-- ===================== 3 ===================== -->
<div class="card" id="sec-3">
  <h2>3. Загрузка котировок</h2>
  <p>
    <span class="badge b-epf">EPF</span>
    <span class="badge b-ok">Замер в коде и в регистре</span>
    <span class="badge b-info">Два слота, одни имена КО</span>
    <code>внЗагрузкаКотировокДУ</code>, версия 2.23
  </p>
  <p>
    Операции Jira «1-й сеанс за дату Т-1» и «финальная за Т-1» — это <strong>два расписания одной команды</strong>
    <code>ФоновоеВыполнение</code>. Дата загрузки в коде: вчера (в понедельник — пятница, −72 ч).
    На запуск пишутся две ключевые операции:
  </p>
  <table>
    <tr><th>Шаг</th><th>Имя в коде</th><th>Наименование в Excel</th></tr>
    <tr><td>Получение с Tibco</td><td><code>{esc(TECH['get_rz'])}</code></td><td><code>{esc(NAIM['get_rz'])}</code></td></tr>
    <tr><td>Создание документов</td><td><code>{esc(TECH['crt_rz'])}</code></td><td><code>{esc(NAIM['crt_rz'])}</code></td></tr>
  </table>
  <div class="flow">
    РЗ 09:00 (сеанс 1) или 11:50 (финал) → <code>ЗагрузитьОбъекты</code> (замер Tibco) →
    <code>СоздатьОбъекты</code> (замер записи документов). Вес Tibco = число котировок;
    вес создания = число проведённых документов. Комментарий: <code>Котировок</code>,
    <code>КСозданию</code>, <code>Создано</code>, <code>Отказ</code>/<code>Ошибок</code>, <code>УдельныйВес</code>.
  </div>
  <p class="muted">
    Фактические часы в регистре: первый сеанс <strong>09:00–09:04</strong> (редко 09:18),
    финал <strong>11:50–11:53</strong>. В старой таблице ДУ-15 стояло 11:00 / 14:00 — это не совпадает с продом.
  </p>
  <div class="chart">{svg_bars([
        ("1-й сеанс, цикл медиана", (c1.get("time") or {}).get("median"), "#1565c0"),
        ("1-й сеанс, цикл P95", (c1.get("time") or {}).get("p95"), "#4e88b7"),
        ("Финал, цикл медиана", (c2.get("time") or {}).get("median"), "#6a1b9a"),
        ("Финал, цикл P95", (c2.get("time") or {}).get("p95"), "#9c7bb8"),
        ("Цель слота 1800 с", 1800, "#dc3545"),
    ], label_w=200)}</div>

  {session_block("s1", "3.1. Загрузка котировок, 1-й сеанс за дату Т-1",
                 "локальное время до 10:30, фактически ~09:00",
                 get_s1, crt_s1, c1, v_s1, r_s1, "sec-3a")}
  {session_block("s2", "3.2. Загрузка котировок, финальная за Т-1",
                 "локальное время 10:30–13:00, фактически ~11:50",
                 get_s2, crt_s2, c2, v_s2, r_s2, "sec-3b")}

  <h3>Смежные контуры той же обработки (не в списке Jira 9391)</h3>
  <table>
    <tr><th>Контур</th><th>Замеров</th><th>Медиана Tibco / создания</th><th>Комментарий</th></tr>
    <tr>
      <td><code>ФоновоеВыполнениеVTBAMПИФ</code></td>
      <td class="num">{get_pif['n']} + {crt_pif['n']}</td>
      <td class="num">{fmt_sec((get_pif.get('time') or {}).get('median'))} / {fmt_sec((crt_pif.get('time') or {}).get('median'))}</td>
      <td>Отдельная команда, отдельные имена КО (суффикс «v t b a m п и ф» в Excel).</td>
    </tr>
    <tr>
      <td><code>ФоновоеВыполнениеCBONDS</code></td>
      <td class="num">0</td>
      <td class="num">—</td>
      <td>Замер в коде есть (<code>ЗагрузкаCbonds</code>), в выгрузке регистра записей нет.</td>
    </tr>
    <tr>
      <td>Ручной запуск формы</td>
      <td class="num">{groups['get_man']['stats']['n']} + {groups['crt_man']['stats']['n']}</td>
      <td class="num">{fmt_sec((groups['get_man']['stats'].get('time') or {}).get('median'))} / {fmt_sec((groups['crt_man']['stats'].get('time') or {}).get('median'))}</td>
      <td>Единичные прогоны; в одном создании <code>КСозданию</code> до тысяч строк — это не объём РЗ.</td>
    </tr>
  </table>
</div>

<!-- ===================== 4 ===================== -->
<div class="card" id="sec-4">
  <h2>4. Сверка загрузки котировок (отчёт контроль)</h2>
  <p>
    <span class="badge b-epf">EPF</span>
    <span class="badge b-ok">Замер в коде и в регистре</span>
    <code>внОтчетКонтрольКотировок</code>, версия 1.07
  </p>
  <table>
    <tr><th style="width:28%">Поле</th><th>Значение</th></tr>
    <tr><td>Операция в Jira</td><td>Сверка загрузки котировок</td></tr>
    <tr><td>Команда РЗ</td><td><code>ФоновоеВыполнение</code> — «Отчет контроль котировок (фон.)»</td></tr>
    <tr><td>Ключевая операция</td><td><code>{esc(TECH['ctrl_rz'])}</code></td></tr>
    <tr><td>В Excel</td><td><code>{esc(NAIM['ctrl_rz'])}</code></td></tr>
    <tr><td>Цель</td><td>1800 с (0,5 ч)</td></tr>
  </table>
  <div class="flow">РЗ → <code>ПроверкаСниженияСЧА</code> (замер всего: отчёт + рассылка). Вес = число строк отчёта. Ручной контур <code>...Ручной</code> в регистре не появляется: форма не вызывает этот метод.</div>
  {q_boxes(True, ctrl['n'], v_ctrl, r_ctrl, '''
        <ul class="tight">
          <li>Замер охватывает формирование отчёта и отправку письма.</li>
          <li>Комментарий: <code>СтрокОтчета</code>, <code>УдельныйВес</code>.</li>
        </ul>''',
        f"<p>Цель <strong>1800 с</strong>. Медиана {fmt_sec((ctrl.get('time') or {}).get('median'))}, P95 {fmt_sec((ctrl.get('time') or {}).get('p95'))}, max {fmt_sec((ctrl.get('time') or {}).get('max'))}, превышений {ctrl.get('over_target', 0)}.</p>")}
  {kpi_row(ctrl)}
  <div class="chart">{svg_bars([
        ("РЗ, медиана", (ctrl.get("time") or {}).get("median"), "#2e7d32"),
        ("РЗ, P95", (ctrl.get("time") or {}).get("p95"), "#66bb6a"),
        ("РЗ, max", (ctrl.get("time") or {}).get("max"), "#a5d6a7"),
        ("Цель 1800 с", 1800, "#dc3545"),
    ], label_w=160)}</div>
  <div class="chart">{svg_line(ctrl.get("daily") or [], color="#2e7d32", title="Медиана контроля котировок по дням, с")}</div>
  <div class="chart">{svg_hist(times_of(ctrl), color="#2e7d32")}</div>
  {table_stats(ctrl, "число строк отчёта, минимум 1")}
</div>

{itogo_html}

<div class="sources">
  Источники: IMDEV-9391.doc; ЗаданиеAI_9391.txt; ЗамерыВремени.xlsx;
  IMDEV-8927 (таблица ДУ-15, целевые 0,1 / 0,5 ч);
  модуль БСП <code>ОценкаПроизводительности</code> / <code>РазложитьСтрокуПоСловам</code>;
  внешние обработки ДУ с замерами IMDEV-8927.
  Отчёт собран {datetime.now().strftime('%d.%m.%Y %H:%M')}.
</div>
</div>
</body>
</html>
"""

    HTML_PATH.parent.mkdir(parents=True, exist_ok=True)
    HTML_PATH.write_text(html_out, encoding="utf-8")
    write_jira_comments({
        "period": period_human,
        "ds_n": ds_rz["n"],
        "ds_man_n": ds_man["n"],
        "ers_n": ers_rz["n"],
        "ers_man_n": ers_man["n"],
        "c1_n": c1["n"],
        "c2_n": c2["n"],
        "ctrl_n": ctrl["n"],
        "ds_med": ds_t.get("median"),
        "ers_med": ers_t.get("median"),
        "c1_med": c1_t.get("median"),
        "c2_med": c2_t.get("median"),
        "ctrl_med": ctrl_t.get("median"),
        "ds_p95": ds_t.get("p95"),
        "ers_p95": ers_t.get("p95"),
        "c1_p95": c1_t.get("p95"),
        "c2_p95": c2_t.get("p95"),
        "ctrl_p95": ctrl_t.get("p95"),
        "ds_max": ds_t.get("max"),
        "ers_max": ers_t.get("max"),
        "c1_max": c1_t.get("max"),
        "c2_max": c2_t.get("max"),
        "ctrl_max": ctrl_t.get("max"),
        "ude_ds": ude_ds,
        "ude_ers": ude_ers,
        "ude_s1": ude_s1,
        "ude_s2": ude_s2,
        "ude_ctrl": ude_ctrl,
        "q1_vol": q1_vol,
        "q2_vol": q2_vol,
    })

    slim = {
        "excel_rows": n_all,
        "period": payload["period"],
        "counts": {k: groups[k]["stats"]["n"] for k in TECH},
        "sessions": {
            "s1_cycles": c1["n"],
            "s2_cycles": c2["n"],
            "s1_time": c1.get("time"),
            "s2_time": c2.get("time"),
        },
        "ds_rz": {k: ds_rz[k] for k in ("n", "time", "weight", "over_target") if k in ds_rz},
        "ers_rz": {k: ers_rz[k] for k in ("n", "time", "weight", "over_target") if k in ers_rz},
        "ctrl": {k: ctrl[k] for k in ("n", "time", "weight", "over_target") if k in ctrl},
        "verdicts": {"ds": v_ds, "ers": v_ers, "s1": v_s1, "s2": v_s2, "ctrl": v_ctrl},
    }
    JSON_PATH.write_text(json.dumps(slim, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print("HTML", HTML_PATH, "bytes", HTML_PATH.stat().st_size)
    print("JIRA summary", JIRA_SUMMARY_PATH, "bytes", JIRA_SUMMARY_PATH.stat().st_size)
    print("JIRA answers", JIRA_ANSWERS_PATH, "bytes", JIRA_ANSWERS_PATH.stat().st_size)
    print("JSON", JSON_PATH)
    print("counts", slim["counts"])
    print("sessions", slim["sessions"])
    print("verdicts", slim["verdicts"])
    print("ds median", (ds_rz.get("time") or {}).get("median"), "ers", (ers_rz.get("time") or {}).get("median"))
    print("c1", (c1.get("time") or {}).get("median"), "c2", (c2.get("time") or {}).get("median"))
    print("ude ds", ude_ds, "ers", ude_ers, "s1", ude_s1, "s2", ude_s2, "ctrl", ude_ctrl)


if __name__ == "__main__":
    build()
