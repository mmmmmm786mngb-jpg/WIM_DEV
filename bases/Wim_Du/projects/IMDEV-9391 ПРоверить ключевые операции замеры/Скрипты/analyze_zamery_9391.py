#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Razbor registra zamerov vremeni dlya IMDEV-9391.
Imena KO v vygruzke 1C idut s probelami; daty starta - ms 1C;
lokalnoe vremya - v kolonke 'Data zapisi lokalnaya'; kommentarij - JSON BSP.
"""

import json
import math
import re
import statistics
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "ЗамерыВремени.xlsx"
OUT_DIR = ROOT / "Тестирование" / "reports"
OUT_JSON = OUT_DIR / "zamery_9391_stats.json"
OUT_TXT = OUT_DIR / "zamery_9391_stats.txt"

# Canonical code names after removing spaces and lowercasing
# Excel: "Д у. сверка денежных средств. сверка. фоновое выполнение"
# Code:  "ДУ.СверкаДенежныхСредств.Сверка.ФоновоеВыполнение"


def norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).lower()


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-", ".", "-."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_dt(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def percentile(sorted_vals, p):
    if not sorted_vals:
        return None
    if len(sorted_vals) == 1:
        return sorted_vals[0]
    k = (len(sorted_vals) - 1) * p / 100.0
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return sorted_vals[int(k)]
    return sorted_vals[f] * (c - k) + sorted_vals[c] * (k - f)


def pack(vals):
    vals = [v for v in vals if v is not None]
    if not vals:
        return None
    s = sorted(vals)
    return {
        "n": len(s),
        "min": round(s[0], 6),
        "p25": round(percentile(s, 25), 6),
        "median": round(percentile(s, 50), 6),
        "mean": round(statistics.fmean(s), 6),
        "p75": round(percentile(s, 75), 6),
        "p95": round(percentile(s, 95), 6),
        "max": round(s[-1], 6),
    }


def parse_comment(raw):
    extra = {}
    flags = []
    dop = ""
    meta = {}
    if not raw:
        return extra, flags, dop, meta
    text = str(raw)
    try:
        obj = json.loads(text)
        if isinstance(obj, dict):
            meta = {k: v for k, v in obj.items() if k != "ДопИнф"}
            dop_val = obj.get("ДопИнф", "")
            if isinstance(dop_val, dict):
                dop = json.dumps(dop_val, ensure_ascii=False)
                extra.update({str(k): str(v) for k, v in dop_val.items()})
            else:
                dop = str(dop_val)
        else:
            dop = text
    except json.JSONDecodeError:
        dop = text

    for part in re.split(r"[;\n]", dop):
        part = part.strip()
        if not part:
            continue
        if "=" in part:
            k, v = part.split("=", 1)
            extra[k.strip()] = v.strip()
        else:
            flags.append(part)
    return extra, flags, dop, meta


def classify_quotes_session(dt):
    """1st session ~09:00, final ~11:50 (local)."""
    if dt is None:
        return "unknown"
    h = dt.hour
    m = dt.minute
    hm = h * 60 + m
    if hm < 10 * 60 + 30:
        return "session1"  # until 10:30
    if hm < 13 * 60:
        return "session2"  # 10:30-13:00 ~ 11:50
    return "other"


# Matching after norm()
KO_GROUPS = {
    "ds_sverka_rz": {
        "title": "Сверка денежных средств / ФоновоеВыполнение",
        "jira": "Сверка денежных средств",
        "epf": "внСверкаДенежныхСредств_обр",
        "code": "ДУ.СверкаДенежныхСредств.Сверка.ФоновоеВыполнение",
        "match": lambda n: n == norm("ДУ.СверкаДенежныхСредств.Сверка.ФоновоеВыполнение"),
        "target": 360,
        "weight_meaning": "число расхождений (минимум 1)",
    },
    "ds_sverka_manual": {
        "title": "Сверка денежных средств / Ручной",
        "jira": "Сверка денежных средств",
        "epf": "внСверкаДенежныхСредств_обр",
        "code": "ДУ.СверкаДенежныхСредств.Сверка.Ручной",
        "match": lambda n: n == norm("ДУ.СверкаДенежныхСредств.Сверка.Ручной"),
        "target": 360,
        "weight_meaning": "число расхождений (минимум 1)",
    },
    "ers_sverka_rz": {
        "title": "Сверка денежных средств ЕРС / ФоновоеВыполнение",
        "jira": "Сверка денежных средств ЕРС",
        "epf": "внСверкаДенежныхСредствНаЕРС_обр",
        "code": "ДУ.СверкаДенежныхСредствЕРС.Сверка.ФоновоеВыполнение",
        "match": lambda n: n == norm("ДУ.СверкаДенежныхСредствЕРС.Сверка.ФоновоеВыполнение"),
        "target": 360,
        "weight_meaning": "число расхождений (минимум 1)",
    },
    "ers_sverka_manual": {
        "title": "Сверка денежных средств ЕРС / Ручной",
        "jira": "Сверка денежных средств ЕРС",
        "epf": "внСверкаДенежныхСредствНаЕРС_обр",
        "code": "ДУ.СверкаДенежныхСредствЕРС.Сверка.Ручной",
        "match": lambda n: n == norm("ДУ.СверкаДенежныхСредствЕРС.Сверка.Ручной"),
        "target": 360,
        "weight_meaning": "число расхождений (минимум 1)",
    },
    "kotir_get_rz": {
        "title": "Загрузка котировок / ПолучениеTibco / ФоновоеВыполнение",
        "jira": "Загрузка котировок 1й сеанс / финальная",
        "epf": "внЗагрузкаКотировокДУ",
        "code": "ДУ.ЗагрузкаКотировок.ПолучениеTibco.ФоновоеВыполнение",
        "match": lambda n: n == norm("ДУ.ЗагрузкаКотировок.ПолучениеTibco.ФоновоеВыполнение"),
        "target": 1800,
        "weight_meaning": "число полученных котировок (минимум 1)",
    },
    "kotir_create_rz": {
        "title": "Загрузка котировок / СозданиеКотировок / ФоновоеВыполнение",
        "jira": "Загрузка котировок 1й сеанс / финальная",
        "epf": "внЗагрузкаКотировокДУ",
        "code": "ДУ.ЗагрузкаКотировок.СозданиеКотировок.ФоновоеВыполнение",
        "match": lambda n: n == norm("ДУ.ЗагрузкаКотировок.СозданиеКотировок.ФоновоеВыполнение"),
        "target": 1800,
        "weight_meaning": "число созданных документов котировок (минимум 1)",
    },
    "kotir_get_pif": {
        "title": "Загрузка котировок / ПолучениеTibco / VTBAM-ПИФ",
        "jira": "смежный контур (не в списке 9391)",
        "epf": "внЗагрузкаКотировокДУ",
        "code": "ДУ.ЗагрузкаКотировок.ПолучениеTibco.ФоновоеВыполнениеVTBAMПИФ",
        "match": lambda n: n == norm("ДУ.ЗагрузкаКотировок.ПолучениеTibco.ФоновоеВыполнениеVTBAMПИФ"),
        "target": 1800,
        "weight_meaning": "число полученных котировок (минимум 1)",
    },
    "kotir_create_pif": {
        "title": "Загрузка котировок / СозданиеКотировок / VTBAM-ПИФ",
        "jira": "смежный контур (не в списке 9391)",
        "epf": "внЗагрузкаКотировокДУ",
        "code": "ДУ.ЗагрузкаКотировок.СозданиеКотировок.ФоновоеВыполнениеVTBAMПИФ",
        "match": lambda n: n == norm("ДУ.ЗагрузкаКотировок.СозданиеКотировок.ФоновоеВыполнениеVTBAMПИФ"),
        "target": 1800,
        "weight_meaning": "число созданных документов котировок (минимум 1)",
    },
    "kotir_get_manual": {
        "title": "Загрузка котировок / ПолучениеTibco / Ручной",
        "jira": "Загрузка котировок (ручной)",
        "epf": "внЗагрузкаКотировокДУ",
        "code": "ДУ.ЗагрузкаКотировок.ПолучениеTibco.Ручной",
        "match": lambda n: n == norm("ДУ.ЗагрузкаКотировок.ПолучениеTibco.Ручной"),
        "target": 1800,
        "weight_meaning": "число полученных котировок (минимум 1)",
    },
    "kotir_create_manual": {
        "title": "Загрузка котировок / СозданиеКотировок / Ручной",
        "jira": "Загрузка котировок (ручной)",
        "epf": "внЗагрузкаКотировокДУ",
        "code": "ДУ.ЗагрузкаКотировок.СозданиеКотировок.Ручной",
        "match": lambda n: n == norm("ДУ.ЗагрузкаКотировок.СозданиеКотировок.Ручной"),
        "target": 1800,
        "weight_meaning": "число созданных документов котировок (минимум 1)",
    },
    "kotir_cbonds": {
        "title": "Загрузка котировок / ЗагрузкаCbonds",
        "jira": "смежный контур CBONDS",
        "epf": "внЗагрузкаКотировокДУ",
        "code": "ДУ.ЗагрузкаКотировок.ЗагрузкаCbonds.ФоновоеВыполнениеCBONDS",
        "match": lambda n: "загрузкаcbonds" in n or "загрузка cbonds" in n.replace(" ", ""),
        "target": 1800,
        "weight_meaning": "1 (фикс.)",
    },
    "kontrol_rz": {
        "title": "Контроль котировок / Отчет / ФоновоеВыполнение",
        "jira": "Сверка загрузки котировок",
        "epf": "внОтчетКонтрольКотировок",
        "code": "ДУ.КонтрольКотировок.Отчет.ФоновоеВыполнение",
        "match": lambda n: n == norm("ДУ.КонтрольКотировок.Отчет.ФоновоеВыполнение"),
        "target": 1800,
        "weight_meaning": "число строк отчёта (минимум 1)",
    },
    "kontrol_manual": {
        "title": "Контроль котировок / Отчет / Ручной",
        "jira": "Сверка загрузки котировок",
        "epf": "внОтчетКонтрольКотировок",
        "code": "ДУ.КонтрольКотировок.Отчет.Ручной",
        "match": lambda n: n == norm("ДУ.КонтрольКотировок.Отчет.Ручной"),
        "target": 1800,
        "weight_meaning": "число строк отчёта (минимум 1)",
    },
}


def stats_for(rows, target):
    times = [r["time"] for r in rows if r["time"] is not None]
    weights = [r["weight"] for r in rows if r["weight"] is not None]
    unit = []
    thr = []
    tw = []
    extras_num = defaultdict(list)
    extras_cat = defaultdict(Counter)
    hours = Counter()
    sessions = Counter()
    users = Counter()
    errors = Counter()
    samples = []
    dates = []
    for r in rows:
        t, w = r["time"], r["weight"]
        if t is not None and w not in (None, 0):
            unit.append(t / w)
            if t > 0:
                thr.append(w / t)
            tw.append((t, w))
        if r.get("local"):
            hours[r["local"].strftime("%H:%M")[:5] if False else f"{r['local'].hour:02d}"] += 1
            sessions[r.get("session") or "unknown"] += 1
            dates.append(r["local"])
        users[r.get("user") or ""] += 1
        errors[r.get("error") or ""] += 1
        for k, v in (r.get("extra") or {}).items():
            num = to_float(v)
            if num is not None:
                extras_num[k].append(num)
            else:
                extras_cat[k][str(v)] += 1
        if len(samples) < 6 and r.get("dop"):
            samples.append({
                "ko": r["ko"],
                "local": r["local"].strftime("%d.%m.%Y %H:%M:%S") if r.get("local") else "",
                "time": r["time"],
                "weight": r["weight"],
                "error": r["error"],
                "dop": r["dop"],
                "user": r["user"],
            })

    sum_t = sum(t for t, w in tw)
    sum_w = sum(w for t, w in tw)
    return {
        "n": len(rows),
        "time": pack(times),
        "weight": pack(weights),
        "sec_per_unit": pack(unit),
        "units_per_sec": pack(thr),
        "weighted_sec_per_unit": round(sum_t / sum_w, 6) if sum_w else None,
        "weighted_units_per_sec": round(sum_w / sum_t, 6) if sum_t else None,
        "over_target": sum(1 for t in times if t > target),
        "over_360": sum(1 for t in times if t > 360),
        "over_1800": sum(1 for t in times if t > 1800),
        "weight_eq_1": sum(1 for w in weights if w == 1),
        "weight_gt_1": sum(1 for w in weights if w and w > 1),
        "errors": errors.most_common(),
        "users": users.most_common(8),
        "hours": hours.most_common(),
        "sessions": sessions.most_common(),
        "date_min": min(dates).strftime("%d.%m.%Y %H:%M") if dates else None,
        "date_max": max(dates).strftime("%d.%m.%Y %H:%M") if dates else None,
        "extra_num": {k: pack(v) for k, v in extras_num.items()},
        "extra_cat": {k: v.most_common(8) for k, v in extras_cat.items()},
        "samples": samples,
    }


def main():
    print("Loading", XLSX)
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["TDSheet"]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    print("headers", header)

    rows_all = []
    ko_all = Counter()
    for row in it:
        ko = str(row[0] or "").strip()
        if not ko:
            continue
        ko_all[ko] += 1
        extra, flags, dop, meta = parse_comment(row[6] if len(row) > 6 else "")
        local = parse_dt(row[10] if len(row) > 10 else None)
        rec = {
            "ko": ko,
            "ko_norm": norm(ko),
            "time": to_float(row[4] if len(row) > 4 else None),
            "weight": to_float(row[5] if len(row) > 5 else None),
            "comment": str(row[6] or "") if len(row) > 6 else "",
            "extra": extra,
            "flags": flags,
            "dop": dop,
            "meta": {k: v for k, v in meta.items() if k in ("Платф", "Конф", "КонфВер")},
            "local": local,
            "session": classify_quotes_session(local),
            "user": str(row[9] or "") if len(row) > 9 else "",
            "error": str(row[11] or "") if len(row) > 11 else "",
            "write_utc": str(row[7] or "") if len(row) > 7 else "",
        }
        rows_all.append(rec)
    wb.close()

    print("rows", len(rows_all), "unique KO", len(ko_all))

    grouped = {}
    assigned = set()
    for gid, spec in KO_GROUPS.items():
        matched = [r for r in rows_all if spec["match"](r["ko_norm"])]
        for r in matched:
            assigned.add(id(r))
        grouped[gid] = {
            "meta": {k: v for k, v in spec.items() if k != "match"},
            "excel_names": Counter(r["ko"] for r in matched).most_common(),
            "stats": stats_for(matched, spec["target"]),
        }
        # session split for quotes RZ
        if gid in ("kotir_get_rz", "kotir_create_rz"):
            for sess in ("session1", "session2", "other"):
                sub = [r for r in matched if r["session"] == sess]
                grouped[gid][f"stats_{sess}"] = stats_for(sub, spec["target"])

        print(f"{gid}: n={len(matched)} names={grouped[gid]['excel_names']}")

    unmatched_related = []
    for r in rows_all:
        if id(r) in assigned:
            continue
        n = r["ko_norm"]
        if any(x in n for x in ("котиров", "денежныхсредств", "контролькотир", "cbonds")):
            unmatched_related.append(r["ko"])
    unmatched_c = Counter(unmatched_related)

    payload = {
        "source": str(XLSX),
        "total_rows": len(rows_all),
        "unique_ko": len(ko_all),
        "all_ko": ko_all.most_common(),
        "groups": grouped,
        "unmatched_related": unmatched_c.most_common(),
        "period_hint": {
            "min": min((r["local"] for r in rows_all if r["local"]), default=None),
            "max": max((r["local"] for r in rows_all if r["local"]), default=None),
        },
    }
    # datetime in period_hint
    if payload["period_hint"]["min"]:
        payload["period_hint"]["min"] = payload["period_hint"]["min"].strftime("%d.%m.%Y %H:%M")
        payload["period_hint"]["max"] = payload["period_hint"]["max"].strftime("%d.%m.%Y %H:%M")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)

    # txt
    lines = [f"total={len(rows_all)} unique_ko={len(ko_all)} period={payload['period_hint']}"]
    for gid, block in grouped.items():
        st = block["stats"]
        t = st["time"] or {}
        w = st["weight"] or {}
        lines.append("")
        lines.append(f"=== {gid}  {block['meta']['code']}  n={st['n']} ===")
        lines.append(f"  excel: {block['excel_names']}")
        if t:
            lines.append(
                f"  time s: min={t['min']} med={t['median']} mean={t['mean']} "
                f"p95={t['p95']} max={t['max']} over_target={st['over_target']}/{st['n']}"
            )
        if w:
            lines.append(
                f"  weight: min={w['min']} med={w['median']} mean={w['mean']} max={w['max']} "
                f"eq1={st['weight_eq_1']} gt1={st['weight_gt_1']}"
            )
        u, thr = st["sec_per_unit"] or {}, st["units_per_sec"] or {}
        if u:
            lines.append(
                f"  sec/unit med={u.get('median')} mean={u.get('mean')}  "
                f"u/s med={thr.get('median')} wmean u/s={st['weighted_units_per_sec']}"
            )
        lines.append(f"  period {st['date_min']} .. {st['date_max']}")
        lines.append(f"  hours {st['hours']}")
        lines.append(f"  sessions {st['sessions']}")
        lines.append(f"  errors {st['errors']} users {st['users']}")
        if st["extra_num"]:
            for k, ev in st["extra_num"].items():
                lines.append(
                    f"  extra {k}: n={ev['n']} med={ev['median']} mean={ev['mean']} "
                    f"min={ev['min']} max={ev['max']}"
                )
        if st["extra_cat"]:
            for k, ev in st["extra_cat"].items():
                lines.append(f"  cat {k}: {ev}")
        for sess_key in ("stats_session1", "stats_session2", "stats_other"):
            if sess_key in block:
                sst = block[sess_key]
                tt = sst["time"] or {}
                ww = sst["weight"] or {}
                lines.append(
                    f"  [{sess_key}] n={sst['n']} time_med={tt.get('median')} "
                    f"time_p95={tt.get('p95')} time_max={tt.get('max')} "
                    f"w_med={ww.get('median')} w_max={ww.get('max')} "
                    f"over={sst['over_target']} hours={sst['hours']}"
                )
        if st["samples"]:
            lines.append("  sample: " + st["samples"][0]["dop"][:220])

    if unmatched_c:
        lines.append("\nUNMATCHED RELATED:")
        for k, c in unmatched_c.most_common():
            lines.append(f"  {c}  {k}")

    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")
    print("Wrote", OUT_JSON)
    print("Wrote", OUT_TXT)
    print("\n".join(lines))


if __name__ == "__main__":
    main()
