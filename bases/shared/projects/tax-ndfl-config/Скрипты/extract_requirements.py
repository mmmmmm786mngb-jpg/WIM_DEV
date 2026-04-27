#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Извлекает структурированные данные из всех листов Excel-файла требований
и сохраняет в текстовый отчёт для последующего анализа.
"""

import os
import sys
import json
from openpyxl import load_workbook

INPUT_XLSX = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "Требования от налогов.xlsx",
)
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Документация", "_extracted")
os.makedirs(OUTPUT_DIR, exist_ok=True)


def truncate(value, limit=300):
    if value is None:
        return ""
    s = str(value)
    if len(s) > limit:
        return s[:limit] + "..."
    return s


def is_row_empty(row):
    return all((c is None or str(c).strip() == "") for c in row)


def export_sheet(ws, out_dir):
    safe_name = ws.title.replace("/", "_").replace("\\", "_")
    out_md = os.path.join(out_dir, safe_name + ".md")
    out_json = os.path.join(out_dir, safe_name + ".json")

    rows = []
    max_col = ws.max_column
    max_row = ws.max_row

    last_col_used = 0
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row):
            if val is not None and str(val).strip() != "":
                if idx + 1 > last_col_used:
                    last_col_used = idx + 1

    if last_col_used == 0:
        last_col_used = max_col

    for row in ws.iter_rows(values_only=True, max_col=last_col_used):
        rows.append(list(row))

    while rows and is_row_empty(rows[-1]):
        rows.pop()

    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=1, default=str)

    with open(out_md, "w", encoding="utf-8") as f:
        f.write(f"# Лист: {ws.title}\n\n")
        f.write(f"Размер: {len(rows)} строк x {last_col_used} колонок\n\n")
        if not rows:
            f.write("(пусто)\n")
            return out_md
        header = rows[0]
        f.write("## Заголовки\n\n")
        for idx, h in enumerate(header, start=1):
            f.write(f"- **Кол {idx}**: {truncate(h, 500)}\n")
        f.write("\n## Данные\n\n")
        for r_idx, row in enumerate(rows[1:], start=2):
            if is_row_empty(row):
                f.write("\n")
                continue
            f.write(f"### Строка {r_idx}\n\n")
            for c_idx, cell in enumerate(row, start=1):
                if cell is None or str(cell).strip() == "":
                    continue
                f.write(f"- **К{c_idx}**: {truncate(cell, 800)}\n")
            f.write("\n")

    return out_md


def main():
    if not os.path.exists(INPUT_XLSX):
        print(f"ERROR: file not found: {INPUT_XLSX}")
        sys.exit(1)

    wb = load_workbook(INPUT_XLSX, data_only=True, read_only=True)
    print(f"Sheets: {wb.sheetnames}")
    summary = []
    for name in wb.sheetnames:
        ws = wb[name]
        out = export_sheet(ws, OUTPUT_DIR)
        summary.append((name, out))
        print(f"OK: {name} -> {os.path.basename(out)}")

    with open(os.path.join(OUTPUT_DIR, "_index.md"), "w", encoding="utf-8") as f:
        f.write("# Извлечённые данные требований\n\n")
        for name, out in summary:
            f.write(f"- [{name}]({os.path.basename(out)})\n")

    print("Done.")


if __name__ == "__main__":
    main()
