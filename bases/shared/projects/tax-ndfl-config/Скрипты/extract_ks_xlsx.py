#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Извлекает содержимое всех листов файла КС.xlsx в markdown
и сохраняет в _extracted/ks/<имя_листа>.md и ks_summary.md.
"""

import os
import sys
import json
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_XLSX = os.path.join(
    BASE_DIR, "КорректировкаОтПользователя_1", "КС.xlsx"
)
OUT_DIR = os.path.join(
    BASE_DIR, "КорректировкаОтПользователя_1", "_extracted", "ks"
)
SUMMARY = os.path.join(
    BASE_DIR, "КорректировкаОтПользователя_1", "_extracted", "ks_summary.md"
)
os.makedirs(OUT_DIR, exist_ok=True)


def safe(v):
    if v is None:
        return ""
    s = str(v).strip()
    s = s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    return s


def export_sheet(ws, out_dir):
    safe_name = ws.title.replace("/", "_").replace("\\", "_").replace(":", "_")
    out_md = os.path.join(out_dir, safe_name + ".md")
    out_json = os.path.join(out_dir, safe_name + ".json")

    rows = []
    last_col_used = 0
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row):
            if val is not None and str(val).strip() != "":
                if idx + 1 > last_col_used:
                    last_col_used = idx + 1
    if last_col_used == 0:
        last_col_used = ws.max_column or 1

    for row in ws.iter_rows(values_only=True, max_col=last_col_used):
        rows.append(list(row))

    while rows and all((c is None or str(c).strip() == "") for c in rows[-1]):
        rows.pop()

    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=1, default=str)

    with open(out_md, "w", encoding="utf-8") as f:
        f.write(f"# Лист: {ws.title}\n\n")
        f.write(f"Размер: {len(rows)} строк x {last_col_used} колонок\n\n")
        if not rows:
            f.write("(пусто)\n")
            return out_md, len(rows), last_col_used

        f.write("## Полный дамп (markdown-таблица)\n\n")
        header = [safe(c) if safe(c) else f"К{i+1}" for i, c in enumerate(rows[0])]
        f.write("| " + " | ".join(header) + " |\n")
        f.write("|" + "|".join(["---"] * last_col_used) + "|\n")
        for r in rows[1:]:
            cells = [safe(c) for c in r]
            while len(cells) < last_col_used:
                cells.append("")
            cells = [c.replace("|", "\\|") for c in cells]
            f.write("| " + " | ".join(cells) + " |\n")

        f.write("\n\n## Построчная развёртка\n\n")
        for r_idx, row in enumerate(rows[1:], start=2):
            if all((c is None or str(c).strip() == "") for c in row):
                continue
            f.write(f"### Строка {r_idx}\n\n")
            for c_idx, cell in enumerate(row, start=1):
                if cell is None or str(cell).strip() == "":
                    continue
                label = safe(rows[0][c_idx - 1]) if c_idx - 1 < len(rows[0]) else ""
                if not label:
                    label = f"К{c_idx}"
                val = safe(cell)
                if len(val) > 800:
                    val = val[:800] + "..."
                f.write(f"- **{label}**: {val}\n")
            f.write("\n")

    return out_md, len(rows), last_col_used


def main():
    if not os.path.exists(INPUT_XLSX):
        print(f"NOT FOUND: {INPUT_XLSX}")
        sys.exit(1)

    wb = load_workbook(INPUT_XLSX, data_only=True, read_only=True)
    print(f"Sheets: {wb.sheetnames}")

    summary_lines = ["# Сводка по КС.xlsx\n", ""]
    summary_lines.append(f"Исходник: `КС.xlsx`")
    summary_lines.append(f"Листов: {len(wb.sheetnames)}\n")
    summary_lines.append("| Лист | Строк | Колонок | Файл |")
    summary_lines.append("|---|---|---|---|")

    for name in wb.sheetnames:
        ws = wb[name]
        out, nrows, ncols = export_sheet(ws, OUT_DIR)
        rel = os.path.join("ks", os.path.basename(out))
        summary_lines.append(f"| {name} | {nrows} | {ncols} | `{rel}` |")
        print(f"OK: {name} -> {os.path.basename(out)}  ({nrows} rows, {ncols} cols)")

    with open(SUMMARY, "w", encoding="utf-8") as f:
        f.write("\n".join(summary_lines) + "\n")

    print(f"Summary: {SUMMARY}")


if __name__ == "__main__":
    main()
